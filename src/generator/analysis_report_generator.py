"""AS-IS 분석서 생성기

이 모듈은 대상 리포지터리의 `.applycrypto` 결과물을 이용해
`대상목록` 시트를 포함하는 AS-IS 분석 엑셀(`artifacts/asis_analysis_report_YYYYMMDD.xlsx`)을 생성합니다.

지원하는 분석 타입(중요):
- "ThreeStep" Type: step1_query_analysis.json, step2_planning.json, table_access_info.json을 사용하여 쿼리, 흐름, 수정지시를 결합.
- "TypeHandler" Type: step2_planning.json을 사용하지 않고 step1_query_analysis.json과 table_access_info.json을 사용해서 분석서 생성.

출력:
- 워크북 내부의 `대상목록` 시트에 레코드를 작성합니다.
"""
import json
import os
import csv
import hashlib
import re
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, TimeoutError
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from config.config_manager import Configuration

# =====================================
# LLM 번역 관련 상수 (artifact_generator 패턴 참고)
# =====================================

# LLM 관련 상수
MAX_TRANSLATION_ITEMS_PER_BATCH = {
    'atype': 10,      # ThreeStep: 10개 항목씩 번역 (3개 필드: Reason, Insertion Point, Sql Summary)
    'btype': 20      # TypeHandler: 20개 항목씩 번역 (1개 필드: Sql Summary)
}
MAX_LLM_RETRIES = 3
LLM_MAX_TOKENS = 10000
LLM_TEMPERATURE = 0.0
LLM_BATCH_TIMEOUT = 30  # 배치당 30초 타임아웃
LLM_BATCH_START_INTERVAL = 1.0 # 배치 sleep : 0.0 지연 없음, 0.2 매우 빠름, 0.5 균형, 1.0 안정적(기본값), 2.0 느림
TRANSLATION_CACHE_TTL = 3600  # 번역 캐시 설정 1시간(=3600초)

# 번역할 필드명 (modification_type별)
TRANSLATION_FIELDS = {
    'atype': ['Reason', 'Insertion Point', 'Sql Summary'],
    'btype': ['Sql Summary']
}

# 배치 길이 경고 임계값 (토큰 기준)
BATCH_LENGTH_WARNING_THRESHOLD = int(LLM_MAX_TOKENS * 0.7)  # 70% 이상 사용 시 경고

# 개요 시트의 항목 설명 - 공통 사용 (atype, btype 모두)
OVERVIEW_ITEMS = [
    ("패키지명", "대상 애플리케이션의 레포지토리 이름"),
    ("SQL ID", "Mapper 인터페이스의 메소드 아이디(예. Mapper XML 내의 SQL ID)"),
    ("Mapper Path", "SQL을 포함하고 있는 소스파일의 Path(XML 또는 java)"),
    ("대상테이블", "암복호화 적용 대상 테이블 이름"),
    ("대상컬럼", "암복호화 적용 대상 테이블의 컬럼"),
    ("암복호화 필요 컬럼", "실제 암복호화 처리가 필요한 컬럼"),
    ("암복호화 필요 Java Field", "실제 암복호화 처리가 필요한 Java 객체의 필드명"),
    ("End Point", "해당 쿼리를 호출하는 최상위의 메서드 시그너처(class_name.method_name)"),
    ("클래스명", "코드 수정이 필요한 클래스"),
    ("메서드명", "코드 수정이 필요한 클래스의 메서드"),
    ("Model", "데이터를 매핑받는 Java 객체 (VO, DTO, Map)"),
    ("ResultMap", "ResultMap 설정 상태\n- ○(정상) : ResultMap 정의 및 컬럼 매핑 존재\n- △(없음) : ResultMap 정의되어 있으나, 컬럼 매핑 정보는 없음\n- X(N/A) : ResultMap이 정의되어 있지 않음"),
    ("Action", "암호화 코드 적용 여부\n- ENCRYPT : 암호화 코드 적용 필요\n- DECRYPT : 복호화 코드 적용 필요\n- ENCRYPT_THEN_DECRYPT : 암호화 후 복호화 코드 적용 필요\n- SKIP : 암호화 코드 적용 필요 없음"),
    ("Reason", "암복호화 적용 혹은 미적용 이유"),
    ("Insertion Point", "암복호화 코드 적용 시 적용해야할 지점"),
    ("Code Pattern Hint", "적용해야할 암복호화 코드"),
    ("Sql Summary", "Sql 쿼리 내용 요약"),
]

def generate_analysis_report(config: Configuration, enable_translation: bool = False, enable_verification: bool = False):
    """
    AS-IS 분석서 생성기 래퍼 함수

    `config`에서 지정한 `target_project`의 `.applycrypto` 디렉터리를 찾아
    `three_step_results`를 이용해 `대상목록` 시트를 채웁니다.

    Args:
        config (Configuration): 설정 객체
        enable_translation (bool): LLM을 사용하여 Reason, Insertion Point, Code Pattern Hint, Sql Summary 항목을 번역할지 여부. 기본값은 False.
        enable_verification (bool): SQL ID 검증 결과를 엑셀 시트로 추가할지 여부. 기본값은 False.

    Returns:
        str: 생성된 엑셀 파일의 경로
    """
    # 설정 파일에서 target_project 경로 및 분석 작업 유형 읽기
    target_project = config.target_project
    modification_type = config.modification_type

    if not target_project:
        raise FileNotFoundError(f"Configuration missing target_project")

    # 레포지터리 이름(마지막 경로)을 패키지명 기본값으로 사용
    try:
        repo_name = os.path.basename(os.path.normpath(target_project))
    except Exception:
        repo_name = ''

    # 대상 프로젝트의 .applycrypto 루트 경로 결정
    applycrypto_root = os.path.join(target_project, '.applycrypto')
    if not os.path.isdir(applycrypto_root):
        raise FileNotFoundError(f".applycrypto가 target_project 아래에 없습니다: {applycrypto_root}")

    # 출력 디렉터리 설정
    out_dir = os.path.join(applycrypto_root, 'artifacts')
    os.makedirs(out_dir, exist_ok=True)

    font_default = Font(name='맑은 고딕', size=10)
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    border = make_border()
    results_paths = []
    wb = Workbook()

    print(f'- AS-IS 분석서 타입 : {modification_type}', flush=True)

    # 개요 시트 생성 (수정 타입에 따라 다른 OVERVIEW_ITEMS 사용)
    # 이는 generate_analysis_report() 처리 후반부에서 실제 선택 후 생성됨
    sheet_title = '개요'
    # 나중에 덮어씀

    # 대상목록 시트 생성 및 채우기
    sheet_title = '대상목록'
    
    # modification_type 리스트 정의 (각 리스트에 속한 타입들은 해당 핸들러를 사용)
    # modification_type 그룹 매핑: 그룹 키 -> 타입 리스트 및 관련 핸들러/컬럼 설정
    MODIFICATION_TYPE_GROUPS = {
        'modification_atype': {
            'types': ['ThreeStep'],
            'widths': [13, 25, 25, 25, 18, 25, 25, 25, 10, 25, 25, 25, 16, 40, 40, 40, 40, 40],
            'headers': [
                "패키지명", "SQL ID", "Mapper Path", "대상테이블", "대상컬럼",
                "암복호화 필요 컬럼", "암복호화 필요 Java Field", "End Point", "ResultMap", "클래스명", "메서드명", "Model", "Action",
                "Reason", "Insertion Point", "Code Pattern Hint", "Sql Summary"
            ],
            'handler': fill_target_list_modification_atype,
        },
        'modification_btype': {
            'types': ['TypeHandler', 'TypeHandler'],
            'widths': [13, 25, 25, 25, 18, 25, 25, 25, 10, 25, 25, 40],
            'headers': [
                "패키지명", "SQL ID", "Mapper Path", "대상테이블", "대상컬럼",
                "암복호화 필요 컬럼", "암복호화 필요 Java Field", "End Point", "ResultMap", "클래스명", "Model", "Sql Summary"
            ],
            'handler': fill_target_list_modification_btype,
        }
    }

    # Allow tolerant matching: accept values like 'Test_ThreeStep_20260130'
    # by matching known type keywords as substrings (case-insensitive).
    canonical_type = None
    for info in MODIFICATION_TYPE_GROUPS.values():
        for t in info.get('types') or []:
            if t and t.lower() in (modification_type or '').lower():
                canonical_type = t
                break
        if canonical_type:
            break

    # If we found a canonical match, use it; otherwise keep original for error message.
    if canonical_type:
        modification_type = canonical_type

    # modification_type에 해당하는 그룹을 찾고 설정을 적용
    group_key = None
    for gk, info in MODIFICATION_TYPE_GROUPS.items():
        if modification_type in (info.get('types') or []):
            group_key = gk
            break
    if group_key is None:
        supported = []
        for info in MODIFICATION_TYPE_GROUPS.values():
            supported.extend(info.get('types') or [])
        raise ValueError(f"지원되지 않는 modification_type: {modification_type}. 지원되는 타입: {supported}")

    cfg = MODIFICATION_TYPE_GROUPS[group_key]
    widths = cfg['widths']
    headers = cfg['headers']
    handler = cfg['handler']
    
    # 개요 시트 생성 (공통 OVERVIEW_ITEMS 사용)
    sh0 = create_overview_sheet(wb, font_default, title='개요', overview_items=OVERVIEW_ITEMS)
    set_row_height(sh0, 16)
    sh0.row_dimensions[13].height = 70
    sh0.row_dimensions[14].height = 80

    # 대상목록 시트 생성 및 채우기
    sheet_title = '대상목록'
    sh2 = create_target_list(wb, font_default, bold_font, border, title=sheet_title, widths=widths, headers=headers)
    sh2, records = handler(
        sh2, applycrypto_root=applycrypto_root, repo_name=repo_name, font_default=font_default, border=border, headers=headers
    )
    set_row_height(sh2, 16)

    # SQL ID 데이터 검증 (--verify 옵션이 있을 때만 실행 및 로그 출력)
    validation_data = None  # 검증 데이터 저장용
    if enable_verification:
        print(f'- SQL ID 데이터 검증 시작', flush=True)
        try:
            out_dir = os.path.join(applycrypto_root, 'artifacts')
            os.makedirs(out_dir, exist_ok=True)
            validation_result, validation_stats, validation_details = _validate_sql_id_existence(applycrypto_root, modification_type, out_dir)
            validation_data = {
                'result': validation_result,
                'stats': validation_stats,
                'details': validation_details
            }
            # 상세 검증 결과 출력
            if validation_stats:
                total_table_ids = validation_stats.get('total_table_ids', 0)
                step1_matched = validation_stats.get('step1_matched', 0)
                step2_matched = validation_stats.get('step2_matched', 0)
                missing_count = validation_stats.get('missing_count', 0)
                
                if modification_type == 'ThreeStep':
                    print(f'- SQL ID 검증 결과: table_access({total_table_ids}개) | step1({step1_matched}개) | step2({step2_matched}개) | 누락({missing_count}개)', flush=True)
                else:
                    print(f'- SQL ID 검증 결과: table_access({total_table_ids}개) | step1({step1_matched}개) | 누락({missing_count}개)', flush=True)
            
            if validation_result:
                print(f'- SQL ID 데이터 검증 완료 (이상 없음)', flush=True)
            else:
                print(f'- SQL ID 데이터 검증 완료 (누락된 항목 있음 - 로그 확인)', flush=True)
        except Exception as e:
            print(f'- SQL ID 데이터 검증 중 오류 (계속 진행): {e}', flush=True)

    # LLM을 사용한 번역 처리 (enable_translation이 True인 경우)
    if enable_translation and records:
        print(f'- LLM 기반 번역 시작', flush=True)
        try:
            from src.modifier.llm.llm_factory import create_llm_provider
            llm_provider = create_llm_provider(config.llm_provider)
            cache = TranslationCache(ttl_seconds=TRANSLATION_CACHE_TTL)
            
            # 번역 처리 (병렬 배치 처리)
            records = translate_records_batch(records, modification_type, llm_provider, cache)
            
            # 엑셀에 번역된 결과 반영 (sh2 시트 업데이트)
            _update_sheet_with_translated_records(sh2, records, headers)
            
            print(f'- LLM 기반 번역 완료', flush=True)
        except Exception as e:
            print(f'- 번역 처리 중 오류 (계속 진행): {e}', flush=True)

    # SQL ID 검증 시트 추가 (enable_verification이 True인 경우)
    if enable_verification and validation_data:
        try:
            add_validation_sheet(wb, validation_data, font_default, border)
        except Exception as e:
            print(f'- SQL ID 검증 시트 추가 오류 (계속 진행): {e}', flush=True)

    # 출력 파일명에 타입 식별자 추가
    date_fragment = datetime.now().strftime('%Y%m%d')
    output_filename_t = f'AsIs_Analysis_Report_{modification_type}_{date_fragment}.xlsx'
    output_path = os.path.join(out_dir, output_filename_t)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    artifact_file = Path(output_path)
    orig_stem = artifact_file.stem

    # 저장 시도: 파일이 잠겨 있으면 타임스탬프를 붙인 임시 파일명으로 폴백
    if artifact_file.exists():
        try:
            artifact_file.unlink()
        except Exception:
            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            artifact_file = artifact_file.with_name(f"{artifact_file.stem}_tmp{ts}{artifact_file.suffix}")

    final_path = artifact_file
    try:
        packages = set(r.get('패키지명', '') for r in records if r.get('패키지명'))
        files = set(r.get('파일명', '') for r in records if r.get('파일명'))
        tables = set(r.get('대상테이블', '') for r in records if r.get('대상테이블'))
        wb.active = sh2  # 두 번째 시트(대상목록)를 활성 시트로 설정
        wb.save(artifact_file)
        print(f'- 분석서 출력 파일 : {artifact_file}', flush=True)
    except PermissionError:
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        fallback = artifact_file.with_name(f"{orig_stem}_tmp{ts}{artifact_file.suffix}")
        wb.save(fallback)
        print(f'- 분석서 출력 파일(임시): {fallback}', flush=True)
        final_path = fallback

    results_paths.append(str(final_path))

    return results_paths[0] if results_paths else ''


def create_overview_sheet(wb, font_default, title='개요', overview_items=None):
    """개요 시트를 생성합니다.
    
    분석 타입(atype/btype)에 따라 다른 항목 설명을 표시합니다.
    - 1행: 헤더 (대상, 설명)
    - 2행부터: 항목별 설명 입력
    - 열 구성: 2열(항목명), 3열(설명)
    """
    if overview_items is None:
        overview_items = OVERVIEW_ITEMS
    
    sh = wb.active
    sh.title = title
    
    # 열 너비 설정
    sh.column_dimensions['A'].width = 3   # 1열 (공백)
    sh.column_dimensions['B'].width = 17  # 2열 (항목명)
    sh.column_dimensions['C'].width = 70  # 3열 (설명)
    
    # 스타일 설정
    header_fill = PatternFill(start_color='DFDFDF', end_color='DFDFDF', fill_type='solid')
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    border = make_border()
    
    # 1행: 헤더 설정
    # B1: 대상
    cell_b1 = sh.cell(row=2, column=2, value='대상')
    cell_b1.font = bold_font
    cell_b1.fill = header_fill
    cell_b1.border = border
    cell_b1.alignment = Alignment(horizontal='center', vertical='center')
    
    # C1: 설명
    cell_c1 = sh.cell(row=2, column=3, value='설명')
    cell_c1.font = bold_font
    cell_c1.fill = header_fill
    cell_c1.border = border
    cell_c1.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2행부터 항목 설명 입력
    for row_idx, (item_name, description) in enumerate(overview_items, start=3):
        # 2열: 항목명
        cell_b = sh.cell(row=row_idx, column=2, value=item_name)
        cell_b.font = font_default
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        # 3열: 설명
        cell_c = sh.cell(row=row_idx, column=3, value=description)
        cell_c.font = font_default
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    return sh
    """개요 시트를 생성합니다. 분석 보고서의 주요 정보를 표시합니다."""
    # 이 함수는 이제 사용되지 않음


def create_target_list(wb, font_default, bold_font, border, title, widths=None, headers=None):
    """`대상목록` 시트를 생성하고 공통 헤더를 설정합니다."""
    sh = wb.create_sheet(title=title)
    # 컬럼 너비: 기본값 설정
    for idx, w in enumerate(widths, start=1):
        sh.column_dimensions[get_column_letter(idx)].width = w

    # 헤더 설정
    if headers is None:
        headers = [
            "패키지명", "파일명", "SQL ID", "대상테이블", "대상컬럼",
            "ResultMap", "클래스명", "메서드명", "Model", "Action",
            "Reason", "Insertion Point", "Code Pattern Hint", "Sql Summary"
        ]
    header_fill = PatternFill(start_color='DFDFDF', end_color='DFDFDF', fill_type='solid')
    for c, h in enumerate(headers, start=1):
        cell = sh.cell(row=2, column=c, value=h)
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill

    # 준비된 데이터 행의 기본 형식을 위해 3행을 비워둡니다.
    for c in range(1, len(headers) + 1):
        cell = sh.cell(row=3, column=c)
        if cell.value is None:
            cell.value = ''
        cell.font = font_default
        cell.border = border

    return sh


def fill_target_list_modification_atype(sh, applycrypto_root=None, repo_name=None, font_default=None, border=None, headers=None):
    """ThreeStep 분석 - table_access 기준으로 대상목록 시트를 채웁니다.

    처리 절차:
    1. table_access_info.json의 모든 SQL ID를 기준 데이터로 추출
    2. step1_query_analysis.json과 step2_planning.json의 데이터를 모두 캐시
    3. 각 SQL ID에 대해:
       - step1에서 query_id 매칭
       - step2에서 sql_query_id 및 modification_instructions 매칭
       - 필터링 없음: table_access의 모든 SQL ID는 행으로 추가됨
       - step1/step2 정보는 보충용 (있으면 추가, 없으면 공란)
    4. 추출된 컬럼값으로 행 추가
    5. 정렬 및 셀 병합 적용

    필터링 효과: 없음 (table_access의 모든 SQL ID를 처리)
    완전성: table_access의 모든 SQL ID가 Excel에 포함됨

    Args:
        sh: openpyxl Worksheet 객체
        applycrypto_root: '.applycrypto' 디렉터리 경로
        repo_name: 패키지명
        font_default: 기본 폰트
        border: 테두리 스타일
        headers: 컬럼 헤더 목록

    Returns:
        tuple: (수정된 시트 객체, 레코드 리스트)
    """

    missing_files = []
    invalid_files = []

    if not applycrypto_root or not os.path.isdir(applycrypto_root):
        print(f"- 지정된 경로에 '.applycrypto'가 없습니다: {applycrypto_root}", flush=True)
        return sh, []

    three_root = os.path.join(applycrypto_root, 'three_step_results')
    if not os.path.isdir(three_root):
        return sh, []

    ts_dirs = [d for d in os.listdir(three_root) if os.path.isdir(os.path.join(three_root, d))]
    if not ts_dirs:
        return sh, []
    latest = sorted(ts_dirs)[-1]
    latest_root = os.path.join(three_root, latest)

    table_access = safe_load_json(
        os.path.join(applycrypto_root, 'results', 'table_access_info.json'),
        missing_files=missing_files,
        invalid_files=invalid_files,
        applycrypto_root=applycrypto_root
    ) or {}

    # table_access의 모든 SQL ID를 기준 데이터로 추출 (동일 SQL ID가 여러 테이블에 존재할 수 있으므로 리스트 사용)
    table_access_sql_ids = {}  # sql_id -> [table_info, table_info, ...]
    if isinstance(table_access, list):
        for tobj in table_access:
            sql_queries = tobj.get('sql_queries', [])
            for sq in sql_queries:
                sql_id = sq.get('id')
                if sql_id:
                    if sql_id not in table_access_sql_ids:
                        table_access_sql_ids[sql_id] = []
                    table_access_sql_ids[sql_id].append(tobj)
    
    # step1/step2 데이터를 사전 캐시 (반복 검색 최소화)
    step_data_cache = {}
    for table in sorted(os.listdir(latest_root)):
        table_dir = os.path.join(latest_root, table)
        if not os.path.isdir(table_dir):
            continue
        for ctrl in sorted(os.listdir(table_dir)):
            ctrl_dir = os.path.join(table_dir, ctrl)
            if not os.path.isdir(ctrl_dir):
                continue
            s1 = safe_load_json(
                os.path.join(ctrl_dir, 'step1_query_analysis.json'),
                missing_files=missing_files,
                invalid_files=invalid_files,
                applycrypto_root=applycrypto_root
            ) or {}
            s2 = safe_load_json(
                os.path.join(ctrl_dir, 'step2_planning.json'),
                missing_files=missing_files,
                invalid_files=invalid_files,
                applycrypto_root=applycrypto_root
            ) or {}
            
            step_data_cache.setdefault(table, {})[ctrl] = {
                's1': s1,
                's2': s2
            }
    
    def qid_match_for_flow(a, b):
        """SQL ID 비교: 전체 일치 또는 마지막 토큰만 비교
        
        예: 'com.example.mapper.UserMapper.findById' vs 'findById' → 일치
        """
        try:
            sa = '' if a is None else str(a)
            sb = '' if b is None else str(b)
            # 둘 다 단순명(점 없음): 전체 비교
            if '.' not in sa and '.' not in sb:
                return sa == sb
            # 한쪽이라도 정규화명(점 있음): 마지막 토큰으로 비교
            return sa.rsplit('.', 1)[-1] == sb.rsplit('.', 1)[-1]
        except Exception:
            return False
    
    def find_step_info_for_qid(qid, table_lookup, step_data_cache):
        """주어진 SQL ID와 테이블명을 step1/step2에서 검색하여 관련 정보를 반환
        
        Args:
            qid: 검색할 SQL ID
            table_lookup: 테이블명 (step1/step2 캐시의 테이블 디렉토리명과 매칭)
            step_data_cache: 캐시된 step1/step2 데이터

        Returns:
            dict: step1/step2 발견 여부 및 추출 정보
                - found_in_step1: step1에서 발견 여부
                - found_in_step2: step2에서 발견 여부
                - s1/s2: step1/step2 전체 데이터
                - meta: step1의 메타정보
                - chosen_flow: step2의 매칭된 flow
                - instr: step2의 매칭된 modification_instructions
        """
        result = {
            'found_in_step1': False,
            'found_in_step2': False,
            'table': None,
            'ctrl': None,
            's1': None,
            's2': None,
            'meta': None,
            'chosen_flow': None,
            'instr': None,
            'q': None  # step1의 매칭된 query 객체 저장
        }
        
        # 테이블명이 일치하는 경우만 검색 (★ 중요: 테이블명 필수 일치)
        if table_lookup not in step_data_cache:
            return result
            
        for ctrl in step_data_cache[table_lookup]:
            cache = step_data_cache[table_lookup][ctrl]
            s1 = cache['s1']
            s2 = cache['s2']
            
            # step1에서 query_id 매칭
            qs = (s1.get('result') or {}).get('queries') or []
            meta = s1.get('metadata') or {}
            
            for q in qs:
                q_qid = q.get('query_id') or ''
                if qid_match_for_flow(q_qid, qid):
                    result['found_in_step1'] = True
                    result['table'] = table_lookup
                    result['ctrl'] = ctrl
                    result['s1'] = s1
                    result['meta'] = meta
                    result['q'] = q  # query 객체 저장
                    break
            
            # step2에서 sql_query_id 매칭
            flows = (s2.get('result') or {}).get('data_flow_analysis', {}).get('flows') or []
            mods = (s2.get('result') or {}).get('modification_instructions') or []
            
            for f in flows:
                f_qid = f.get('sql_query_id') or ''
                if qid_match_for_flow(f_qid, qid):
                    result['found_in_step2'] = True
                    result['s2'] = s2
                    result['chosen_flow'] = f
                    
                    # flow_id로 modification_instructions 매칭
                    flow_id = f.get('flow_id')
                    for mi in mods:
                        if flow_id is not None and str(mi.get('flow_id')) == str(flow_id):
                            result['instr'] = mi
                            break
                    if not result['instr']:
                        for mi in mods:
                            if mi.get('query_id') == qid or mi.get('query') == qid:
                                result['instr'] = mi
                                break
                    break
        
        return result
    
    records = []
    
    # table_access의 모든 SQL ID를 기준으로 처리 (동일 SQL ID가 여러 테이블에 있을 수 있음)
    for sql_id in sorted(table_access_sql_ids.keys()):
        table_info_list = table_access_sql_ids[sql_id]  # 리스트 (같은 SQL ID의 모든 테이블)
        
        # 각 테이블별로 개별 레코드 생성
        for table_info in table_info_list:
            table_lookup = table_info.get('table_name', '')
            
            # step1/step2에서 관련 정보 검색 (테이블명과 SQL ID 모두 일치하는 경우만)
            step_info = find_step_info_for_qid(sql_id, table_lookup, step_data_cache)
            
            # ★ 중요: table_access의 모든 SQL ID는 행으로 추가됨
            # step1/step2 정보는 보충용 (있으면 추가, 없으면 공란)
            
            qid = sql_id
            instr = step_info['instr']
            chosen = step_info['chosen_flow']
            meta = step_info['meta'] or {}
            q = step_info['q']  # step1의 query 객체

            # 1. SQL ID 표시 (마지막 토큰만 사용)
            display_qid = ''
            try:
                display_qid = qid.rsplit('.', 1)[-1] if qid else ''
            except Exception:
                display_qid = qid

            # 2. Mapper Path 추출
            mapper_path = ''
            try:
                _mp = _find_mapper_path_for_qid(table_access, table_lookup, qid) or ''
                mapper_path = os.path.basename(_mp) if _mp else ''
            except Exception:
                mapper_path = ''

            # 3. 대상테이블: metadata에서 추출
            # (table_lookup 사용)

            # 4. 대상컬럼: table_access에서 현재 테이블의 모든 columns 추출
            target_cflds = _extract_table_columns_from_table_access(table_access, table_lookup)

            # 5. 암복호화 필요 컬럼과 6. 암복호화 필요 Java Field
            # step1의 query 객체에서 input/output_mapping의 crypto_fields 추출
            crypto_cflds = extract_crypto_fields_from_query(q, 'column_name')
            java_flds = extract_crypto_fields_from_query(q, 'java_field')

            # 7. End Point: table_access의 call_stacks에서 첫 항목 추출
            end_point = ''
            try:
                if isinstance(table_access, list):
                    for tobj in table_access:
                        if tobj.get('table_name') == table_lookup:
                            sqs = tobj.get('sql_queries') or []
                            for sq in sqs:
                                sid = sq.get('id')
                                if sid and (str(sid) == str(qid) or 
                                           str(sid).rsplit('.', 1)[-1] == str(qid).rsplit('.', 1)[-1]):
                                    call_stacks = sq.get('call_stacks') or []
                                    end_point = _extract_end_point_from_call_stacks(call_stacks)
                                    break
                            break
            except Exception:
                end_point = ''

            # 8. ResultMap: table_access에서 result_map 여부 판정
            rm = get_result_map(table_lookup, qid, table_access)

            # 9. 클래스명: modification_instructions에서 file_name 추출
            cls = ''
            if instr and isinstance(instr, dict):
                try:
                    file_name = instr.get('file_name') or ''
                    cls = file_name.replace('.java', '').strip() if file_name else ''
                except Exception:
                    cls = ''

            # 10. 메서드명: modification_instructions에서 target_method 추출
            method_name = ''
            if instr and isinstance(instr, dict):
                try:
                    method_name = instr.get('target_method') or ''
                except Exception:
                    method_name = ''

            # 11. Model: step1에서 input/output_mapping의 타입 정보 추출
            model_val = ''
            if q:
                try:
                    model_val = derive_model_common(step_info['s1'], q)
                except Exception:
                    model_val = ''

            # 12. Action: modification_instructions에서 실행 액션 추출
            action = ''
            if instr and isinstance(instr, dict):
                try:
                    action = instr.get('action') or ''
                except Exception:
                    action = ''

            # 13. Reason: modification_instructions에서 변경 사유 추출
            reason_val = ''
            if instr and isinstance(instr, dict):
                try:
                    reason_val = instr.get('reason') or ''
                except Exception:
                    reason_val = ''

            # 14. Insertion Point: modification_instructions에서 코드 삽입 위치 추출
            insertion_point = ''
            if instr and isinstance(instr, dict):
                try:
                    insertion_point = instr.get('insertion_point') or ''
                except Exception:
                    insertion_point = ''

            # 15. Code Pattern Hint: modification_instructions에서 코드 패턴 힌트 추출
            code_pattern_hint = ''
            if instr and isinstance(instr, dict):
                try:
                    code_pattern_hint = instr.get('code_pattern_hint') or ''
                except Exception:
                    code_pattern_hint = ''

            # 16. Sql Summary: step1에서 SQL 요약 추출
            sql_summary = ''
            if q:
                try:
                    sql_summary = q.get('sql_summary') or ''
                except Exception:
                    sql_summary = ''

            # 추출된 모든 정보로 레코드 생성
            records.append({
                '패키지명': repo_name,
                'SQL ID': display_qid,
                'Mapper Path': mapper_path,
                '대상테이블': table_lookup,
                '대상컬럼': target_cflds,
                '암복호화 필요 컬럼': crypto_cflds,
                '암복호화 필요 Java Field': java_flds,
                'End Point': end_point,
                'ResultMap': rm,
                '클래스명': cls,
                '메서드명': method_name,
                'Model': model_val,
                'Action': action,
                'Reason': reason_val,
                'Insertion Point': insertion_point,
                'Code Pattern Hint': code_pattern_hint,
                'Sql Summary': sql_summary,
            })

    # 정렬
    _sort_records(records)

    _print_summary(missing_files, invalid_files)

    # 시트에 데이터 작성
    last_row = _write_to_sheet(sh, records, headers, font_default, border)

    # 연속된 같은 값의 셀 병합
    _merge_columns(sh, last_row)

    return sh, records


def fill_target_list_modification_btype(sh, applycrypto_root=None, repo_name=None, font_default=None, border=None, headers=None):
    """TypeHandler 분석 - table_access 기준으로 대상목록 시트를 채웁니다.

    처리 절차:
    1. table_access_info.json의 모든 SQL ID를 기준 데이터로 추출
    2. step1_query_analysis.json의 데이터를 미리 캐시
    3. 각 SQL ID에 대해:
       - step1에서 query_id 매칭 (보충정보)
       - 필터링 없음: table_access의 모든 SQL ID는 행으로 추가됨
    4. 추출된 컬럼값으로 행 추가
    5. 정렬 및 셀 병합 적용

    필터링 효과: 없음 (table_access의 모든 SQL ID를 처리)
    완전성: table_access에 존재하는 SQL ID만 Excel에 포함됨 (step1은 보충정보)

    주의: ThreeStep과 달리 TypeHandler은 modification_instructions이 없으므로
    '메서드명', '클래스명', 'Action', 'Reason' 등은 공란입니다.

    Args:
        sh: openpyxl Worksheet 객체
        applycrypto_root: '.applycrypto' 디렉터리 경로
        repo_name: 패키지명
        font_default: 기본 폰트
        border: 테두리 스타일
        headers: 컬럼 헤더 목록

    Returns:
        tuple: (수정된 시트 객체, 레코드 리스트)
    """
    missing_files = []
    invalid_files = []

    if not applycrypto_root or not os.path.isdir(applycrypto_root):
        print(f"- 지정된 경로에 '.applycrypto'가 없습니다: {applycrypto_root}", flush=True)
        return sh, []

    three_root = os.path.join(applycrypto_root, 'three_step_results')
    if not os.path.isdir(three_root):
        return sh, []

    ts_dirs = [d for d in os.listdir(three_root) if os.path.isdir(os.path.join(three_root, d))]
    if not ts_dirs:
        return sh, []
    latest = sorted(ts_dirs)[-1]
    latest_root = os.path.join(three_root, latest)

    table_access = safe_load_json(
        os.path.join(applycrypto_root, 'results', 'table_access_info.json'),
        missing_files=missing_files,
        invalid_files=invalid_files,
        applycrypto_root=applycrypto_root
    ) or {}

    # table_access의 모든 SQL ID를 기준 데이터로 추출 (동일 SQL ID가 여러 테이블에 존재할 수 있으므로 리스트 사용)
    table_access_sql_ids = {}  # sql_id -> [table_info, table_info, ...]
    if isinstance(table_access, list):
        for tobj in table_access:
            sql_queries = tobj.get('sql_queries', [])
            for sq in sql_queries:
                sql_id = sq.get('id')
                if sql_id:
                    if sql_id not in table_access_sql_ids:
                        table_access_sql_ids[sql_id] = []
                    table_access_sql_ids[sql_id].append(tobj)
    
    # step1 데이터를 사전 캐시
    step1_data_cache = {}
    for table in sorted(os.listdir(latest_root)):
        table_dir = os.path.join(latest_root, table)
        if not os.path.isdir(table_dir):
            continue
        for ctrl in sorted(os.listdir(table_dir)):
            ctrl_dir = os.path.join(table_dir, ctrl)
            if not os.path.isdir(ctrl_dir):
                continue
            s1 = safe_load_json(
                os.path.join(ctrl_dir, 'step1_query_analysis.json'),
                missing_files=missing_files,
                invalid_files=invalid_files,
                applycrypto_root=applycrypto_root
            ) or {}
            
            step1_data_cache.setdefault(table, {})[ctrl] = s1
    
    def qid_match_for_table_access(a, b):
        """SQL ID 비교: 전체 일치 또는 마지막 토큰만 비교
        
        예: 'com.example.mapper.UserMapper.findById' vs 'findById' → 일치
        """
        try:
            sa = '' if a is None else str(a)
            sb = '' if b is None else str(b)
            # 둘 다 단순명(점 없음): 전체 비교
            if '.' not in sa and '.' not in sb:
                return sa == sb
            # 한쪽이라도 정규화명(점 있음): 마지막 토큰으로 비교
            return sa.rsplit('.', 1)[-1] == sb.rsplit('.', 1)[-1]
        except Exception:
            return False
    
    def find_step1_info_for_qid(qid, table_lookup, step1_data_cache):
        """주어진 SQL ID와 테이블명을 step1에서 검색하여 관련 정보를 반환
        
        Args:
            qid: 검색할 SQL ID
            table_lookup: 테이블명 (step1 캐시의 테이블 디렉토리명과 매칭)
            step1_data_cache: 캐시된 step1 데이터

        Returns:
            dict: step1 발견 여부 및 추출 정보
                - found: step1에서 발견 여부
                - s1: step1 전체 데이터
                - q: 매칭된 query 객체
        """
        result = {
            'found': False,
            's1': None,
            'q': None
        }
        
        # 테이블명이 일치하는 경우만 검색 (★ 중요: 테이블명 필수 일치)
        if table_lookup not in step1_data_cache:
            return result
            
        for ctrl in step1_data_cache[table_lookup]:
            s1 = step1_data_cache[table_lookup][ctrl]
            qs = (s1.get('result') or {}).get('queries') or []
            
            for q in qs:
                q_qid = q.get('query_id') or ''
                if qid_match_for_table_access(q_qid, qid):
                    result['found'] = True
                    result['s1'] = s1
                    result['q'] = q
                    return result
        
        return result
    
    records = []
    
    # table_access의 모든 SQL ID를 기준으로 처리 (동일 SQL ID가 여러 테이블에 있을 수 있음)
    for sql_id in sorted(table_access_sql_ids.keys()):
        table_info_list = table_access_sql_ids[sql_id]  # 리스트 (같은 SQL ID의 모든 테이블)
        
        # 각 테이블별로 개별 레코드 생성
        for table_info in table_info_list:
            table_lookup = table_info.get('table_name', '')
        
            # step1에서 정보 찾기 (테이블명과 SQL ID 모두 일치하는 경우만)
            step1_info = find_step1_info_for_qid(sql_id, table_lookup, step1_data_cache)
        
            qid = sql_id
            q = step1_info['q'] if step1_info['found'] else None
        
            # 1. SQL ID 표시 (마지막 토큰만 사용)
            try:
                display_qid = qid.rsplit('.', 1)[-1] if qid else ''
            except Exception:
                display_qid = qid

            # 2. Mapper Path 추출
            mapper_path = ''
            try:
                _mp = _find_mapper_path_for_qid(table_access, table_lookup, qid) or ''
                mapper_path = os.path.basename(_mp) if _mp else ''
            except Exception:
                mapper_path = ''

            # 3. 대상테이블: metadata에서 추출
            # (table_lookup 사용)

            # 4. 대상컬럼: table_access에서 현재 테이블의 모든 columns 추출
            target_cflds = _extract_table_columns_from_table_access(table_access, table_lookup)

            # 5. 암복호화 필요 컬럼과 6. 암복호화 필요 Java Field
            # step1의 query 객체에서 input/output_mapping의 crypto_fields 추출
            crypto_cflds = extract_crypto_fields_from_query(q, 'column_name')
            java_flds = extract_crypto_fields_from_query(q, 'java_field')

            # 7. End Point: table_access의 call_stacks에서 첫 항목 추출
            match_sq = None
            try:
                if isinstance(table_access, list):
                    for tobj in table_access:
                        if tobj.get('table_name') == table_lookup:
                            sqs = tobj.get('sql_queries') or []
                            for sq in sqs:
                                if qid_match_for_table_access(sq.get('id'), qid):
                                    match_sq = sq
                                    break
                            break
            except Exception:
                pass

            call_stacks = match_sq.get('call_stacks') or [] if match_sq else []
            
            end_point = _extract_end_point_from_call_stacks(call_stacks)
            
            # 8. ResultMap: table_access에서 result_map 여부 판정
            rm = get_result_map(table_lookup, qid, table_access)

            # 8. 클래스명: call_stacks[0]의 마지막 항목에서 추출 (btype용)
            cls = ''
            method_name = ''
            try:
                if call_stacks and isinstance(call_stacks, list) and len(call_stacks) > 0:
                    first_stack = call_stacks[0]
                    if isinstance(first_stack, list) and len(first_stack) > 0:
                        last_item = first_stack[-1]
                        if isinstance(last_item, str) and '.' in last_item:
                            parts = last_item.rsplit('.', 1)
                            cls = parts[0]  # ClassName
                            method_name = parts[1]  # methodName
            except Exception:
                cls = ''
                method_name = ''

            # 9. 메서드명: call_stacks[0]의 마지막 항목에서 추출 (btype용)
            # 위의 8번 섹션에서 이미 추출됨

            # 10. Model: step1에서 input/output_mapping의 타입 정보 추출
            model_val = ''
            if q and step1_info['found']:
                try:
                    model_val = derive_model_common(step1_info['s1'], q)
                except Exception:
                    model_val = ''

            # 11. Action: modification_instructions에서 실행 액션 추출
            # btype 해당없음

            # 12. Reason: modification_instructions에서 변경 사유 추출
            # btype 해당없음

            # 13. Insertion Point: modification_instructions에서 코드 삽입 위치 추출
            # TypeHandler 미사용

            # 14. Code Pattern Hint: modification_instructions에서 코드 패턴 힌트 추출
            # btype 해당없음

            # 15. Sql Summary: step1에서 SQL 요약 추출
            sql_summary = ''
            if q:
                try:
                    sql_summary = q.get('sql_summary') or ''
                except Exception:
                    sql_summary = ''

            # TypeHandler은 modification_instructions가 없으므로 메서드명 포함하여 13개 필드 추가
            records.append({
                '패키지명': repo_name,
                'SQL ID': display_qid,
                'Mapper Path': mapper_path,
                '대상테이블': table_lookup,
                '대상컬럼': target_cflds,
                '암복호화 필요 컬럼': crypto_cflds,
                '암복호화 필요 Java Field': java_flds,
                'End Point': end_point,
                'ResultMap': rm,
                '클래스명': cls,
                '메서드명': method_name,
                'Model': model_val,
                'Sql Summary': sql_summary,
            })

    # 정렬
    _sort_records(records)

    _print_summary(missing_files, invalid_files)

    # 시트에 데이터 작성
    last_row = _write_to_sheet(sh, records, headers, font_default, border)

    # 연속된 같은 값의 셀 병합
    _merge_columns(sh, last_row)

    return sh, records


def get_result_map(table, qid, table_access):
    """ResultMap 존재 여부를 조회합니다.

    매개변수:
        table: 테이블명
        qid: 쿼리 ID
        table_access: table_access_info.json 데이터

    반환:
        ResultMap 상태: 'X', '△', '○'
    """

    rm = 'X'
    try:
        # 비교 함수: 전체 ID 일치 또는 마지막 토큰(마지막 '.') 일치 허용
        def qid_eq(a, b):
            try:
                sa = '' if a is None else str(a)
                sb = '' if b is None else str(b)
                if sa == sb:
                    return True
                if sa.rsplit('.', 1)[-1] == sb.rsplit('.', 1)[-1]:
                    return True
            except Exception:
                pass
            return False

        if not isinstance(table_access, list):
            return 'X'
        
        found = False
        for tobj in table_access:
            if tobj.get('table_name') == table:
                # sql_queries의 형제 노드인 columns (테이블 레벨)
                table_columns = tobj.get('columns') or []
                for sq in (tobj.get('sql_queries') or []):
                    if qid_eq(sq.get('id'), qid):
                        rs = (sq.get('strategy_specific') or {})
                        rm_val = rs.get('result_map')
                        rfm = rs.get('result_field_mappings') or []
                        
                        if rm_val is None or rm_val == '':
                            rm = 'X'
                        else:
                            # result_map이 있음
                            # result_field_mappings이 있는지
                            has_result_field_mappings = bool(rfm)
                            
                            if not has_result_field_mappings:
                                rm = '△'
                            else:
                                rm = '○'
                                # 2024-06-11: 중복 체크 로직 **임시** 제거(□와 ○구분 없음)
                                # result_field_mappings이 있음
                                # columns의 name과 result_field_mappings 중복 체크
                                # names = set(
                                #     col.get('name') 
                                #     for col in table_columns 
                                #     if isinstance(col, dict) and col.get('name')
                                # )
                                
                                # # result_field_mappings에서 매핑 이름들 추출
                                # # 형식: [["col1", "mapping1"], ["col2", "mapping2"], ...]
                                # mappings = set()
                                # for item in rfm:
                                #     if isinstance(item, (list, tuple)) and len(item) >= 1:
                                #         # 첫 번째 요소가 매핑 이름
                                #         if item[0]:
                                #             mappings.add(str(item[0]))
                                #     elif isinstance(item, str):
                                #         # 문자열인 경우 직접 추가
                                #         mappings.add(item)
                                
                                # overlap = names & mappings
                                # if overlap:
                                #     overlap_list = sorted(overlap)
                                #     rm = '○ ' + ', '.join(overlap_list)
                                # else:
                                #     rm = '□'
                        found = True
                        break
            if found:
                break
    except Exception:
        print(f'ResultMap 조회 실패: 테이블 {table}, qid {qid}', flush=True)
        rm = 'X'
    return rm


def _find_mapper_path_for_qid(table_access, table_name, qid):
    """주어진 테이블과 qid에 매칭되는 sql_queries 항목에서 source_file_path에서 mapper정보를 반환합니다."""
    try:
        if not isinstance(table_access, list):
            return ''
        
        for tobj in table_access:
            if tobj.get('table_name') != table_name:
                continue
            for sq in (tobj.get('sql_queries') or []):
                sid = sq.get('id')
                # quick equality with last token fallback
                if sid is None:
                    continue
                try:
                    s_sid = str(sid)
                    s_qid = '' if qid is None else str(qid)
                    if s_sid == s_qid or s_sid.rsplit('.', 1)[-1] == s_qid.rsplit('.', 1)[-1]:
                        # look for known fields
                        if 'source_file_path' in sq and sq.get('source_file_path'):
                            return sq.get('source_file_path')
                        # fallback to access_files
                        if 'access_files' in sq and sq.get('access_files'):
                            af = sq.get('access_files')
                            if isinstance(af, list) and af:
                                return af[0]
                        # fallback: call_stacks may contain paths
                        if 'call_stacks' in sq and sq.get('call_stacks'):
                            cs = sq.get('call_stacks')
                            if isinstance(cs, list) and cs:
                                return cs[0]
                except Exception:
                    continue
    except Exception:
        pass
    return ''


def _extract_end_point_from_call_stacks(call_stacks):
    """call_stacks에서 End Point (첫 번째 항목)를 추출합니다.
    
    지원 형식:
    1. 문자열 리스트: ["com.example.UserController.login", "com.example.UserService.login"]
    2. 리스트의 리스트: [["UserController.login", ...], ["UserService.login", ...]]
    
    반환: 첫 번째 항목들을 쉼표로 결합 (중복 제거)
    """
    if not call_stacks or not isinstance(call_stacks, list):
        return ''
    
    endpoints = []
    for item in call_stacks:
        # Case 1: 문자열 직접 입력
        if isinstance(item, str) and item.strip():
            endpoints.append(item.strip())
        # Case 2: 리스트 입력 (각 리스트의 첫 번째 항목 추출)
        elif isinstance(item, list) and item:
            first_item = item[0]
            if isinstance(first_item, str) and first_item.strip():
                endpoints.append(first_item.strip())
    
    # 중복 제거 (순서 유지)
    seen = set()
    unique_endpoints = []
    for ep in endpoints:
        if ep not in seen:
            seen.add(ep)
            unique_endpoints.append(ep)
    
    return ',\n'.join(unique_endpoints)


def _extract_table_columns_from_table_access(table_access, table_name):
    """table_access_info.json에서 특정 테이블의 모든 컬럼명을 추출합니다.
    
    처리:
    1. table_access_info.json 배열에서 table_name 일치하는 객체 찾기
    2. 해당 객체의 columns[] 추출
    3. 각 column의 name 수집
    4. 쉼표+공백으로 구분
    
    Returns:
        컬럼명을 ', '로 구분한 문자열
    """
    if not isinstance(table_access, list):
        return ''
    
    try:
        for tobj in table_access:
            if isinstance(tobj, dict) and tobj.get('table_name') == table_name:
                columns = tobj.get('columns') or []
                col_names = []
                for col in columns:
                    if isinstance(col, dict):
                        col_name = col.get('name')
                        if col_name and isinstance(col_name, str):
                            col_names.append(col_name.strip())
                return ', '.join(col_names)
    
    except Exception:
        pass
    
    return ''


def extract_crypto_fields_from_query(q, field_key):
    """암복호화 필드를 step1의 query 객체에서 추출합니다.
    
    query의 input_mapping과 output_mapping에서 crypto_fields 배열을 순회하며,
    지정된 필드명(field_key)의 값들을 추출하여 중복을 제거한 후 쉼표로 구분하여 반환합니다.
    
    Args:
        q (dict): step1의 query 객체
        field_key (str): 추출할 필드명 ('column_name' 또는 'java_field' 등)
    
    Returns:
        str: 쉼표와 공백으로 구분된 필드값 (중복 제거, 순서 유지)
             예외 발생 시 빈 문자열
    
    예시:
        >>> q = {...}
        >>> extract_crypto_fields_from_query(q, 'column_name')  # '컬럼1, 컬럼2'
        >>> extract_crypto_fields_from_query(q, 'java_field')   # '필드1, 필드2'
    """
    if not q:
        return ''
    
    try:
        in_cf = (q.get('input_mapping') or {}).get('crypto_fields') or []
        out_cf = (q.get('output_mapping') or {}).get('crypto_fields') or []
        items = []
        
        # input_mapping과 output_mapping의 crypto_fields를 모두 순회
        for c in (in_cf or []) + (out_cf or []):
            if isinstance(c, dict):
                item = c.get(field_key)
                if item and isinstance(item, str):
                    items.append(item.strip())
        
        # 중복 제거 (순서 유지)
        seen = set()
        uniq = []
        for item in items:
            if item not in seen:
                seen.add(item)
                uniq.append(item)
        
        return ', '.join(uniq)
    except Exception:
        return ''


def derive_model_common(s1, q):
    """공통 Model 추출 로직: `input_mapping`과 `output_mapping`에서 `type_category`가
    'VO' 또는 'MAP'인 `class_name`들을 모두 수집하여 `, `로 구분해 반환합니다."""
    models = []
    try:
        if isinstance(q, dict):
            im = q.get('input_mapping')
            if isinstance(im, dict):
                if im.get('type_category') in ('VO', 'MAP'):
                    cn = im.get('class_name') or q.get('class_name')
                    if cn:
                        models.append(cn)

        if isinstance(q, dict):
            om = q.get('output_mapping')
            if isinstance(om, dict):
                if om.get('type_category') in ('VO', 'MAP'):
                    cn = om.get('class_name')
                    if cn:
                        models.append(cn)

        if isinstance(q, dict):
            if q.get('type_category') in ('VO', 'MAP') and q.get('class_name'):
                models.append(q.get('class_name'))

        if isinstance(s1, dict):
            res = s1.get('result') or {}
            for key in ('objects', 'classes', 'types'):
                arr = res.get(key) or []
                if isinstance(arr, list):
                    for it in arr:
                        if not isinstance(it, dict):
                            continue
                        tc = it.get('type_category')
                        cn = it.get('class_name') or it.get('name')
                        if tc in ('VO', 'MAP') and cn:
                            models.append(cn)

            meta = s1.get('metadata') or {}
            if meta:
                if meta.get('type_category') in ('VO', 'MAP') and meta.get('class_name'):
                    models.append(meta.get('class_name'))
    except Exception:
        pass
    # 중복 제거
    unique_models = list(set(models))
    return ', '.join(unique_models) if unique_models else ''


def safe_load_json(path, missing_files=None, invalid_files=None, applycrypto_root=None):
    """로컬: JSON 파일을 안전하게 로드합니다. 누락/무효 목록에 경로를 추가할 수 있습니다.

    반환값:
        파싱된 JSON 객체 또는 None을 반환합니다. 파일이 없거나(또는 0바이트) 파싱 오류가 발생하면
        각각 `missing_files` 또는 `invalid_files` 리스트에 경로를 추가합니다.
    """
    try:
        try:
            if os.path.exists(path) and os.path.getsize(path) == 0:
                if isinstance(missing_files, list):
                    missing_files.append(path)
                try:
                    rel = os.path.relpath(path, applycrypto_root) if applycrypto_root else path
                except Exception:
                    rel = path
                print(f"- 빈 JSON 파일 무시: {rel}", flush=True)
                return None
        except Exception:
            pass
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except FileNotFoundError:
        if isinstance(missing_files, list):
            missing_files.append(path)
        return None
    except json.JSONDecodeError:
        if isinstance(invalid_files, list):
            invalid_files.append(path)
        try:
            rel = os.path.relpath(path, applycrypto_root) if applycrypto_root else path
        except Exception:
            rel = path
        print(f"- 파싱 오류 : {rel}", flush=True)
        return None
    except Exception:
        print(f"- JSON 로드 실패: {path}", flush=True)
        return None


def _extract_class_from_callstack(call_str):
    """호출 스택 문자열에서 클래스명을 추출합니다."""
    try:
        parts = [p for p in str(call_str).split('.') if p]
        if len(parts) >= 2:
            return parts[-2]
        if parts:
            return parts[0]
    except Exception:
        pass
    return ''


def make_border():
    """얇은 테두리(Border)를 생성하여 반환합니다."""
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def add_validation_sheet(wb, validation_data, font_default, border):
    """SQL ID 검증 결과를 시트로 추가합니다.
    
    Args:
        wb: openpyxl Workbook 객체
        validation_data: 검증 결과 데이터 {'result': bool, 'stats': dict, 'details': list}
        font_default: 기본 폰트
        border: 테두리 스타일
    
    Returns:
        Worksheet: 생성된 검증 시트
    """
    if not validation_data or not validation_data.get('details'):
        return None
    
    sh = wb.create_sheet(title='SQL ID 검증')
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    header_fill = PatternFill(start_color='DFDFDF', end_color='DFDFDF', fill_type='solid')
    
    # 컬럼 너비 설정
    sh.column_dimensions['A'].width = 20  # 테이블명
    sh.column_dimensions['B'].width = 40  # SQL ID
    sh.column_dimensions['C'].width = 15  # step1 존재여부
    sh.column_dimensions['D'].width = 15  # step2 존재여부
    sh.column_dimensions['E'].width = 15  # 상태
    
    # 헤더 설정 (1행)
    headers = ['테이블명', 'SQL ID', 'step1 존재여부', 'step2 존재여부', '상태']
    for c, h in enumerate(headers, start=1):
        cell = sh.cell(row=1, column=c, value=h)
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
    
    # 데이터 행 추가
    details = validation_data.get('details', [])
    for row_idx, detail in enumerate(details, start=2):
        sh.cell(row=row_idx, column=1, value=detail.get('table_name', ''))
        sh.cell(row=row_idx, column=2, value=detail.get('sql_id', ''))
        sh.cell(row=row_idx, column=3, value=detail.get('step1_status', ''))
        sh.cell(row=row_idx, column=4, value=detail.get('step2_status', ''))
        sh.cell(row=row_idx, column=5, value=detail.get('status', ''))
        
        # 각 셀에 기본 스타일 적용
        for col_idx in range(1, 6):
            cell = sh.cell(row=row_idx, column=col_idx)
            cell.font = font_default
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 상태가 'MISSING'인 경우 배경색 표시
            if col_idx == 5 and detail.get('status') == 'MISSING':
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # 통계 정보 추가 (마지막 행 아래에 빈 행 후 표시)
    stats_row = len(details) + 3
    stats = validation_data.get('stats', {})
    
    total_ids = stats.get('total_table_ids', 0)
    step1_matched = stats.get('step1_matched', 0)
    step2_matched = stats.get('step2_matched', 0)
    missing = stats.get('missing_count', 0)
    
    # 통계 헤더
    stat_headers = ['항목', '개수']
    for c, h in enumerate(stat_headers, start=1):
        cell = sh.cell(row=stats_row, column=c, value=h)
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
    
    # 통계 데이터
    stat_rows = [
        ('총 SQL ID', total_ids),
        ('Step1 매칭', step1_matched),
        ('Step2 매칭', step2_matched),
        ('누락', missing),
        ('상태', 'OK' if missing == 0 else 'MISSING')
    ]
    
    for idx, (label, value) in enumerate(stat_rows, start=1):
        row = stats_row + idx
        sh.cell(row=row, column=1, value=label)
        sh.cell(row=row, column=2, value=value)
        
        for col_idx in range(1, 3):
            cell = sh.cell(row=row, column=col_idx)
            cell.font = font_default
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 누락 행 강조
            if label == '누락' and value > 0:
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif label == '상태' and value == 'MISSING':
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    return sh


def set_row_height(ws, height=16):
    """워크시트의 모든 행에 대해 높이를 설정합니다.

    매개변수:
        wb: openpyxl Worksheet 객체
        height: 행 높이(기본 16)
    """
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = height


def _sort_records(records):
    """레코드를 SQL ID, Mapper Path, 대상테이블 순으로 정렬합니다."""
    try:
        def sort_key(r):
            def kv(v):
                s = (v or '')
                return s.lower() if s else '\uffff'
            return (kv(r.get('SQL ID')), kv(r.get('Mapper Path')), kv(r.get('대상테이블')))
        records.sort(key=sort_key)
    except Exception:
        pass


def _print_summary(missing_files, invalid_files):
    """누락된 파일과 파싱 오류 파일 요약을 출력합니다."""
    try:
        if missing_files or invalid_files:
            if missing_files:
                print(f"- 누락된 JSON 건수: {len(missing_files)}. 첫 파일: {missing_files[0]}", flush=True)
            if invalid_files:
                print(f"- 파싱 오류 JSON 건수: {len(invalid_files)}. 첫 파일: {invalid_files[0]}", flush=True)
    except Exception:
        pass


def _write_to_sheet(sh, records, headers, font_default, border):
    """레코드를 워크시트에 기록하고 마지막 행 번호를 반환합니다."""
    cols = headers if headers else [
        "패키지명", "파일명", "SQL ID", "대상테이블", "대상컬럼",
        "ResultMap", "클래스명", "메서드명", "Model", "Action",
        "Reason", "Insertion Point", "Code Pattern Hint", "Sql Summary"
    ]
    r = 3
    for rec in records:
        for c in range(1, len(cols) + 1):
            col_name = cols[c - 1]
            cell = sh.cell(row=r, column=c, value=rec.get(col_name, ''))
            if font_default:
                cell.font = font_default
            if border:
                cell.border = border
            
            # 컬럼별 정렬 설정
            if col_name == 'ResultMap':
                # ResultMap: 가운데 정렬
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            elif col_name == '패키지명':
                # 패키지명: 위쪽 정렬 + 줄바꿈
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            elif col_name in ('Reason', 'Insertion Point', 'Code Pattern Hint', 'Sql Summary'):
                # 이 항목들: 줄바꿈 활성화
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            else:
                # 기타: 중앙 정렬
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        r += 1
    return r - 1  # last_row


def _merge_columns(sh, last_row):
    """병합 전략:
    1. 패키지명(col 1): 연속된 같은 값만 병합
    2. SQL ID(col 2) + Mapper Path(col 3): SQL ID와 Mapper Path가 모두 같은 행들끼리만 병합
       - 대상테이블(col 4)이 다르더라도 상관없음
    """
    def norm_val(v):
        try:
            if v is None:
                return ''
            return str(v).strip()
        except Exception:
            return ''
    
    if last_row < 3:
        return
    
    try:
        # 1. 패키지명(col 1) - 일반 병합
        seg_start = 3
        prev = norm_val(sh.cell(row=3, column=1).value)
        
        for rr in range(4, last_row + 1):
            cur = norm_val(sh.cell(row=rr, column=1).value)
            if cur != prev:
                if prev not in ('', None) and rr - 1 > seg_start:
                    try:
                        sh.merge_cells(start_row=seg_start, start_column=1, end_row=rr - 1, end_column=1)
                    except Exception:
                        pass
                seg_start = rr
                prev = cur
        
        if prev not in ('', None) and last_row > seg_start:
            try:
                sh.merge_cells(start_row=seg_start, start_column=1, end_row=last_row, end_column=1)
            except Exception:
                pass
        
        # 2. SQL ID(col 2)와 Mapper Path(col 3) - 둘 다 같을 때만 병합
        # 먼저 SQL ID와 Mapper Path가 모두 같은 행들을 식별
        seg_start = 3
        prev_sql_id = norm_val(sh.cell(row=3, column=2).value)
        prev_mapper = norm_val(sh.cell(row=3, column=3).value)
        
        for rr in range(4, last_row + 1):
            cur_sql_id = norm_val(sh.cell(row=rr, column=2).value)
            cur_mapper = norm_val(sh.cell(row=rr, column=3).value)
            
            # SQL ID 또는 Mapper Path가 다르면 새 세그먼트 시작
            if cur_sql_id != prev_sql_id or cur_mapper != prev_mapper:
                # 이전 세그먼트 병합
                if (prev_sql_id not in ('', None) and prev_mapper not in ('', None) and 
                    rr - 1 > seg_start):
                    try:
                        # SQL ID 병합
                        sh.merge_cells(start_row=seg_start, start_column=2, end_row=rr - 1, end_column=2)
                    except Exception:
                        pass
                    try:
                        # Mapper Path 병합
                        sh.merge_cells(start_row=seg_start, start_column=3, end_row=rr - 1, end_column=3)
                    except Exception:
                        pass
                
                seg_start = rr
                prev_sql_id = cur_sql_id
                prev_mapper = cur_mapper
        
        # 마지막 세그먼트 병합
        if (prev_sql_id not in ('', None) and prev_mapper not in ('', None) and 
            last_row > seg_start):
            try:
                sh.merge_cells(start_row=seg_start, start_column=2, end_row=last_row, end_column=2)
            except Exception:
                pass
            try:
                sh.merge_cells(start_row=seg_start, start_column=3, end_row=last_row, end_column=3)
            except Exception:
                pass
    
    except Exception:
        pass


# =====================================
# 번역 관련 클래스 및 함수 (artifact_generator 패턴 참고)
# =====================================

def _update_sheet_with_translated_records(sh: 'Worksheet', records: list, headers: list):
    """번역된 레코드를 엑셀 시트에 반영합니다.
    
    Args:
        sh: 대상 워크시트
        records: 번역된 레코드 리스트
        headers: 헤더 리스트
    
    Returns:
        None
    """
    try:
        # 헤더 행(2행) 다음부터 데이터 시작 (3행)
        data_start_row = 3
        
        # 각 레코드의 셀 값 업데이트
        for row_idx, record in enumerate(records, start=data_start_row):
            for col_idx, header in enumerate(headers, start=1):
                try:
                    cell = sh.cell(row=row_idx, column=col_idx)
                    # 병합된 셀이 아닌 경우만 값 설정
                    # MergedCell 체크: cell.data_type 확인
                    if not hasattr(cell, 'is_merged') or not cell.is_merged:
                        cell.value = record.get(header, '')
                except Exception:
                    # 병합된 셀은 건너뛰고 계속 진행
                    pass
    except Exception as e:
        print(f'- 시트 업데이트 중 오류: {e}', flush=True)


class TranslationCache:
    """MD5 해시 기반 번역 결과 캐시 (TTL 관리)"""
    
    def __init__(self, ttl_seconds=TRANSLATION_CACHE_TTL):
        self.cache = {}
        self.ttl = ttl_seconds
    
    def get_key(self, text: str) -> str:
        """텍스트의 MD5 해시를 캐시 키로 생성"""
        return hashlib.md5(text.encode('utf-8')).hexdigest()
    
    def get(self, text: str) -> str:
        """캐시에서 번역된 텍스트 조회 (TTL 확인)"""
        key = self.get_key(text)
        if key in self.cache:
            translated, timestamp = self.cache[key]
            if (datetime.now() - timestamp).total_seconds() < self.ttl:
                return translated
            else:
                del self.cache[key]
        return None
    
    def set(self, text: str, translated: str):
        """번역 결과를 캐시에 저장"""
        key = self.get_key(text)
        self.cache[key] = (translated, datetime.now())


def translate_batch_atype(records: list, llm_provider, cache: 'TranslationCache' = None) -> list:
    """ThreeStep(atype) 레코드 배치 번역: Reason, Insertion Point, Code Pattern Hint, Sql Summary
    
    Args:
        records: 번역할 레코드 리스트 (최대 5개)
        llm_provider: LLM 프로바이더 객체
        cache: 번역 캐시 객체
    
    Returns:
        번역된 레코드 리스트
    """
    if not records or not llm_provider:
        return records
    
    # 번역 필요 여부 확인
    needs_translation = False
    for record in records:
        for field in TRANSLATION_FIELDS['atype']:
            if record.get(field) and isinstance(record.get(field), str):
                needs_translation = True
                break
        if needs_translation:
            break
    
    if not needs_translation:
        return records
    
    # 배치 프롬프트 구성
    batch_items = []
    for idx, record in enumerate(records, start=1):
        item = {'idx': idx}
        for field in TRANSLATION_FIELDS['atype']:
            text = record.get(field, '')
            if text and isinstance(text, str) and len(text) > 0:
                # 캐시 확인
                if cache:
                    cached = cache.get(text)
                    if cached:
                        item[field] = cached
                        continue
                item[field] = text
        batch_items.append(item)
    
    if not batch_items:
        return records
    
    # LLM 배치 호출
    batch_prompt = {
        'task': 'translate_analysis_fields',
        'instruction': """
# Java 코드 마이그레이션 분석 결과 번역

당신은 Java 애플리케이션 마이그레이션 프로젝트의 분석 결과를 한국어로 번역하는 전문가입니다.

## 번역 대상 필드

### 1. Reason (변경 사유)
- **목적**: 소스 코드의 특정 부분이 왜 변경되었는지 설명
- **내용 예시**:
  - "암호화 정책 P019 적용 (주민등록번호 전체 암호화)"
  - "보안 강화를 위한 SSL/TLS 버전 업그레이드"
  - "데이터베이스 마이그레이션에 따른 연결 정보 변경"
  - "레거시 API 제거 및 신규 API 호출로 변경"
  - "개인정보보호법 준수를 위한 데이터 처리 로직 개선"
  
### 2. Insertion Point (삽입 위치)
- **목적**: 변경된 코드가 어느 위치에 추가/변경/삭제되었는지 지정
- **내용 예시**:
  - "메서드 시작 부분에 Null Check 로직 추가"
  - "SQL 쿼리 실행 전 암호화 처리 추가"
  - "try-catch 블록 내부에서 오류 처리 강화"
  - "메서드 매개변수 유효성 검증 로직 추가"
  - "데이터 바인딩 후 암호화 함수 호출 지점"
  
### 3. Sql Summary (SQL 요약)
- **목적**: SQL 쿼리가 무엇을 조회/수정하는지, 어떤 테이블/컬럼을 대상으로 하는지 설명
- **내용 예시**:
  - "사용자 기본정보 조회 (name, email, phone 컬럼 포함)"
  - "주민등록번호를 조건으로 사용자 정보 검색"
  - "암호화된 전화번호 업데이트"
  - "결제 정보 및 개인정보를 JOIN하여 조회"
  - "배치 작업으로 암호화되지 않은 레거시 데이터 마이그레이션"

## 번역 규칙

1. **정확성**: 마이그레이션 분석의 기술적 정확성을 유지하세요
2. **명확성**: 개발자가 즉시 이해할 수 있는 명확한 용어를 사용하세요
3. **일관성**: Java/SQL 기술 용어는 개발 업계에서 일반적으로 사용되는 표현을 사용하세요
4. **간결성**: 불필요한 설명은 제외하고 핵심만 전달하세요
5. **컨텍스트 유지**: 원문의 의도와 기술적 맥락을 손상시키지 마세요

## 응답 형식

JSON 배열로만 응답하세요 (마크다운, 설명, 추가 텍스트 제외):

[
  {"idx": 1, "Reason": "번역된 변경 사유", "Insertion Point": "번역된 삽입 위치", "Sql Summary": "번역된 SQL 요약"},
  {"idx": 2, ...},
  ...
]
        """,
        'context': {
            'items': batch_items
        }
    }
    
    # LLM 호출 전 메트릭 계산 및 출력
    prompt_str = json.dumps(batch_prompt, ensure_ascii=False)
    prompt_length = len(prompt_str)
    estimated_tokens = prompt_length // 4
    
    # 배치 항목의 세부 메트릭
    item_total_length, item_tokens, field_lengths = _calculate_llm_metrics(batch_items, 'atype')
    field_info = ', '.join([f"{field}: {length}자" for field, length in field_lengths.items() if length > 0])
    
    print(f"  [LLM 호출] atype 배치 - 예상 토큰: {estimated_tokens} (전체: {prompt_length}자, 항목: {item_total_length}자)")
    if field_info:
        print(f"             필드별: {field_info}")
    
    try:
        batch_resp = llm_provider.call(
            prompt_str,
            max_tokens=LLM_MAX_TOKENS,
            temperature=LLM_TEMPERATURE
        )
        
        batch_content = batch_resp.get('content', '[]') if isinstance(batch_resp, dict) else str(batch_resp)
        batch_results = []
        
        if batch_content.strip():
            try:
                batch_results = json.loads(batch_content)
                if isinstance(batch_results, dict):
                    batch_results = [batch_results]
            except json.JSONDecodeError:
                # 마크다운 코드블록 제거 시도
                s = batch_content
                s = re.sub(r'```(?:json)?\\n?', '', s)
                s = s.replace('```', '')
                
                start = s.find('[')
                end = s.rfind(']')
                if start != -1 and end != -1 and end > start:
                    try:
                        batch_results = json.loads(s[start:end+1])
                    except Exception:
                        batch_results = []
        
        # 결과 병합 및 캐시 업데이트
        result_map = {}
        for obj in batch_results or []:
            if isinstance(obj, dict) and 'idx' in obj:
                result_map[obj['idx']] = obj
        
        # 원본 레코드 업데이트 (실패한 필드는 원본 유지)
        for idx, record in enumerate(records, start=1):
            if idx in result_map:
                result = result_map[idx]
                for field in TRANSLATION_FIELDS['atype']:
                    if field in result and result[field]:
                        # 번역 결과가 있으면 사용
                        translated_text = str(result[field]).strip()
                        if translated_text:
                            record[field] = translated_text
                            # 캐시에 저장
                            if cache:
                                original_text = batch_items[idx-1].get(field, '')
                                if original_text:
                                    cache.set(original_text, translated_text)
                        # else: 번역이 비어있으면 원본 유지
                    # else: 번역 결과 없으면 원본 유지
            # else: 배치 결과에 없는 idx는 원본 유지
        
        return records
    except Exception as e:
        print(f"- atype 배치 번역 실패: {e} (원본 텍스트 유지)", flush=True)
        return records


def translate_batch_btype(records: list, llm_provider, cache: 'TranslationCache' = None) -> list:
    """TypeHandler(btype) 레코드 배치 번역: Sql Summary만 번역
    
    Args:
        records: 번역할 레코드 리스트 (최대 20개)
        llm_provider: LLM 프로바이더 객체
        cache: 번역 캐시 객체
    
    Returns:
        번역된 레코드 리스트
    """
    if not records or not llm_provider:
        return records
    
    # Sql Summary 필드만 번역
    field = 'Sql Summary'
    needs_translation = any(record.get(field) and isinstance(record.get(field), str) for record in records)
    
    if not needs_translation:
        return records
    
    # 배치 프롬프트 구성
    batch_items = []
    for idx, record in enumerate(records, start=1):
        text = record.get(field, '')
        if text and isinstance(text, str) and len(text) > 0:
            # 캐시 확인
            if cache:
                cached = cache.get(text)
                if cached:
                    record[field] = cached
                    continue
            batch_items.append({'idx': idx, field: text})
    
    if not batch_items:
        return records
    
    # LLM 배치 호출
    batch_prompt = {
        'task': 'translate_sql_summary',
        'instruction': """
# Java 코드 마이그레이션 분석 - SQL 요약 번역

당신은 Java 애플리케이션 마이그레이션 프로젝트의 SQL 분석 결과를 한국어로 번역하는 전문가입니다.

## 번역 대상: Sql Summary (SQL 요약)

SQL 쿼리가 무엇을 조회/수정하는지, 어떤 테이블과 컬럼을 대상으로 하는지 명확하게 설명합니다.

### 번역 시 포함해야 할 정보:
1. **주요 작업**: SELECT, INSERT, UPDATE, DELETE 등의 작업 유형
2. **대상 테이블**: 쿼리가 조회/수정하는 테이블명
3. **주요 컬럼**: 개인정보, 보안 관련 컬럼 등 중요 컬럼 언급
4. **조건/필터**: WHERE 절의 주요 조건 (있는 경우)
5. **JOIN/서브쿼리**: 다중 테이블 조회인 경우 관계 설명 (있는 경우)
6. **특수 처리**: 암호화, 해시, 포맷팅 등 특수한 데이터 처리

### 번역 예시:
- "회원 기본정보 조회 (이름, 이메일, 전화번호, 주민등록번호 포함)"
- "주민등록번호를 조건으로 사용자 계정 및 개인정보 검색"
- "암호화된 신용카드 정보 업데이트"
- "사용자, 결제정보, 배송주소를 JOIN하여 주문 내역 조회"
- "배치 처리로 레거시 데이터(암호화되지 않음)를 신규 형식으로 마이그레이션"
- "장기 미사용 회원의 개인정보(암호화된 데이터 포함) 일괄 삭제"

## 번역 규칙

1. **정확성**: SQL의 기술적 정확성을 유지하고, 테이블명/컬럼명은 영문으로 표기
2. **명확성**: 개발자가 SQL의 목적을 즉시 파악할 수 있도록 작성
3. **간결성**: 핵심 정보만 전달하고 과도한 세부사항은 제외
4. **컨텍스트**: 마이그레이션 프로젝트의 보안/암호화 관련 변경사항 반영
5. **용어**: JOIN, WHERE, GROUP BY 등 SQL 예약어는 영문 사용

## 응답 형식

JSON 배열로만 응답하세요 (마크다운, 설명, 추가 텍스트 제외):

[
  {"idx": 1, "Sql Summary": "번역된 SQL 요약 설명"},
  {"idx": 2, "Sql Summary": "번역된 SQL 요약 설명"},
  ...
]
        """,
        'context': {
            'items': batch_items
        }
    }
    
    # LLM 호출 전 메트릭 계산 및 출력
    prompt_str = json.dumps(batch_prompt, ensure_ascii=False)
    prompt_length = len(prompt_str)
    estimated_tokens = prompt_length // 4
    
    # 배치 항목의 세부 메트릭
    item_total_length, item_tokens, field_lengths = _calculate_llm_metrics(batch_items, 'btype')
    field_info = ', '.join([f"{field}: {length}자" for field, length in field_lengths.items() if length > 0])
    
    print(f"  [LLM 호출] btype 배치 - 예상 토큰: {estimated_tokens} (전체: {prompt_length}자, 항목: {item_total_length}자)")
    if field_info:
        print(f"             필드별: {field_info}")
    
    try:
        batch_resp = llm_provider.call(
            prompt_str,
            max_tokens=LLM_MAX_TOKENS,
            temperature=LLM_TEMPERATURE
        )
        
        batch_content = batch_resp.get('content', '[]') if isinstance(batch_resp, dict) else str(batch_resp)
        batch_results = []
        
        if batch_content.strip():
            try:
                batch_results = json.loads(batch_content)
                if isinstance(batch_results, dict):
                    batch_results = [batch_results]
            except json.JSONDecodeError:
                # 마크다운 코드블록 제거 시도
                s = batch_content
                s = re.sub(r'```(?:json)?\\n?', '', s)
                s = s.replace('```', '')
                
                start = s.find('[')
                end = s.rfind(']')
                if start != -1 and end != -1 and end > start:
                    try:
                        batch_results = json.loads(s[start:end+1])
                    except Exception:
                        batch_results = []
        
        # 결과 병합 및 캐시 업데이트
        result_map = {}
        for obj in batch_results or []:
            if isinstance(obj, dict) and 'idx' in obj:
                result_map[obj['idx']] = obj
        
        # 원본 레코드 업데이트 (번역 실패 시 원본 유지)
        for idx in result_map:
            if idx <= len(records):
                record = records[idx - 1]
                result = result_map[idx]
                if field in result and result[field]:
                    # 번역 결과가 있으면 사용
                    translated_text = str(result[field]).strip()
                    if translated_text:
                        original_text = record.get(field, '')
                        record[field] = translated_text
                        # 캐시에 저장
                        if cache:
                            cache.set(original_text, translated_text)
                    # else: 번역이 비어있으면 원본 유지
                # else: 번역 결과 없으면 원본 유지
        
        return records
    except Exception as e:
        print(f"- btype 배치 번역 실패: {e} (원본 텍스트 유지)", flush=True)
        return records


def translate_records_batch(records: list, modification_type: str, llm_provider, cache: 'TranslationCache' = None) -> list:
    """전체 레코드 배치 번역 (병렬 처리)
    
    Args:
        records: 번역할 레코드 전체 리스트
        modification_type: 'ThreeStep' 또는 'TypeHandler'
        llm_provider: LLM 프로바이더 객체
        cache: 번역 캐시 객체
    
    Returns:
        번역된 레코드 리스트
    """
    if not records or not llm_provider:
        return records
    
    # modification_type에 따른 배치 크기 결정
    if modification_type == 'ThreeStep':
        batch_size = MAX_TRANSLATION_ITEMS_PER_BATCH['atype']
        translate_func = translate_batch_atype
    elif modification_type in ('TypeHandler', 'TypeHandler'):
        batch_size = MAX_TRANSLATION_ITEMS_PER_BATCH['btype']
        translate_func = translate_batch_btype
    else:
        return records
    
    # 병렬 처리를 위해 ThreadPoolExecutor 사용 (max_workers=2)
    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            futures = []
            batch_info = []  # 배치 정보 저장 (순번, 시작행, 종료행, 크기)
            
            # 배치별로 번역 작업 제출
            batch_num = 0
            for i in range(0, len(records), batch_size):
                batch_num += 1
                batch = records[i:i + batch_size]
                batch_end = min(i + batch_size, len(records))
                
                # 배치 길이 검사
                param_length, is_warning = _calculate_batch_param_length(batch, modification_type)
                
                # 배치 시작 로그
                warning_msg = ', 경고: 길이 초과' if is_warning else ''
                print(f"  [배치 {batch_num}] 시작 - 행 {i+1}~{batch_end} ({len(batch)}개 항목{warning_msg})", flush=True)
                
                future = executor.submit(
                    translate_func,
                    batch,
                    llm_provider,
                    cache
                )
                futures.append((batch_num, i, batch_end, future))
                batch_info.append((batch_num, i, batch_end, len(batch)))
                
                # 배치 시작 사이 인터벌
                if batch_num < (len(records) // batch_size + (1 if len(records) % batch_size else 0)):
                    if LLM_BATCH_START_INTERVAL > 0:
                        time.sleep(LLM_BATCH_START_INTERVAL)
            
            # 결과 수집
            for batch_num, batch_start_idx, batch_end_idx, future in futures:
                try:
                    batch_result = future.result(timeout=LLM_BATCH_TIMEOUT)
                    # 결과를 원본 리스트에 병합
                    for j, record in enumerate(batch_result):
                        records[batch_start_idx + j] = record
                    print(f"  [배치 {batch_num}] 완료 - 행 {batch_start_idx+1}~{batch_end_idx} ✓", flush=True)
                except TimeoutError:
                    print(f"  [배치 {batch_num}] 실패 - 행 {batch_start_idx+1}~{batch_end_idx} (타임아웃 {LLM_BATCH_TIMEOUT}초) ✗", flush=True)
                except Exception as e:
                    print(f"  [배치 {batch_num}] 실패 - 행 {batch_start_idx+1}~{batch_end_idx} ({type(e).__name__}: {str(e)[:50]}) ✗", flush=True)
        
        print(f"- {len(records)}개 레코드 번역 완료", flush=True)
        return records
    except Exception as e:
        print(f"- 전체 번역 처리 실패: {e}", flush=True)
        return records


def _calculate_batch_param_length(batch, modification_type):
    """배치의 예상 파라미터 길이(문자 수)를 계산합니다.
    
    Args:
        batch: 번역할 레코드 리스트
        modification_type: 'atype' 또는 'btype'
    
    Returns:
        (총_길이, 경고_여부) 튜플
    """
    if not batch:
        return 0, False
    
    fields = TRANSLATION_FIELDS.get(modification_type, [])
    if not fields:
        return 0, False
    
    total_length = 0
    for record in batch:
        for field in fields:
            text = record.get(field, '') or ''
            total_length += len(str(text))
    
    # 임계값 체크 (토큰 추정: 평균 4문자 = 1토큰)
    estimated_tokens = total_length // 4
    is_warning = estimated_tokens > BATCH_LENGTH_WARNING_THRESHOLD
    
    return total_length, is_warning


def _calculate_llm_metrics(batch_items, modification_type):
    """배치 항목의 LLM 호출 메트릭을 계산합니다.
    
    Args:
        batch_items: 배치 프롬프트에 포함될 항목 리스트
        modification_type: 'atype' 또는 'btype'
    
    Returns:
        (총_길이, 예상_토큰, 필드별_길이_dict) 튜플
    """
    if not batch_items:
        return 0, 0, {}
    
    fields = TRANSLATION_FIELDS.get(modification_type, [])
    total_length = 0
    field_lengths = {field: 0 for field in fields}
    
    for item in batch_items:
        for field in fields:
            text = item.get(field, '') or ''
            length = len(str(text))
            total_length += length
            field_lengths[field] += length
    
    # 토큰 추정 (평균 4문자 = 1토큰)
    estimated_tokens = total_length // 4
    
    return total_length, estimated_tokens, field_lengths


def _validate_sql_id_existence(applycrypto_root, modification_type, output_dir=None):
    """SQL ID 검증: table_access_info의 모든 SQL ID가 step1/step2에 존재하는지 확인합니다.
    
    Args:
        applycrypto_root: .applycrypto 루트 경로
        modification_type: 'ThreeStep' 또는 'TypeHandler'
        output_dir: 로그 파일을 저장할 디렉터리 (엑셀 artifacts 디렉터리)
    
    Returns:
        tuple: (validation_result: bool, validation_stats: dict, validation_details: list)
               validation_stats = {
                   'total_table_ids': int,    # table_access_info.json의 총 SQL ID 개수
                   'step1_matched': int,      # step1_query_analysis.json에서 매칭된 개수
                   'step2_matched': int,      # step2_planning.json에서 매칭된 개수 (ThreeStep만)
                   'missing_count': int       # 누락된 개수
               }
               validation_details = [
                   {'table_name': str, 'sql_id': str, 'step1_status': str, 'step2_status': str, 'status': str},
                   ...
               ]
    """
    
    # 검증 상세 정보 저장용
    validation_details = []
    
    def _log_info(msg):
        """로그 메시지를 콘솔에 기록"""
        # print(f"  {msg}", flush=True)
    
    # 검증 통계
    total_table_ids = 0
    step1_matched_count = 0
    step2_matched_count = 0
    total_checked = 0
    total_missing = 0
    
    try:
        # table_access_info.json 로드
        table_access_path = os.path.join(applycrypto_root, 'results', 'table_access_info.json')
        if not os.path.exists(table_access_path):
            _log_info(f"[SQL ID 검증] table_access_info.json을 찾을 수 없습니다: {table_access_path}")
            return True, {'total_table_ids': 0, 'step1_matched': 0, 'step2_matched': 0, 'missing_count': 0}, []
        
        with open(table_access_path, 'r', encoding='utf-8') as f:
            table_access = json.load(f)
        
        # three_step_results 최신 폴더 찾기
        three_root = os.path.join(applycrypto_root, 'three_step_results')
        if not os.path.isdir(three_root):
            _log_info(f"[SQL ID 검증] three_step_results 디렉터리를 찾을 수 없습니다")
            return True, {'total_table_ids': 0, 'step1_matched': 0, 'step2_matched': 0, 'missing_count': 0}, []
        
        ts_dirs = sorted([d for d in os.listdir(three_root) if os.path.isdir(os.path.join(three_root, d))])
        if not ts_dirs:
            _log_info(f"[SQL ID 검증] three_step_results 내 결과 디렉터리를 찾을 수 없습니다")
            return True, {'total_table_ids': 0, 'step1_matched': 0, 'step2_matched': 0, 'missing_count': 0}, []
        
        latest = ts_dirs[-1]
        latest_root = os.path.join(three_root, latest)
        
        # table_access_info에서 테이블별 SQL ID 추출
        # 순수 배열 형식: [{table_name, columns[], sql_queries[]}, ...]
        tables = table_access if isinstance(table_access, list) else []
        
        _log_info(f"[SQL ID 검증] modification_type: {modification_type}")
        
        for table_obj in tables:
            table_name = table_obj.get('table_name', '')
            sql_queries = table_obj.get('sql_queries', [])
            
            if not sql_queries:
                continue
            
            total_table_ids += len(sql_queries)
            
            # 해당 테이블의 step1_query_analysis.json 로드
            step1_path = os.path.join(latest_root, table_name, '*', 'step1_query_analysis.json')
            import glob
            step1_files = glob.glob(step1_path)
            
            step1_query_ids = set()
            step2_query_ids = set()
            
            # Step1에서 query_id 수집
            for s1_file in step1_files:
                try:
                    with open(s1_file, 'r', encoding='utf-8') as f:
                        s1_data = json.load(f)
                        queries = (s1_data.get('result', {}).get('queries') or [])
                        for q in queries:
                            qid = q.get('query_id', '')
                            if qid:
                                step1_query_ids.add(qid)
                except Exception:
                    pass
            
            # Step2에서 sql_query_id 수집 (atype=ThreeStep인 경우만)
            if modification_type == 'ThreeStep':
                step2_path = os.path.join(latest_root, table_name, '*', 'step2_planning.json')
                step2_files = glob.glob(step2_path)
                
                for s2_file in step2_files:
                    try:
                        with open(s2_file, 'r', encoding='utf-8') as f:
                            s2_data = json.load(f)
                            flows = (s2_data.get('result', {}).get('data_flow_analysis', {}).get('flows') or [])
                            for flow in flows:
                                if isinstance(flow, dict) and flow.get('sql_query_id'):
                                    step2_query_ids.add(flow['sql_query_id'])
                    except Exception:
                        pass
            
            # SQL ID별 검증
            for sql_query in sql_queries:
                sql_id = sql_query.get('id', '')
                if not sql_id:
                    continue
                
                total_checked += 1
                
                # 마지막 토큰(점 기준)으로 비교
                def _match_qid(a, b):
                    try:
                        if not a or not b:
                            return False
                        a_last = str(a).rsplit('.', 1)[-1]
                        b_last = str(b).rsplit('.', 1)[-1]
                        return a_last == b_last
                    except Exception:
                        return False
                
                step1_exists = any(_match_qid(sql_id, qid) for qid in step1_query_ids)
                if step1_exists:
                    step1_matched_count += 1
                
                # atype=ThreeStep일 때만 step2 검증, btype=TypeHandler일 때는 N/A
                if modification_type == 'ThreeStep':
                    step2_exists = any(_match_qid(sql_id, qid) for qid in step2_query_ids)
                    if step2_exists:
                        step2_matched_count += 1
                else:
                    step2_exists = 'N/A'
                
                # 로그 출력
                step2_status = '[OK]' if step2_exists == True else ('[NG]' if step2_exists == False else '[N/A]')
                step1_status = '[OK]' if step1_exists else '[NG]'
                
                # 검증 기준: 
                # - atype(ThreeStep): step1과 step2 모두 확인
                # - btype(TypeHandler): step1만 확인 (step2는 N/A)
                if modification_type == 'ThreeStep':
                    # atype: step1과 step2 모두 확인
                    if not step1_exists or not step2_exists:
                        total_missing += 1
                        status = "MISSING"
                    else:
                        status = "OK"
                else:
                    # btype: step1만 확인
                    if not step1_exists:
                        total_missing += 1
                        status = "MISSING"
                    else:
                        status = "OK"
                
                # CSV 행 추가 및 상세 데이터 저장
                validation_details.append({
                    'table_name': table_name,
                    'sql_id': sql_id,
                    'step1_status': step1_status,
                    'step2_status': step2_status,
                    'status': status
                })
        
        _log_info(f"[SQL ID 검증] 총 검사: {total_checked}개, 누락: {total_missing}개")
        
        validation_stats = {
            'total_table_ids': total_table_ids,
            'step1_matched': step1_matched_count,
            'step2_matched': step2_matched_count if modification_type == 'ThreeStep' else 0,
            'missing_count': total_missing
        }
        
        return total_missing == 0, validation_stats, validation_details
    
    except Exception as e:
        _log_info(f"[SQL ID 검증] 검증 중 오류: {e}")
        return False, {'total_table_ids': 0, 'step1_matched': 0, 'step2_matched': 0, 'missing_count': 0}, []
