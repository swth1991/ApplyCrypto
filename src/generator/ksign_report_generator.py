"""
KSIGN 호출 예측보고서 생성기
(docs\\ksign_call_estimation_report.md 참조)
"""

import ast
import difflib
import json
import logging
import os
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from parser.java_ast_parser import JavaASTParser
from analyzer.callgraph_endpoint_finder import (
    find_all_endpoints_for_method,
    find_endpoints_that_call_method
)

@dataclass
class EndpointDetail:
    """Step 4: 엔드포인트 상세 정보"""
    method_signature: str  # Controller 메서드 (call_stacks[0])
    end_point: str
    class_path: str  # Controller 클래스 경로
    
    # Mapper 정보 추출용 (call_stacks[-1])
    mapper_method_signature: Optional[str] = None  # Mapper 메서드 (Step 5에서 output type 추출용)
    
    # Call stack 전체 저장 (Step 3-5 추적용)
    call_stack: List[str] = field(default_factory=list)  # 전체 call_stacks


@dataclass
class EndpointResult:
    """Step 5: 엔드포인트 가중치 결과"""
    method_signature: str
    end_point: str
    class_path: str
    input_parameter_type: Optional[str] = None
    output_parameter_type: Optional[str] = None
    data_type: str = 'single'  # 새 필드: 'single' | 'paged_list' | 'unpaged_list'
    input_parameter_type_weight: float = 1.0
    output_parameter_type_weight: float = 1.0
    crypto_weight: float = 0.0


@dataclass
class QueryWeight:
    """Step 3-5: 쿼리 중심 가중치 정보 (여러 엔드포인트 포함)"""
    table_name: str
    query_id: str
    sql_query: str = ""  # SQL 쿼리 (페이징 시나리오 판단용)
    input_fields_count: int = 0
    output_fields_count: int = 0
    endpoints: List[str] = field(default_factory=list)  # Step 3: Controller 메서드 시그니처 리스트
    call_stacks_map: Dict[str, List[str]] = field(default_factory=dict)  # 각 endpoint의 전체 call_stack (controller → mapper → ...)
    endpoint_details: List[EndpointDetail] = field(default_factory=list)  # Step 4: 엔드포인트 상세 정보
    endpoint_results: List[EndpointResult] = field(default_factory=list)  # Step 5: 엔드포인트 가중치 결과
    # === Step 3에서 쿼리 분석 정보 저장 (Step 5에서 JavaASTParser 대신 사용) ===
    input_parameter_type: Optional[str] = None  # step1_query_analysis.parameter_type (Controller 첫 파라미터)
    output_parameter_type: Optional[str] = None  # step1_query_analysis.result_type (Mapper 반환 타입)
    inferred_data_type: str = 'single'  # === NEW: result_type으로부터 추론된 data_type ('single' | 'paged_list' | 'unpaged_list') ===


@dataclass
class CryptoWeight:
    """최종 가중치 정보 (각 메서드별 1행) - LLM 응답 형식과 통일"""
    table_name: str
    query_id: str
    method_signature: str
    end_point: str
    class_path: str
    input_fields_count: int
    output_fields_count: int
    input_parameter_type: Optional[str] = None
    output_parameter_type: Optional[str] = None
    # === LLM 응답과 통일된 필드 (Anyframe) ===
    data_type: str = 'single'  # 'single' | 'paged_list' | 'unpaged_list'
    loop_depth: int = 0  # 0-2
    loop_structure: str = ''
    multiplier: str = '1'  # 루프 배수: single(loop_depth=0)='1', loop_depth≥1=실제 배수식 (e.g., "items.size()", "orders.size() × items.size()")
    dep0_crypto_count: int = 0  # loop 밖 crypto 호출 수 (루프 깊이=0)
    dep1_crypto_count: int = 0  # depth=1 loop 안 crypto 호출 수
    dep2_crypto_count: int = 0  # depth=2 loop 안 crypto 호출 수
    # === Spring 호환성 필드 ===
    input_parameter_type_weight: float = 1.0
    output_parameter_type_weight: float = 1.0
    crypto_weight: float = 0.0  # 최종 가중치 (base_weight × ksignutil_count × data_type_multiplier)
    access_cnt: int = 0  # 런타임 호출 횟수


@dataclass
class EndpointAccess:
    """Endpoint 호출빈도 정보"""
    end_point: str
    access_cnt: int


class KSIGNReportGenerator:
    """KSIGN 호출 예측보고서 생성기"""
    
    CARDINALITY_WEIGHTS = {
        'multiple': 10.0,   # List, Collection
        'single': 1.0,      # VO, DTO
        'none': 0.0,        # void
    }
    
    # New Weight Policy (Spring & Anyframe unified)
    # data_type에 따른 가중치 (LLM이 detect한 데이터 타입)
    DATA_TYPE_WEIGHTS = {
        'single': 1.0,            # Single record: weight = outside counts
        'paged_list': 20.0,       # Paged list: weight = 20 × (inside counts)
        'unpaged_list': 100.0,    # Unpaged list: weight = 100 × (inside counts)
    }
      
    # Pagination detection patterns and keywords
    PAGINATION_TYPE_PATTERN = re.compile(r'\bpage(?:list)?<', re.IGNORECASE)
    PAGINATION_KEYWORDS = frozenset({'page', 'paging', 'pagenum', 'pageno', 'pagesize'})
    SQL_PAGINATION_KEYWORDS = frozenset({'LIMIT', 'OFFSET', 'ROWNUM'})
    COLLECTION_INDICATORS = frozenset({'list', 'collection', '[]', 'set'})
    
    def __init__(self, config, applycrypto_dir: Optional[str] = None):
        """
        생성기 초기화
        
        Args:
            config: Configuration 객체 (또는 dict - 호환성 유지)
            applycrypto_dir: .applycrypto 디렉토리 절대 경로 (선택, 없으면 자동 계산)
        """
        # Logger 초기화 (cli_controller의 "applycrypto" logger 사용)
        self.logger = logging.getLogger("applycrypto")
        
        # Configuration 객체 또는 딕셔너리 처리
        self.logger.debug(f"KSIGNReportGenerator.__init__ 시작 (config 타입: {type(config).__name__})")
        
        if isinstance(config, dict):
            self.config = config
            self.logger.debug("config는 이미 dict 형태")
        else:
            # Configuration 객체 → 딕셔너리 변환
            if hasattr(config, 'model_dump'):
                # mode='python'을 사용하여 중첩된 Pydantic 객체도 dict로 변환
                self.config = config.model_dump(mode='python', exclude_none=False)
                self.logger.debug("config.model_dump(mode='python', exclude_none=False) 호출 성공")
            else:
                self.config = dict(config)
                self.logger.debug("dict(config) 호출")
        
        # applycrypto_dir 결정 (전달받은 값 또는 target_project/.applycrypto)
        if applycrypto_dir:
            self.applycrypto_dir = Path(applycrypto_dir)
        else:
            target_project = self.config.get("target_project", "")
            self.applycrypto_dir = Path(target_project) / ".applycrypto"
        
        self.target_project = Path(self.config.get("target_project", ""))
        self.results_dir = self.applycrypto_dir / "results"
        self.artifacts_dir = self.applycrypto_dir / "artifacts"
        self.prompt_logs_dir = self.artifacts_dir / "prompt_logs"  # LLM 프롬프트 로깅용
        self.sanity_reports_dir = self.artifacts_dir / "sanity_reports"  # sanity check 요약 리포트용
        
        # 타임스탐프 기반 실행 식별자
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        
        # 필수 디렉토리 생성
        self.results_dir.mkdir(parents=True, exist_ok=True)
        self.artifacts_dir.mkdir(parents=True, exist_ok=True)
        self.prompt_logs_dir.mkdir(parents=True, exist_ok=True)  # prompt_logs 디렉토리 생성
        self.sanity_reports_dir.mkdir(parents=True, exist_ok=True)
        
        # 데이터 저장소
        self.table_access: List[Dict[str, Any]] = []
        self.query_analysis: Dict[str, Any] = {}
        self.query_weights: List[QueryWeight] = []   # 쿼리 중심 분석 데이터 (Step 3-5)
        self.crypto_weights: List[CryptoWeight] = [] # 평탄화된 최종 리포트 행 데이터 (Step 7)
        self.endpoint_access: List[EndpointAccess] = []
        self.endpoint_weights: Dict[str, Any] = {}   # 엔드포인트별 집계 데이터 (Step 8)
        self.call_graph: Dict[str, Any] = {}
        
        # 성능 최적화용 인덱스 및 캐시
        self.class_name_to_file_index: Dict[str, str] = {} # 클래스명 -> 소스 파일 경로
        self.class_info_cache: Dict[str, Optional[Any]] = {} # 메서드 시그니처 -> 분석된 클래스 정보
        
        # Java AST Parser 초기화 (tree-sitter 기반)
        self.java_ast_parser = None
        try:
            self.java_ast_parser = JavaASTParser()
        except Exception as e:
            self.logger.warning(f"JavaASTParser 초기화 실패: {e}")
        
        # config에서 필드 추출
        try:
            self.framework_type = self.config.get("framework_type", "Spring")
            
            # artifact_generation 로드 (dict 또는 객체 모두 처리)
            artifact_gen = self.config.get("artifact_generation")
            
            # artifact_generation이 dict가 아니면 (Pydantic 객체면) dict로 변환
            if artifact_gen and not isinstance(artifact_gen, dict):
                if hasattr(artifact_gen, 'model_dump'):
                    artifact_gen = artifact_gen.model_dump(exclude_none=False)
                else:
                    artifact_gen = vars(artifact_gen) if hasattr(artifact_gen, '__dict__') else {}
            
            artifact_gen = artifact_gen or {}
            
            # artifact_generation 로드 완료
            self.logger.debug(f"artifact_generation 로드 완료 (키: {list(artifact_gen.keys())})")
            
            # 이제 artifact_gen은 dict이므로안전하게 접근 가능 (항상 list 보장)
            self.ksignutil_patterns = artifact_gen.get("ksignUtils_pattern", [])
            self.policy_ids = artifact_gen.get("policyId", [])
            
            # Step 1: 초기화 완료 메시지 (console + log file)
            self.logger.info(f"Step 0 (Config 로드): ksignutil_patterns {len(self.ksignutil_patterns)}개")
                
        except (KeyError, AttributeError) as e:
            self.logger.error(f"KSIGNReportGenerator 초기화 실패: {e}")
            raise
    

    def _validate_step1_config(self) -> bool:
        """Step 1: config.json의 필수 정보 검증
        
        검증 항목:
        1. framework_type이 "Anyframe"으로 시작하는지 확인
        2. ksignutils_pattern과 policyId가 있는지 확인
        3. 필수 필드가 모두 있는지 확인
        
        Returns:
            bool: 모든 검증을 통과하면 True, 실패하면 False
        """
        # 검증 1: framework_type이 "Anyframe"으로 시작하는지 확인
        if not self.framework_type.startswith("Anyframe"):
            self.logger.error(f"framework_type이 'Anyframe'으로 시작하지 않음 (현재값: {self.framework_type})")
            return False
        
        # 검증 2: ksignutils_pattern 또는 policyId 존재 여부
        if not self.ksignutil_patterns and not self.policy_ids:
            self.logger.error("ksignUtils_pattern 또는 policyId가 정의되지 않음")
            self.logger.error(f"  파일: config.json, 필드: artifact_generation.ksignUtils_pattern 또는 policyId")
            return False
        
        # 검증 4: 필수 config 필드 확인
        required_fields = ["target_project", "artifact_generation", "modification_type"]
        missing_fields = []
        for field in required_fields:
            if field not in self.config:
                missing_fields.append(field)
        
        if missing_fields:
            self.logger.error(f"config.json에 필수 필드가 없음: {', '.join(missing_fields)}")
            return False
        
        return True
    

    def load_table_access(self, file_path: Optional[str] = None) -> bool:
        """table_access_info.json 로드 및 access_files 인덱스 생성"""
        if file_path is None:
            file_path = self.results_dir / "table_access_info.json"
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 리스트 또는 'tables' 키를 가진 딕셔너리 지원
            if isinstance(data, dict):
                self.table_access = data.get('tables', [])
            else:
                self.table_access = data
            
            # === 성능 개선: access_files 기반 인덱스 생성 ===
            # table_access_info.json의 모든 access_files[] 수집하여 클래스명→파일경로 매핑 생성
            for table_entry in self.table_access:
                access_files = table_entry.get('access_files', [])
                for file_path_str in access_files:
                    # 파일 경로에서 클래스명 추출
                    # 예: "src/main/java/com/example/EmployeeDao.java" → "EmployeeDao"
                    file_obj = Path(file_path_str)
                    class_name = file_obj.stem  # 파일명 (확장자 없음)
                    
                    # 경로 검증: 파일 경로가 target_project로 시작하는지 확인
                    target_root_str = str(self.target_project).replace("\\", "/")
                    normalized_file_path = file_path_str.replace("\\", "/")
                    
                    if not normalized_file_path.startswith(target_root_str):
                        self.logger.error(f"파일 경로가 target_project와 불일치: {file_path_str} (기준: {target_root_str})")
                    
                    # 인덱스에 저장 (첫 번째 발견한 경로만 저장)
                    if class_name and class_name not in self.class_name_to_file_index:
                        self.class_name_to_file_index[class_name] = file_path_str
            
            self.logger.info(f"Step 1 (Table Access 로드): {len(self.table_access)}개 테이블, {len(self.class_name_to_file_index)}개 클래스 인덱싱")
            self.logger.debug(f"table_access_info.json 로드 완료: {list(self.class_name_to_file_index.keys())[:5]}...")
            return True
        except Exception as e:
            self.logger.error(f"table_access_info.json 로드 실패: {e}")
            return False
    

    def _split_signature(self, signature: Optional[str]) -> Tuple[str, str]:
        """메소드 시그니처를 클래스명과 메소드명으로 분리"""
        if not signature:
            return "", ""
        if '.' not in signature:
            return signature, ""
        parts = signature.rsplit('.', 1)
        return parts[0], parts[1]


    def extract_method_signature(self, call_stack: List[str]) -> str:
        """Call stack에서 메소드 시그니처 추출 (첫 번째 항목) - 배열 정규화"""
        if call_stack and len(call_stack) > 0:
            sig = call_stack[0]
            # 정규화: 배열이면 첫 요소, 아니면 그대로
            if isinstance(sig, list):
                return sig[0] if sig else ""
            return sig if sig else ""
        return ""
    

    def calculate_crypto_weights(self) -> bool:
        """
        Step 3: QueryID별 암복호화 가중치 계산
        
        쿼리 중심 구조: 같은 쿼리가 여러 엔드포인트에서 호출될 수 있으므로,
        각 쿼리별로 endpoint[] 배열 생성
        
        필드 개수 계산:
        - input_fields_count: input_mapping.crypto_fields의 컬럼 개수
        - output_fields_count: output_mapping.crypto_fields의 컬럼 개수
        """
        # 테이블별 쿼리 분석 데이터 구성 (대소문자 무시를 위해 uppercase/lowercase로 키 구성)
        analysis_by_table = {}
        for table_result in self.query_analysis.get('results', []):
            table_name = table_result.get('table_name', '').upper()
            analysis_by_table[table_name] = {
                str(q.get('query_id', '')).lower(): q 
                for q in table_result.get('queries', [])
            }
        
        # 쿼리별 가중치 (table_name + query_id를 key로)
        query_weight_dict = {}  # key: (table_name.upper(), query_id.lower())
        
        # table_access_info.json의 각 테이블 처리
        for table_entry in self.table_access:
            raw_table_name = table_entry.get('table_name', '')
            table_name_upper = raw_table_name.upper()
            query_list = table_entry.get('sql_queries', [])
            analysis_queries = analysis_by_table.get(table_name_upper, {})
            
            for query in query_list:
                query_id = query.get('id', '')
                query_id_lower = query_id.lower()
                sql_query = query.get('sql', '')
                call_stacks = query.get('call_stacks', [])
                
                # 쿼리 분석 데이터 찾기 - 시그니처 매칭 (소문자로 변환하여 매칭)
                # 정확히 일치하거나, 후방 일치 (패키지명 차이 해결)
                query_analysis = analysis_queries.get(query_id_lower)
                if not query_analysis:
                    # 후방 일치 확인 (예: TestMapper.findAll vs com.example.TestMapper.findAll)
                    for q_id, q_data in analysis_queries.items():
                        if q_id.endswith("." + query_id_lower) or query_id_lower.endswith("." + q_id):
                            query_analysis = q_data
                            break
                            
                if not query_analysis:
                    continue
                
                input_mapping = query_analysis.get('input_mapping', {})
                output_mapping = query_analysis.get('output_mapping', {})
                
                input_crypto_fields = input_mapping.get('crypto_fields', [])
                output_crypto_fields = output_mapping.get('crypto_fields', [])
                
                # crypto_fields 컬럼 개수 (사용 횟수가 아닌 정의된 필드 개수)
                input_fields_count = len(input_crypto_fields)
                output_fields_count = len(output_crypto_fields)
                
                # 이 쿼리에 대한 모든 엔드포인트 수집
                key = (table_name_upper, query_id_lower)
                if key not in query_weight_dict:
                    # === NEW: result_type을 query (table_access_info)에서 추출 및 data_type 추론! ===
                    # query는 table_entry.get('sql_queries', [])에서 나온 것이고,
                    # strategy_specific을 가지고 있습니다 (table_access_info.json에서)
                    output_parameter_type = None
                    inferred_data_type = 'single'  # 기본값
                    
                    # ✅ query['strategy_specific']에서 result_type 추출 (query_analysis 아님!)
                    strategy_specific = query.get('strategy_specific', {})
                    if strategy_specific:
                        output_parameter_type = strategy_specific.get('result_type')
                    
                    # Fallback: strategy_specific에 result_type이 없으면 AST parser로 추출
                    if not output_parameter_type and strategy_specific:
                        mapper_method_name = strategy_specific.get('method_name', '')
                        mapper_file_path = strategy_specific.get('file_path', '')
                        if mapper_method_name and mapper_file_path:
                            # AST parser를 통해 mapper 메서드의 return type 추출
                            _, output_parameter_type = self.extract_parameter_types(mapper_file_path, mapper_method_name)
                    
                    # === data_type 추론: Mapper의 return type 기반 (우선순위 1) ===
                    # 원칙: return type으로 데이터 구조 결정, SQL의 LIMIT는 보조 정보만 사용
                    if output_parameter_type:
                        output_lower = output_parameter_type.lower()
                        
                        # 1. Page, PageList 타입 → paged_list (반드시 페이징)
                        if 'page<' in output_lower or 'pagelist<' in output_lower or 'pagelist' in output_lower:
                            inferred_data_type = 'paged_list'
                        # 2. List, Collection, 배열 타입 → unpaged_list (전체 목록, SQL의 LIMIT과 무관)
                        elif 'list<' in output_lower or 'collection<' in output_lower or '[]' in output_lower:
                            inferred_data_type = 'unpaged_list'
                        # 3. 그 외 (DTO, primitive 등) → single
                        else:
                            inferred_data_type = 'single'
                    else:
                        # output_parameter_type을 알 수 없는 경우: SQL의 LIMIT/OFFSET을 보조 힌트로 사용
                        sql_upper = query.get('sql', '').upper()
                        if 'LIMIT' in sql_upper or 'OFFSET' in sql_upper or 'ROWNUM' in sql_upper:
                            # SQL에 LIMIT/OFFSET이 있으면서 return type이 불명확 → paged_list 추정
                            inferred_data_type = 'paged_list'
                    
                    query_weight_dict[key] = {
                        'table_name': raw_table_name, # 원본 테이블명 유지
                        'query_id': query_id,         # 원본 쿼리ID 유지
                        'input_fields_count': input_fields_count,
                        'output_fields_count': output_fields_count,
                        'endpoints': [],
                        'call_stacks_map': {},  # controller_method → call_stack 매핑
                        'input_parameter_type': None,  # Step 5에서 추출
                        'output_parameter_type': output_parameter_type,  # ✅ result_type 저장
                        'inferred_data_type': inferred_data_type  # ✅ NEW: 추론된 data_type
                    }
                
                # 모든 call_stack 처리
                if call_stacks:
                    for call_stack in call_stacks:
                        method_signature = self.extract_method_signature(call_stack)
                        if method_signature and method_signature not in query_weight_dict[key]['endpoints']:
                            query_weight_dict[key]['endpoints'].append(method_signature)
                            # call_stack 맵에 저장 (controller_method → 전체 call_stack)
                            query_weight_dict[key]['call_stacks_map'][method_signature] = call_stack
                else:
                    # call_stacks가 없는 경우 fallback: query_id 자체를 endpoint로 등록 (Step 4에서 역추적 시도)
                    if query_id and query_id not in query_weight_dict[key]['endpoints']:
                        query_weight_dict[key]['endpoints'].append(query_id)
                        query_weight_dict[key]['call_stacks_map'][query_id] = [query_id]
        
        # QueryWeight 객체 생성
        for query_data in query_weight_dict.values():
            # SQL 쿼리 찾기
            sql_query = ""
            for table_entry in self.table_access:
                if table_entry.get('table_name') == query_data['table_name']:
                    for query in table_entry.get('sql_queries', []):
                        if query.get('id') == query_data['query_id']:
                            sql_query = query.get('sql', '')
                            break
            
            weight = QueryWeight(
                table_name=query_data['table_name'],
                query_id=query_data['query_id'],
                sql_query=sql_query,
                input_fields_count=query_data['input_fields_count'],
                output_fields_count=query_data['output_fields_count'],
                endpoints=query_data['endpoints'],
                call_stacks_map=query_data.get('call_stacks_map', {}),  # call_stacks_map 전달
                input_parameter_type=query_data.get('input_parameter_type'),  # === NEW: 분석된 입력 타입 저장 ===
                output_parameter_type=query_data.get('output_parameter_type'),  # === NEW: 분석된 출력 타입 저장 ===
                inferred_data_type=query_data.get('inferred_data_type', 'single')  # === NEW: 추론된 data_type 전달 ===
            )
            self.query_weights.append(weight)
        
        self.logger.info(f"Step 2 (Calculate Crypto Weights): {len(self.query_weights)}개 항목")
        return True
    

    def load_call_graph(self) -> bool:
        """Call Graph JSON 로드"""
        try:
            call_graph_file = self.applycrypto_dir / "results" / "call_graph.json"
            if not call_graph_file.exists():
                self.logger.warning(f"call_graph.json 없음: {call_graph_file}")
                return False
            
            with open(call_graph_file, 'r', encoding='utf-8') as f:
                self.call_graph = json.load(f)
            self.logger.debug(f"call_graph.json 로드 완료: {len(self.call_graph)}개 항목")
            return True
        except Exception as e:
            self.logger.warning(f"call_graph.json 로드 실패: {e}")
            return False
    

    def find_endpoints_for_method(self, method_signature: str) -> List[str]:
        """
        Call Graph를 사용하여 메소드에 대응하는 엔드포인트 찾기
        """
        if not self.call_graph:
            return []
        
        all_paths = []
        
        # Step 1: endpoints에서 직접 검색 (Controller 메소드)
        results = find_all_endpoints_for_method(method_signature, self.call_graph)
        for res in results:
            paths = res.get("path", [])
            if isinstance(paths, list):
                all_paths.extend(paths)
            else:
                all_paths.append(paths)
        
        # Step 2: call_trees에서 역추적 (이 메소드를 호출하는 엔드포인트 찾기)
        matching_endpoints = find_endpoints_that_call_method(
            method_signature, self.call_graph
        )
        for ep in matching_endpoints:
            paths = ep.get("path", [])
            if isinstance(paths, list):
                all_paths.extend(paths)
            else:
                all_paths.append(paths)
        
        # 중복 제거 및 정렬
        return sorted(list(set(all_paths)))
    

    def find_java_file_for_class(self, class_name: str) -> Optional[str]:
        """
        Target 프로젝트에서 클래스에 해당하는 Java 파일 경로 찾기
        
        최적화 전략 (3단계):
        1. access_files 인덱스 확인 (table_access_info.json 기반) - 빠름 O(1)
        2. 캐시 확인 (이전 검색 결과)
        3. rglob으로 전체 프로젝트 검색 (느림 O(n)) - 최후의 수단
        
        Args:
            class_name: 클래스명 (e.g., "Employee", "EmployeeController")
        
        Returns:
            str: 프로젝트 루트 기준 상대 경로 (e.g., "src/main/java/com/example/Employee.java")
                 찾지 못하면 None
        """
        if not self.target_project.exists():
            return None
        
        # === Step 1: access_files 인덱스에서 먼저 확인 (빠름!) ===
        if class_name in self.class_name_to_file_index:
            return self.class_name_to_file_index[class_name]
        
        # === Step 2+3: rglob으로 검색 (느림, 캐시에 저장) ===
        # table_access_info.json에 없는 경우 (call_graph의 클래스 등)
        for java_file in self.target_project.rglob("*.java"):
            if java_file.stem == class_name:
                try:
                    # 프로젝트 루트 기준 상대 경로로 정규화
                    rel_path = java_file.relative_to(self.target_project)
                    file_path_str = str(rel_path).replace("\\", "/")
                    # 캐시에 저장 (다음 검색에서 빠르게)
                    self.class_name_to_file_index[class_name] = file_path_str
                    return file_path_str
                except ValueError:
                    continue
        
        return None
    

    def get_method_info(self, class_name: str, method_name: str, file_path: Optional[Path] = None):
        """ClassInfo에서 특정 메서드의 정보를 가져옴"""
        if not file_path:
            java_file = self.find_java_file_for_class(class_name)
            if not java_file:
                return None
            file_path = self.target_project / java_file
            
        if not file_path.exists():
            return None
            
        try:
            if self.java_ast_parser:
                tree, _ = self.java_ast_parser.parse_file(str(file_path), remove_comments=True)
                if tree:
                    class_infos = self.java_ast_parser.extract_class_info(tree, file_path)
                    for class_info in class_infos:
                        if class_info.name == class_name:
                            for method in class_info.methods:
                                if method.name == method_name:
                                    return method
        except Exception as e:
            self.logger.debug(f"Error parsing {file_path}: {e}")
            
        return None

    def _is_collection_type(self, return_type_lower: str) -> bool:
        """
        반환 타입이 Collection 타입인지 정확하게 판단
        
        부분 문자열이 아닌 정확한 패턴 매칭 수행
        예: "List<T>", "ArrayList<T>", "Collection<T>" 등
        반대로: "guestlist", "blacklist" 등은 감지 안 됨
        
        Args:
            return_type_lower: 소문자로 변환된 반환 타입 문자열
        
        Returns:
            bool: Collection 타입 여부
        """
        # Collection 타입 키워드와 대각괄호 조합으로 확인
        collection_patterns = [
            'list<', 'collection<', 'set<', 'deque<', 'queue<',
            'arraylist<', 'linkedlist<', 'hashset<', 'linkedhashset<',
            'map<', 'hashmap<', 'linkedhashmap<', 'treemap<',
            'vector<', 'stack<',  # Java legacy types
            '[]',  # 배열 표기
        ]
        
        return any(pattern in return_type_lower for pattern in collection_patterns)

    def _has_pagination_param(self, method_params: Optional[List]) -> bool:
        """
        파라미터에 페이징 관련 정보가 있는지 확인
        
        주의: "page"는 부분 문자열이 아닌 정확한 매칭만 수행
        예: "pageSize"는 'paged'로 감지되지만, "homePageFlag"는 감지되지 않음
        
        Args:
            method_params: 메서드 파라미터 리스트
        
        Returns:
            bool: 페이징 파라미터 존재 여부
        """
        if not method_params:
            return False
        
        for param in method_params:
            # ClassInfo.Parameter 객체와 dict 모두 처리
            if isinstance(param, dict):
                param_type = str(param.get('type', '')).lower()
                param_name = str(param.get('name', '')).lower()
            else:
                # ClassInfo.Parameter 객체
                param_type = str(getattr(param, 'type', '')).lower()
                param_name = str(getattr(param, 'name', '')).lower()
            
            # Pageable 타입 확인 (정확한 매칭)
            if 'pageable' in param_type:
                return True
            
            # 파라미터명 확인 (정확한 단어 경계 매칭)
            # "page"를 포함하되, "page"로 시작하거나 "_page"로 구분되어야 함
            if any(
                param_name == keyword or  # 정확한 매칭
                param_name.startswith(keyword + '_') or  # page_num, page_size
                '_' + keyword + '_' in param_name or  # xxx_page_yyy
                param_name.startswith(keyword.capitalize()) or  # PageNum
                '_' + keyword.capitalize() in param_name  # _PageNum
                for keyword in self.PAGINATION_KEYWORDS
            ):
                return True
        
        return False


    def _sanitize_json_string(self, text: str) -> str:
        """JSON 파싱을 위해 제어 문자를 제거합니다.
        
        Args:
            text: 정제할 문자열
            
        Returns:
            제어 문자가 제거된 문자열
        """
        result = []
        for char in text:
            code = ord(char)
            # 제어 문자 제외 (탭, 줄바꿈, 캐리지 리턴은 유지)
            if code < 0x20 and code not in (0x09, 0x0A, 0x0D):  # \t, \n, \r
                result.append(' ')
            else:
                result.append(char)
        return ''.join(result)

    def _extract_json_from_response(self, text: str) -> str:
        """응답에서 JSON 부분을 추출합니다.
        
        마크다운 형식이나 설명 텍스트가 포함된 응답에서
        JSON 배열 또는 객체만 정확히 추출합니다.
        
        Args:
            text: 응답 텍스트
            
        Returns:
            추출된 JSON 문자열
        """
        # 마크다운 코드블록 제거
        text = re.sub(r'```(?:json)?\s*\n?', '', text)
        text = re.sub(r'```\s*', '', text)
        
        # JSON 배열의 시작 [ 찾기
        array_start = text.find('[')
        if array_start != -1:
            # JSON 배열 형식 → 끝에서부터 ] 찾기
            end_idx = text.rfind(']')
            if end_idx != -1 and end_idx > array_start:
                json_str = text[array_start:end_idx + 1]
                # 추출한 JSON 정제
                return self._sanitize_json_string(json_str)
        
        # JSON 객체 { } 찾기
        obj_start = text.find('{')
        if obj_start != -1:
            # 일치하는 닫는 괄호를 찾음 (중첩된 {} 고려)
            bracket_count = 0
            in_string = False
            escape_next = False
            
            for i in range(obj_start, len(text)):
                char = text[i]
                
                if escape_next:
                    escape_next = False
                    continue
                
                if char == '\\':
                    escape_next = True
                    continue
                
                if char == '"' and not escape_next:
                    in_string = not in_string
                    continue
                
                if in_string:
                    continue
                
                if char == '{':
                    bracket_count += 1
                elif char == '}':
                    bracket_count -= 1
                    if bracket_count == 0:
                        json_str = text[obj_start:i + 1]
                        return self._sanitize_json_string(json_str)
        
        # JSON을 찾지 못한 경우 원본 반환
        return text.strip()

    def determine_data_type(
        self,
        return_type: Optional[str],
        method_params: Optional[List] = None,
        sql_query: str = ""
    ) -> str:
        """
        메서드의 반환 타입 및 파라미터 분석하여 data_type 결정
        
        원칙: Return Type 우선, SQL은 보조 정보만 사용
        
        Rules (우선순위):
        1. Page<T>, PageList<T> → "paged_list" (반드시 페이징)
        2. Pageable 파라미터 존재 → "paged_list" (반드시 페이징)
        3. List<T>, Collection<T>, [] → "unpaged_list" (전체 목록, SQL의 LIMIT과 무관)
        4. SQL에 LIMIT/OFFSET 있고 return_type 불명확 → "paged_list" (추정)
        5. 기타 (DTO, primitive, void) → "single"
        
        Args:
            return_type: 메서드 반환 타입 (e.g., "Page<Employee>", "List<Employee>")
            method_params: 메서드 파라미터 리스트
            sql_query: 쿼리 문자열
        
        Returns:
            str: "single" | "paged_list" | "unpaged_list"
        """
        if not return_type:
            # return_type이 없는 경우: SQL 기반 추정만 가능
            sql_upper = sql_query.upper() if sql_query else ""
            if sql_upper and any(kw in sql_upper for kw in self.SQL_PAGINATION_KEYWORDS):
                return 'paged_list'
            return 'single'
        
        # Cache lowercased/uppercased strings once
        return_type_lower = return_type.lower()
        
        # Check 1: Page<T>, PageList<T> 타입 → 반드시 paged_list
        if self.PAGINATION_TYPE_PATTERN.search(return_type):
            return 'paged_list'
        
        # Check 2: Pageable 파라미터 확인 → 반드시 paged_list
        if self._has_pagination_param(method_params):
            return 'paged_list'
        
        # Check 3: List<T>, Collection<T>, [] → 반드시 unpaged_list
        # (SQL의 LIMIT/OFFSET은 단순 쿼리 최적화일 뿐, 데이터 구조와 무관)
        if self._is_collection_type(return_type_lower):
            return 'unpaged_list'
        
        # Check 4: 기타 타입이면서 SQL에 LIMIT/OFFSET → 추정 paged_list
        sql_upper = sql_query.upper() if sql_query else ""
        if sql_upper and any(kw in sql_upper for kw in self.SQL_PAGINATION_KEYWORDS):
            return 'paged_list'
        
        # 기본값: single
        return 'single'


    def enrich_endpoint_and_class_path(self) -> bool:
        """
        Step 4: Endpoint 정보 수집 (endpoint_details[] 채우기)
        
        Call Graph 기반으로 method_signature로부터 엔드포인트 추출하고
        call_stack을 통해 controller의 ClassInfo 추출
        mapper_method_signature는 Step 5에서 output type 추출용으로만 저장
        
        - call_stack[0]: Controller (method_signature)
        - call_stack[-1]: Mapper (output type 추출용)
        """
        # Call Graph 로드
        if not self.load_call_graph():
            self.logger.warning("Call Graph 없이 진행 (엔드포인트 정보 추가 불가)")
        
        details_found = 0
        
        # query_weights의 각 항목 처리
        for query_weight in self.query_weights:
            # 각 endpoint(method_signature)에 대해 endpoint_details 생성
            for method_signature in query_weight.endpoints:
                # 모든 매핑된 경로 가져오기
                endpoints = []
                if self.call_graph:
                    endpoints = self.find_endpoints_for_method(method_signature)
                
                # 매핑된 경로가 없으면 빈 문자열 하나라도 생성
                if not endpoints:
                    endpoints = [""]
                
                # 각 경로마다 별도의 상세 정보 생성 (행 분리)
                for path in endpoints:
                    # === call_stack 정보 추출 ===
                    call_stack = query_weight.call_stacks_map.get(method_signature, [])
                    
                    # === Controller ClassPath 추출 ===
                    class_path = ""
                    if call_stack and len(call_stack) > 0:
                        controller_signature = call_stack[0]
                        controller_class_name = controller_signature.split('.')[0] if '.' in controller_signature else ""
                        class_path = self.class_name_to_file_index.get(controller_class_name, "")
                    
                    # === Mapper 메서드 시그니처 추출 ===
                    mapper_method_signature = call_stack[-1] if call_stack else None
                    
                    # EndpointDetail 생성
                    endpoint_detail = EndpointDetail(
                        method_signature=method_signature,
                        end_point=path,  # 개별 경로 할당
                        class_path=class_path,
                        mapper_method_signature=mapper_method_signature,
                        call_stack=call_stack
                    )
                    query_weight.endpoint_details.append(endpoint_detail)
                    details_found += 1
                details_found += 1
        
        self.logger.info(f"Step 3 (Enrich Endpoint): {details_found}개 항목")
        return True
    

    def extract_parameter_types(self, class_path: Optional[str], method_name: str) -> Tuple[Optional[str], Optional[str]]:
        """메소드의 입출력 파라미터 타입 추출 (AST -> Regex fallback)"""
        if not class_path or not method_name:
            return None, None
        
        # Step 1: AST 기반 추출 시도
        class_name = Path(class_path).stem
        full_path = self.target_project / class_path if not Path(class_path).is_absolute() else Path(class_path)
        
        method_info = self.get_method_info(class_name, method_name, full_path)
        if method_info:
            input_type = method_info.parameters[0].type if method_info.parameters else 'void'
            return input_type, method_info.return_type
            
        # Step 2: Regex 기반 Fallback
        try:
            if not full_path.exists():
                return None, None
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 메서드 시그니처 매칭: 반환타입(제네릭포함) + methodName + 파라미터
            # public List<Dto> getName(...) → List<Dto> 추출
            pattern = rf'([\w<>.,\[\]]+)\s+{re.escape(method_name)}\s*\(([^)]*)\)'
            match = re.search(pattern, content)
            if match:
                output_type = match.group(1).strip()
                params = match.group(2).strip()
                input_type = 'void'
                if params:
                    first_param = params.split(',')[0].strip()
                    param_match = re.match(r'^([\w<>.,\[\]]*?)\s+\w+', first_param)
                    input_type = param_match.group(1).strip() if param_match else first_param.split()[0]
                return input_type, output_type
        except Exception as e:
            self.logger.debug(f"Regex extraction failed for {method_name} in {class_path}: {e}")
            
        return None, None
    

    def get_cardinality(self, type_str: Optional[str]) -> str:
        """
        파라미터/반환 타입의 기수(Cardinality) 판별
        
        기수란: 데이터가 반환되는 구조상의 개수
        - 'multiple': 여러 개 항목 반환 가능 (List, Collection, 배열 등)
        - 'single': 단일 항목만 반환 (VO, DTO, primitive 등)
        - 'none': 데이터 반환 없음 (void, null 등)
        
        주의: 기수와 페이징의 차이
        - 기수 'multiple': 메서드 정의상 여러 개를 반환할 수 있음 (List<T>, T[] 등)
        - 페이징: 런타임에 한 번에 가져오는 개수 (조회 쿼리의 LIMIT나 page size)
        - 예: List<Employee> 반환 = 기수 'multiple'이지만,
            페이징으로 10개씩 가져가면 총 10번 호출 필요 (별도 추적 필요)
        
        Args:
            type_str: 파라미터/반환 타입 문자열 (예: 'List', 'String', 'void')
        
        Returns:
            str: 'multiple' | 'single' | 'none'
        """
        if not type_str:
            return 'none'
        
        type_lower = type_str.lower()
        
        # 'none' 판별 (void, null 등)
        if type_str == 'void' or type_lower == 'null':
            return 'none'
        
        # 'multiple' 판별 (Collection 타입)
        collection_keywords = [
            'list', 'arraylist', 'collection', 'set', 'hashset', 'linkedhashset',
            'map', 'hashmap', 'linkedhashmap', 'treemap',
            'queue', 'deque', 'linkedlist', 'stack', 'vector',
            'iterable', 'stream',
            '[]', 'array',  # 배열 표기
        ]
        
        if any(keyword in type_lower for keyword in collection_keywords):
            return 'multiple'
        
        # 기본: 'single' (VO, DTO, primitive 등)
        return 'single'
    

    def calculate_input_weight(self, input_type: Optional[str]) -> float:
        """
        Input Type에 따른 가중치 계산 (Controller 입력 파라미터)
        
        Rule (사용자 요청 기준):
        - MultipartHttpServletRequest 포함 → 1000 (파일 업로드)
        - 그 외 → 1 (단건/다건 입력)
        
        Args:
            input_type: Controller 메서드의 첫 번째 파라미터 타입
        
        Returns:
            float: 입력 가중치
        """
        if not input_type:
            return 1.0
        
        input_lower = input_type.lower()
        
        # Rule 1: MultipartHttpServletRequest (파일 업로드)
        if 'multipart' in input_lower or 'multiparthttp' in input_lower:
            return 1000.0
        
        # Rule 2: 그 외 (단건/다건 입력 모두 1)
        return 1.0
    

    def simplify_type_name(self, type_str: Optional[str]) -> Optional[str]:
        """
        제네릭 타입을 단순화 (Generic parameter만 제거, 외부 타입명은 유지)
        
        예:
        - 'List<QgBoardDto>' → 'List' (외부 타입 유지)
        - 'PageList<Entity>' → 'PageList' (외부 타입 유지)
        - 'Map<String, Object>' → 'Map' (외부 타입 유지)
        - 'QgBoardDto' → 'QgBoardDto' (이미 단순)
        - 'void' → 'void'
        
        주의: 외부 타입(List, PageList 등)은 반드시 유지되어야 데이터 구조 판단 가능
        """
        if not type_str:
            return type_str
        
        # Generic parameter 제거 (< ... > 부분)
        simplified = re.sub(r'<.*?>', '', type_str).strip()
        
        # List<QgBoardDto> → List로 올바르게 변환되는지 검증
        # 만약 generic이 사라져도 외부 타입이 남아있는지 확인
        if not simplified or simplified.endswith('>'):
            # 정규식이 실패하면 대체 방법 시도 (중첩 generic의 경우)
            # 가장 앞의 < 이전 부분만 추출
            bracket_pos = type_str.find('<')
            if bracket_pos > 0:
                simplified = type_str[:bracket_pos].strip()
        
        return simplified
    

    def apply_parameter_type_weights(self) -> bool:
        """Step 5: Endpoint별 파라미터 타입 가중치 부여
        
        핵심 원칙:
        - Input Type: Controller 메서드의 Input Parameter
        - Output Type: Mapper의 Return Type  
        - data_type: Mapper의 Return Type 기반으로 결정
        """
        results_created = 0
        # 분석 캐시: (class_name, method_name) -> input_type
        controller_cache = {}
        
        for query_weight in self.query_weights:
            for endpoint_detail in query_weight.endpoint_details:
                call_stack = endpoint_detail.call_stack
                controller_sig = call_stack[0] if call_stack else endpoint_detail.method_signature
                c_class, c_method = self._split_signature(controller_sig)
                
                # === 1단계: Controller 분석 (Input Type 추출) ===
                if (c_class, c_method) not in controller_cache:
                    method_info = self.get_method_info(c_class, c_method)
                    if method_info:
                        has_multipart = any('multipart' in p.type.lower() for p in method_info.parameters)
                        input_type = 'MultipartHttpServletRequest' if has_multipart else \
                                    (method_info.parameters[0].type if method_info.parameters else 'void')
                    else:
                        input_type, _ = self.extract_parameter_types(endpoint_detail.class_path, c_method)
                        input_type = input_type or 'void'
                    controller_cache[(c_class, c_method)] = input_type
                
                input_type = controller_cache[(c_class, c_method)]
                input_type_simplified = self.simplify_type_name(input_type)
                
                # === 2단계: Mapper 분석 (Output Type 및 data_type 결정) ===
                # ✅ 우선순위 1: query_weight에 저장된 output_parameter_type 사용 (Step 3에서 추출)
                output_type = query_weight.output_parameter_type
                
                if not output_type:
                    # ✅ Fallback: mapper_method_signature로부터 추출 시도
                    mapper_sig = endpoint_detail.mapper_method_signature or controller_sig
                    m_class, m_method = self._split_signature(mapper_sig)
                    mapper_file = self.find_java_file_for_class(m_class)
                    _, output_type = self.extract_parameter_types(mapper_file, m_method)
                
                output_type_simplified = self.simplify_type_name(output_type) if output_type else None
                
                # === 3단계: data_type 결정 ===
                # 우선순위 1: Step 3에서 이미 Mapper의 return type으로 결정한 inferred_data_type 사용
                # (Step 3에서 정확하게 판단했으므로 다시 판단할 필요 없음)
                if query_weight.inferred_data_type and query_weight.inferred_data_type != 'single':
                    data_type = query_weight.inferred_data_type
                else:
                    # Fallback: Step 3에서 result_type을 얻지 못한 경우 Step 5에서 다시 시도
                    data_type = self.determine_data_type(
                        return_type=output_type,  # Mapper의 return type
                        method_params=None,
                        sql_query=query_weight.sql_query if hasattr(query_weight, 'sql_query') else ""
                    )
                
                # output_type_simplified 정의 (이미 위에서 정의됨)
                # output_type_simplified = self.simplify_type_name(output_type) if output_type else None
                
                # 가중치 계산
                input_weight = self.calculate_input_weight(input_type_simplified)
                
                # output_weight: data_type 기반 계산 (새 정책)
                # === UPDATED: 모든 경우에 data_type 기반 weight 적용 (output_type과 무관) ===
                output_weight = self.DATA_TYPE_WEIGHTS.get(data_type, 1.0)
                # 단, void인 경우는 0
                if output_type_simplified and output_type_simplified.lower() == 'void':
                    output_weight = 0.0
                
                crypto_weight = (query_weight.input_fields_count * input_weight) + \
                                (query_weight.output_fields_count * output_weight)

                endpoint_result = EndpointResult(
                    method_signature=endpoint_detail.method_signature,
                    end_point=endpoint_detail.end_point,
                    class_path=endpoint_detail.class_path,
                    input_parameter_type=input_type_simplified,
                    output_parameter_type=output_type_simplified,
                    data_type=data_type,  # 새 필드: data_type 저장
                    input_parameter_type_weight=input_weight,
                    output_parameter_type_weight=output_weight,
                    crypto_weight=crypto_weight
                )
                
                query_weight.endpoint_results.append(endpoint_result)
                results_created += 1
        
        self.logger.info(f"Step 4 (Apply Parameter Weights): {results_created}개 항목")
        return True
    
    
    def save_crypto_weights_json(self, crypto_weights: Optional[List[Dict[str, Any]]] = None) -> bool:
        """암복호화 가중치를 JSON으로 저장
        
        Spring 파이프라인용: crypto_weights가 None이면 self.crypto_weights를 Dict로 변환하여 사용
        Anyframe 파이프라인용: crypto_weights가 제공되면 그 값을 직접 사용
        
        Args:
            crypto_weights: 암복호화 가중치 리스트 (None이면 self.crypto_weights 사용)
        
        Returns:
            bool: 성공 여부
        """
        # Spring 파이프라인: self.crypto_weights 사용
        if crypto_weights is None:
            if not self.crypto_weights:
                self.logger.warning("crypto_weights가 비어있음")
                return True
            # CryptoWeight 객체를 Dict로 변환
            crypto_weights = [asdict(w) for w in self.crypto_weights]
        
        # 공통 로직: Dict 리스트를 JSON으로 저장
        try:
            output_file = self.artifacts_dir / f"crypto_weight_{self.timestamp}.json"
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(crypto_weights, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Step 5 (Save Crypto Weights): {len(crypto_weights)}개 항목")
            return True
        
        except Exception as e:
            self.logger.error(f"crypto_weight.json 저장 실패: {e}")
            return False


    def load_step1_query_analysis_hierarchical(self) -> bool:
        """
        Step 1-2에서 저장된 계층적 구조의 쿼리 분석 결과 로드
        (최신 타임스탐프 폴더 중 데이터가 있는 폴더 선택)
        """
        three_step_dir = self.applycrypto_dir / "three_step_results"
        if not three_step_dir.exists():
            self.logger.error("three_step_results 디렉토리 없음")
            return False
        
        # 모든 타임스탐프 디렉토리를 최신순으로 정렬
        timestamp_dirs = sorted([d for d in three_step_dir.iterdir() if d.is_dir()], reverse=True)
        if not timestamp_dirs:
            self.logger.error("타임스탐프 디렉토리 없음")
            return False
        
        # 데이터가 있는 최신 폴더 찾기
        latest_dir_with_data = None
        for latest_dir in timestamp_dirs:
            # 폴더에 step1_query_analysis.json 파일이 있는지 확인
            query_files = list(latest_dir.glob("*/*/step1_query_analysis.json"))
            if query_files:
                latest_dir_with_data = latest_dir
                break
        
        if not latest_dir_with_data:
            self.logger.error(f"데이터가 있는 타임스탐프 폴더 없음 (검색: {[d.name for d in timestamp_dirs[:5]]})")
            return False
        
        latest_dir = latest_dir_with_data
        self.logger.debug(f"최신 결과 사용: {latest_dir.name}")
        
        # 모든 step1_query_analysis.json 수집
        all_queries = []
        for table_dir in latest_dir.iterdir():
            if not table_dir.is_dir():
                continue
            
            table_name = table_dir.name
            
            for class_dir in table_dir.iterdir():
                if not class_dir.is_dir():
                    continue
                
                query_file = class_dir / "step1_query_analysis.json"
                if query_file.exists():
                    try:
                        with open(query_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            # 쿼리 수집
                            queries = data.get('result', {}).get('queries', [])
                            for query in queries:
                                query['table_name'] = table_name  # 테이블명 추가
                            all_queries.extend(queries)
                    except Exception as e:
                        self.logger.warning(f"{query_file} 로드 실패: {e}")
        
        # query_analysis 구조로 변환
        queries_by_table = {}
        for query in all_queries:
            table = query.get('table_name', {})
            if table not in queries_by_table:
                queries_by_table[table] = []
            queries_by_table[table].append(query)
        
        self.query_analysis = {
            'metadata': {
                'timestamp': datetime.now().isoformat(),
                'step_number': 1,
                'phase': 'query_analysis',
                'source': 'hierarchical'
            },
            'results': [
                {'table_name': table, 'queries': queries}
                for table, queries in queries_by_table.items()
            ]
        }
        
        total_queries = sum(len(r.get('queries', [])) for r in self.query_analysis.get('results', []))
        self.logger.info(f"{len(queries_by_table)} 테이블, {total_queries} 쿼리 로드 완료")
        return True
    

    def save_ksign_report_excel(self) -> bool:
        """Step 11 (Anyframe) / Step 9 (Spring): KSIGN 예측 리포트를 Excel로 저장"""
        try:
            import openpyxl
        except ImportError:
            self.logger.error("openpyxl 라이브러리 필요: pip install openpyxl")
            return False
        
        # 파일명 구성: 프로젝트명 + 날짜 (spec_generator 방식)
        date_str = datetime.now().strftime('%Y%m%d')
        
        tp = Path(self.target_project)
        if not tp.exists():
            raise FileNotFoundError(f"타겟 프로젝트를 찾을 수 없습니다: {self.target_project}")        
        
        output_file = self.artifacts_dir / f"{tp.name}_ksign_call_report_{date_str}.xlsx"
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            self._create_wb_styles(wb)
            self._add_crypto_weight_sheet(wb)
            
            # Framework 타입에 따라 다른 시트 추가
            if self.framework_type.startswith("Anyframe"):
                self._add_anyframe_estimation_detail_sheet(wb)
                self._add_anyframe_final_report_sheet(wb)
            else:
                # Spring 타입 (기본값)
                self._add_spring_final_report_sheet(wb)
                self._add_spring_estimation_detail_sheet(wb)
            
            try:
                wb.save(output_file)
                self.logger.info(f"{output_file.name} 저장 완료")
            except PermissionError:
                output_file_tmp = output_file.with_name(f"{output_file.stem}_tmp.xlsx")
                wb.save(output_file_tmp)
                self.logger.info(f"{output_file_tmp.name} 저장 완료 (기존 파일 잠김)")
            
            return True
        except Exception as e:
            self.logger.error(f"Excel 저장 실패: {e}", exc_info=True)
            self.logger.error(f"Excel 저장 실패: {e}")
            import traceback
            traceback.print_exc()
            return False


    def _create_wb_styles(self, wb):
        """Excel 스타일 정의"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        self.excel_styles = {
            'header_font': Font(name="맑은 고딕", size=10, bold=True),
            'header_fill': PatternFill(start_color="DFDFDF", end_color="DFDFDF", fill_type="solid"),
            'header_align': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'data_font': Font(name="맑은 고딕", size=10),
            'data_align_left': Alignment(horizontal="left", vertical="center"),
            'data_align_center': Alignment(horizontal="center", vertical="center"),
            'thin_border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }


    def _apply_cell_style(self, cell, style_name='data', align='left'):
        """셀에 스타일 적용 공통 함수"""
        if style_name == 'header':
            cell.font = self.excel_styles['header_font']
            cell.fill = self.excel_styles['header_fill']
            cell.alignment = self.excel_styles['header_align']
        else:
            cell.font = self.excel_styles['data_font']
            cell.alignment = self.excel_styles['data_align_left'] if align == 'left' else self.excel_styles['data_align_center']
        cell.border = self.excel_styles['thin_border']


    def _add_crypto_weight_sheet(self, wb):
        """가중치 상수 시트 - 산출 근거 참고용"""
        ws = wb.create_sheet("Crypto Weight")
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        
        row = 2
        ws[f'B{row}'] = "# Output Type Weights"
        ws[f'B{row}'].font = self.excel_styles['header_font']
        row += 1
        
        headers = ["Cardinality", "Weight", "Description"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i+2, value=h)
            self._apply_cell_style(cell, 'header')
        row += 1
        
        desc_map = {'multiple': 'List, Collection', 'single': 'VO, String', 'none': 'void'}
        for card, weight in sorted(self.CARDINALITY_WEIGHTS.items()):
            ws.cell(row=row, column=2, value=card)
            ws.cell(row=row, column=3, value=weight)
            ws.cell(row=row, column=4, value=desc_map.get(card, ''))
            for i in range(3):
                self._apply_cell_style(ws.cell(row=row, column=i+2))
            row += 1

        ws[f'B{row}'] = "# Input Type Weight"; ws[f'B{row}'].font = self.excel_styles['header_font']
        row += 1
        for i, h in enumerate(headers): self._apply_cell_style(ws.cell(row=row, column=i+2, value=h), 'header')
        row += 1
        
        input_weights = [('MultipartHttpServletRequest', 1000, '파일 업로드'), ('List', 10, '다건 입력 처리'), ('Other', 1, '단건 입력 처리')]
        for typ, weight, desc in input_weights:
            ws.cell(row=row, column=2, value=typ); ws.cell(row=row, column=3, value=weight); ws.cell(row=row, column=4, value=desc)
            for i in range(3): self._apply_cell_style(ws.cell(row=row, column=i+2))
            row += 1
            
        # Data Type Weights (New Policy)
        row += 1
        ws[f'B{row}'] = "# DATA TYPE WEIGHTS (Spring & Anyframe unified)"; ws[f'B{row}'].font = self.excel_styles['header_font']
        row += 1
        headers_dtype = ["Data Type", "Multiplier", "Description"]
        for i, h in enumerate(headers_dtype): self._apply_cell_style(ws.cell(row=row, column=i+2, value=h), 'header')
        row += 1
        
        dtype_weights = [('single', 1.0, 'Single record: weight = outside counts'), ('paged_list', 20.0, 'Paged list: weight = 20 × inside counts'), ('unpaged_list', 100.0, 'Unpaged list: weight = 100 × inside counts'), ('★Nested loop★', 10.0, 'Paged or Unpaged llop x 10')]
        for dtype, weight, desc in dtype_weights:
            ws.cell(row=row, column=2, value=dtype); ws.cell(row=row, column=3, value=weight); ws.cell(row=row, column=4, value=desc)
            for i in range(3): self._apply_cell_style(ws.cell(row=row, column=i+2))
            row += 1


    def _add_spring_final_report_sheet(self, wb):
        """Spring 타입 최종 리포트"""
        import openpyxl.utils
        ws = wb.create_sheet("Final Report")
        headers = ['End Point', 'Query Count', 'input_fields_count', 'output_fields_count', 'Total Crypto Weight', 'Jenifer Access Count', 'Estimated KSIGN Calls']
        
        ws.cell(row=1, column=1, value="")
        for i, h in enumerate(headers):
            cell = ws.cell(row=1, column=i+2, value=h)
            self._apply_cell_style(cell, 'header')

        sorted_data = sorted(
            self.endpoint_weights.values(),
            key=lambda x: (
                ((x.get('endpoint') or '') == ''),
                x.get('endpoint') or ''
            )
        )
        for r_idx, data in enumerate(sorted_data, 2):
            # endpoint가 list인 경우 첫 번째 요소 추출 (Excel cell 쓰기 전 정규화)
            endpoint = data['endpoint']
            endpoint_str = endpoint[0] if isinstance(endpoint, list) else (endpoint or '')
            
            values = [endpoint_str, data['query_count'], data.get('input_fields_count', 0), data.get('output_fields_count', 0), data['total_crypto_weight'], data['access_count'], data['estimated_ksign_calls']]
            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                self._apply_cell_style(cell, 'data')
        
        col_widths = [3, 40, 20, 18, 18, 20, 15, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width


    def _add_spring_estimation_detail_sheet(self, wb):
        """Spring 타입 예측 시트 - 쿼리별 상세"""
        import openpyxl.utils
        ws = wb.create_sheet("Ksign Call Estimation")
        headers = ['End Point', 'Method', 'Table', 'Query', 'Input Fields', 'Output Fields', 'Input Type', 'Output Type', 'data_type', 'In Weight', 'Out Weight', 'Total Weight', 'Jenifer Access Count']
        
        ws.cell(row=1, column=1, value="")
        for i, h in enumerate(headers):
            cell = ws.cell(row=1, column=i+2, value=h)
            self._apply_cell_style(cell, 'header')
        
        sorted_weights = sorted(
            self.crypto_weights,
            key=lambda x: (
                ((x.end_point or '') == ''),
                x.end_point or '',
                x.method_signature or ''
            )
        )
        
        for r_idx, w in enumerate(sorted_weights, 2):
            # 각 필드를 정규화 (list → str로 변환)
            end_point = w.end_point[0] if isinstance(w.end_point, list) else (w.end_point or '')
            method_sig = w.method_signature[0] if isinstance(w.method_signature, list) else (w.method_signature or '')
            
            values = [
                end_point,
                method_sig,
                w.table_name,
                w.query_id,
                w.input_fields_count,
                w.output_fields_count,
                w.input_parameter_type or '',
                w.output_parameter_type or '',
                w.data_type,
                w.input_parameter_type_weight,
                w.output_parameter_type_weight,
                w.crypto_weight,
                w.access_cnt
            ]
            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                self._apply_cell_style(cell, 'data')
        
        col_widths = [3, 40, 30, 20, 15, 15, 15, 15, 15, 12, 12, 12, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    
    def _add_anyframe_estimation_detail_sheet(self, wb):
        """Anyframe 타입 예측 시트 - endpoint별 요약 (4컬럼)"""
        import openpyxl.utils
        
        ws = wb.create_sheet("Final Report")
        headers = ['End Point', 'Total Crypto Weight', 'Jenifer Access Count', 'Estimated KSIGN Calls']
        
        ws.cell(row=1, column=1, value="")
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i+1, value=h)
            self._apply_cell_style(cell, 'header')
        
        if not self.crypto_weights:
            return
        
        endpoint_weights = {}
        for w in self.crypto_weights:
            # Handle both CryptoWeight objects and dict
            is_obj = hasattr(w, 'end_point')
            
            ep = w.end_point if is_obj else w.get('end_point', '')
            
            # ep가 list인 경우 첫 번째 요소 추출 (정규화)
            if isinstance(ep, list):
                ep = ep[0] if ep else ''
            
            # blank endpoint인 경우 access_cnt를 0으로 강제 설정
            if not ep:
                access = 0
            else:
                access = w.access_cnt if is_obj else w.get('Jenifer Access Count', 1)
            
            # Anyframe Dict: 'Total Weight', Spring CryptoWeight: 'crypto_weight'
            weight = w.crypto_weight if is_obj else w.get('Total Weight', w.get('crypto_weight', 0))
            
            # blank endpoint는 별도 그룹화하지 말고 그냥 포함 (나중에 정렬으로 하단으로)
            if ep not in endpoint_weights:
                endpoint_weights[ep] = {'total_weight': 0, 'access_cnt': access}
            else:
                # 동일 endpoint에서 access_count 최대값 사용 (모두 같아야 함)
                endpoint_weights[ep]['access_cnt'] = max(endpoint_weights[ep]['access_cnt'], access)
            
            endpoint_weights[ep]['total_weight'] += weight
        
        self.logger.debug(f"Anyframe endpoint별 집계: {len(endpoint_weights)}개 엔드포인트, 총 weight= {sum(v['total_weight'] for v in endpoint_weights.values())}")
        
        sorted_endpoints = sorted(
            endpoint_weights.items(),
            key=lambda x: (
                ((x[0] or '') == ''),
                x[0] or ''
            )
        )
        
        for r_idx, (endpoint, data) in enumerate(sorted_endpoints, 2):
            total_weight = data['total_weight']
            access_cnt = data['access_cnt']
            estimated_calls = total_weight * access_cnt
            
            # endpoint가 list인 경우 첫 번째 요소 추출 (Excel cell 쓰기 전 정규화)
            endpoint_str = endpoint[0] if isinstance(endpoint, list) else (endpoint or '')
            
            values = [endpoint_str, total_weight, access_cnt, estimated_calls]
            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                self._apply_cell_style(cell, 'data')
        
        col_widths = [3, 40, 20, 20, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    

    def _add_anyframe_final_report_sheet(self, wb):
        """Anyframe 타입 최종 리포트 - method별 상세 (LLM 응답 형식 통일, data_type 추가)"""
        import openpyxl.utils
        
        ws = wb.create_sheet("Ksign Call Estimation")
        headers = [
            'End Point', 'Method', 'loop_depth', 'loop_structure', 'multiplier',
            'dep0_crypto_count', 'dep1_crypto_count', 'dep2_crypto_count',
            'data_type', 'Total Weight', 'Jenifer Access Count'
        ]
        
        ws.cell(row=1, column=1, value="")
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i+1, value=h)
            self._apply_cell_style(cell, 'header')
        
        if not self.crypto_weights:
            return
        
        # Helper to safely get attribute from object or dict
        def get_attr(obj, attr):
            if isinstance(obj, dict):
                return obj.get(attr) or ''
            return getattr(obj, attr, '') or ''
        
        sorted_weights = sorted(
            self.crypto_weights,
            key=lambda x: (
                (get_attr(x, 'end_point') == ''),
                get_attr(x, 'end_point'),
                get_attr(x, 'method_signature')
            )
        )
        
        for r_idx, w in enumerate(sorted_weights, 2):
            # Handle both CryptoWeight objects and dict
            is_obj = hasattr(w, 'end_point')
            
            end_point = w.end_point if is_obj else w.get('end_point', '')
            method = w.method_signature if is_obj else w.get('method_signature', '')
            loop_depth = w.loop_depth if is_obj else w.get('loop_depth', 0)
            loop_structure = w.loop_structure if is_obj else w.get('loop_structure', '')
            multiplier = w.multiplier if is_obj else w.get('multiplier', '1')
            dep0 = w.dep0_crypto_count if is_obj else w.get('dep0_crypto_count', 0)
            dep1 = w.dep1_crypto_count if is_obj else w.get('dep1_crypto_count', 0)
            dep2 = w.dep2_crypto_count if is_obj else w.get('dep2_crypto_count', 0)
            # Anyframe Dict: 'crypto_weight' (최종 가중치)
            total_weight = w.crypto_weight if is_obj else w.get('crypto_weight', w.get('Total Weight', 0))
            data_type = w.data_type if is_obj else w.get('data_type', 'single')
            # Step 6에서 'Jenifer Access Count'로 업데이트됨. Dict는 이를 우선 확인
            access_cnt = w.access_cnt if is_obj else w.get('Jenifer Access Count', w.get('access_cnt', 0))
            
            # Excel cell 쓰기 전 값 정규화 (list → str로 변환)
            end_point = end_point[0] if isinstance(end_point, list) else (end_point or '')
            method = method[0] if isinstance(method, list) else (method or '')
            loop_structure = loop_structure[0] if isinstance(loop_structure, list) else (loop_structure or '')
            multiplier = str(multiplier[0] if isinstance(multiplier, list) else (multiplier or '1'))
            
            values = [
                end_point, method,
                loop_depth, loop_structure, multiplier,
                dep0, dep1, dep2,
                data_type, total_weight, access_cnt
            ]
            
            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                self._apply_cell_style(cell, 'data')
        
        col_widths = [3, 40, 30, 12, 15, 15, 15, 15, 15, 15, 12, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    def load_endpoint_access(self, endpoint_dict: Optional[Dict[str, int]] = None) -> bool:
        """엔드포인트 호출 데이터 로드 (Spring/Anyframe 공통)
        
        .applycrypto/endpoint_access.txt에서 데이터를 로드하여
        EndpointAccess 객체 리스트로 self.endpoint_access에 저장합니다.
        
        Args:
            endpoint_dict: 이미 로드된 Dict (전달되면 다시 로드하지 않음)
        """
        # Dict가 전달되지 않으면 로드
        if endpoint_dict is None:
            endpoint_dict = self._load_endpoint_access_dict()
        
        if not endpoint_dict:
            self.logger.warning("endpoint_dict이 비어있음")
            return False
        
        # Dict를 EndpointAccess 객체로 변환하여 저장
        self.endpoint_access = [
            EndpointAccess(end_point=ep, access_cnt=cnt)
            for ep, cnt in endpoint_dict.items()
        ]
        self.logger.info(f"EndpointAccess 객체 {len(self.endpoint_access)}개 생성 완료")
        # 첫 3개 샘플
        for i, ea in enumerate(self.endpoint_access[:3]):
            self.logger.debug(f"EndpointAccess(end_point='{ea.end_point}', access_cnt={ea.access_cnt})")
        return len(self.endpoint_access) > 0
    

    def flatten_query_weights_to_crypto_weights(self) -> bool:
        """
        Step 7 전단계: query_weights의 endpoint_results[]를 평탄화하여 crypto_weights 생성
        
        각 쿼리의 endpoint_results[] 배열을 순회하며,
        각 엔드포인트별로 1개의 CryptoWeight 행 생성
        """
        self.crypto_weights.clear()
        
        has_endpoint_access = bool(self.endpoint_access)
        
        for query_weight in self.query_weights:
            if not query_weight.endpoint_results:
                continue
            
            # 각 endpoint_result에 대해 1개의 CryptoWeight 행 생성
            for endpoint_result in query_weight.endpoint_results:
                access_cnt = self._get_endpoint_access_count(endpoint_result.end_point)
                
                crypto_weight = CryptoWeight(
                    table_name=query_weight.table_name,
                    query_id=query_weight.query_id,
                    method_signature=endpoint_result.method_signature,
                    end_point=endpoint_result.end_point,
                    class_path=endpoint_result.class_path,
                    input_fields_count=query_weight.input_fields_count,
                    output_fields_count=query_weight.output_fields_count,
                    input_parameter_type=endpoint_result.input_parameter_type,
                    output_parameter_type=endpoint_result.output_parameter_type,
                    data_type=endpoint_result.data_type,  # 새 필드: data_type 추가
                    input_parameter_type_weight=endpoint_result.input_parameter_type_weight,
                    output_parameter_type_weight=endpoint_result.output_parameter_type_weight,
                    crypto_weight=endpoint_result.crypto_weight,
                    access_cnt=access_cnt
                )
                self.crypto_weights.append(crypto_weight)
        
        self.logger.info(f"{len(self.query_weights)} 개 쿼리 → {len(self.crypto_weights)} 개 항목 평탄화")
        return True


    def _get_endpoint_access_count(self, target_ep: str) -> int:
        """엔드포인트 경로에 따른 호출 빈도 계산 (바인딩 경로 및 일반 경로 대응)"""
        # target_ep가 list인 경우 첫 번째 요소 추출 (보안 대책)
        if isinstance(target_ep, list):
            target_ep = target_ep[0] if target_ep else ''
        
        # target_ep가 string이 아닌 경우 문자열로 변환
        if not isinstance(target_ep, str):
            target_ep = str(target_ep) if target_ep else ''
        
        target_ep_normalized = target_ep  # 정규화된 버전 저장 (디버깅용)
        
        if not self.endpoint_access:
            self.logger.warning(f"_get_endpoint_access_count('{target_ep_normalized}'): endpoint_access가 비어있음")
            return 1 # 데이터가 없으면 기본값 1
            
        # 1. 바인딩 경로 ({}) 처리 로직
        if "{" in target_ep:
            # {no}와 같은 부분을 .+로 치환하여 정규표현식 생성
            parts = re.split(r'\{[^}]+\}', target_ep)
            regex_parts = [re.escape(p) for p in parts]
            pattern_str = ".+".join(regex_parts)
            
            # 1-1. 패턴 매칭 (Exact: ^...$) - 합계 계산
            exact_pattern = re.compile(f"^{pattern_str}$")
            total_cnt = sum(a.access_cnt for a in self.endpoint_access 
                          if exact_pattern.search(
                              a.end_point[0] if isinstance(a.end_point, list) else (a.end_point or '')
                          ))
            if total_cnt > 0:
                return total_cnt
                
            # 1-2. 패턴 매칭 (Contains) - 합계 계산
            contains_pattern = re.compile(pattern_str)
            total_cnt = sum(a.access_cnt for a in self.endpoint_access 
                          if contains_pattern.search(
                              a.end_point[0] if isinstance(a.end_point, list) else (a.end_point or '')
                          ))
            return total_cnt

        # 2. 일반 경로 처리 로직
        # 2-1. Exact match (우선순위 1)
        exact_match_found = False
        for access in self.endpoint_access:
            # access.end_point가 list인 경우 첫 번째 요소 추출
            access_ep = access.end_point[0] if isinstance(access.end_point, list) else (access.end_point or '')
            if access_ep == target_ep_normalized:
                self.logger.debug(f"_get_endpoint_access_count('{target_ep_normalized}'): Exact match found! → {access.access_cnt}")
                return access.access_cnt
            # 비교 로그 (처음 3개만)
            # if not exact_match_found and self.endpoint_access.index(access) < 3:
            #     print(f"    [CMP] '{access_ep}' (len={len(access_ep)}) == '{target_ep_normalized}' (len={len(target_ep_normalized)})? {access_ep == target_ep_normalized}", flush=True)
        
        # 2-2. EndsWith match (우선순위 2)
        for access in self.endpoint_access:
            # access.end_point가 list인 경우 첫 번째 요소 추출
            access_ep = access.end_point[0] if isinstance(access.end_point, list) else (access.end_point or '')
            if access_ep.endswith(target_ep_normalized):
                self.logger.debug(f"_get_endpoint_access_count('{target_ep_normalized}'): EndsWith match found! → {access.access_cnt}")
                return access.access_cnt
        
        self.logger.warning(f"_get_endpoint_access_count('{target_ep_normalized}'): No match found (0 반환)")
        return 0
    
    
    def generate_ksign_report(self) -> bool:
        """
        Step 8: 엔드포인트 별 KSIGN 호출 예측 집계 및 JSON 리포트 생성
        
        각 엔드포인트별로 암복호화 가중치를 합산하고,
        런타임 호출빈도를 곱하여 예측 KSIGN 호출 수를 계산합니다.
        결과를 JSON 리포트로 저장하고, self.endpoint_weights에도 저장합니다.
        """
        # 데이터 통합 (평탄화된 CryptoWeight로부터 집계)
        self.endpoint_weights = {}
        for weight in self.crypto_weights:
            key = weight.end_point
            if key not in self.endpoint_weights:
                # blank endpoint는 access_count를 0으로 강제 설정
                access_count = 0 if not key else weight.access_cnt
                
                self.endpoint_weights[key] = {
                    'endpoint': key,
                    'total_crypto_weight': 0.0,
                    'query_count': 0,
                    'input_fields_count': weight.input_fields_count,  # 첫 번째 쿼리의 값 저장
                    'output_fields_count': weight.output_fields_count,  # 첫 번째 쿼리의 값 저장
                    'data_type': weight.data_type,
                    'access_count': access_count,  # Step 6에서 이미 계산됨, blank면 0
                    'estimated_ksign_calls': 0.0,
                    'queries': []
                }
            
            self.endpoint_weights[key]['total_crypto_weight'] += weight.crypto_weight
            self.endpoint_weights[key]['query_count'] += 1
            self.endpoint_weights[key]['queries'].append({
                'table': weight.table_name,
                'query_id': weight.query_id,
                'weight': weight.crypto_weight
            })
        
        # 최종 집계 결과 계산 (암복호화 가중치 합계 × 호출 빈도)
        for key in self.endpoint_weights:
            # KSIGN 예측 호출 = 암복호화 가중치 합계 × 엔드포인트 호출빈도
            self.endpoint_weights[key]['estimated_ksign_calls'] = (
                self.endpoint_weights[key]['total_crypto_weight'] * self.endpoint_weights[key]['access_count']
            )
        
        # 결과 정렬 (예측 호출 수 기준 내림차순)
        sorted_endpoints = sorted(
            self.endpoint_weights.values(),
            key=lambda x: x['estimated_ksign_calls'],
            reverse=True
        )
        
        # JSON 저장
        report_file = self.artifacts_dir / f"ksign_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        try:
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'report_metadata': {
                        'generated_at': datetime.now().isoformat(),
                        'total_endpoints': len(sorted_endpoints),
                        'total_ksign_calls': sum(e['estimated_ksign_calls'] for e in sorted_endpoints)
                    },
                    'endpoints': sorted_endpoints
                }, f, indent=2, ensure_ascii=False)
            self.logger.info(f"ksign_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json 저장 완료")
        except Exception as e:
            self.logger.error(f"ksign_report JSON 저장 실패: {e}")
            return False
        
        return True
    

    def run_full_pipeline(self) -> bool:
        """
        전체 파이프라인 실행 (dispatcher)
        
        framework_type에 따라 적절한 파이프라인 선택:
        - Spring* : Spring Type (mapper.xml 기반)
        - Anyframe* : Anyframe Type (SVC/SVCImpl/BIZ 코드 기반)
        """
        self.logger.info("=" * 60)
        self.logger.info(f"KSIGN 호출 예측보고서 생성 파이프라인 [시작!] - {self.framework_type}")
        self.logger.info("=" * 60)
        
        # framework_type에 따라 파이프라인 선택
        if self.framework_type.startswith("Anyframe"):
            return self.run_full_pipeline_anyframe()
        else:
            # Spring으로 시작하거나 명시되지 않은 경우 기본값으로 Spring 타입 사용
            return self.run_full_pipeline_spring()
    

    def run_full_pipeline_spring(self) -> bool:
        """
        Spring Type 파이프라인 실행
        
        수정 전 코드를 analyze, modify 수행 후 generate_ksign 수행
        mapper.xml에 암복호화 TypeHandler를 적용하는 방식
        """
        self.logger.info("[Framework] Spring Type - mapper.xml 기반 처리")
        
        # Step 1: table_access_info.json 로드
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 1] 테이블 접근 정보 로드 중...")
        if not self.load_table_access():
            return False
        self.logger.info("[Step 1] 완료 ✓")
        
        # Step 2: 최신 타임스탐프 폴더의 계층적 쿼리 분석 결과 로드
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 2] 계층적 쿼리 분석 결과 로드 중...")
        if not self.load_step1_query_analysis_hierarchical():
            return False
        self.logger.info("[Step 2] 완료 ✓")
        
        # Step 3: QueryWeight 생성 (쿼리별 endpoint[] 배열)
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 3] QueryID별 암복호화 가중치 계산 중...")
        if not self.calculate_crypto_weights():
            return False
        self.logger.info("[Step 3] 완료 ✓")
        
        # Step 4: Endpoint 정보 수집 (endpoint_details[] 채우기)
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 4] Endpoint 정보 수집 중...")
        if not self.enrich_endpoint_and_class_path():
            return False
        self.logger.info("[Step 4] 완료 ✓")
        
        # Step 5: 파라미터 타입 가중치 부여 (endpoint_results[] 채우기)
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 5] 파라미터 타입 가중치 부여 중...")
        if not self.apply_parameter_type_weights():
            return False
        self.logger.info("[Step 5] 완료 ✓")
        
        # Step 6: 런타임 호출빈도 데이터 로드
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 6] 런타임 호출빈도 데이터 로드 중...")
        endpoint_access_dict = self._load_endpoint_access_dict()
        self.load_endpoint_access(endpoint_access_dict)  # 런타임 데이터 로드 (중복 로드 방지)
        self.logger.info("[Step 6] 완료 ✓")
        
        # Step 7: 평탄화 (endpoint_results[] → CryptoWeight[] 변환)
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 7] 평탄화 중...")
        if not self.flatten_query_weights_to_crypto_weights():
            return False
        self.logger.info("[Step 7] 완료 ✓")
        
        # Step 8: 암복호화 가중치 JSON 저장
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 8] 암복호화 가중치 JSON 저장 중...")
        if not self.save_crypto_weights_json():
            return False
        self.logger.info("[Step 8] 완료 ✓")
        
        # Step 9: 엔드포인트 별 집계 및 최종 리포트 생성
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 9] 엔드포인트 별 KSIGN 호출 예측 생성 중...")
        if not self.generate_ksign_report():
            return False
        self.logger.info("[Step 9] 완료 ✓")
        
        # Step 10: Excel 리포트 생성 (3개 시트: 가중치 상수, 상세분석, 최종리포트)
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 10] KSIGN 호출 예측보고서 Excel 생성 중...")
        if not self.save_ksign_report_excel():
            return False
        self.logger.info("[Step 10] 완료 ✓")
        
        self.logger.info("")
        self.logger.info("=" * 60)
        self.logger.info("KSIGN 호출 예측보고서 생성 파이프라인 [완료!]")
        self.logger.info("=" * 60)
        return True
    

    def run_full_pipeline_anyframe(self) -> bool:
        """
        Anyframe Type 전체 파이프라인 (Step 1~9)
        
        SVC/SVCImpl/BIZ layer 암복호화 코드 분석 + call_graph 기반 endpoint 매핑
        
        Step 별 처리:
        - Step 1: config 검증
        - Step 2: difflib로 변경 파일 추출
        - Step 3: call_graph.json 로드 + endpoint_reachable_sigs 추출
        - Step 4: 암복호화 가중치 계산 (endpoint 필터링 후 LLM 호출)
        - Step 5: endpoint_access 데이터 로드
        - Step 6: call_trees 역탐색 → end_point + access_count 설정
        - Step 7: access_count 최종화 + JSON 저장
        - Step 8: endpoint별 집계 + ksign_report.json
        - Step 9: Excel 리포트 생성
        
        Returns:
            bool: 성공 여부
        """
        self.logger.info(f"[Framework] Anyframe Type - SVC/SVCImpl/BIZ 코드 기반 처리")
        self.logger.info(f"Step 0 (Config 로드): ksignutil_patterns {len(self.ksignutil_patterns)}개")
        
        # Step 1: 추출 대상 ksignUtil 항목 검증
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 1] 추출 대상 ksignUtil 항목 검증 중...")
        if not self._validate_step1_config():
            return False
        
        # Step 1 검증 성공 - 로드된 값 확인
        self.logger.debug(f"framework_type: {self.framework_type}")
        if self.ksignutil_patterns:
            self.logger.debug(f"ksignUtils_pattern ({len(self.ksignutil_patterns)}개)")
            for pattern in self.ksignutil_patterns[:5]:  # 처음 5개만 출력
                self.logger.debug(f"  {pattern}")
            if len(self.ksignutil_patterns) > 5:
                self.logger.debug(f"  ... 그 외 {len(self.ksignutil_patterns) - 5}개")
        if self.policy_ids:
            self.logger.debug(f"policyId ({len(self.policy_ids)}개): {self.policy_ids}")
        self.logger.info("[Step 1] 검증 완료 ✓")
        
        # Step 2: 원본, 수정본 비교 - 암복호화를 적용한 코드의 파일/메서드 추출
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 2] 암복호화 적용 코드 추출 중...")
        if not self._extract_changed_methods():
            return False
        self.logger.info(f"Step 1 (Table Access 로드): 변경된 파일 추출 완료")
        self.logger.info("[Step 2] 완료 ✓")
        
        # Step 3: call_graph.json, table_access_info.json 로드
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 3] call_graph.json 로드 중...")
        if not self._load_call_graph():
            return False
        self.logger.info(f"Step 2 (Calculate Crypto Weights): call_graph 로드 완료")
        self.logger.info("[Step 3] 완료 ✓")
        
        # Step 4: 변경된 클래스에서 ksignUtil weight 계산 및 crypto_weight.json 생성
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 4] 암복호화 가중치 계산 중...")
        if not self._calculate_anyframe_weights():
            return False
        self.logger.info(f"Step 3 (Enrich Endpoint): 가중치 계산 완료")
        self.logger.info("[Step 4] 완료 ✓")
        
        # Step 5: 런타임 호출빈도 데이터 로드
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 5] 런타임 호출빈도 데이터 로드 중...")
        endpoint_access_dict = self._load_endpoint_access_dict()
        if endpoint_access_dict:
            self.load_endpoint_access(endpoint_access_dict)  # Dict 전달 → 중복 로드 방지
            self.logger.debug(f"endpoint_access 로드됨: {len(endpoint_access_dict)}개 엔드포인트")
        else:
            self.logger.warning("endpoint_access 데이터 없음 (빈 Dict) - Step 6 스킵됨")
            endpoint_access_dict = {}  # None 대신 빈 Dict 유지
        self.logger.info(f"Step 4 (Apply Parameter Weights): endpoint_access 로드 완료")
        self.logger.info("[Step 5] 완료 ✓")
        
        # Step 6: call_trees 역탐색으로 메서드 → endpoint 경로 설정 + access_count 매핑
        # - end_point 설정: call_graph.call_trees 기준 (endpoint_access 유무와 무관)
        # - access_count 설정: endpoint_access_dict가 있을 때만 채워짐 (없으면 0)
        self.logger.info("")
        self.logger.info("-" * 60)
        if self.call_graph is not None:
            self.logger.info("[Step 6] call_graph를 통한 메서드-엔드포인트 매핑 중...")
            if not endpoint_access_dict:
                self.logger.info("endpoint_access_dict 없음 - end_point 경로는 설정되지만 access_count는 0으로 유지됨")
            self.crypto_weights = self._enrich_crypto_weights_with_endpoint_access(
                self.crypto_weights,
                endpoint_access_dict,
                self.call_graph
            )
            self.logger.info("Step 6 완료: 메서드별 end_point + 호출빈도 업데이트")
            self.logger.info("[Step 6] 완료 ✓")
        else:
            self.logger.warning("Step 6 스킵: call_graph 없음 (Step 3 실패?)")
        
        # Step 7: access_count 최종화 및 crypto_weight_{timestamp}.json 저장
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 7] access_count 최종화 중...")
        if not self._enrich_crypto_weights_with_access_count():
            return False
        self.logger.info("[Step 7] 완료 ✓")
        
        # Step 8: 최종 엔드포인트별 집계 및 ksign_report.json 생성
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 8] 최종 KSIGN 호출 예측 리포트 생성 중...")
        if not self._generate_anyframe_ksign_report():
            return False
        self.logger.info("[Step 8] 완료 ✓")
        
        # Step 9: Excel 리포트 생성
        self.logger.info("")
        self.logger.info("-" * 60)
        self.logger.info("[Step 9] KSIGN 호출 예측보고서 Excel 생성 중...")
        if not self.save_ksign_report_excel():
            return False
        self.logger.info("[Step 9] 완료 ✓")
        
        self.logger.info("")
        self.logger.info("=" * 60)
        self.logger.info("KSIGN 호출 예측보고서 생성 파이프라인 [완료!]")
        self.logger.info("=" * 60)
        return True
    
    def _extract_changed_methods(self) -> bool:
        """Step 2: 암복호화 적용 코드 추출 (difflib + AST)
        
        구현 방식:
        1. difflib로 변경된 파일 목록 추출 (원본 vs 수정본 비교)
        2. AST Parser로 해당 파일의 메서드 추출
        3. 각 메서드에서 ksignUtil 호출 여부 검증
        4. method_signature 목록 생성 (ClassName.methodName)
        
        데이터 출처:
        - old_code_path: config.json의 artifact_generation.old_code_path
        - target_project: 현재 프로젝트
        - ksignutil_patterns: Step 1에서 로드된 패턴 리스트 (예: ["com.example.Util.encrypt(String) -> String"])
        
        출력 데이터:
        - self.files_with_ksignutil: {rel_path → full_path}
        - self.methods_with_ksignutil: [method_signature, ...]
        """
        try:
            # 1. difflib로 변경된 Java 파일 목록 추출
            changed_files = self._get_changed_java_files_flexible()
            self.logger.debug(f"변경된 파일: {len(changed_files)}개")
            
            # 2. 변경된 파일에서 AST로 메서드 추출
            files_with_ksignutil = {}
            methods_with_ksignutil = []
            
            for file_path in changed_files:
                try:
                    # AST로 메서드 추출
                    methods = self._extract_methods_with_ast(file_path)
                    if not methods:
                        continue
                    
                    # ksignUtil 호출 메서드 필터링
                    filtered_methods = self._filter_methods_by_ksignutil(file_path, methods)
                    
                    if filtered_methods:
                        # 상대 경로 계산
                        rel_path = os.path.relpath(file_path, str(self.target_project))
                        files_with_ksignutil[rel_path] = file_path
                        methods_with_ksignutil.extend(filtered_methods)
                
                except Exception as e:
                    self.logger.warning(f"메서드 추출 오류: {file_path}: {e}")
                    continue
            
            self.files_with_ksignutil = files_with_ksignutil
            self.methods_with_ksignutil = list(set(methods_with_ksignutil))  # 중복 제거
            
            self.logger.info(f"ksignUtil 적용 파일: {len(files_with_ksignutil)}개, 메서드: {len(self.methods_with_ksignutil)}개")
            return True
        except Exception as e:
            self.logger.error(f"변경된 메서드 추출 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_changed_java_files_flexible(self) -> List[str]:
        """difflib로 변경된 Java 파일 목록 추출 (경로 구조 무관)
        
        원본(old_code_path)과 수정본(target_project)을 비교하여:
        - 파일 MD5 해시로 변경 감지
        - Java 파일만 선별
        - 새 파일도 포함
        
        Returns:
            List[str]: 변경된 Java 파일의 전체 경로
        """
        import hashlib
        from pathlib import Path
        
        changed_files = []
        
        # target_project의 모든 Java 파일 수집 (상대경로 기준)
        target_java_files = {}  # {rel_path: full_path}
        target_src_dir = self._detect_src_directory(str(self.target_project))
        self.logger.debug(f"target_project src 디렉토리: {target_src_dir}")
        
        for root, dirs, files in os.walk(target_src_dir):
            dirs[:] = [d for d in dirs if d != '.applycrypto']
            
            for file in files:
                if file.endswith('.java'):
                    full_path = os.path.join(root, file)
                    # src 기준 상대 경로로 정규화
                    rel_path = os.path.relpath(full_path, target_src_dir)
                    target_java_files[rel_path] = full_path
        
        # old_code_path의 Java 파일 수집
        old_java_files = {}
        old_code_path = self.config.get("artifact_generation", {}).get("old_code_path")
        if not old_code_path or not os.path.exists(old_code_path):
            # old_code_path 없으면 target_project의 전체 Java 파일 모두 변경된 것으로 간주
            self.logger.warning(f"old_code_path 없음: {old_code_path}")
            return list(target_java_files.values())
        
        old_src_dir = self._detect_src_directory(old_code_path)
        self.logger.debug(f"old_code_path src 디렉토리: {old_src_dir}")
        for root, dirs, files in os.walk(old_src_dir):
            dirs[:] = [d for d in dirs if d != '.applycrypto']
            
            for file in files:
                if file.endswith('.java'):
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, old_src_dir)
                    old_java_files[rel_path] = full_path
        
        # 파일 변경 비교
        for rel_path, target_path in target_java_files.items():
            if rel_path in old_java_files:
                old_path = old_java_files[rel_path]
                try:
                    # 파일 내용 비교 (MD5 해시)
                    with open(target_path, 'rb') as f:
                        target_hash = hashlib.md5(f.read()).hexdigest()
                    with open(old_path, 'rb') as f:
                        old_hash = hashlib.md5(f.read()).hexdigest()
                    
                    if target_hash != old_hash:
                        changed_files.append(target_path)
                except Exception:
                    changed_files.append(target_path)
            else:
                # 새로 생성된 파일
                changed_files.append(target_path)
        
        self.logger.debug(f"파일 비교 결과: {len(target_java_files)}개 target, {len(old_java_files)}개 old, {len(changed_files)}개 변경")
        if changed_files:
            for f in changed_files[:5]:
                self.logger.debug(f"  {os.path.relpath(f, self.target_project)}")
        
        return changed_files
    
    
    def _detect_src_directory(self, project_path: str) -> str:
        """프로젝트 구조에 따라 Java 소스 디렉토리 감지
        
        우선순위:
        1. src/main/java
        2. src
        3. 프로젝트 루트
        """
        candidates = [
            os.path.join(project_path, "src", "main", "java"),
            os.path.join(project_path, "src"),
            project_path
        ]
        
        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate
        
        return project_path
    
    
    def _extract_methods_with_ast(self, file_path: str) -> Dict[str, str]:
        """AST Parser로 Java 파일의 모든 메서드 추출
        
        Args:
            file_path: Java 파일 경로
        
        Returns:
            Dict[str, str]: {class_name: method_body} 형식으로 메서드정보
                           또는 {method_name: method_body} 형식 (여러 메서드)
        """
        from parser.java_ast_parser import JavaASTParser
        from pathlib import Path
        
        try:
            parser = JavaASTParser()
            tree, error = parser.parse_file(Path(file_path), remove_comments=True)
            
            if error or tree is None:
                self.logger.warning(f"AST 파싱 실패: {file_path}: {error}")
                return {}
            
            class_infos = parser.extract_class_info(tree, Path(file_path))
            
            methods_data = {}
            for class_info in class_infos:
                for method in class_info.methods:
                    # method_signature = "ClassName.methodName"
                    method_sig = f"{class_info.name}.{method.name}"
                    methods_data[method_sig] = method  # Method 객체 저장
            
            return methods_data
        
        except Exception as e:
            self.logger.warning(f"AST 메서드 추출 오류: {file_path}: {e}")
            return {}
    
    
    def _filter_methods_by_ksignutil(self, file_path: str, methods_data: Dict[str, any]) -> List[str]:
        """ksignUtil 호출 메서드만 필터링
        
        각 메서드의 바디를 읽어서 ksignUtil 호출 여부 확인
        
        Args:
            file_path: Java 파일 경로
            methods_data: {method_sig: method_object}
        
        Returns:
            List[str]: ksignUtil 호출 메서드 시그니처 리스트
        """
        filtered_methods = []
        
        # [DEBUG] 입력 매개변수 검증
        self.logger.debug(f"_filter_methods_by_ksignutil 호출: file_path={file_path}, methods_data 개수={len(methods_data)}, patterns 개수={len(self.ksignutil_patterns)}")
        
        # 패턴에서 메서드명 추출 (예: "SliEncryptionUtil.encrypt(...)" → "encrypt")
        ksignutil_names = set()
        for pattern in self.ksignutil_patterns:
            # 패턴에서 괄호 앞의 메서드명 추출하는 정규표현식
            import re
            match = re.search(r'\.(\w+)\s*\(', pattern)
            if match:
                ksignutil_names.add(match.group(1))
        
        self.logger.debug(f"추출된 메서드명: {ksignutil_names}")
        
        if not ksignutil_names:
            self.logger.warning(f"ksignutil_names 추출 실패. 패턴: {self.ksignutil_patterns}")
            return filtered_methods
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # 파일에 메서드 이름이 포함되어 있는지 먼저 확인
            file_has_any_ksignutil = any(name in content for name in ksignutil_names)
            self.logger.debug(f"파일에 포함된 메서드명: {file_has_any_ksignutil}")
            
            if not file_has_any_ksignutil:
                return []
            
            for method_sig, method in methods_data.items():
                # 메서드 바디에서 ksignUtil 호출 확인
                has_ksignutil = False
                
                for ksignutil_name in ksignutil_names:
                    # 메서드 바디에서 호출 검색
                    if hasattr(method, 'body') and method.body:
                        if ksignutil_name in method.body:
                            has_ksignutil = True
                            break
                
                if has_ksignutil:
                    filtered_methods.append(method_sig)
            
            if filtered_methods:
                self.logger.debug(f"{os.path.basename(file_path)}: {len(filtered_methods)}개 메서드")
            
            return filtered_methods
        
        except Exception as e:
            self.logger.warning(f"ksignUtil 필터링 오류: {file_path}: {e}")
            return []
    

    def _load_call_graph(self) -> bool:
        """
        Step 3: call_graph.json 로드 및 endpoint 도달 가능 메서드 Set 추출
        
        처리:
        1. call_graph.json 로드
        2. call_trees[].method_chain에서 모든 메서드 시그니처 수집
           → self.endpoint_reachable_sigs: set으로 저장
           → 이 Set에 속하지 않는 메서드는 endpoint 호출 체인 없음
        3. Step 4에서 LLM 호출 필터링에 사용 (불필요한 LLM 호출 방지)
        
        Returns:
            bool: 성공 여부
        """
        try:
            file_path = self.results_dir / "call_graph.json"
            if not file_path.exists():
                self.logger.error(f"call_graph.json 파일을 찾을 수 없음: {file_path}")
                return False
            
            with open(file_path, 'r', encoding='utf-8') as f:
                self.call_graph = json.load(f)
            
            endpoints_count = len(self.call_graph.get('endpoints', []))
            self.logger.info(f"call_graph.json 로드 완료: {endpoints_count}개 엔드포인트")
            
            # === Anyframe 전용: call_trees에서 endpoint에 도달 가능한 method_signature Set 추출 ===
            # call_graph 구조:
            #   endpoints[]: {path, method_signature, ...}  ← HTTP endpoint 목록
            #   call_trees[]: {method_signature, children: [{method_signature, children: [...]}]}
            #                  ↑ Controller 루트부터 Service/DAO까지 트리 형태
            # Step 4에서 LLM 호출 대상을 이 Set에 속한 메서드로만 한정하여 불필요한 LLM 호출 방지
            def _collect_sigs_recursive(node: dict, result: set):
                sig_raw = node.get('method_signature', '')
                # method_signature가 리스트인 경우 첫 번째 요소 사용
                sig = sig_raw[0] if isinstance(sig_raw, list) else sig_raw
                if sig:
                    result.add(sig)
                for child in node.get('children', []):
                    _collect_sigs_recursive(child, result)            
            self.endpoint_reachable_sigs: set = set()
            for call_tree in self.call_graph.get('call_trees', []):
                _collect_sigs_recursive(call_tree, self.endpoint_reachable_sigs)
            self.logger.info(f"endpoint 도달 가능 메서드: {len(self.endpoint_reachable_sigs)}개 (call_trees 재귀 순회)")
            if self.endpoint_reachable_sigs:
                sample = list(self.endpoint_reachable_sigs)[:5]
                for s in sample:
                    self.logger.debug(f"{s}")
            
            return True
        except Exception as e:
            self.logger.error(f"call_graph.json 로드 실패: {e}")
            return False
    

    def _enrich_crypto_weights_with_endpoints(self) -> bool:
        """
        Step 8: call_graph.json의 endpoints 배열과 매핑하여
        method_signature와 일치하는 end_point 찾기
        
        Returns:
            bool: 성공 여부
        """
        try:
            # 1. 메모리에서 crypto_weights 로드 (Step 4/7에서 저장됨)
            if hasattr(self, 'crypto_weights_with_endpoints') and self.crypto_weights_with_endpoints:
                crypto_weights = self.crypto_weights_with_endpoints
            elif self.crypto_weights:
                crypto_weights = self.crypto_weights if isinstance(self.crypto_weights, list) else list(self.crypto_weights)
            else:
                self.logger.warning("crypto_weights가 메모리에 없음 (Step 4에서 계산 결과 0건이거나 이전 단계 실패)")
                return False
            
            # 2. call_graph.endpoints 인덱싱 (method_signature → endpoint path 매핑)
            endpoint_map = {}  # {method_signature → path}
            for endpoint_entry in self.call_graph.get('endpoints', []):
                method_sig_raw = endpoint_entry.get('method_signature')
                # method_signature 정규화: 배열이면 첫 요소, None이면 건너뛰기
                if isinstance(method_sig_raw, list):
                    method_sig = method_sig_raw[0] if method_sig_raw else None
                else:
                    method_sig = method_sig_raw
                
                # path 정규화: 배열이면 첫 요소, 없으면 None
                path_raw = endpoint_entry.get('path', [])
                if isinstance(path_raw, list):
                    path = path_raw[0] if path_raw else None
                else:
                    path = path_raw if path_raw else None
                
                if method_sig and path:
                    endpoint_map[method_sig] = path
            
            self.logger.debug(f"call_graph endpoints: {len(endpoint_map)}개 매핑")
            
            # 3. crypto_weights의 method_signature와 매칭하여 end_point 추가
            matched_count = 0
            for crypto_weight in crypto_weights:
                method_sig_raw = crypto_weight.get('method_signature')
                # method_signature 정규화: 배열이면 첫 요소, 문자열이면 그대로
                if isinstance(method_sig_raw, list):
                    method_sig = method_sig_raw[0] if method_sig_raw else None
                else:
                    method_sig = method_sig_raw
                
                # 정확한 매칭 시도
                if method_sig in endpoint_map:
                    crypto_weight['end_point'] = endpoint_map[method_sig]
                    matched_count += 1
                else:
                    # 부분 매칭 시도 (클래스명.메서드명 대소문자 무시)
                    for mapped_sig, path in endpoint_map.items():
                        if method_sig and mapped_sig and method_sig.split('.')[-1] == mapped_sig.split('.')[-1]:
                            crypto_weight['end_point'] = path
                            matched_count += 1
                            break
            
            self.logger.debug(f"매핑 완료: {matched_count}/{len(crypto_weights)}개 엔드포인트")
            
            # 4. 매핑된 crypto_weights를 메모리에만 저장 (중간 JSON 파일 생성하지 않음)
            self.crypto_weights_with_endpoints = crypto_weights
            self.crypto_weights = crypto_weights
            
            self.logger.info("Step 8 완료: endpoint 경로 매핑 완료")
            return True
        
        except Exception as e:
            self.logger.error(f"call_graph 매핑 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    

    def _enrich_crypto_weights_with_access_count(self) -> bool:
        """
        Step 7: access_count 최종화 및 최종 JSON 파일 저장
        
        처리:
        1. self.crypto_weights (또는 self.crypto_weights_with_endpoints) 로드
        2. end_point가 있는 항목만 Jenifer Access Count 검증/유지
           - Step 6에서 이미 설정되었으므로 여기서는 final check만 수행
        3. end_point 없는 항목은 access_count = 0으로 설정
        4. 최종 JSON 파일 저장: artifacts/crypto_weight_{timestamp}.json
           (모든 필드 포함: data_type, loop_depth, multiplier, encrypt/decrypt counts)
        
        Returns:
            bool: 성공 여부
        """
        try:
            # Anyframe: self.crypto_weights 직접 사용, Spring: self.crypto_weights_with_endpoints
            if hasattr(self, 'crypto_weights_with_endpoints') and self.crypto_weights_with_endpoints:
                crypto_weights = self.crypto_weights_with_endpoints
            elif hasattr(self, 'crypto_weights') and self.crypto_weights:
                crypto_weights = self.crypto_weights
            else:
                self.logger.warning("매핑할 crypto_weights 데이터 없음")
                return False
            
            # 1. endpoint_access 데이터 확인
            if not self.endpoint_access:
                self.logger.warning("endpoint_access 데이터 없음. 기본값 1 사용")
            
            # 2. 각 crypto_weight에 access_count 추가
            updated_count = 0
            for crypto_weight in crypto_weights:
                end_point = crypto_weight.get('end_point') if isinstance(crypto_weight, dict) else getattr(crypto_weight, 'end_point', None)
                
                # end_point가 list인 경우 첫 번째 요소 추출 (정규화)
                if isinstance(end_point, list):
                    end_point = end_point[0] if end_point else ''
                
                if end_point:
                    access_cnt = self._get_endpoint_access_count(end_point)
                    if isinstance(crypto_weight, dict):
                        crypto_weight['Jenifer Access Count'] = access_cnt
                    else:
                        crypto_weight.access_cnt = access_cnt
                    updated_count += 1
                else:
                    # 엔드포인트가 공백이면 0으로 설정
                    if isinstance(crypto_weight, dict):
                        crypto_weight['Jenifer Access Count'] = 0
                    else:
                        crypto_weight.access_cnt = 0
            
            self.logger.info(f"{updated_count}개 항목의 Jenifer Access Count 업데이트 완료")
            
            # 업데이트된 데이터를 self.crypto_weights에 저장 (Excel 생성에서 사용)
            self.crypto_weights = crypto_weights
            
            # 최종 JSON 파일 저장 (모든 정보 포함)
            output_file = self.artifacts_dir / f"crypto_weight_{self.timestamp}.json"
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(crypto_weights, f, indent=2, ensure_ascii=False)
            self.logger.info(f"crypto_weight_{self.timestamp}.json 저장 완료 ({len(crypto_weights)}개 항목)")
            
            return True
        
        except Exception as e:
            self.logger.error(f"runtime 데이터 매핑 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _load_endpoint_access_dict(self) -> Dict[str, int]:
        """엔드포인트 호출 데이터를 Dict로 로드 (내부 함수)
        
        .applycrypto/endpoint_access.txt에서 엔드포인트별 호출 횟수 로드
        형식: endpoint,access_count (쉼표 구분)
        
        Returns:
            Dict[str, int]: {endpoint: access_count, ...}
        """
        endpoint_access = {}
        
        try:
            # .applycrypto/endpoint_access.txt 로드
            txt_path = os.path.join(self.target_project, '.applycrypto', 'endpoint_access.txt')
            if not os.path.exists(txt_path):
                self.logger.warning(f"endpoint_access.txt 없음: {txt_path}")
                return {}
            
            self.logger.debug(f"endpoint_access.txt 로드 중: {txt_path}")
            
            with open(txt_path, 'r', encoding='utf-8') as f:
                line_count = 0
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    line_count += 1
                    # 쉼표(,) 또는 whitespace(탭/공백)로 split
                    if ',' in line:
                        parts = [p.strip() for p in line.split(',')]
                    else:
                        parts = line.split()
                    if len(parts) >= 2:
                        endpoint = parts[0]
                        try:
                            access_count = int(parts[1])
                            endpoint_access[endpoint] = access_count
                            # 처음 3개만 로깅
                            if line_count <= 3:
                                self.logger.debug(f"endpoint: '{endpoint}', access_count: {access_count}")
                        except ValueError as ve:
                            self.logger.warning(f"Line {line_count} ValueError: {line} -> {parts}")
                    else:
                        self.logger.warning(f"Line {line_count} format error: {line}")
            
            if endpoint_access:
                self.logger.info(f"endpoint_access 로드 완료: {len(endpoint_access)}개 엔드포인트")
                # 첫 3개 샘플 로깅
                for i, (ep, cnt) in enumerate(list(endpoint_access.items())[:3]):
                    self.logger.debug(f"'{ep}' → {cnt}")
            else:
                self.logger.warning("endpoint_access.txt에서 유효한 데이터를 찾을 수 없음")
            
            return endpoint_access
        
        except Exception as e:
            self.logger.error(f"endpoint_access 로드 실패: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def _calculate_anyframe_weights(self) -> bool:
        """
        Step 4: 암복호화 가중치 계산
        
        처리 순서 (파일별):
        1. ksignUtil 호출이 있는 메서드만 추출
        2. call_trees 기반 endpoint_reachable_sigs 필터링 (완전일치)
           - endpoint 호출 체인에 없는 메서드는 LLM 호출 없이 SKIP
           - endpoint_reachable_sigs가 비어있으면 (call_trees 없음) 필터링 생략 + WARN 출력
        3. 필터링 통과한 메서드만 LLM 호출 → weight 계산
        4. LLM 호출/파싱 실패 시 해당 파일 SKIP (전체 중단 안 함)
        
        Returns:
            bool: 성공 여부 (모든 파일 실패 시에만 False)
        """
        try:
            crypto_weights = []
            
            # 1. LLM provider 초기화 (필수)
            llm_provider = None
            if self.config and self.config.get('llm_provider'):
                try:
                    from src.modifier.llm.llm_factory import create_llm_provider
                    llm_provider = create_llm_provider(self.config.get('llm_provider'))
                    if llm_provider:
                        self.logger.info(f"LLM 프로바이더 초기화: {type(llm_provider).__name__}")
                    else:
                        self.logger.error("LLM 프로바이더 초기화 실패 - llm_provider가 None")
                        return False
                except Exception as e:
                    self.logger.error(f"LLM 프로바이더 초기화 실패: {e}")
                    return False
            else:
                self.logger.error("config에 llm_provider 설정 필수")
                return False
            
            # 2. ksignUtil이 사용된 파일 순회 (파일별 그룹화)
            if not hasattr(self, 'files_with_ksignutil'):
                self.logger.warning("files_with_ksignutil 정의 안 됨. Step 2 먼저 실행 필요")
                return False
            
            if not self.files_with_ksignutil:
                self.logger.warning("ksignUtil이 포함된 파일 없음")
                return True  # 정상 종료 (항목 없음)
            
            total_files = len(self.files_with_ksignutil)
            self.logger.info(f"{total_files}개 파일에서 weight 계산 시작...")
            
            # === NEW: 파일별 성공/실패 통계 ===
            successful_files = []
            failed_files = []  # (file_path, class_name, reason)
            
            for file_idx, (rel_path, file_path) in enumerate(self.files_with_ksignutil.items(), 1):
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        file_content = f.read()
                    
                    # 3. 파일에서 클래스명 추출
                    class_name_match = re.search(r'public\s+(?:class|interface)\s+(\w+)', file_content)
                    if not class_name_match:
                        self.logger.warning(f"{file_idx}/{total_files} - 클래스명 추출 불가: {rel_path} (SKIP)")
                        failed_files.append((rel_path, "N/A", "클래스명 추출 실패"))
                        continue
                    
                    class_name = class_name_match.group(1)
                    
                    # 4. 파일의 메서드 추출
                    method_blocks = self._extract_method_blocks(file_path)
                    
                    # 5. ksignUtil 호출이 있는 메서드만 필터링
                    methods_with_ksignutil = []
                    for method_block in method_blocks:
                        method_content = method_block.get('content', '')
                        
                        # 이 메서드에서 ksignUtil 호출 개수 확인
                        ksignutil_calls = self._count_ksignutil_calls(method_content)
                        if ksignutil_calls > 0:
                            method_block['ksignutil_count'] = ksignutil_calls
                            methods_with_ksignutil.append(method_block)
                    
                    if not methods_with_ksignutil:
                        self.logger.warning(f"{file_idx}/{total_files} - {class_name}: ksignUtil 메서드 없음(SKIP)")
                        failed_files.append((rel_path, class_name, "ksignUtil 메서드 없음"))
                        continue
                    
                    # === endpoint 도달 가능 메서드만 LLM 분석 대상으로 한정 ===
                    # call_trees에 없는 메서드 = endpoint 호출 체인에 없음 = 가중치 계산해도 보고서에 무의미
                    if self.endpoint_reachable_sigs:
                        methods_with_endpoint = []
                        methods_skipped_no_endpoint = []
                        for method_block in methods_with_ksignutil:
                            method_name = method_block.get('name', '')
                            if isinstance(method_name, list):
                                method_name = method_name[0] if method_name else ''
                            method_sig_candidate = f"{class_name}.{method_name}"
                            if method_sig_candidate in self.endpoint_reachable_sigs:
                                methods_with_endpoint.append(method_block)
                            else:
                                methods_skipped_no_endpoint.append(method_sig_candidate)
                        
                        if methods_skipped_no_endpoint:
                            self.logger.debug(f"{file_idx}/{total_files} - {class_name}: endpoint 없어 제외 {len(methods_skipped_no_endpoint)}개: {methods_skipped_no_endpoint[:3]}")
                        
                        if not methods_with_endpoint:
                            self.logger.warning(f"{file_idx}/{total_files} - {class_name}: 모든 메서드가 endpoint 체인에 없음(SKIP)")
                            continue
                        
                        methods_with_ksignutil = methods_with_endpoint
                    else:
                        # call_trees가 비어있으면 필터링 불가 → 모든 메서드 LLM 전송 (비효율적)
                        self.logger.warning("Endpoint 못찾음 (LLM call SKIP)")
                    
                    self.logger.info(f"{file_idx}/{total_files} - {class_name}: {len(methods_with_ksignutil)}개 메서드 LLM 분석 중...")
                    
                    # 6. 파일 단위로 LLM 호출 (용허적: 실패해도 다음 파일로 계속)
                    file_weights = self._calculate_file_weights_with_llm(
                        llm_provider,
                        file_path,
                        class_name,
                        file_content,
                        methods_with_ksignutil
                    )
                    
                    # === NEW: LLM 호출 또는 파싱 실패해도 계속 진행 ===
                    if not file_weights:
                        self.logger.warning(f"{file_idx}/{total_files} - LLM 호출 또는 파싱 실패: {class_name} ({len(methods_with_ksignutil)}개 메서드 전송됨, SKIP)")
                        failed_files.append((rel_path, class_name, "LLM 호출/파싱 실패"))
                        continue  # 다음 파일로 계속
                    
                    successful_files.append(rel_path)
                    
                    # 7. 각 메서드의 weight를 토대로 crypto_weight 생성
                    for weight_info in file_weights:
                        method_name = weight_info.get('method_name', '')
                        # method_name 정규화: 배열이면 첫 요소
                        if isinstance(method_name, list):
                            method_name = method_name[0] if method_name else ''
                        
                        method_sig = f"{class_name}.{method_name}"
                        
                        # 해당 메서드의 ksignUtil 호출 개수 찾기
                        ksignutil_count = 0
                        for method in methods_with_ksignutil:
                            method_from_list = method.get('name')
                            # method name 정규화: 배열이면 첫 요소
                            if isinstance(method_from_list, list):
                                method_from_list = method_from_list[0] if method_from_list else ''
                            
                            if method_from_list == method_name:
                                ksignutil_count = method.get('ksignutil_count', 0)
                                break
                        
                        if ksignutil_count == 0:
                            continue
                        
                        # Total Weight 계산 = base_weight(LLM이 이미 data_type 적용) × ksignUtil 호출 개수
                        # 주의: LLM은 이미 "Base Weight = (encrypt + decrypt) × data_type_multiplier"로 계산해서 반환
                        # 따라서 여기서는 ksignutil_count만 곱하면 됨 (data_type_multiplier는 불필요)
                        base_weight = weight_info.get('base_weight', weight_info.get('Base Weight', 1))
                        
                        # CryptoWeight 객체 생성 - LLM 응답 형식과 통일 (Anyframe용)
                        crypto_weight = {
                            'method_signature': method_sig,
                            'class_path': weight_info.get('class_path', str(file_path)),
                            'table_name': '',  # Anyframe에서는 사용하지 않음
                            'query_id': '',    # Anyframe에서는 사용하지 않음
                            'end_point': None,  # Step 6: call_graph 매핑에서 채워짐
                            'input_fields_count': 0,  # Spring과의 호환성
                            'output_fields_count': 0,  # Spring과의 호환성
                            # === LLM 응답 형식 필드 (통일됨) ===
                            'data_type': weight_info.get('data_type', 'single'),
                            'loop_depth': weight_info.get('loop_depth', 0),
                            'loop_structure': weight_info.get('loop_structure', ''),
                            'multiplier': weight_info.get('multiplier', '1'),  # 문자열로 통일
                            'dep0_crypto_count': weight_info.get('dep0_crypto_count', 0),  # 루프 밖
                            'dep1_crypto_count': weight_info.get('dep1_crypto_count', 0),  # depth=1 루프 안
                            'dep2_crypto_count': weight_info.get('dep2_crypto_count', 0),  # depth=2 루프 안
                            # === 최종 계산 필드 ===
                            'crypto_weight': base_weight,  # 최종 가중치 = base_weight(LLM이 data_type 적용)
                            'access_cnt': 0,  # Step 6: 런타임 호출빈도 (기본값 0, endpoint 매핑 후 업데이트)
                            'input_parameter_type': None,  # Spring과의 호환성
                            'output_parameter_type': None,  # Spring과의 호환성
                            'input_parameter_type_weight': 1.0,  # Spring과의 호환성
                            'output_parameter_type_weight': 1.0  # Spring과의 호환성
                        }
                        
                        crypto_weights.append(crypto_weight)
                
                except Exception as e:
                    self.logger.warning(f"{file_idx}/{total_files} - 파일 분석 예외 발생: {rel_path}: {e} (SKIP)")
                    failed_files.append((rel_path, "N/A", str(e)))
                    continue  # 다음 파일로 계속
            
            # === NEW: 파일별 성공/실패 통계 출력 ===
            self.logger.info("Step 4 처리 결과")
            self.logger.info(f"성공: {len(successful_files)}개 파일")
            if successful_files:
                for f in successful_files[:5]:  # 최대 5개만 표시
                    self.logger.debug(f"  - {f}")
                if len(successful_files) > 5:
                    self.logger.debug(f"  ... 외 {len(successful_files)-5}개")
            
            if failed_files:
                self.logger.warning(f"실패: {len(failed_files)}개 파일 (무시됨)")
                for rel_path, class_name, reason in failed_files[:5]:  # 최대 5개만 표시
                    self.logger.debug(f"  - {rel_path} ({class_name}): {reason}")
                if len(failed_files) > 5:
                    self.logger.debug(f"  ... 외 {len(failed_files)-5}개")
                
                # 실패 파일 리스트를 파일로 저장 (감사/추적용)
                failed_files_log = self.artifacts_dir / f"step4_failed_files_{self.timestamp}.json"
                with open(failed_files_log, 'w', encoding='utf-8') as f:
                    json.dump([
                        {
                            'file': rel_path,
                            'class': class_name,
                            'reason': reason
                        }
                        for rel_path, class_name, reason in failed_files
                    ], f, indent=2, ensure_ascii=False)
                self.logger.info(f"실패 파일 리스트 저장: step4_failed_files_{self.timestamp}.json")
            
            # 8. crypto_weights를 메모리에 저장 (중간 JSON 파일 생성하지 않음)
            if crypto_weights:
                self.logger.info(f"총 {len(crypto_weights)}개 weight 계산 완료 ({len(successful_files)}개 파일에서 추출)")
                self.crypto_weights = crypto_weights  # Step 4에서 사용할 수 있도록 저장
                
                # === NEW: Anyframe Dict에 대해 access_cnt 업데이트 ===
                # endpoint_access가 로드되어 있으면 Dict의 access_cnt를 업데이트
                if self.endpoint_access:
                    for crypto_weight_dict in crypto_weights:
                        # method_signature로부터 end_point를 추론할 수 없으므로,
                        # Step 7에서 call_graph 매핑 후 업데이트하도록 미루기
                        # (여기서는 placeholder 유지)
                        pass
            else:
                # === NEW: 일부 파일 실패해도 계속 진행 ===
                if len(successful_files) == 0 and len(failed_files) > 0:
                    # 모든 파일이 실패한 경우만 return False
                    self.crypto_weights = []
                    self.logger.error(f"Step 4 결과가 비어있음 - 계산된 crypto_weight 항목이 없습니다")
                    self.logger.error(f"모든 파일이 실패하였습니다. 원인: AST 메서드 추출 실패, target utility exact match 불일치, LLM 응답 누락")
                    return False
                else:
                    # 일부 파일이 성공한 경우 계속 진행 (부분 성공)
                    self.crypto_weights = crypto_weights if crypto_weights else []
                    if not crypto_weights:
                        self.logger.warning(f"계산된 crypto_weight 항목이 없음 (일부 파일만 처리됨, 계속 진행)")
            
            self.logger.info(f"Step 4 완료: {len(crypto_weights)}개 weight 계산 완료 (성공: {len(successful_files)}개, 실패: {len(failed_files)}개)")
            return True
        
        except Exception as e:
            self.logger.error(f"weight 계산 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _enrich_crypto_weights_with_endpoint_paths(self, crypto_weights: List[Dict], call_graph: Dict) -> List[Dict]:
        """
        Step 6: call_graph.endpoints에서 method_signature별 end_point 경로 추출
        
        call_graph.endpoints 배열을 순회하여:
        1. method_signature로 crypto_weight 항목 찾기
        2. endpoint의 path를 end_point 필드에 할당
        - path가 배열이면 첫 번째 요소, 문자열이면 직접 사용
        - method_signature 포맷 차이 커버 (완전일치 먼저, 부분일치 나중)
        
        Args:
            crypto_weights: Step 3-1에서 생성된 crypto_weight 리스트
            call_graph: call_graph.json 데이터
        
        Returns:
            List[Dict]: "end_point" 필드가 채워진 crypto_weights
        """
        try:
            if not call_graph or not call_graph.get('endpoints'):
                self.logger.warning(f"call_graph.endpoints가 비어있음")
                return crypto_weights
            
            # 1. crypto_weights를 method_signature로 인덱싱
            method_sig_to_crypto_weight = {}
            total_crypto_weights = len(crypto_weights)
            for crypto_weight in crypto_weights:
                method_sig_raw = crypto_weight.get('method_signature', '')
                # method_signature 정규화: 배열이면 첫 요소, 문자열이면 그대로
                if isinstance(method_sig_raw, list):
                    method_sig = method_sig_raw[0] if method_sig_raw else ''
                else:
                    method_sig = method_sig_raw
                if method_sig:
                    method_sig_to_crypto_weight[method_sig] = crypto_weight
            
            self.logger.debug(f"crypto_weights 총 {total_crypto_weights}개, 인덱싱된 method_sig 총 {len(method_sig_to_crypto_weight)}개")
            if method_sig_to_crypto_weight:
                self.logger.debug(f"crypto_weights method_sig 샘플: {list(method_sig_to_crypto_weight.keys())[:3]}")
            
            self.logger.debug(f"call_graph.endpoints 총 {len(call_graph.get('endpoints', []))}개")
            
            # 2. call_graph.endpoints에서 method_signature → path 매핑
            endpoints_matched = 0
            endpoints_not_matched = []
            
            for endpoint in call_graph.get('endpoints', []):
                method_sig_raw = endpoint.get('method_signature', '')
                # method_signature 정규화: 배열이면 첫 요소, 문자열이면 그대로
                if isinstance(method_sig_raw, list):
                    method_sig = method_sig_raw[0] if method_sig_raw else ''
                else:
                    method_sig = method_sig_raw
                
                # path 필드: 배열 또는 문자열 모두 처리
                path_raw = endpoint.get('path', [])
                end_point_path = None
                
                # path가 배열인 경우
                if isinstance(path_raw, list) and len(path_raw) > 0:
                    end_point_path = path_raw[0]
                # path가 문자열인 경우
                elif isinstance(path_raw, str) and path_raw:
                    end_point_path = path_raw
                
                if not method_sig or not end_point_path:
                    continue
                
                # 1단계: 정확한 매칭 시도
                if method_sig in method_sig_to_crypto_weight:
                    method_sig_to_crypto_weight[method_sig]['end_point'] = end_point_path
                    endpoints_matched += 1
                    self.logger.debug(f"정확매칭: {method_sig} → {end_point_path}")
                else:
                    # 2단계: 부분 매칭 시도 (메서드명 마지막 부분 비교)
                    method_name = method_sig.split('.')[-1] if '.' in method_sig else method_sig
                    partial_matched = False
                    
                    for indexed_sig, crypto_weight in method_sig_to_crypto_weight.items():
                        # 메서드명 끝부분 비교
                        indexed_method_name = indexed_sig.split('.')[-1] if '.' in indexed_sig else indexed_sig
                        if method_name == indexed_method_name and not crypto_weight.get('end_point'):
                            crypto_weight['end_point'] = end_point_path
                            endpoints_matched += 1
                            self.logger.debug(f"부분매칭: {indexed_sig} ({method_name}) ← {method_sig} → {end_point_path}")
                            partial_matched = True
                            break
                    
                    if not partial_matched:
                        endpoints_not_matched.append(method_sig)
            
            self.logger.info(f"Step 6 완료: {endpoints_matched}개 메서드 엔드포인트 경로 매핑")
            
            # 3. end_point가 여전히 None인 항목 확인
            not_matched_in_weights = sum(1 for w in crypto_weights if not w.get('end_point'))
            if not_matched_in_weights > 0:
                self.logger.warning(f"{not_matched_in_weights}/{total_crypto_weights}개 항목이 아직 end_point 없음 (다른 단계에서 채워질 예정)")
            
            return crypto_weights
        
        except Exception as e:
            self.logger.error(f"Step 6 엔드포인트 경로 매핑 실패: {e}")
            import traceback
            traceback.print_exc()
            return crypto_weights

    def _enrich_crypto_weights_with_endpoint_access(self, crypto_weights: List[Dict], endpoint_access: Dict[str, int], call_graph: Dict) -> List[Dict]:
        """
        Step 6: call_graph call_trees 역탐색으로 메서드→엔드포인트 매핑 및 호출빈도 설정
        
        처리 (역방향 탐색):
        1. call_trees 순회: endpoint_path ← method_chain의 각 메서드
        2. 각 method_signature에 대해:
           - 이 메서드를 호출하는 endpoint 목록 찾기
           - endpoint_access에서 해당 endpoint별 접근 횟수 조회
           - 총합을 crypto_weight의 접근 횟수로 설정
        3. 필드 업데이트:
           - end_point: 호출 endpoint (첫 번째)
           - Jenifer Access Count: 총 접근 횟수
           - access_cnt: Anyframe Dict용 (호환성)
        
        Args:
            crypto_weights: LLM이 생성한 weight 정보
            endpoint_access: {endpoint: access_count} Dict
            call_graph: call_trees 포함 (역탐색용)
        
        Returns:
            List[Dict]: end_point + 호출빈도 업데이트된 crypto_weights
        """
        try:
            if not call_graph or not call_graph.get('call_trees'):
                self.logger.warning("call_graph가 비어있거나 call_trees 없음")
                if call_graph:
                    self.logger.debug(f"call_graph keys: {list(call_graph.keys())}")
                return crypto_weights
            
            call_trees = call_graph.get('call_trees', [])
            self.logger.debug(f"call_trees 개수: {len(call_trees)}")
            
            if not endpoint_access:
                self.logger.info("endpoint_access Dict 없음 - access_count는 0으로 유지, end_point는 정상 설정됨")
            else:
                self.logger.debug(f"endpoint_access Dict 크기: {len(endpoint_access)}")
            
            # 1. endpoints 배열에서 method_sig → endpoint_path 매핑 구축
            # call_graph 구조: endpoints[]{path, method_signature} + call_trees[]{method_signature, children[]}
            # call_trees 루트(Controller)의 method_signature가 endpoints 배열의 method_signature와 대응됨
            # 리스트가 중첩될 수 있으므로 문자열이 나올 때까지 언래핑
            def _unwrap_to_str(val) -> str:
                while isinstance(val, list):
                    val = val[0] if val else ''
                return val or ''
            
            endpoint_path_map = {}  # controller_method_sig → endpoint_path
            for ep_entry in call_graph.get('endpoints', []):
                sig = _unwrap_to_str(ep_entry.get('method_signature', ''))
                path = _unwrap_to_str(ep_entry.get('path', ''))
                if sig and path:
                    endpoint_path_map[sig] = path
            
            self.logger.debug(f"endpoint_path_map 구축 완료: {len(endpoint_path_map)}개 Controller endpoint")
            
            # 2. call_trees를 재귀 순회하여 method_sig → [endpoint_path] 역매핑 구축
            # 트리 루트(Controller)의 endpoint_path를 모든 하위 자식들에게 전파
            method_to_endpoints = {}  # method_sig → [endpoint_paths]
            
            def _collect_method_endpoints(node: dict, endpoint_path: str):
                sig = _unwrap_to_str(node.get('method_signature', ''))
                # endpoint_path도 혹시 리스트면 언래핑 (방어)
                ep = _unwrap_to_str(endpoint_path) if not isinstance(endpoint_path, str) else endpoint_path
                if sig and ep:
                    if sig not in method_to_endpoints:
                        method_to_endpoints[sig] = []
                    if ep not in method_to_endpoints[sig]:
                        method_to_endpoints[sig].append(ep)
                for child in node.get('children', []):
                    _collect_method_endpoints(child, ep or endpoint_path)
            
            for call_tree in call_trees:
                root_sig_raw = call_tree.get('method_signature', '')
                root_sig = root_sig_raw[0] if isinstance(root_sig_raw, list) else root_sig_raw
                ep_path = endpoint_path_map.get(root_sig, '')
                if not ep_path:
                    ep_path = root_sig
                _collect_method_endpoints(call_tree, ep_path)
            
            self.logger.debug(f"method_to_endpoints 구축 완료: {len(method_to_endpoints)}개 메서드")
            if method_to_endpoints:
                sample = list(method_to_endpoints.items())[:2]
                for method, endpoints in sample:
                    self.logger.debug(f"{method} → {endpoints}")
            else:
                self.logger.warning("method_to_endpoints가 비어있음 - Step 6 매핑 불가!")
            
            # 2. crypto_weights에서 각 메서드의 "Jenifer Access Count" 업데이트
            matched_count = 0
            for crypto_weight in crypto_weights:
                method_sig_raw = crypto_weight.get('method_signature', '')
                # method_signature 정규화: 배열이면 첫 요소, 문자열이면 그대로
                if isinstance(method_sig_raw, list):
                    method_sig = method_sig_raw[0] if method_sig_raw else ''
                else:
                    method_sig = method_sig_raw
                
                # 이 메서드를 호출하는 엔드포인트들 찾기
                related_endpoints = method_to_endpoints.get(method_sig, [])
                
                if related_endpoints:
                    matched_count += 1
                    # 호출하는 엔드포인트들의 access_count 합산
                    total_access_count = 0
                    for endpoint in related_endpoints:
                        # 최종 방어: endpoint가 리스트인 경우 언래핑
                        if isinstance(endpoint, list):
                            endpoint = endpoint[0] if endpoint else ''
                        access_count = endpoint_access.get(endpoint, 0) if endpoint else 0
                        total_access_count += access_count
                    
                    # Dict와 CryptoWeight 객체 모두 호환되도록 업데이트
                    crypto_weight['Jenifer Access Count'] = total_access_count
                    crypto_weight['access_cnt'] = total_access_count  # Anyframe Dict도 함께 업데이트
                    crypto_weight['end_point'] = related_endpoints[0]  # 주 엔드포인트 (첫번째)
                else:
                    # 관련 엔드포인트 없음 (call_graph에 없는 메서드)
                    crypto_weight['Jenifer Access Count'] = 0
                    crypto_weight['access_cnt'] = 0  # Anyframe Dict도 함께 업데이트
                    crypto_weight['end_point'] = None
            
            matched_with_access = len([x for x in crypto_weights if x.get('Jenifer Access Count', 0) > 0])
            self.logger.info(f"Step 6 완료: {matched_count}/{len(crypto_weights)}개 메서드 매핑, {matched_with_access}개 메서드가 엔드포인트 호출빈도 설정됨")
            return crypto_weights
        
        except Exception as e:
            self.logger.error(f"엔드포인트 접근 횟수 처리 실패: {e}")
            import traceback
            traceback.print_exc()
            return crypto_weights

    def _extract_method_blocks(self, file_path: str) -> List[Dict[str, str]]:
        """
        AST Parser를 사용하여 파일의 모든 메서드 추출
        
        Args:
            file_path: Java 파일 경로
        
        Returns:
            List[Dict]: 메서드 정보 리스트 ({'name': 'methodName', 'content': '메서드 바디'})
        """
        method_blocks = []
        
        try:
            parser = JavaASTParser()
            # LLM이 실제 brace/indent 구조를 그대로 볼 수 있도록 원본 소스로 파싱한다.
            tree, error = parser.parse_file(Path(file_path), remove_comments=True)
            
            if error or tree is None:
                self.logger.warning(f"AST 파싱 실패 (_extract_method_blocks): {file_path}: {error}")
                return []
            
            # 파일 내용 로드 (메서드 바디 추출용)
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()
            except Exception as e:
                self.logger.warning(f"파일 읽기 실패: {file_path}: {e}")
                return []
            
            class_infos = parser.extract_class_info(tree, Path(file_path))
            
            for class_info in class_infos:
                for method in class_info.methods:
                    method_name = method.name
                    
                    # 1. AST에서 추출된 body가 있다면 우선 사용 (권장)
                    if method.body:
                        method_content = method.body
                    # 2. 없으면 라인 번호를 사용하여 파일에서 추출 (Fallback)
                    else:
                        try:                           
                            start_idx = method.line_number - 1 if method.line_number > 0 else 0
                            end_idx = method.end_line_number if method.end_line_number > 0 else len(lines)
                            
                            if start_idx < len(lines) and end_idx <= len(lines):
                                method_content = ''.join(lines[start_idx:end_idx])
                            else:
                                method_content = f"// {method_name} code not extracted"
                        except Exception:
                            method_content = f"// {method_name} code extraction failed"
                    
                    method_blocks.append({
                        'name': method_name,
                        'content': method_content
                    })
            
            return method_blocks
        
        except Exception as e:
            self.logger.warning(f"메서드 블록 추출 오류 (AST): {e}")
            return []

    def _find_matching_brace(self, text: str, open_brace_index: int) -> int:
        """주어진 여는 중괄호의 대응 닫는 중괄호 위치를 찾습니다."""
        brace_depth = 0

        for index in range(open_brace_index, len(text)):
            char = text[index]
            if char == '{':
                brace_depth += 1
            elif char == '}':
                brace_depth -= 1
                if brace_depth == 0:
                    return index

        return -1

    def _sanitize_java_code_for_analysis(self, method_code: str) -> str:
        """주석/문자열을 공백으로 치환해 구조 분석 시 오탐을 줄입니다."""
        if not method_code:
            return ''

        sanitized_code = method_code
        for pattern, flags in [
            (r'/\*.*?\*/', re.DOTALL),
            (r'//[^\r\n]*', 0),
            (r'"(?:\\.|[^"\\])*"', 0),
            (r"'(?:\\.|[^'\\])*'", 0),
        ]:
            sanitized_code = re.sub(
                pattern,
                lambda match: ' ' * len(match.group(0)),
                sanitized_code,
                flags=flags,
            )

        return sanitized_code

    def _collect_loop_ranges(self, sanitized_code: str) -> List[Tuple[int, int]]:
        """for/while loop 의 brace 범위를 수집합니다."""
        loop_ranges = []

        for loop_match in re.finditer(r'\b(for|while)\s*\(', sanitized_code):
            open_brace_index = sanitized_code.find('{', loop_match.end())
            if open_brace_index == -1:
                continue

            close_brace_index = self._find_matching_brace(sanitized_code, open_brace_index)
            if close_brace_index == -1:
                continue

            loop_ranges.append((open_brace_index, close_brace_index))

        return loop_ranges

    def _build_target_crypto_patterns(self, ksignutil_methods: Optional[List[str]] = None) -> Dict[str, List[re.Pattern]]:
        """config 로 전달된 Encryption Utilities 만 정확히 매칭하는 패턴을 생성합니다.

        ksignUtils_pattern + policyId 가 config 에 정의된 경우 policyId-aware 정밀 패턴을 사용합니다.
        명시적으로 ksignutil_methods 를 전달한 경우(sanity check 제외)는 단순 매칭을 유지합니다.
        """
        # policyId-aware 정밀 패턴: self.config 기반 + ksignUtils_pattern + policyId 모두 있을 때만
        if ksignutil_methods is None and self.ksignutil_patterns and self.policy_ids:
            return self._build_policyid_aware_patterns()

        patterns = {
            'encrypt': [],
            'decrypt': [],
        }

        # 패턴 기반 처리: ksignutil_methods가 None이고 self.ksignutil_patterns가 있으면 패턴에서 메서드명 추출
        if ksignutil_methods is None and self.ksignutil_patterns:
            # 패턴에서 메서드명만 추출
            # 예: "com.example.SliEncryptionUtil.encrypt(String) -> String" → "SliEncryptionUtil.encrypt"
            configured_methods = []
            for pattern in self.ksignutil_patterns:
                # 패턴 문자열에서 마지막 괄호 전까지 추출
                # "com.example.Package.Method(...)" → "Package.Method"
                m = re.match(r'^([a-zA-Z0-9_$.]+)\(', pattern)
                if m:
                    full_method = m.group(1)  # "com.example.Package.Method"
                    # 마지막 두 부분만 유지
                    parts = full_method.split('.')
                    if len(parts) >= 2:
                        configured_methods.append(f"{parts[-2]}.{parts[-1]}")
                    elif len(parts) == 1:
                        configured_methods.append(parts[0])
        else:
            configured_methods = ksignutil_methods if ksignutil_methods is not None else []
        
        for utility in configured_methods:
            if not utility:
                continue

            qualifier, _, method_name = utility.rpartition('.')
            if not method_name:
                method_name = qualifier
                qualifier = ''

            method_name_lower = method_name.lower()
            if 'encrypt' in method_name_lower:
                bucket = 'encrypt'
            elif 'decrypt' in method_name_lower:
                bucket = 'decrypt'
            else:
                continue

            if qualifier:
                qualifier_pattern = r'\b' + r'\s*\.\s*'.join(re.escape(part) for part in qualifier.split('.'))
                pattern_text = qualifier_pattern + r'\s*\.\s*' + re.escape(method_name) + r'\s*\('
            else:
                pattern_text = r'\b' + re.escape(method_name) + r'\s*\('

            patterns[bucket].append(re.compile(pattern_text))

        return patterns

    def _build_policyid_aware_patterns(self) -> Dict[str, List[re.Pattern]]:
        """ksignUtils_pattern + policyId 기반 정밀 매칭 패턴을 생성합니다.

        [String/policyId 변형] 설정된 policyId 값을 가진 호출만 매칭
          - '.' 포함 → 상수 참조 (e.g., SliEncryptionConstants.Policy.NAME)
          - 그 외 → 소스코드 내 문자열 리터럴 (e.g., "P017")
          - 원본 코드(unsanitized)에서 매칭 → sanitize에 의해 리터럴 "P017"이 지워지기 때문

        [List 변형] 인자가 Java 식별자로 시작하는 호출만 매칭 (문자열 리터럴 및 policyId 상수 제외)
          - [a-zA-Z_$] 시작 → 변수명(List 인자)
          - policyId 상수는 negative lookahead로 제외 → 이중 카운트 방지
          - sanitize 코드에서도 안전하게 동작 (리터럴 "P017" → 공백 → [a-zA-Z_$] 불일치)
        """
        patterns: Dict[str, List[re.Pattern]] = {'encrypt': [], 'decrypt': []}

        # --- policyId 대안 패턴 구성 ---
        policyid_alts = []
        for pid in self.policy_ids:
            if not pid:
                continue
            if '.' in pid:
                # 상수 참조 (e.g., SliEncryptionConstants.Policy.NAME)
                policyid_alts.append(re.escape(pid))
            else:
                # 문자열 리터럴 값 → 소스코드에서 "P017" 형태로 등장
                policyid_alts.append(r'"' + re.escape(pid) + r'"')

        if not policyid_alts:
            return patterns

        policyid_re = r'(?:' + r'|'.join(policyid_alts) + r')'

        # List 변형용: policyId 상수만 negative lookahead로 제외 (리터럴은 [a-zA-Z_$] 체크로 자동 제외)
        # 점(.)이 오지 않도록 보장: 변수명 다음에는 반드시 쉼표/괄호만 올 수 있음
        const_alts = [alt for alt in policyid_alts if not alt.startswith('"')]
        if const_alts:
            # 설정된 상수는 명시적으로 배제, 순수 변수명만 허용 (점 포함 불가)
            list_start = r'(?!(?:' + r'|'.join(const_alts) + r')\s*[,)])(?=[a-zA-Z_$][a-zA-Z0-9_$]*\s*[,)])[a-zA-Z_$][a-zA-Z0-9_$]*'
        else:
            # 변수명 다음에 반드시 쉼표/괄호 → 점(.)이 올 수 없음
            list_start = r'(?=[a-zA-Z_$][a-zA-Z0-9_$]*\s*[,)])[a-zA-Z_$][a-zA-Z0-9_$]*'

        # --- ksignUtils_pattern 분석: 버킷별 String/List 변형 및 policyId 위치 파악 ---
        bucket_info: Dict[str, Dict] = {}

        for pat_str in self.ksignutil_patterns:
            paren_idx = pat_str.find('(')
            if paren_idx == -1:
                continue
            method_part = pat_str[:paren_idx].strip()
            params_part = pat_str[paren_idx + 1:].rstrip(')').strip()

            qualifier, _, method_name = method_part.rpartition('.')
            if not method_name:
                continue

            ml = method_name.lower()
            if 'encrypt' in ml:
                bucket = 'encrypt'
            elif 'decrypt' in ml:
                bucket = 'decrypt'
            else:
                continue

            if bucket not in bucket_info:
                bucket_info[bucket] = {
                    'qualifier': qualifier,
                    'method_name': method_name,
                    'has_string': False,
                    'has_list': False,
                    'policyid_pos': None,
                }

            params = [p.strip() for p in params_part.split(',') if p.strip()]
            this_has_policyid = False
            for i, param in enumerate(params):
                parts = param.split()
                if not parts:
                    continue
                pname = parts[-1].lower() if len(parts) > 1 else parts[0].lower()
                if pname == 'policyid':
                    this_has_policyid = True
                    bucket_info[bucket]['has_string'] = True
                    if bucket_info[bucket]['policyid_pos'] is None:
                        bucket_info[bucket]['policyid_pos'] = i
                    break

            if not this_has_policyid:
                # policyId 없는 변형 → List 파라미터 여부 확인
                has_list_param = any(
                    p.strip().split()[0].lower() == 'list'
                    for p in params if p.strip() and p.strip().split()
                )
                if has_list_param:
                    bucket_info[bucket]['has_list'] = True

        # --- 버킷별 regex 패턴 컴파일 ---
        for bucket, binfo in bucket_info.items():
            qualifier = binfo['qualifier']
            method_name = binfo['method_name']
            qual_re = (
                r'\b' + r'\s*\.\s*'.join(re.escape(part) for part in qualifier.split('.'))
                if qualifier else r'\b'
            )
            base = qual_re + r'\s*\.\s*' + re.escape(method_name) + r'\s*\('

            # [1] policyId 필터 패턴 (String 변형)
            #     원본(unsanitized) 코드에서 탐색해야 정확 → _analyze_target_crypto_calls에서 처리
            if binfo['has_string']:
                pos = binfo['policyid_pos'] or 0
                if pos == 0:
                    # e.g., encrypt("P017", targetStr)
                    str_pattern = base + r'\s*' + policyid_re + r'\s*,'
                else:
                    # e.g., decrypt(0, "P017", targetStr) → pos개 앞 인자 건너뜀
                    skip = r'\s*[^,()]+\s*,\s*' * pos
                    str_pattern = base + skip + policyid_re + r'\s*,'
                patterns[bucket].append(re.compile(str_pattern))

            # [2] List 변형 패턴
            #     인자가 Java 식별자([a-zA-Z_$]) 시작 → 리터럴 문자열/policyId 상수 자동 제외
            #     sanitized/원본 코드 모두에서 안전하게 동작
            if binfo['has_list']:
                pos = binfo['policyid_pos']
                if pos and pos > 0:
                    # decrypt 형태: int 앞 인자 건너뜀 후 list_start
                    skip = r'\s*[^,()]+\s*,\s*' * pos
                    list_pattern = base + skip + list_start
                else:
                    # encrypt 형태
                    list_pattern = base + r'\s*' + list_start
                patterns[bucket].append(re.compile(list_pattern))

            # String/List 모두 미정의 시 단순 메서드명 매칭으로 폴백
            if not binfo['has_string'] and not binfo['has_list']:
                patterns[bucket].append(re.compile(base))

        return patterns

    def _analyze_target_crypto_calls(self, method_code: str, ksignutil_methods: Optional[List[str]] = None) -> Dict[str, int]:
        """Encryption Utilities 에 포함된 호출을 루프 depth별로 집계합니다.

        dep0: 루프 밖 (loop depth = 0)
        dep1: 첫 번째 루프 안 (loop depth = 1)
        dep2: 두 번째 이상 루프 안 (loop depth >= 2, 상한선 2로 고정)
        """
        counts = {
            'dep0_crypto_count': 0,
            'dep1_crypto_count': 0,
            'dep2_crypto_count': 0,
            'max_loop_depth': 0,
            'total_calls': 0,
        }
        if not method_code:
            return counts

        sanitized_code = self._sanitize_java_code_for_analysis(method_code)
        loop_ranges = self._collect_loop_ranges(sanitized_code)
        target_patterns = self._build_target_crypto_patterns(ksignutil_methods)

        # policyId-aware 모드: String 변형 패턴은 원본 코드에서 탐색 (sanitizer가 "P017" 리터럴을 제거하기 때문)
        # List 변형 패턴은 sanitized 코드에서 탐색해도 안전 ([a-zA-Z_$] 시작 조건으로 구분)
        use_policyid_mode = (ksignutil_methods is None and bool(self.ksignutil_patterns) and bool(self.policy_ids))

        for crypto_kind, patterns in target_patterns.items():
            for pattern in patterns:
                # policyId-aware 모드에서 String 변형 패턴(policyId 인자 포함)은 원본 코드 사용
                # 판별 기준: 패턴에 '"' 또는 상수 참조가 포함되어 있으면 원본 코드 필요
                if use_policyid_mode and ('"' in pattern.pattern or any(
                    re.escape(pid) in pattern.pattern for pid in self.policy_ids if '.' in pid
                )):
                    search_code = method_code
                else:
                    search_code = sanitized_code

                for crypto_match in pattern.finditer(search_code):
                    # 원본 코드 탐색 시: sanitize 처리된 위치(문자열/주석 내부)는 건너뜀
                    if search_code is method_code:
                        pos = crypto_match.start()
                        if pos < len(sanitized_code) and sanitized_code[pos] == ' ' and method_code[pos] != ' ':
                            continue
                    crypto_index = crypto_match.start()
                    current_depth = sum(
                        1 for loop_start, loop_end in loop_ranges if loop_start < crypto_index < loop_end
                    )
                    dep_key = f'dep{min(current_depth, 2)}_crypto_count'
                    counts[dep_key] += 1
                    counts['max_loop_depth'] = max(counts['max_loop_depth'], current_depth)
                    counts['total_calls'] += 1

        return counts

    def _estimate_crypto_loop_depth(self, method_code: str) -> int:
        """메서드 코드에서 crypto call 이 실제 위치한 최대 loop depth를 추정합니다."""
        return self._analyze_target_crypto_calls(method_code).get('max_loop_depth', 0)

    def _recalculate_base_weight(self, result: Dict[str, Any]) -> int:
        """dep0/dep1/dep2 절대 depth 카운트 기준으로 Base Weight를 재계산합니다.

        dep0 × 1
        dep1 × 10          (max_loop=2 일 때 고정, paged/unpaged 무관)
        dep1 × 20|100      (max_loop=1 일 때 paged|unpaged)
        dep2 × 20|100      (max_loop=2 일 때 paged|unpaged)
        """
        dep0 = int(result.get('dep0_crypto_count', 0) or 0)
        dep1 = int(result.get('dep1_crypto_count', 0) or 0)
        dep2 = int(result.get('dep2_crypto_count', 0) or 0)
        loop_depth = int(result.get('loop_depth', 0) or 0)
        data_type  = result.get('data_type', 'single') or 'single'

        if loop_depth <= 0:
            return dep0

        if loop_depth == 1:
            if data_type == 'paged_list':
                return dep0 + 20 * dep1
            if data_type == 'unpaged_list':
                return dep0 + 100 * dep1
            return dep0 + dep1

        # loop_depth >= 2: dep1 × 10 고정, dep2 에 paged/unpaged multiplier
        if data_type == 'paged_list':
            return dep0 + 10 * dep1 + 20 * dep2
        if data_type == 'unpaged_list':
            return dep0 + 10 * dep1 + 100 * dep2
        return dep0 + dep1 + dep2

    def _apply_llm_weight_sanity_check(self, result: Dict[str, Any], method_code: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """LLM 응답을 실제 target utility 호출 수와 loop depth 기준으로 교정합니다."""
        audit = {
            'method_name': result.get('method_name', '<unknown>') if result else '<unknown>',
            'corrected': False,
            'reported_loop_depth': 0,
            'actual_loop_depth': 0,
            'reported_counts': {},
            'actual_counts': {},
            'fields_corrected': [],
        }
        if not method_code or not result:
            return result, audit

        try:
            reported_depth = int(result.get('loop_depth', 0) or 0)
        except (TypeError, ValueError):
            reported_depth = 0

        audit['reported_loop_depth'] = reported_depth

        actual_counts = self._analyze_target_crypto_calls(method_code)
        actual_depth = actual_counts.get('max_loop_depth', 0)
        audit['actual_loop_depth'] = actual_depth
        audit['actual_counts'] = {
            field: actual_counts.get(field, 0)
            for field in ['dep0_crypto_count', 'dep1_crypto_count', 'dep2_crypto_count']
        }

        for field in ['dep0_crypto_count', 'dep1_crypto_count', 'dep2_crypto_count']:
            try:
                reported_count = int(result.get(field, 0) or 0)
            except (TypeError, ValueError):
                reported_count = 0

            audit['reported_counts'][field] = reported_count

            actual_count = actual_counts.get(field, 0)
            if reported_count != actual_count:
                method_name = result.get('method_name', '<unknown>')
                self.logger.warning(f"{method_name}: {field}={reported_count} → exact target-call count corrected to {actual_count}")
                result[field] = actual_count
                audit['corrected'] = True
                audit['fields_corrected'].append(field)

        if reported_depth <= actual_depth:
            if actual_depth > reported_depth:
                result['loop_depth'] = actual_depth
                audit['corrected'] = True
                audit['fields_corrected'].append('loop_depth')
            result['Base Weight'] = self._recalculate_base_weight(result)
            return result, audit

        method_name = result.get('method_name', '<unknown>')
        self.logger.warning(f"{method_name}: LLM loop_depth={reported_depth} → sanity check corrected to {actual_depth}")

        result['loop_depth'] = actual_depth
        audit['corrected'] = True
        audit['fields_corrected'].append('loop_depth')

        if actual_depth == 0:
            result['data_type'] = 'single'
            result['loop_structure'] = ''
            result['multiplier'] = '1'
            audit['fields_corrected'].extend(['data_type', 'loop_structure', 'multiplier'])
        elif actual_depth == 1:
            loop_structure = result.get('loop_structure', '') or ''
            if '>' in loop_structure:
                result['loop_structure'] = loop_structure.split('>')[0].strip()
                audit['fields_corrected'].append('loop_structure')

            multiplier = result.get('multiplier', '') or ''
            multiplier_parts = re.split(r'\s*[×*?]\s*', multiplier, maxsplit=1)
            if len(multiplier_parts) > 1:
                result['multiplier'] = multiplier_parts[0].strip()
                audit['fields_corrected'].append('multiplier')

        result['Base Weight'] = self._recalculate_base_weight(result)
        audit['fields_corrected'].append('Base Weight')
        audit['fields_corrected'] = sorted(set(audit['fields_corrected']))
        return result, audit

    

    def _count_ksignutil_calls(self, content: str) -> int:
        """
        메서드 바디에서 ksignUtil 메서드 호출 개수 카운트
        
        Args:
            content: 메서드 내용
        
        Returns:
            int: 호출 개수
        """
        return self._analyze_target_crypto_calls(content).get('total_calls', 0)
    

    def _save_prompt_log(self, class_name: str, file_path: str, prompt_data: Dict, response_data: str, result_data: Any):
        """
        LLM 프롬프트 및 응답을 로깅 파일에 저장
        
        Args:
            class_name: 클래스명
            file_path: 파일 경로
            prompt_data: 프롬프트 딕셔너리
            response_data: LLM 응답 (원본)
            result_data: 파싱된 결과 (JSON)
        """
        try:
            # 파일명 생성 (클래스명 기반, 특수문자 제거)
            safe_class_name = re.sub(r'[<>:"/\\|?*]', '_', class_name)
            log_file = self.prompt_logs_dir / f"{safe_class_name}_{self.timestamp}.json"
            
            # 로그 데이터 구성
            log_data = {
                "timestamp": datetime.now().isoformat(),
                "class_name": class_name,
                "file_path": file_path,
                "prompt": {
                    "instruction": prompt_data.get('instruction', ''),
                    "target_methods": prompt_data.get('target_methods', []),
                    "file_content_preview": prompt_data.get('file_content', '')[:200] + "..." if len(prompt_data.get('file_content', '')) > 200 else prompt_data.get('file_content', '')
                },
                "llm_response_raw": response_data,
                "parsed_result": result_data,
                "sanity_audit": prompt_data.get('sanity_audit', []),
                "sanity_summary_file": prompt_data.get('sanity_summary_file', ''),
                "parse_status": "success" if result_data else "failed"
            }
            
            # JSON 저장
            with open(log_file, 'w', encoding='utf-8') as f:
                json.dump(log_data, f, indent=2, ensure_ascii=False)
            
            self.logger.debug(f"{log_file.name} 저장")
            return True
        except Exception as e:
            self.logger.warning(f"프롬프트 로그 저장 실패: {e}")
            return False

    def _save_sanity_audit_summary(self, class_name: str, file_path: str, sanity_audit: List[Dict[str, Any]], result_data: Any) -> Optional[str]:
        """보정이 발생한 메서드만 별도 요약 JSON으로 저장합니다."""
        corrected_audits = [audit for audit in (sanity_audit or []) if audit.get('corrected')]
        if not corrected_audits:
            return None

        try:
            safe_class_name = re.sub(r'[<>:"/\\|?*]', '_', class_name)
            summary_file = self.sanity_reports_dir / f"{safe_class_name}_{self.timestamp}_sanity_summary.json"
            result_lookup = {}
            if isinstance(result_data, list):
                for item in result_data:
                    if isinstance(item, dict):
                        method_name = item.get('method_name')
                        if method_name:
                            result_lookup[method_name] = item

            summary_data = {
                'timestamp': datetime.now().isoformat(),
                'class_name': class_name,
                'file_path': file_path,
                'total_methods_analyzed': len(result_data) if isinstance(result_data, list) else 0,
                'corrected_methods_count': len(corrected_audits),
                'corrected_methods': [],
            }

            for audit in corrected_audits:
                method_name = audit.get('method_name', '<unknown>')
                summary_data['corrected_methods'].append({
                    'method_name': method_name,
                    'reported_loop_depth': audit.get('reported_loop_depth', 0),
                    'actual_loop_depth': audit.get('actual_loop_depth', 0),
                    'reported_counts': audit.get('reported_counts', {}),
                    'actual_counts': audit.get('actual_counts', {}),
                    'fields_corrected': audit.get('fields_corrected', []),
                    'final_result': result_lookup.get(method_name, {}),
                })

            with open(summary_file, 'w', encoding='utf-8') as f:
                json.dump(summary_data, f, indent=2, ensure_ascii=False)

            self.logger.debug(f"{summary_file.name} 저장")
            return str(summary_file)
        except Exception as e:
            self.logger.warning(f"sanity summary 저장 실패: {e}")
            return None

    def _load_ksign_weight_prompt(self) -> str:
        """
        Load LLM prompt template from src/generator/ksign_weight_prompt.md
        
        Returns:
            str: Prompt template content or None
        """
        try:
            prompt_file = Path(__file__).parent / 'ksign_weight_prompt.md'
            if not prompt_file.exists():
                self.logger.warning(f"ksign_weight_prompt.md not found: {prompt_file}")
                return None
            
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            self.logger.warning(f"Failed to load ksign_weight_prompt.md: {e}")
            return None

    def _calculate_file_weights_with_llm(self, 
                                         llm_provider,
                                         file_path: str,
                                         class_name: str,
                                         file_content: str,
                                         methods_with_ksignutil: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        파일(클래스) 단위로 LLM 호출하여 모든 메서드의 weight 계산
        
        Args:
            llm_provider: LLM 프로바이더
            file_path: 파일 경로
            class_name: 클래스명
            file_content: 파일 전체 내용
            methods_with_ksignutil: ksignUtil이 포함된 메서드 리스트
            
        Returns:
            List[Dict]: 각 메서드의 weight 정보
        """
        try:
            if not methods_with_ksignutil:
                return []
            
            # 메서드 이름 리스트 생성 (정규화: 배열이면 첫 요소)
            method_names = []
            for m in methods_with_ksignutil:
                name = m['name']
                if isinstance(name, list):
                    name = name[0] if name else ''
                if name:
                    method_names.append(name)
            
            # 각 메서드의 코드 추출 (AST parser에서 이미 추출한 content 직접 사용)
            method_codes = {}
            for method in methods_with_ksignutil:
                method_name = method['name']
                # method_name 정규화: 배열이면 첫 요소
                if isinstance(method_name, list):
                    method_name = method_name[0] if method_name else ''
                
                if not method_name:
                    continue
                
                # AST parser에서 이미 추출한 content 사용
                method_content = method.get('content', '')
                if method_content:
                    # 불완전한 내용을 피하기 위해 최대 길이 제한 (10000자까지)
                    method_codes[method_name] = method_content[:10000]
                else:
                    # Fallback: 정규식으로 재추출 (AST 실패한 경우)
                    pattern = rf'(public|private|protected)?\s+\w+\s+{re.escape(method_name)}\s*\([^)]*\)\s*\{{[^}}]*\}}'
                    match = re.search(pattern, file_content, re.DOTALL)
                    if match:
                        method_codes[method_name] = match.group(0)[:10000]
                    else:
                        # 최후 대체: 메서드명을 포함한 주변 코드 추출
                        lines = file_content.split('\n')
                        for i, line in enumerate(lines):
                            if method_name in line and ('public' in line or 'private' in line):
                                method_codes[method_name] = '\n'.join(lines[max(0, i-2):min(len(lines), i+50)])
                                break
            
            # LLM 프롬프트 작성 (개선된 버전)
            methods_list = "\n".join([f"- {name}" for name in method_names])

            readable_method_sections = []
            for method_name in method_names:
                method_code = method_codes.get(method_name, '')
                if not method_code:
                    continue
                readable_method_sections.append(
                    f"[Method: {method_name}]\n```java\n{method_code}\n```"
                )
            method_codes_text = "\n\n".join(readable_method_sections)
            
            # Build valid signatures block (from ksignUtils_pattern)
            signatures_block = ''
            if self.ksignutil_patterns:
                sigs = '\n'.join(f'  {p}' for p in self.ksignutil_patterns)
                signatures_block = f'Valid Signatures (count ONLY these overloads):\n{sigs}'

            # Build policyId filter block
            policyid_block = ''
            if self.policy_ids:
                pids = ', '.join(self.policy_ids)
                policyid_block = (
                    f'Policy ID Filter: {pids}\n'
                    f'  - String-parameter calls: count ONLY when policyId argument matches one of the above\n'
                    f'    (as string literal, constant reference, or variable holding one of these values).\n'
                    f'  - List-parameter calls: count ALL unconditionally (no policyId to filter).\n'
                    f'  - SKIP calls using any other policyId not listed above.'
                )

            # Load prompt template from src/generator/ksign_weight_prompt.md
            base_prompt = self._load_ksign_weight_prompt()
            
            if base_prompt:
                # Use prompt from file with readable method code samples.
                prompt_instruction = f"""=== Class Information ===
Class Name: {class_name}
File: {file_path}
{signatures_block}
{policyid_block}
Count ONLY calls that exactly match the signatures and policy ID filters listed above.
Ignore any other encrypt/decrypt methods on different classes, objects, wrappers, or helper names.

=== Target Methods ===
{methods_list}

=== Method Code Samples ===
The following are raw Java method code blocks. Analyze the brace structure exactly as written.
Do NOT treat them as JSON strings. Determine loop depth from the actual code block structure.

{method_codes_text}

{base_prompt}
"""
            else:
                # Final fallback: inline minimal prompt
                self.logger.warning(f"{class_name}: Both prompts failed, using inline fallback")
                
                # Build signatures and policyId info for fallback
                signatures_info = ''
                if self.ksignutil_patterns:
                    signatures_info = 'Valid Signatures (count ONLY these): ' + ' | '.join(self.ksignutil_patterns)
                
                policyid_info = ''
                if self.policy_ids:
                    policyid_info = f'Policy ID Filter: count String-param calls ONLY where policyId is one of: {", ".join(self.policy_ids)}. List-param calls are ALWAYS counted.'
                
                prompt_instruction = f"""You are an expert analyzing Java source code for encryption/decryption weights.

Analyze methods and return ONLY a JSON array. Complete these 9 fields for each method:
- method_name, loop_depth (0-2), loop_structure (string), multiplier (string)
- data_type ("single"|"paged_list"|"unpaged_list")
- dep0_crypto_count, dep1_crypto_count, dep2_crypto_count
- Base Weight

Count ONLY calls that exactly match the valid signatures below.
Do NOT count any other encrypt/decrypt methods on different classes, objects, or helpers.

{signatures_info}

{policyid_info}

dep0: crypto calls outside all loops (depth=0)
dep1: crypto calls inside depth-1 loops
dep2: crypto calls inside depth-2 or deeper loops

Weight formula:
- loop_depth=0: dep0
- loop_depth=1, paged: dep0 + 20*dep1
- loop_depth=1, unpaged: dep0 + 100*dep1
- loop_depth=2, paged: dep0 + 10*dep1 + 20*dep2
- loop_depth=2, unpaged: dep0 + 10*dep1 + 100*dep2

Class: {class_name}
Methods: {', '.join(method_names)}
Code blocks:
{method_codes_text}

Return ONLY valid JSON array. No markdown, text, or code blocks."""

            
            prompt_data = {
                'instruction': prompt_instruction,
                'target_methods': method_names
            }
            
            # LLM 호출
            self.logger.info(f"{class_name}: LLM 호출 중...")
            response = llm_provider.call(
                prompt_instruction,
                max_tokens=100000,
                temperature=0.1
            )
            
            # 응답 파싱
            response_content = response.get('content', '') if isinstance(response, dict) else str(response)
            
            # 응답에서 JSON 부분 추출 및 정제
            content = self._extract_json_from_response(response_content)
            
            # JSON 배열 파싱
            try:
                results = json.loads(content)
                if not isinstance(results, list):
                    results = [results]
            except json.JSONDecodeError as e:
                # 파싱 실패시 로그 저장하고 반환
                self._save_prompt_log(class_name, str(file_path), prompt_data, response_content, None)
                self.logger.error(f"{class_name}: LLM 응답 파싱 실패 - {e}")
                self.logger.debug(f"응답 내용: {content[:300]}...")
                return []

            if not results:
                self._save_prompt_log(class_name, str(file_path), prompt_data, response_content, results)
                self.logger.error(f"{class_name}: LLM 응답이 빈 배열입니다 ({len(methods_with_ksignutil)}개 메서드 전송됨)")
                return []
            
            sanity_audit = []

            # 파일 경로 추가 및 필드 정규화
            for result in results:
                result['class_path'] = str(file_path)
                
                # 문자열 필드 정규화 (배열 → 스트링)
                for field in ['method_name', 'loop_structure']:
                    if field in result:
                        value = result[field]
                        if isinstance(value, list):
                            result[field] = value[0] if value else ''
                        elif value is None:
                            result[field] = ''

                method_name = result.get('method_name', '')
                method_code = method_codes.get(method_name, '')
                _, audit = self._apply_llm_weight_sanity_check(result, method_code)
                sanity_audit.append(audit)

            prompt_data['sanity_audit'] = sanity_audit
            prompt_data['sanity_summary_file'] = self._save_sanity_audit_summary(
                class_name,
                str(file_path),
                sanity_audit,
                results,
            ) or ''
            
            # 성공 로그 저장
            self._save_prompt_log(class_name, str(file_path), prompt_data, response_content, results)
            
            self.logger.info(f"{class_name}: LLM으로 {len(results)}개 메서드 weight 계산 완료")
            return results
        
        except Exception as e:
            self.logger.error(f"{class_name}: LLM 호출 실패 - {e}")
            import traceback
            traceback.print_exc()
            return []
    


    def _generate_anyframe_ksign_report(self) -> bool:
        """
        Step 8: 최종 엔드포인트별 집계 및 ksign_report.json 생성
        
        처리:
        1. crypto_weights를 end_point별로 그룹화
        2. 각 endpoint별 계산:
           - 총 weight = sum of crypto_weight for all methods in this endpoint
           - 접근 횟수 = endpoint의 Jenifer Access Count
           - 예상 KSIGN 호출 = total_weight × access_count
        3. 하위항목으로 method별 세부 정보 포함
        4. 최종 JSON 생성: artifacts/ksign_report_{timestamp}.json
        
        Returns:
            bool: 성공 여부
        """
        try:
            if not self.crypto_weights:
                self.logger.warning(f"crypto_weights가 비어있음")
                return True
            
            # 1. endpoint별로 crypto_weights 그룹화
            endpoint_groups = {}  # {endpoint: [crypto_weight, ...]}
            
            for crypto_weight in self.crypto_weights:
                endpoint = crypto_weight.get('end_point')
                
                # endpoint가 list인 경우 첫 번째 요소 추출 (정규화)
                if isinstance(endpoint, list):
                    endpoint = endpoint[0] if endpoint else ''
                
                # endpoint가 None인 경우 빈 문자열로 처리
                if not endpoint:
                    endpoint = ''
                
                if not endpoint:
                    self.logger.warning(f"end_point 없는 항목 스킵: {crypto_weight.get('method_signature')}")
                    continue
                
                if endpoint not in endpoint_groups:
                    endpoint_groups[endpoint] = []
                endpoint_groups[endpoint].append(crypto_weight)
            
            self.logger.info(f"엔드포인트: {len(endpoint_groups)}개 그룹화 완료")
            
            # 2. 각 endpoint별 총 weight 계산
            report_endpoints = []
            total_ksign_calls = 0
            
            for endpoint, weights_list in endpoint_groups.items():
                # 총 weight 계산
                total_weight = sum(w.get('Total Weight', 0) for w in weights_list)
                
                # 접근 횟수 (첫 번째 항목에서 Jenifer Access Count 사용)
                access_count = weights_list[0].get('Jenifer Access Count', 0) if weights_list else 0
                
                # 예상 KSIGN 호출 = total_weight × access_count
                estimated_ksign_calls = total_weight * access_count if access_count > 0 else total_weight
                total_ksign_calls += estimated_ksign_calls
                
                # 엔드포인트 데이터 생성
                endpoint_data = {
                    'endpoint': endpoint,
                    'total_crypto_weight': total_weight,
                    'access_count': access_count,
                    'estimated_ksign_calls': estimated_ksign_calls,
                    'children': [
                        {
                            'method_signature': w.get('method_signature'),
                            'Total Weight': w.get('Total Weight', 0)
                        } for w in weights_list
                    ]
                }
                report_endpoints.append(endpoint_data)
            
            # endpoint별 estimated_ksign_calls로 정렬 (내림차순)
            report_endpoints.sort(key=lambda x: x['estimated_ksign_calls'], reverse=True)
            
            # 3. ksign_report.json 생성
            from datetime import datetime
            report_data = {
                'report_metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'total_endpoints': len(report_endpoints),
                    'total_ksign_calls': total_ksign_calls
                },
                'endpoints': report_endpoints
            }
            
            # JSON 파일 저장
            report_file = os.path.join(self.applycrypto_dir, 'results', f'ksign_report_{self.timestamp}.json')
            os.makedirs(os.path.dirname(report_file), exist_ok=True)
            
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Step 6 완료: ksign_report.json 저장 - {len(report_endpoints)}개 엔드포인트, 예상 KSIGN 호출: {total_ksign_calls:,.0f}회")
            
            return True
        
        except Exception as e:
            self.logger.error(f"ksign_report 생성 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    

def main():
    """메인 함수"""
    import sys
    
    applycrypto_dir = Path(__file__).parent / ".applycrypto"
    target_project = Path.cwd()
    
    if len(sys.argv) > 1:
        applycrypto_dir = sys.argv[1]
    if len(sys.argv) > 2:
        target_project = sys.argv[2]
    
    generator = KSIGNReportGenerator(str(applycrypto_dir), str(target_project))
    success = generator.run_full_pipeline()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
