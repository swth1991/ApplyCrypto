"""
이관산출물 생성기 (Artifact Generator)

클래스별로 Excel 이관산출물를 생성합니다.
"""

import os
import re
import json
import glob
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import difflib
from config.config_manager import Configuration

# =====================================
# 상수 정의
# =====================================

# 기본 설정
DEFAULT_FONT_NAME = '맑은 고딕'
DEFAULT_FONT_SIZE = 10
DATE_FORMAT = '%Y%m%d'
DATE_TIME_FORMAT = '%Y%m%d%H%M%S'
CHANGED_FILE_LIST_PATTERN = 'ChangedFileList_*.txt'
DEFAULT_ARTIFACTS_DIR = '.applycrypto/artifacts'

# LLM 관련 상수
MAX_BLOCKS_PER_LLM_CALL = 10
MAX_LLM_RETRIES = 3
LLM_MAX_TOKENS = 10000
LLM_TEMPERATURE = 0.0

# Excel 색상 코드
COLOR_HEADER = 'BFBFBF'
COLOR_SUBHEADER = 'D9D9D9'
COLOR_LINK = '0022EE'
COLOR_AS_IS = 'FFD966'
COLOR_TO_BE = 'A9D08E'
COLOR_YELLOW_LIGHT = 'FFF2CC'
COLOR_GREEN_LIGHT = 'E2EFDA'
COLOR_GRAY_LIGHT = 'DFDFDF'

# 정규식 패턴
PATTERN_METHOD_SIGNATURE = r'^\s*(public|protected|private)\s+[^\(]+\([^\)]*\)\s*\{'
PATTERN_GET_MAPPING = r'@GetMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_POST_MAPPING = r'@PostMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_PUT_MAPPING = r'@PutMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_DELETE_MAPPING = r'@DeleteMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_PATCH_MAPPING = r'@PatchMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_REQUEST_MAPPING = r'@RequestMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']\s*(?:,\s*method\s*=\s*RequestMethod\.([A-Z]+))?'
PATTERN_CLASS_REQUEST_MAPPING = r'@RequestMapping\s*\(\s*(?:value\s*=\s*)?["\']([^"\']+)["\']'
PATTERN_ENCRYPT_EVIDENCE = r'(CryptoService\.[a-zA-Z_]+|encrypt\(|decrypt\(|@Encrypted)'

# 파일 처리
FILE_ENCODINGS = ['utf-8', 'euc-kr', 'cp949', 'latin-1']
FILE_ENCODING_WITH_BOM = 'utf-8-sig'
BOM_MARKER = '\ufeff'

# 길이 제한
MAX_API_CHANGES_TEXT_LENGTH = 1200
MAX_PRIVACY_TEXT_LENGTH = 2000

# 텍스트 상수
NO_VALUE_TEXT = '해당 사항 없음'

# 프라이버시 관련 키워드
PRIVACY_KEYWORDS = [
    'email', 'lastName', 'dayOfBirth', 'juminNumber', 'sex', 'name',
    '이름', '이메일', '주민번호', '생년월일'
]

# 딕셔너리 필드명
FIELD_FILENAME = 'filename'
FIELD_FULL_PATH = 'full_path'
FIELD_RELATIVE_PATH = 'relative_path'
FIELD_GOAL = 'goal'
FIELD_ORIGINAL = 'original'
FIELD_MODIFIED = 'modified'
FIELD_ORIGINAL_LINES = 'original_lines'
FIELD_MODIFIED_LINES = 'modified_lines'

# 열 인덱스
COLUMN_NO = 2
COLUMN_FILENAME = 3
COLUMN_FULL_PATH = 4
COLUMN_REMARK = 5

# =====================================
# 진입점(Entry point)
# =====================================

def generate_artifact(config: Configuration, use_llm: bool = False):
    """타겟 프로젝트의 ChangedFileList를 이용해 이관용 Excel 워크북을 생성합니다.

    Args:
        config (Configuration): ApplyCrypto 설정 객체
        use_llm (bool): LLM을 사용하여 분석 고도화 여부. 기본값은 False.

    Returns:
        None: 워크북을 파일로 저장합니다.
    """
    target_project = config.target_project
    old_code_path = config.artifact_generation.old_code_path if config.artifact_generation else None

    if not target_project:
        raise ValueError('config must include target_project')
    
    if not old_code_path:
        raise ValueError('config must include artifact_generation.old_code_path')

    # LLM 프로바이더 초기화 (LLM 옵션 사용 시)
    llm_provider = None
    if use_llm:
        from src.modifier.llm.llm_factory import create_llm_provider
        llm_provider = create_llm_provider(config.llm_provider)

    # 출력 디렉터리 설정
    out_dir = os.path.join(target_project, DEFAULT_ARTIFACTS_DIR)
    os.makedirs(out_dir, exist_ok=True)

    changelog_pattern_txt = os.path.join(target_project, '.applycrypto', CHANGED_FILE_LIST_PATTERN)
    changelog_files = glob.glob(changelog_pattern_txt)
    if not changelog_files:
        raise FileNotFoundError(f"ChangeLog TXT 파일을 찾을 수 없습니다: {changelog_pattern_txt}")
    changelog_file = changelog_files[0]

    tp = Path(target_project)
    op = Path(old_code_path) if old_code_path else None

    if not tp.exists():
        raise FileNotFoundError(f"타겟 프로젝트를 찾을 수 없습니다: {target_project}")
    if op is None or not op.exists():
        raise FileNotFoundError(f"원본 백업을 찾을 수 없습니다: {old_code_path}")

    print(f"\n[이관 산출물 생성]")
    print(f"  target_project: {target_project}")
    print(f"  tp (Path 객체): {tp}")
    print(f"  tp.name: {tp.name}")
    print(f"  changelog_file: {changelog_file}")

    files = read_changedFileList(changelog_file, tp.name)

    artifact_dir = Path(out_dir)
    artifact_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime(DATE_FORMAT)
    artifact_file = artifact_dir / f"{tp.name} 이관 산출물 - {today}.xlsx"

    # 시트 생성
    wb = openpyxl.Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass
    
    create_business_requirements_sheet(wb)
    create_development_docs_sheet(wb)
    create_interface_sheet(wb)
    create_db_privacy_sheet(wb)
    create_test_cases_sheet(wb)
    create_source_code_sheet(wb)
    
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = False
    wb.iso_dates = False

    for file_info in files:
        try:
            # dict 형태 보장
            if isinstance(file_info, str):
                file_info = {
                    FIELD_FILENAME: os.path.basename(file_info),
                    FIELD_FULL_PATH: file_info,
                    FIELD_RELATIVE_PATH: normalize_path(file_info, tp.name)
                }

            try:
                fill_development_docs(wb, file_info, str(tp), str(op), append=True, use_llm=use_llm, llm_provider=llm_provider)
            except Exception:
                pass

            try:
                fill_source_code(wb, file_info, str(tp), str(op), append=True)
            except Exception:
                pass

        except Exception:
            continue

    try:
        remove_all_formulas(wb)
    except Exception:
        pass

    # '1.개발관련문서' 시트를 앞쪽으로 이동(존재하는 경우)
    try:
        if '1.개발관련문서' in wb.sheetnames:
            sheet = wb['1.개발관련문서']
            try:
                wb._sheets.remove(sheet)
            except Exception:
                pass
            wb._sheets.insert(1, sheet)
    except Exception:
        pass

    wb.active = 0

    try:
        wb.calculation.calcMode = 'manual'
    except Exception:
        pass

    # 폰트 정규화 적용
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    try:
                        cur = cell.font
                        cell.font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE,
                                         bold=getattr(cur, 'bold', False),
                                         italic=getattr(cur, 'italic', False),
                                         underline=getattr(cur, 'underline', None),
                                         strike=getattr(cur, 'strike', False),
                                         color=getattr(cur, 'color', None))
                    except Exception:
                        continue
    except Exception:
        pass

    # DB 개인정보 셀이 작은따옴표로 시작하도록 보장
    try:
        ws_db_dbg = wb['3.DB_파일_로그변경 여부 및 개인정보 여부']
        cur_val = ws_db_dbg.cell(row=7, column=2).value
        ws_db_dbg.cell(row=7, column=2).value = ensure_leading_quote(cur_val)
    except Exception:
        pass

    # 워크북 저장
    try:
        if artifact_file.exists():
            try:
                artifact_file.unlink()
            except Exception:
                ts = datetime.now().strftime(DATE_TIME_FORMAT)
                artifact_file = artifact_file.with_name(f"{artifact_file.stem}_tmp{ts}{artifact_file.suffix}")
        wb.save(artifact_file)
        print(f'- 출력 파일: {artifact_file}', flush=True)
    except PermissionError:
        ts = datetime.now().strftime(DATE_TIME_FORMAT)
        fallback = artifact_file.with_name(f"{artifact_file.stem}_fallback{ts}{artifact_file.suffix}")
        wb.save(fallback)
        print(f'- 출력 파일(대체): {fallback}', flush=True)
    return


# -----------------------------
# 핵심 Helpers (entry 바로 아래)
# -----------------------------

def create_business_requirements_sheet(wb):
    """업무요건 시트(0.업무요건)를 생성하고 기본 컬럼 너비와 샘플 값을 설정합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    from openpyxl.styles import Font, Border, Side
    ws = wb.create_sheet("0.업무요건")
    default_font = Font(name='맑은 고딕', size=10)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws['B2'] = "1. 업무요건"
    ws['B3'] = "1) 성명, 생년월일 컬럼의 경우 암호화 적용"
    ws['B4'] = "2) 생년월일 컬럼인 경우는 암호화 적용 대상이 아님 (문서상 명기된 내용)"
    ws['B5'] = "3) 주민등록번호의 경우 기존 뒷자리 6자리만 암호화하던 방식에서전체 주민등록번호 암호화 방식으로 변경"
    ws['B6'] = "4) 주민등록번호 앞 6자리를 파싱하여 생년월일을 처리하는 경우, 별도의 생년월일 컬럼을 신규 생성하여 처리"
    ws['B8'] = "2. 변경정책"
    ws['B9'] = "구분"
    ws['C9'] = "정책코드"
    ws['D9'] = "최소 길이"
    ws['E9'] = "최대 길이"
    ws['F9'] = "데이터 타입"
    ws['B10'] = "성명"
    ws['C10'] = "P017"
    ws['D10'] = "1"
    ws['E10'] = "255"
    ws['F10'] = "숫자+문자"
    ws['B11'] = "생년월일"
    ws['C11'] = "P018"
    ws['D11'] = "6"
    ws['E11'] = "8"
    ws['F11'] = "숫자"
    ws['B12'] = "주민등록번호(신규, FULL 암호화)"
    ws['C12'] = "P019"
    ws['D12'] = "13"
    ws['E12'] = "13"
    ws['F12'] = "숫자"
    ws['B13'] = "※ 비고: 주민번호의 경우 ERP 시스템은 P001 정책을 사용하며, 그 외 시스템은 신규 P019 정책을 사용함."
    for row in range(1, 20):
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.alignment = Alignment(vertical='center')
            if cell.value is not None:
                cell.font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    bold_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    ws['B2'].font = bold_font
    ws['B8'].font = bold_font
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(9, 13):
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].alignment = center_alignment


def create_development_docs_sheet(wb):
    """개발관련문서 시트(1.개발관련문서)를 생성하고 기본 컬럼 너비를 설정합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    from openpyxl.styles import Font, Alignment
    ws = wb.create_sheet("1.개발관련문서")
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 2
    ws.column_dimensions['C'].width = 100


def create_interface_sheet(wb):
    """통신 인터페이스 관련 시트(2.통신인터페이스추가및변경여부)를 생성합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    from openpyxl.styles import Font, Alignment
    ws = wb.create_sheet("2.통신인터페이스추가및변경여부")
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 100
    ws.row_dimensions[2].height = 20
    ws['B2'] = "2. 통신 인터페이스 추가 및 변경 여부"
    ws['B2'].alignment = Alignment(vertical='center')
    ws.row_dimensions[3].height = 16
    ws['B3'] = "1) 통신 인터페이스 문서"
    ws['B3'].alignment = Alignment(vertical='center')
    ws.row_dimensions[4].height = 50
    ws['B4'] = "해당 사항 없음"
    ws['B4'].font = Font(color=COLOR_LINK)
    ws['B4'].alignment = Alignment(vertical='center')
    ws.row_dimensions[5].height = 16
    ws['B5'] = "2) 주요 개인정보(비밀번호, 주민등록번호) E2E 암호화 여부"
    ws['B5'].alignment = Alignment(vertical='center')
    ws.row_dimensions[6].height = 50
    ws['B6'] = "해당 사항 없음"
    ws['B6'].font = Font(color=COLOR_LINK)
    ws['B6'].alignment = Alignment(vertical='center')
    ws.row_dimensions[7].height = 16
    ws['B7'] = "3) TLS 버전과 알고리즘 확인"
    ws['B7'].alignment = Alignment(vertical='center')
    ws.row_dimensions[8].height = 50
    ws['B8'] = "해당 사항 없음"
    ws['B8'].font = Font(color=COLOR_LINK)
    ws['B8'].alignment = Alignment(vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')


def create_db_privacy_sheet(wb):
    """DB 및 개인정보 여부 시트(3.DB_파일_로그변경 여부 및 개인정보 여부)를 생성합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    from openpyxl.styles import Font
    ws = wb.create_sheet("3.DB_파일_로그변경 여부 및 개인정보 여부")
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 100
    ws.row_dimensions[2].height = 20
    ws['B2'] = "3. DB, 파일, 로그 변경 여부 및 개인정보 여부"
    ws['B2'].alignment = Alignment(vertical='center')
    ws.row_dimensions[3].height = 16
    ws['B3'] = "1) 변경내역"
    ws['B3'].alignment = Alignment(vertical='center')
    ws.row_dimensions[4].height = 50
    ws['B4'] = "DB 마이그레이션 운영이관 관련 산출물 참조."
    ws['B4'].font = Font(color=COLOR_LINK)
    ws['B4'].alignment = Alignment(vertical='center')
    ws.row_dimensions[5].height = 16
    ws['B5'].alignment = Alignment(vertical='center')
    ws.row_dimensions[6].height = 16
    ws['B6'] = "2) 개인정보 암호화 대상"
    ws['B6'].alignment = Alignment(vertical='center')
    ws.row_dimensions[7].height = 50
    ws['B7'] = ensure_leading_quote(NO_VALUE_TEXT)
    ws['B7'].font = Font(color=COLOR_LINK)
    ws['B7'].alignment = Alignment(vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')


def create_test_cases_sheet(wb):
    """테스트케이스 시트(5.테스트케이스)를 생성하고 기본 레이블을 설정합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    ws = wb.create_sheet("5.테스트케이스")
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 70
    ws.row_dimensions[2].height = 20
    ws['B2'] = "5. 테스트 케이스"
    ws['B2'].alignment = Alignment(vertical='center')


def create_source_code_sheet(wb):
    """소스코드 목록 시트(6.소스코드)를 생성하고 컬럼/헤더 포맷을 설정합니다.

    Args:
        wb: 대상 openpyxl.Workbook 객체

    Returns:
        None
    """
    ws = wb.create_sheet("6.소스코드")
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 90
    ws.column_dimensions['E'].width = 20
    ws.row_dimensions[2].height = 20
    ws['B2'] = "6. 변경 소스 목록"
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.row_dimensions[3].height = 16
    ws['B3'] = "No"
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].font = Font(bold=True)
    ws['B3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws['B3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['C3'] = "파일명"
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].font = Font(bold=True)
    ws['C3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws['C3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['D3'] = "전체 경로"
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D3'].font = Font(bold=True)
    ws['D3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws['D3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['E3'] = "비고"
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E3'].font = Font(bold=True)
    ws['E3'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws['E3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in ['B', 'C', 'D', 'E']:
        cell = ws[f'{col}4']
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    for row in range(1, 100):
        if row == 2:
            continue
        cell = ws[f'B{row}']
        if cell.alignment:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')


# -----------------------------
# 모듈 수준 API 래퍼들
# -----------------------------

def fill_development_docs(wb: openpyxl.Workbook, file_info: Dict, target_project: str, old_code_path: str, append: bool = False, use_llm: bool = False, llm_provider = None):
    """단일 파일의 AS-IS / TO-BE 문서 블록을 `1.개발관련문서` 시트에 채웁니다.

    Args:
        wb: 대상 워크북 객체
        file_info: 변경 파일 정보 딕셔너리 (filename, full_path, relative_path 등)
        target_project: 대상 프로젝트 루트 경로
        old_code_path: 원본 백업 경로
        append: 기존 내용 뒤에 추가할지 여부
        use_llm: LLM을 사용하여 변경 분석 고도화 여부
        llm_provider: LLM 프로바이더 객체

    Returns:
        None
    """
    ws = wb['1.개발관련문서']

    try:
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['C'].width = 100
    except Exception:
        pass

    if not append:
        ws.row_dimensions[3].height = 16
        for row in range(2, ws.max_row + 1):
            for col in [1, 3]:
                try:
                    ws.cell(row=row, column=col).value = None
                except Exception:
                    pass

    relative_path = file_info['relative_path']
    original_file = Path(old_code_path) / relative_path
    modified_file = Path(target_project) / relative_path

    # [DEBUG] 파일 경로 검증
    print(f"  [DEBUG] 파일 경로 검증:")
    print(f"    - relative_path: {relative_path}")
    print(f"    - original_file: {original_file}")
    print(f"    - modified_file: {modified_file}")
    print(f"    - original 존재: {original_file.exists()}")
    print(f"    - modified 존재: {modified_file.exists()}")
    
    if not original_file.exists():
        print(f"    [⚠️ WARN] 원본 파일을 찾을 수 없음: {original_file}")
    if not modified_file.exists():
        print(f"    [⚠️ WARN] 수정본 파일을 찾을 수 없음: {modified_file}")

    original_content = read_file_safe(str(original_file))
    modified_content = read_file_safe(str(modified_file))

    if not original_content and not modified_content:
        return

    original_lines = original_content.splitlines() if original_content else []
    modified_lines = modified_content.splitlines() if modified_content else []

    blocks = extract_logical_change_blocks(original_lines, modified_lines)

    print(f"- 파일: {file_info['filename']}, blocks 수: {len(blocks)}")

    # LLM 호출인 경우 변경 블록 분석 수행
    if use_llm and llm_provider and blocks:
       # LLM에 보낼 데이터로 경량화 (필요한 정보만 추림)
        prompt_input_data = []
        for idx, block in enumerate(blocks, start=1):
            prompt_input_data.append({
                "idx": idx,
                "original_lines": block.get('original_lines', ''),
                "modified_lines": block.get('modified_lines', ''),
                "original_code": block.get('original', ''),   # 분석을 위해 코드 내용 전달
                "modified_code": block.get('modified', '')    # 분석을 위해 코드 내용 전달
            })
        
        # User Prompt 구성
        instruction = """
            # Role
            당신은 Java 프로젝트의 코드 변경 사항을 분석하여 WinMerge 스타일의 요약 보고서를 작성하는 AI 전문가입니다.

            # Task
            입력으로 주어지는 [코드 변경 블록 리스트]를 분석하여, 각 블록의 변경 유형과 목적을 파악하고 지정된 JSON 형식으로 출력하십시오.

            # Analysis Guidelines (분석 가이드)
            1. **변경 유형 판단**:
            - **코드 추가**: 원본(As-Is)이 비어있거나 없고, 수정본(To-Be)만 있는 경우.
            - **코드 삭제**: 수정본(To-Be)이 비어있거나 없고, 원본(As-Is)만 있는 경우.
            - **코드 변경**: 기능 수정, 변수명 변경, 로직 개선 등.
            - **주석 처리**: JavaDoc(`/** */`), 블록 주석(`/* */`), 한줄 주석(`//`) 등의 추가/변경. 단순 주석 변경은 "주석 추가" 또는 "주석 수정"으로 명시.

            2. **Goal 작성 규칙**:
            - 변경된 코드의 **핵심 기능**이나 **변경 목적**을 한글로 요약하십시오.
            - 문장은 간결하게 작성하며 **공백 포함 50자 이내**로 제한합니다.
            - 예시: "암호화 유틸리티 메소드 추가", "불필요한 로그 삭제", "Null 체크 로직 보강", "변수명 가독성 개선"

            3. **Line Formatting**:
            - 입력받은 라인 번호 정보를 이용하여 "Line {시작}-{끝}" 또는 "Line {번호}" 형식으로 작성하십시오.

            # Output Format (JSON)
            응답은 오직 아래의 JSON 리스트 포맷으로만 출력해야 하며, 마크다운 태그(```json)나 부가적인 설명은 포함하지 마십시오.

            [
            {
                "idx" : 1,
                "goal": "{분석된 변경 목적이나 기능 요약}"
            },
            ...
            ]
        """

        # LLM 분석 수행 (blocks가 많으면 나누어 호출)
        llm_results = []
        for i in range(0, len(prompt_input_data), MAX_BLOCKS_PER_LLM_CALL):
            batch = prompt_input_data[i:i + MAX_BLOCKS_PER_LLM_CALL]
            batch_prompt = {
                'task': 'analyze_code_changes',
                'instruction': instruction,
                'context': {
                    'changed_blocks': batch
                }
            }
            
            batch_results = None
            llm_attempt = 0
            for attempt in range(MAX_LLM_RETRIES):
                try:
                    llm_attempt += 1
                    batch_resp = llm_provider.call(json.dumps(batch_prompt, ensure_ascii=False), max_tokens=LLM_MAX_TOKENS, temperature=LLM_TEMPERATURE)
                    batch_content = batch_resp.get('content', '[]') if isinstance(batch_resp, dict) else str(batch_resp)
                    batch_results = []
                    if batch_content.strip():
                        try:
                            batch_results = json.loads(batch_content)
                            if isinstance(batch_results, dict):
                                batch_results = [batch_results]
                        except json.JSONDecodeError as je:
                            # 시도적 복구: 마크다운 코드블록, 텍스트 앞뒤를 제거하고 JSON 배열/객체 추출 시도
                            s = batch_content
                            try:
                                # 마크다운 ```json``` 또는 ``` 제거
                                s = re.sub(r'```(?:json)?\\n?', '', s)
                                s = s.replace('```', '')
                                # 배열 형태 추출: 첫 '[' 와 마지막 ']' 사이
                                start = s.find('[')
                                end = s.rfind(']')
                                if start != -1 and end != -1 and end > start:
                                    candidate = s[start:end+1]
                                    try:
                                        batch_results = json.loads(candidate)
                                    except Exception:
                                        batch_results = []
                                else:
                                    # 객체들만 연속으로 반환된 경우, 각 객체를 찾아 파싱하여 리스트로 변환
                                    objs = re.findall(r'\{.*?\}', s, flags=re.DOTALL)
                                    parsed_objs = []
                                    for o in objs:
                                        try:
                                            parsed_objs.append(json.loads(o))
                                        except Exception:
                                            continue
                            except Exception as ex2:
                                batch_results = []
                        except Exception as ex:
                            batch_results = []
                    # 응답에 포함된 객체들을 'idx'로 매핑하고, 원본 배치 순서대로 재구성
                    parsed_map = {}
                    try:
                        for obj in batch_results or []:
                            if not isinstance(obj, dict):
                                continue
                            raw_idx = obj.get('idx')
                            if raw_idx is None:
                                continue
                            try:
                                key = int(raw_idx)
                            except Exception:
                                try:
                                    key = int(str(raw_idx).strip())
                                except Exception:
                                    continue
                            parsed_map[key] = obj
                    except Exception:
                        parsed_map = {}

                    expected_indices = [b.get('idx') for b in batch if isinstance(b, dict) and 'idx' in b]
                    # 재구성: 누락된 idx는 빈 객체로 대체
                    batch_results = [parsed_map.get(eidx, {}) for eidx in expected_indices]
                    break  # 성공 시 루프 탈출
                except Exception as e:
                    if attempt == MAX_LLM_RETRIES - 1:
                        print(f"- 배치 {i//MAX_BLOCKS_PER_LLM_CALL + 1} 최종 실패 (재시도 {MAX_LLM_RETRIES}회): {e}")
                        batch_results = [{}] * len(batch)
                    else:
                        print(f"- 배치 {i//MAX_BLOCKS_PER_LLM_CALL + 1} 재시도 {attempt + 1}/{MAX_LLM_RETRIES}: {e}")
            
            llm_results.extend(batch_results) 

        # print(f"  LLM 호출 시도: {llm_attempt} , 코드 블럭 응답: {len(llm_results)} ")

        # LLM 결과를 blocks에 병합
        for i, block in enumerate(blocks):
            if i < len(llm_results):
                block['goal'] = llm_results[i].get('goal', '')

    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')

    if append:
        current_row = ws.max_row + 1
    else:
        current_row = 2

    if append:
        try:
            current_row = ws.max_row + 1
        except Exception:
            current_row = current_row

    try:
        try:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        except Exception:
            pass

        filename_cell = ws.cell(row=current_row, column=1, value=f"{file_info['filename']}  ({file_info['relative_path']})")
        filename_cell.font = Font(name=DEFAULT_FONT_NAME, bold=True)
        filename_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 16
        for col_idx in range(1, 4):
            try:
                c = ws.cell(row=current_row, column=col_idx)
                c.fill = PatternFill(start_color=COLOR_GRAY_LIGHT, end_color=COLOR_GRAY_LIGHT, fill_type='solid')
                c.border = border
                c.alignment = Alignment(horizontal='left', vertical='center')
                c.font = Font(name=DEFAULT_FONT_NAME, bold=True)
            except Exception:
                pass
        current_row += 1
    except Exception:
        pass

    try:
        h_as = ws.cell(row=current_row, column=1, value='AS-IS')
        h_tb = ws.cell(row=current_row, column=3, value='TO-BE')
        h_as.font = Font(name=DEFAULT_FONT_NAME, bold=True)
        h_tb.font = Font(name=DEFAULT_FONT_NAME, bold=True)
        h_as.alignment = Alignment(horizontal='center', vertical='center')
        h_tb.alignment = Alignment(horizontal='center', vertical='center')
        h_as.border = border
        h_tb.border = border
        h_as.fill = PatternFill(start_color=COLOR_AS_IS, end_color=COLOR_AS_IS, fill_type='solid')
        h_tb.fill = PatternFill(start_color=COLOR_TO_BE, end_color=COLOR_TO_BE, fill_type='solid')
        ws.row_dimensions[current_row].height = 16
        current_row += 1
    except Exception:
        pass

    for blk in blocks:
        orig = blk.get('original') or ''
        mod = blk.get('modified') or ''

        orig_line = f"Line {blk.get('original_lines', '')}" if blk.get('original_lines') else ''
        mod_line = f"Line {blk.get('modified_lines', '')}" if blk.get('modified_lines') else ''

        # LLM goal 추가 (LLM 옵션 사용 시)
        goal = blk.get('goal', '').strip('{}')
        try:
            if goal:
                rt = CellRichText()
                rt.append(TextBlock(text=mod_line + " (", font=InlineFont(rFont=DEFAULT_FONT_NAME, sz=DEFAULT_FONT_SIZE)))
                rt.append(TextBlock(text=goal, font=InlineFont(rFont=DEFAULT_FONT_NAME, sz=DEFAULT_FONT_SIZE, color="0000FF")))
                rt.append(TextBlock(text=")", font=InlineFont(rFont=DEFAULT_FONT_NAME, sz=DEFAULT_FONT_SIZE)))
                mod_line_display_value = rt
            else:
                mod_line_display_value = mod_line
        except Exception as e:
            print(f"Error creating rich text for goal: {e}")
            # fallback to plain text
            if goal:
                mod_line = f"{mod_line} ({goal})"
            else:
                mod_line_display_value = mod_line 

        line_cell_as = ws.cell(row=current_row, column=1, value=orig_line)
        line_cell_to = ws.cell(row=current_row, column=3, value=mod_line_display_value)
        line_cell_as.border = border
        line_cell_to.border = border
        line_cell_as.alignment = wrap
        line_cell_to.alignment = wrap
        line_cell_as.fill = PatternFill(start_color=COLOR_YELLOW_LIGHT, end_color=COLOR_YELLOW_LIGHT, fill_type='solid')
        line_cell_to.fill = PatternFill(start_color=COLOR_GREEN_LIGHT, end_color=COLOR_GREEN_LIGHT, fill_type='solid')
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        orig_lines = orig.split('\n') if orig else []
        mod_lines = mod.split('\n') if mod else []
        diff_lines = list(difflib.unified_diff(orig_lines, mod_lines, fromfile='original', tofile='modified', lineterm='', n=0))
        diff_content = '\n'.join(diff_lines[3:]) if len(diff_lines) > 3 else ''
        orig_diff = '\n'.join(line.replace('-', '- ', 1) for line in diff_content.split('\n') if line.startswith('-'))
        mod_diff = '\n'.join(line.replace('+', '+ ', 1) for line in diff_content.split('\n') if line.startswith('+'))

        acell = ws.cell(row=current_row, column=1, value=ensure_leading_quote(orig_diff.strip() if orig_diff.strip() else ''))
        if mod_diff.strip():
            mod_display = mod_diff.strip()
        elif orig_diff.strip():
            mod_display = ''
        else:
            mod_display = ''

        bcell = ws.cell(row=current_row, column=3, value=ensure_leading_quote(mod_display))
        acell.alignment = wrap
        acell.border = border
        bcell.alignment = wrap
        bcell.border = border
        acell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        bcell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        source_row = current_row
        spacer_row = current_row + 1
        ws.row_dimensions[source_row].height = 100
        ws.row_dimensions[spacer_row].height = 16
        current_row += 2

    try:
        spacer_idx = ws.max_row + 1
        ws.insert_rows(spacer_idx)
        ws.row_dimensions[spacer_idx].height = 16
        ws.cell(row=spacer_idx, column=1).value = None
        ws.cell(row=spacer_idx, column=3).value = None
    except Exception:
        try:
            ws.row_dimensions[current_row].height = 16
        except Exception:
            pass

    for r in range(1, ws.max_row + 1):
        if ws.row_dimensions[r].height == 100:
            continue
        for col in [1, 3]:
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                current_alignment = cell.alignment
                if current_alignment:
                    cell.alignment = current_alignment.copy(vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')

 
def remove_all_formulas(wb: openpyxl.Workbook):
    """워크북 내 모든 시트에서 수식(=로 시작하거나 데이터 타입이 formula)을 제거합니다.

    Args:
        wb: openpyxl 워크북 객체

    Returns:
        None
    """
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                try:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.strip().startswith('=')):
                        try:
                            v = cell.value
                            if isinstance(v, str) and v.strip().startswith('='):
                                nv = v.lstrip('=')
                                cell.value = nv
                                cell.data_type = 's'
                            else:
                                cell.value = str(v) if v is not None else None
                                cell.data_type = 's' if cell.value else 'n'
                        except Exception:
                            cell.value = None
                            cell.data_type = 'n'
                except Exception:
                    try:
                        cell.value = None
                    except Exception:
                        pass


# -----------------------------
# Diff / 소스 관련 헬퍼
# -----------------------------

def extract_logical_change_blocks(original_lines: List[str], modified_lines: List[str]) -> List[Dict]:
    """원본과 수정본의 라인 리스트를 비교하여 의미 있는 변경 블록(예: 메서드 단위)을 추출합니다.

    Args:
        original_lines: 원본 파일의 라인 리스트
        modified_lines: 수정된 파일의 라인 리스트

    Returns:
        blocks: 변경 블록 리스트. 각 블록은 dict로 { 'title','original','modified','original_lines','modified_lines' }
    """
    blocks: List[Dict] = []

    def find_method_starts(lines: List[str]):
        """라인 리스트에서 메서드 시그니처가 시작되는 인덱스와 시그니처 텍스트를 찾습니다.

        Args:
            lines: 파일의 라인 문자열 리스트

        Returns:
            starts: (index, signature) 튜플 리스트
        """
        starts = []
        sig_re = re.compile(PATTERN_METHOD_SIGNATURE)
        for idx, line in enumerate(lines):
            if sig_re.search(line):
                starts.append((idx, line.strip()))
        return starts

    orig_methods = find_method_starts(original_lines)
    mod_methods = find_method_starts(modified_lines)

    sm = difflib.SequenceMatcher(None, original_lines, modified_lines)
    opcodes = sm.get_opcodes()

    def find_enclosing_method(index: int, methods: List):
        """지정한 인덱스가 속하는(또는 바로 이전의) 메서드 시작 정보를 반환합니다.

        Args:
            index: 라인 인덱스
            methods: (index, signature) 튜플 리스트

        Returns:
            (index, signature) 또는 None
        """
        last = None
        for sidx, sig in methods:
            if sidx <= index:
                last = (sidx, sig)
            else:
                break
        return last

    for tag, i1, i2, j1, j2 in opcodes:
        if tag == 'equal':
            continue

        title = None
        if i1 < len(original_lines):
            m = find_enclosing_method(i1, orig_methods)
            if m:
                title = m[1]
        if not title and j1 < len(modified_lines):
            m = find_enclosing_method(j1, mod_methods)
            if m:
                title = m[1]
        if not title:
            title = f"Lines {i1+1}-{max(i2, i1)} / {j1+1}-{max(j2, j1)}"

        orig_block = '\n'.join(original_lines[i1:i2]).rstrip('\n')
        mod_block = '\n'.join(modified_lines[j1:j2]).rstrip('\n')

        blocks.append({
            'title': title,
            'original': orig_block,
            'modified': mod_block,
            'original_lines': f"{i1+1}-{i2}" if i2 > i1 else f"{i1+1}",
            'modified_lines': f"{j1+1}-{j2}" if j2 > j1 else f"{j1+1}",
            'i1': i1, 'i2': i2, 'j1': j1, 'j2': j2,
        })

    def split_on_blank_outside_block(text: str) -> List[str]:
        """블록 주석 내부가 아닌 빈 줄 기준으로 텍스트를 분할합니다.

        Args:
            text: 입력 텍스트 블록

        Returns:
            분할된 텍스트 조각 리스트
        """
        if not text:
            return []
        parts = []
        cur: List[str] = []
        in_block = False
        for ln in text.splitlines():
            s = ln.strip()
            if '/*' in s and '*/' not in s:
                in_block = True
            if s == '' and not in_block:
                if cur:
                    parts.append('\n'.join(cur))
                    cur = []
                else:
                    continue
            else:
                cur.append(ln)
            if '*/' in s and in_block:
                in_block = False
        if cur:
            parts.append('\n'.join(cur))
        return parts

    split_blocks: List[Dict] = []
    for blk in blocks:
        orig_parts = split_on_blank_outside_block(blk.get('original') or '')
        mod_parts = split_on_blank_outside_block(blk.get('modified') or '')

        if mod_parts and len(mod_parts) > 1:
            if orig_parts and len(orig_parts) == len(mod_parts):
                use_orig = orig_parts
            else:
                use_orig = [''] * len(mod_parts)

            cur_i = blk['i1']
            cur_j = blk['j1']
            for oi, mj in zip(use_orig, mod_parts):
                o_lines = oi.count('\n') + 1 if oi else 0
                m_lines = mj.count('\n') + 1 if mj else 0
                sub = {
                    'title': blk['title'],
                    'original': oi,
                    'modified': mj,
                    'original_lines': f"{cur_i+1}-{cur_i+o_lines}" if o_lines > 0 else '',
                    'modified_lines': f"{cur_j+1}-{cur_j+m_lines}" if m_lines > 0 else '',
                    'i1': cur_i, 'i2': cur_i + o_lines, 'j1': cur_j, 'j2': cur_j + m_lines,
                }
                split_blocks.append(sub)
                cur_i += o_lines
                cur_j += m_lines
        else:
            split_blocks.append(blk)

    def is_comment_text(text: str) -> bool:
        """주어진 텍스트가 주석(JavaDoc/block/line)으로만 구성되었는지 검사합니다.

        Args:
            text: 검사할 텍스트

        Returns:
            True이면 주석 텍스트, 아니면 False
        """
        if not text:
            return False
        lines = [ln for ln in text.splitlines() if ln.strip() != '']
        if not lines:
            return False
        s = text.strip()
        if s.startswith('/*') and s.endswith('*/'):
            return True
        for ln in lines:
            if not re.match(r'^\s*(//|/\*|\*/|\*).*', ln):
                return False
        return True

    def comment_style(text: str) -> Optional[str]:
        """주석 텍스트의 스타일을 판별합니다: 'javadoc','block','line' 또는 None.

        Args:
            text: 주석 또는 텍스트

        Returns:
            스타일 문자열 또는 None
        """
        if not text:
            return None
        s = text.strip()
        if s.startswith('/**'):
            return 'javadoc'
        if s.startswith('/*'):
            return 'block'
        for ln in text.splitlines():
            if ln.strip().startswith('//'):
                return 'line'
        return None

    def expand_modified_range(j2: int) -> int:
        """수정된 블록의 끝 인덱스를 주석/빈 라인 범위까지 확장합니다.

        Args:
            j2: 원래 수정 블록의 끝 인덱스

        Returns:
            확장된 끝 인덱스
        """
        limit = min(len(modified_lines), j2 + 8)
        if j2 - 1 >= 0:
            last = modified_lines[j2-1]
            if '/*' in last and '*/' not in last:
                for k in range(j2, limit):
                    if '*/' in modified_lines[k]:
                        j2 = k + 1
                        break
        while j2 < len(modified_lines):
            ln = modified_lines[j2]
            if ln.strip() == '' or re.match(r'^\s*(//|/\*|\*/|\*).*', ln):
                j2 += 1
            else:
                break
        return j2

    merged: List[Dict] = []
    for blk in split_blocks:
        blk_j2 = blk.get('j2', 0) or 0
        blk['j2'] = expand_modified_range(blk_j2)
        blk['modified_lines'] = f"{blk.get('j1',0)+1}-{blk.get('j2',0)}" if blk.get('j2',0) > blk.get('j1',0) else f"{blk.get('j1',0)+1}"

        if not merged:
            merged.append(blk)
            continue

        prev = merged[-1]

        adjacent_orig = prev.get('i2') == blk.get('i1')
        adjacent_mod = prev.get('j2') == blk.get('j1')

        def gap_is_only_comments(side_lines, a_end, b_start):
            """두 인덱스 사이의 갭이 오직 빈줄 및 주석으로만 구성되었는지 검사합니다.

            Args:
                side_lines: 라인 리스트
                a_end: 시작 인덱스
                b_start: 종료 인덱스

            Returns:
                True/False
            """
            if b_start <= a_end:
                return True
            gap = side_lines[a_end:b_start]
            if not gap:
                return True
            for ln in gap:
                if ln.strip() == '':
                    continue
                if not re.match(r'^\s*(//|/\*|\*/|\*).*', ln):
                    return False
            return True

        is_prev_comment = is_comment_text(prev.get('modified') or prev.get('original'))
        is_blk_comment = is_comment_text(blk.get('modified') or blk.get('original'))

        prev_style = comment_style(prev.get('modified') or prev.get('original'))
        blk_style = comment_style(blk.get('modified') or blk.get('original'))

        should_merge = False
        if prev_style and blk_style and prev_style == blk_style:
            if adjacent_orig and is_prev_comment and is_blk_comment and gap_is_only_comments(original_lines, prev.get('i2',0), blk.get('i1',0)):
                should_merge = True
            if adjacent_mod and is_prev_comment and is_blk_comment and gap_is_only_comments(modified_lines, prev.get('j2',0), blk.get('j1',0)):
                should_merge = True

        def is_closer_only(text: str) -> bool:
            """블록 닫힘 토큰(예: '}', ');', '*/' 등)만 포함하는지 검사합니다.

            Args:
                text: 검사할 텍스트

            Returns:
                True이면 닫힘 토큰만 포함
            """
            if not text:
                return False
            s = ''.join([ln.strip() for ln in text.splitlines() if ln.strip()])
            if re.match(r'^[\s\)\]\}\*\/;,\.]+$', s):
                return True
            if '*/' in s and re.sub(r'[\s\)\]\}\*\/;,\.]+', '', s) == '':
                return True
            return False

        if should_merge or is_closer_only(blk.get('modified') or blk.get('original')):
            if blk.get('original'):
                if prev.get('original'):
                    prev['original'] = prev['original'] + '\n\n' + blk.get('original')
                else:
                    prev['original'] = blk.get('original')
            if blk.get('modified'):
                if prev.get('modified'):
                    prev['modified'] = prev['modified'] + '\n\n' + blk.get('modified')
                else:
                    prev['modified'] = blk.get('modified')
            prev['i2'] = max(prev.get('i2',0), blk.get('i2',0))
            prev['j2'] = max(prev.get('j2',0), blk.get('j2',0))
            prev['original_lines'] = f"{prev.get('i1',0)+1}-{prev.get('i2',0)}"
            prev['modified_lines'] = f"{prev.get('j1',0)+1}-{prev.get('j2',0)}"
        else:
            merged.append(blk)

    return merged


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, cols: List[int]):
    """워크시트의 한 행에서 다른 행으로 지정한 열들의 스타일을 복사합니다.

    Args:
        ws: 대상 워크시트
        src_row: 원본 행 번호
        dst_row: 대상 행 번호
        cols: 스타일을 복사할 열 인덱스 리스트

    Returns:
        None
    """
    from copy import copy

    for col in cols:
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)

        try:
            if src_cell.font:
                dst_cell.font = copy(src_cell.font)
            if src_cell.border:
                dst_cell.border = copy(src_cell.border)
            if src_cell.fill:
                dst_cell.fill = copy(src_cell.fill)
            if src_cell.alignment:
                dst_cell.alignment = copy(src_cell.alignment)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
            if src_cell.protection:
                dst_cell.protection = copy(src_cell.protection)
        except Exception:
            continue


def fill_source_code(wb: openpyxl.Workbook, file_info: Dict, target_project: str, old_code_path: str, append: bool = False):
    """변경된 소스 파일의 전체 소스 텍스트를 `6.소스코드` 시트에 채웁니다.

    Args:
        wb: 대상 워크북
        file_info: 변경 파일 정보
        target_project: 대상 프로젝트 루트
        old_code_path: 원본 백업 경로
        append: 기존 내용 뒤에 추가 여부

    Returns:
        None
    """
    from copy import copy

    ws = wb['6.소스코드']
    template_row_idx = 4
    if ws.max_row < template_row_idx:
        for col in range(1, 6):
            _ = ws.cell(row=template_row_idx, column=col)

    if not append:
        template_row = {}
        for col in range(1, 6):
            template_row[col] = ws.cell(row=template_row_idx, column=col)
        if ws.max_row >= template_row_idx:
            ws.delete_rows(template_row_idx, ws.max_row - (template_row_idx - 1))

        for col in range(1, 6):
            new_cell = ws.cell(row=template_row_idx, column=col)
            template_cell = template_row[col]
            try:
                if template_cell.has_style:
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.fill = copy(template_cell.fill)
                    new_cell.number_format = copy(template_cell.number_format)
                    new_cell.protection = copy(template_cell.protection)
                    new_cell.alignment = copy(template_cell.alignment)
            except Exception:
                pass

        ws.cell(row=template_row_idx, column=1, value='')
        ws.cell(row=template_row_idx, column=2, value=1)
        ws.cell(row=template_row_idx, column=3, value=file_info['filename'])
        ws.cell(row=template_row_idx, column=4, value=file_info['full_path'])
        ws.cell(row=template_row_idx, column=5, value='')
        return

    try:
        ws.insert_rows(template_row_idx)
    except Exception:
        template_row_idx = ws.max_row + 1

    try:
        copy_row_style(ws, src_row=template_row_idx+1, dst_row=template_row_idx, cols=list(range(1, 6)))
    except Exception:
        pass

    ws.cell(row=template_row_idx, column=1, value='')
    ws.cell(row=template_row_idx, column=3, value=file_info.get('filename') or '')
    ws.cell(row=template_row_idx, column=4, value=file_info.get('full_path') or file_info.get('relative_path') or '')
    ws.cell(row=template_row_idx, column=5, value='')

    try:
        no = 1
        for r in range(template_row_idx, ws.max_row + 1):
            fname = (ws.cell(row=r, column=3).value or '')
            if fname and str(fname).strip():
                ws.cell(row=r, column=2, value=no)
                try:
                    ws.cell(row=r, column=1, value='')
                except Exception:
                    pass
                no += 1
            else:
                break
    except Exception:
        pass


# -----------------------------
# 기타 작은 유틸들 (이미 위에 `read_file_safe` 포함)
# -----------------------------

def ensure_leading_quote(text: Optional[str]) -> Optional[str]:
    """문자열이 작은따옴표로 시작하도록 보장합니다.

    Args:
        text: 입력 문자열 또는 None

    Returns:
        작은따옴표로 시작하는 문자열 또는 None
    """
    try:
        if text is None:
            return text
        s = str(text)
        s = s.lstrip("'")
        return "'" + s
    except Exception:
        return text


def normalize_path(path: str, target_project: str) -> str:
    """경로 문자열을 정규화합니다.
    
    다양한 입력 형태를 처리:
    - "book/ssm1/src/..." → "ssm1/src/..." (프로젝트명 제거)
    - "ssm1/src/..." → "ssm1/src/..." (이미 정규화됨, 그대로 반환)
    - "ssm1\\ssm2\\src\\..." → "ssm1/ssm2/src/..." (백슬래시 → 슬래시)

    Args:
        path: 원본 경로 문자열
        target_project: 대상 프로젝트 루트명 (예: "book")

    Returns:
        정규화된 상대경로 문자열
    """
    if path is None:
        return ''
    
    try:
        # 1. 백슬래시를 슬래시로 통일
        p = str(path).replace('\\', '/')
        
        # 2. 이중 슬래시 제거
        while '//' in p:
            p = p.replace('//', '/')
        
        # 3. 선행 슬래시 제거
        p = p.lstrip('/')
        
        # 4. target_project가 경로의 시작에 있으면 제거
        # 예: path="book/ssm1/src/...", target_project="book" → "ssm1/src/..."
        prefix = target_project + '/'
        if p.startswith(prefix):
            p = p[len(prefix):]
        
        return p
    except Exception as e:
        return str(path or '')


def read_file_safe(file_path: str) -> Optional[str]:
    """파일을 여러 인코딩으로 안전하게 읽어와 문자열을 반환합니다.

    Args:
        file_path: 읽을 파일 경로

    Returns:
        파일 내용 문자열 또는 None(읽기 실패)
    """
    p = Path(file_path)
    if not p.exists():
        return None

    for encoding in FILE_ENCODINGS:
        try:
            with open(p, 'r', encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, LookupError, FileNotFoundError):
            continue
    return None


def read_changedFileList(changedFileList_file: str, target_project_name: str) -> List[Dict]:
    """`.applycrypto` 내 ChangedFileList_*.txt 파일을 읽어 변경 파일 목록을 반환합니다.

    Args:
        changedFileList_file: ChangedFileList 파일 경로
        target_project_name: 대상 프로젝트 이름

    Returns:
        files: 변경 파일 정보의 딕셔너리 리스트 (필드: filename, full_path, relative_path 등)
    """
    print(f"  [DEBUG] read_changedFileList 시작:")
    print(f"    - file: {changedFileList_file}")
    print(f"    - target_project_name: {target_project_name}")
    
    if changedFileList_file.endswith('.txt'):
        try:
            # 파일에 BOM이 있을 경우 `utf-8-sig`로 자동 제거
            with open(changedFileList_file, 'r', encoding=FILE_ENCODING_WITH_BOM) as f:
                lines = f.readlines()

            files = []
            for idx, line in enumerate(lines):
                # 공백을 트림하고 남아있는 BOM 문자를 방어적으로 제거
                raw = line.strip()
                if raw and not raw.startswith('#'):
                    # 남아있는 선행 BOM 및 선행 슬래시 제거
                    cleaned = raw.lstrip(BOM_MARKER).lstrip('/')
                    # 이중 슬래시 정규화
                    while '//' in cleaned:
                        cleaned = cleaned.replace('//', '/')
                    filename = os.path.basename(cleaned)
                    relative = normalize_path(cleaned, target_project_name)
                    
                    # [상세 로깅]
                    # print(f"  [DEBUG] 라인 {idx+1}: '{raw}'")
                    # print(f"           클린: '{cleaned}'")
                    # print(f"           상대경로: '{relative}'")
                    
                    files.append({
                        'filename': filename,
                        'full_path': cleaned,
                        'relative_path': relative
                    })
            print(f"  [DEBUG] 총 {len(files)}개 파일 파싱 완료")
            return files
        except Exception as e:
            raise ValueError(f"TXT 파일 읽기 실패: {changedFileList_file} - {e}")

    # Excel 파일 읽기
    wb = load_workbook(changedFileList_file, data_only=True)
    ws = wb.active

    files = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        full_path = row[1] if len(row) > 1 else None
        if full_path:
            filename = full_path.split('/')[-1]
            files.append({
                'filename': filename,
                'full_path': full_path,
                'relative_path': normalize_path(full_path, target_project_name)
            })
    return files


# -----------------------------
# 모듈 수준 개인정보(privacy) 헬퍼
# -----------------------------

def format_endpoint_label(endpoint: str, method: Optional[str]) -> str:
    """엔드포인트 경로와 HTTP 메서드를 결합해 표시용 라벨을 생성합니다.

    Args:
        endpoint: 엔드포인트 경로 문자열
        method: HTTP 메서드(예: GET, POST) 또는 None

    Returns:
        라벨 문자열(예: "GET /api/foo")
    """
    try:
        s = str(endpoint or '').strip()
        meth = (method or '').upper().strip()

        if re.search(r'RequestMapping|GetMapping|PostMapping|PutMapping|DeleteMapping|RequestMethod|value=', s, re.IGNORECASE) or '"' in s:
            quotes = re.findall(r'"([^"]+)"', s)
            path = ''
            if quotes:
                for q in quotes:
                    if q.startswith('/') or '/' in q:
                        path = q
                        break
                if not path:
                    path = quotes[0]

            if not path:
                m = re.search(r'(/[^",\s\)]+)', s)
                if m:
                    path = m.group(1)

            m2 = re.search(r'RequestMethod\.([A-Z]+)', s)
            if m2:
                meth = m2.group(1)

        else:
            m = re.match(r'https?://[^/]+(.*)', s)
            if m:
                path = m.group(1) or '/'
            else:
                path = s if s.startswith('/') else ('/' + s.lstrip('/'))

        if not path:
            path = '/'

        if meth:
            return f'"{path}", method=RequestMethod.{meth.upper()}'
        return f'"{path}"'
    except Exception:
        ep = str(endpoint or '')
        if not ep.startswith('/') and '://' not in ep:
            ep = '/' + ep.lstrip('/')
        return f'"{ep}"'


def normalize_api_changes_text(text: str) -> str:
    """API 변경 설명 텍스트를 정리하고 길이·특수문자 제한을 적용합니다.

    Args:
        text: 원본 텍스트

    Returns:
        정리된 텍스트
    """
    if not text:
        return ''
    try:
        if '\n' in text:
            parts = [p.strip() for p in text.splitlines() if p.strip()]
        else:
            parts = [p.strip() for p in re.split(r'[;,\|]', text) if p.strip()]

        seen = set()
        out = []
        for p in parts:
            p = re.sub(r'^(추가|삭제|변경|수정):\s*', '', p).strip()
            method_match = re.search(r'\b(GET|POST|PUT|DELETE|PATCH)\b', p, re.IGNORECASE)
            path_match = re.search(r'(/[^\s,]+)', p)
            if method_match and path_match:
                method = method_match.group(1).upper()
                path = path_match.group(1)
                key = f"{method} {path}"
            else:
                key = p
            if key not in seen:
                seen.add(key)
                out.append(p)

        result = '\n'.join(out)
        if len(result) > MAX_API_CHANGES_TEXT_LENGTH:
            result = result[:MAX_API_CHANGES_TEXT_LENGTH].rsplit('\n', 1)[0] + '\n... (생략)'
        return result
    except Exception:
        return text if len(text) <= MAX_API_CHANGES_TEXT_LENGTH else text[:MAX_API_CHANGES_TEXT_LENGTH] + '\n... (생략)'


def normalize_privacy_text(original_text: str, filename: str) -> str:
    """개인정보 필드 설명을 간결한 불릿 형태로 정규화합니다."""
    if not original_text:
        return ''
    try:
        stripped = original_text.strip()
        if stripped.startswith('-'):
            parts = [p.strip() for p in original_text.splitlines() if p.strip()]
            seen = set()
            out = []
            for p in parts:
                if p not in seen:
                    seen.add(p)
                    out.append(p)
            return '\n'.join(out)

        text = re.sub(r'\s{2,}', ' ', original_text)
        segments = re.split(r'\s+-\s+|(?=\b[A-Za-z0-9_]+Controller\.java\b)|(?=\b[A-Za-z0-9_]+Service\.java\b)', text)
        bullets = []
        for seg in segments:
            seg = seg.strip()
            if not seg:
                continue
            m = re.search(r'([A-Za-z0-9_]+Controller\.java|[A-Za-z0-9_]+Service\.java)', seg)
            title = m.group(1) if m else filename
            # privacy_keywords에서 패턴 생성
            privacy_pattern = '|'.join(PRIVACY_KEYWORDS)
            tokens = re.findall(rf'\b({privacy_pattern})\b', seg, flags=re.I)
            if not tokens:
                tokens = re.findall(r'\b([a-z][a-zA-Z0-9_]{2,})\b', seg)
            evidence = None
            em = re.search(PATTERN_ENCRYPT_EVIDENCE, seg, flags=re.I)
            if em:
                evidence = em.group(0).strip('()')

            seen_tokens = set()
            out_tokens = []
            for t in tokens:
                tt = t.strip()
                if not tt or tt in seen_tokens:
                    continue
                seen_tokens.add(tt)
                if evidence:
                    out_tokens.append(f"{tt} (근거: {evidence})")
                else:
                    out_tokens.append(f"{tt}")

            if out_tokens:
                bullets.append(f"- {title} - {', '.join(out_tokens)}")

        if not bullets:
            return ''
        result = "\n".join(bullets)
        if len(result) > MAX_PRIVACY_TEXT_LENGTH:
            result = result[:MAX_PRIVACY_TEXT_LENGTH] + '\n... (생략)'
        return result
    except Exception:
        return ''


# -----------------------------
# API/텍스트 헬퍼:
# -----------------------------

def extract_api_endpoints_with_method(file_path: str) -> List[Dict]:
    """Controller 소스에서 엔드포인트 경로와 HTTP 메서드를 추출합니다.

    Args:
        file_path: 컨트롤러 소스 파일 경로

    Returns:
        endpoints: {'path': str, 'method': Optional[str]} 형태의 딕셔너리 리스트
    """
    content = read_file_safe(str(file_path))
    if not content:
        return []

    endpoints = []
    class_mapping = ''
    class_match = re.search(PATTERN_CLASS_REQUEST_MAPPING, content)
    if class_match:
        class_mapping = class_match.group(1).rstrip('/')

    method_patterns = [
        (PATTERN_GET_MAPPING, 'GET'),
        (PATTERN_POST_MAPPING, 'POST'),
        (PATTERN_PUT_MAPPING, 'PUT'),
        (PATTERN_DELETE_MAPPING, 'DELETE'),
        (PATTERN_PATCH_MAPPING, 'PATCH'),
        (PATTERN_REQUEST_MAPPING, None),
    ]

    for pattern, default_method in method_patterns:
        for match in re.findall(pattern, content):
            if isinstance(match, tuple):
                path, m = match
                method = m if m else default_method
            else:
                path = match
                method = default_method
            if not path:
                continue
            if class_mapping:
                cm = class_mapping if class_mapping.startswith('/') else '/' + class_mapping
                pm = path if path.startswith('/') else '/' + path
                if pm.startswith(cm):
                    full_path = pm
                else:
                    if pm.startswith('/api') and cm.startswith('/api'):
                        full_path = pm
                    else:
                        full_path = cm + pm
            else:
                full_path = path
            if not full_path.startswith('/'):
                full_path = '/' + full_path
            full_path = re.sub(r'/+', '/', full_path)
            if method is None and pattern.startswith(r'@RequestMapping'):
                method = 'GET'
            method = method.upper() if method else None
            endpoints.append({'path': full_path, 'method': method})

    seen = set()
    deduped = []
    for ep in endpoints:
        key = (ep['method'], ep['path'])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(ep)
    return deduped

