"""
check_join 기능 구현

- analyze 결과(.applycrypto/results/table_access_info.json)와 config(access_tables)를 이용해
  source table/column 기준 JOIN 대상(target table/column)을 LLM으로 추출합니다.
- 결과는 target_project/.applycrypto/results/check_join_results.json 에 저장합니다.
- --export 옵션으로 결과 JSON을 openpyxl로 엑셀로 내보냅니다.
"""

from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from config.config_manager import Configuration
from models.table_access_info import TableAccessInfo
from persistence.data_persistence_manager import DataPersistenceManager, PersistenceError


logger = logging.getLogger(__name__)


def _normalize_table_name(name: str) -> str:
    return (name or "").strip().lower()


def _extract_json_object(text: str) -> Dict[str, Any]:
    """
    LLM 응답 텍스트에서 JSON 오브젝트를 추출하여 dict로 반환합니다.
    - ```json ... ``` 코드블록이 있으면 그 내부를 우선 사용
    - 없으면 첫 '{'부터 마지막 '}'까지를 JSON으로 간주
    """
    if not text:
        raise ValueError("빈 응답입니다.")

    fenced = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text, re.IGNORECASE)
    candidate = fenced.group(1) if fenced else text

    start = candidate.find("{")
    end = candidate.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("응답에서 JSON 오브젝트를 찾을 수 없습니다.")

    json_str = candidate[start : end + 1]
    return json.loads(json_str)


def _read_prompt_template() -> str:
    template_path = Path(__file__).with_name("check_join_prompt.md")
    return template_path.read_text(encoding="utf-8")


def _render_prompt(template: str, *, source_table: str, source_column: str, sql_query: str) -> str:
    return (
        template.replace("{{source_table}}", source_table)
        .replace("{{source_column}}", source_column)
        .replace("{{sql_query}}", sql_query)
    )


def _dedupe_joins(joins: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """query_id, mapper_file 포함해 동일한 join이 다른 쿼리/매퍼에서 온 경우 별도 행 유지"""
    seen: set[Tuple[str, str, str, str, str, str, str]] = set()
    out: List[Dict[str, Any]] = []
    for j in joins or []:
        key = (
            str(j.get("target_table", "")).strip(),
            str(j.get("alias", "")).strip(),
            str(j.get("target_column", "")).strip(),
            str(j.get("join_type", "")).strip(),
            str(j.get("condition", "")).strip(),
            str(j.get("query_id", "")).strip(),
            str(j.get("mapper_file", "")).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(j)
    return out


def _filter_joins_by_known_tables(
    joins: List[Dict[str, Any]],
    known_table_names_normalized: set[str],
) -> List[Dict[str, Any]]:
    """
    table_access_info에 포함된 테이블명과 이름이 일치하는 target_table은 제외합니다.
    일치하지 않는(외부/다른 스키마 등) target_table만 남깁니다.
    """
    if not known_table_names_normalized:
        return list(joins) if joins else []
    out: List[Dict[str, Any]] = []
    for j in joins or []:
        if not isinstance(j, dict):
            continue
        target = _normalize_table_name(str(j.get("target_table", "")))
        if target in known_table_names_normalized:
            continue
        out.append(j)
    return out


def _normalize_column_name(name: str) -> str:
    return (name or "").strip().lower()


def _build_table_to_columns(table_access_infos: List[Any]) -> Dict[str, set[str]]:
    """
    table_access_info 리스트로부터 (정규화된 테이블명 -> 컬럼명 set) 맵을 생성합니다.
    _filter_joins_by_known_tables_columns에 넘길 table_to_columns를 만들 때 사용합니다.
    """
    table_to_columns: Dict[str, set[str]] = {}
    for t in table_access_infos or []:
        if isinstance(t, TableAccessInfo):
            table_name = t.table_name
            columns_raw = t.columns
        elif isinstance(t, dict):
            table_name = t.get("table_name") or ""
            columns_raw = t.get("columns") or []
        else:
            continue
        key = _normalize_table_name(table_name)
        if not key:
            continue
        col_set: set[str] = set()
        for c in columns_raw or []:
            if isinstance(c, str):
                col_set.add(_normalize_column_name(c))
            elif isinstance(c, dict):
                name = c.get("name")
                if name is not None:
                    col_set.add(_normalize_column_name(str(name)))
            else:
                name = getattr(c, "name", None)
                if name is not None:
                    col_set.add(_normalize_column_name(str(name)))
        table_to_columns[key] = col_set
    return table_to_columns


def _filter_joins_by_known_tables_columns(
    joins: List[Dict[str, Any]],
    table_to_columns: Dict[str, set[str]],
) -> List[Dict[str, Any]]:
    """
    table_access_info에 포함된 (테이블, 컬럼) 조합에 해당하는 join은 제외합니다.
    - target_table이 table_to_columns 키에 있고,
    - target_column이 해당 테이블의 컬럼 set에 있으면 → 제외(filter out).
    그 외의 join은 출력 목록에 포함합니다.
    table_to_columns는 _build_table_to_columns(table_access_infos)로 생성합니다.
    """
    out: List[Dict[str, Any]] = []
    for j in joins or []:
        if not isinstance(j, dict):
            continue
        target_table = _normalize_table_name(str(j.get("target_table", "")))
        target_column = _normalize_column_name(str(j.get("target_column", "")))
        if target_table in table_to_columns and target_column in table_to_columns[target_table]:
            continue
        out.append(j)
    return out


def _load_existing_results(results_path: Path) -> Dict[str, Any]:
    if not results_path.exists():
        return {"results": []}
    try:
        data = json.loads(results_path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("results", None), list):
            return data
        return {"results": []}
    except Exception:
        return {"results": []}


def _upsert_table_result(existing: Dict[str, Any], table_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    existing["results"] 내에서 source_table이 같은 항목이 있으면 교체, 없으면 append.
    """
    results = existing.get("results", [])
    if not isinstance(results, list):
        results = []

    source_table = table_result.get("source_table")
    replaced = False
    for i, item in enumerate(results):
        if isinstance(item, dict) and item.get("source_table") == source_table:
            results[i] = table_result
            replaced = True
            break
    if not replaced:
        results.append(table_result)

    existing["results"] = results
    return existing


def _iter_config_access_tables(config: Configuration) -> Iterable[Tuple[str, List[str]]]:
    """
    config.access_tables에서 (table_name, [column_name...])를 반환합니다.
    ColumnDetail(str/obj) 혼재를 안전하게 처리합니다.
    """
    for t in config.access_tables or []:
        table_name = getattr(t, "table_name", None) or (t.get("table_name") if isinstance(t, dict) else None)
        columns_raw = getattr(t, "columns", None) if not isinstance(t, dict) else t.get("columns")
        col_names: List[str] = []
        for c in columns_raw or []:
            if isinstance(c, str):
                col_names.append(c)
            elif isinstance(c, dict):
                name = c.get("name")
                if name:
                    col_names.append(str(name))
            else:
                name = getattr(c, "name", None)
                if name:
                    col_names.append(str(name))
        if table_name:
            yield str(table_name), col_names


@dataclass
class CheckJoinRunner:
    config: Configuration

    def run(self) -> Path:
        """
        check_join 실행: 결과 JSON 저장 후 저장 경로를 반환합니다.
        """
        target_project = Path(self.config.target_project)
        pm = DataPersistenceManager(target_project=target_project)

        try:
            raw_list = pm.load_from_file("table_access_info.json")
        except PersistenceError as e:
            raise RuntimeError(f"table_access_info.json 로드 실패: {e}") from e

        table_access_infos: List[TableAccessInfo] = []
        for item in raw_list or []:
            if isinstance(item, TableAccessInfo):
                table_access_infos.append(item)
            elif isinstance(item, dict):
                table_access_infos.append(TableAccessInfo.from_dict(item))

        by_table: Dict[str, TableAccessInfo] = {
            _normalize_table_name(t.table_name): t for t in table_access_infos if t and t.table_name
        }
        table_to_columns = _build_table_to_columns(table_access_infos)

        template = _read_prompt_template()
        from modifier.llm.llm_factory import create_llm_provider

        llm = create_llm_provider(provider_name=self.config.llm_provider)

        results_path = pm.output_dir / "check_join_results.json"
        # 매 실행 시 기존 파일이 있어도 내용을 비우고 이번 실행 결과만 새로 씁니다.
        existing: Dict[str, Any] = {"results": []}

        access_tables_list = list(_iter_config_access_tables(self.config))
        total_tables = len(access_tables_list)

        for table_index, (source_table, source_columns) in enumerate(access_tables_list, start=1):
            logger.info(
                f"source table: {source_table} ({table_index}/{total_tables}) 처리중....."
            )
            table_key = _normalize_table_name(source_table)
            table_info = by_table.get(table_key)
            if not table_info:
                logger.warning(f"table_access_info에 테이블이 없습니다: {source_table}")
                table_result = {
                    "source_table": source_table,
                    "source_columns": [
                        {"column_name": col, "joins": []} for col in (source_columns or [])
                    ],
                }
                existing = _upsert_table_result(existing, table_result)
                results_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
                continue

            sql_queries = table_info.sql_queries or []

            source_columns_results: List[Dict[str, Any]] = []
            for col in source_columns or []:
                logger.info(f"  source column: {col} 처리중.....")
                all_joins: List[Dict[str, Any]] = []

                for q in sql_queries:
                    sql = ""
                    query_id = ""
                    if isinstance(q, dict):
                        sql = q.get("sql", "") or ""
                        query_id = str(q.get("id", "") or "").strip() or "(no id)"
                    if not sql.strip():
                        continue

                    logger.info(f"    query id: {query_id} 점검중.....")

                    prompt = _render_prompt(
                        template,
                        source_table=source_table,
                        source_column=col,
                        sql_query=sql,
                    )

                    try:
                        resp = llm.call(prompt)
                        content = str(resp.get("content", "") if isinstance(resp, dict) else resp)
                        parsed = _extract_json_object(content)
                    except Exception as e:
                        logger.warning(
                            f"LLM/파싱 실패 (table={source_table}, col={col}): {e}"
                        )
                        continue

                    # 단일 컬럼 결과를 기대하지만, 안전하게 joins만 수집
                    # target_table+target_column이 table_access_info에 있으면 제외
                    mapper_file = ""
                    if isinstance(q, dict) and q.get("source_file_path"):
                        mapper_file = Path(q["source_file_path"]).name
                    try:
                        parsed_cols = parsed.get("source_columns", [])
                        if parsed_cols and isinstance(parsed_cols, list):
                            for pc in parsed_cols:
                                if not isinstance(pc, dict):
                                    continue
                                joins = pc.get("joins", [])
                                if isinstance(joins, list):
                                    filtered_joins = _filter_joins_by_known_tables_columns(
                                        [j for j in joins if isinstance(j, dict)],
                                        table_to_columns,
                                    )
                                    for j in filtered_joins:
                                        j_with_meta = dict(j)
                                        j_with_meta["query_id"] = query_id
                                        j_with_meta["mapper_file"] = mapper_file
                                        all_joins.append(j_with_meta)
                    except Exception:
                        continue

                source_columns_results.append(
                    {"column_name": col, "joins": _dedupe_joins(all_joins)}
                )

            table_result = {
                "source_table": source_table,
                "source_columns": source_columns_results,
            }

            existing = _upsert_table_result(existing, table_result)
            results_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

        return results_path

    def export_to_excel(self) -> Path:
        """
        check_join_results.json을 읽어 openpyxl로 엑셀로 내보냅니다.
        셀 테두리, 제목 대문자, 자동 필터, 헤더 회색 음영, 열 너비 자동 조정을 적용합니다.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        target_project = Path(self.config.target_project)
        pm = DataPersistenceManager(target_project=target_project)

        results_path = pm.output_dir / "check_join_results.json"
        if not results_path.exists():
            raise RuntimeError(f"결과 파일이 없습니다: {results_path}")

        data = json.loads(results_path.read_text(encoding="utf-8"))
        results = data.get("results", []) if isinstance(data, dict) else []
        if not isinstance(results, list):
            results = []

        rows: List[Dict[str, Any]] = []
        for table_item in results:
            if not isinstance(table_item, dict):
                continue
            source_table = table_item.get("source_table", "")
            source_columns = table_item.get("source_columns", [])
            if not isinstance(source_columns, list):
                continue
            for col_item in source_columns:
                if not isinstance(col_item, dict):
                    continue
                column_name = col_item.get("column_name", "")
                joins = col_item.get("joins", [])
                if not joins:
                    rows.append(
                        {
                            "source_table": source_table,
                            "column_name": column_name,
                            "target_table": "",
                            "alias": "",
                            "target_column": "",
                            "join_type": "",
                            "condition": "",
                            "query_id": "",
                            "mapper_file": "",
                        }
                    )
                    continue
                if isinstance(joins, list):
                    for j in joins:
                        if not isinstance(j, dict):
                            continue
                        rows.append(
                            {
                                "source_table": source_table,
                                "column_name": column_name,
                                "target_table": j.get("target_table", ""),
                                "alias": j.get("alias", ""),
                                "target_column": j.get("target_column", ""),
                                "join_type": j.get("join_type", ""),
                                "condition": j.get("condition", ""),
                                "query_id": j.get("query_id", ""),
                                "mapper_file": j.get("mapper_file", ""),
                            }
                        )

        headers = [
            "source_table",
            "column_name",
            "target_table",
            "alias",
            "target_column",
            "join_type",
            "condition",
            "query_id",
            "mapper_file",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "check_join"

        # 제목 행: 대문자로 쓰고, 회색 음영 적용
        header_row = [h.upper() for h in headers]
        ws.append(header_row)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        max_row = ws.max_row
        max_col = len(headers)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row == 1:
                    cell.fill = gray_fill

        # 자동 필터 (A1 ~ 마지막 셀)
        if max_row >= 1 and max_col >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # 열 너비: 데이터에 맞게 확장 (최소 8, 최대 50)
        for col in range(1, max_col + 1):
            width = 8
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is not None:
                    # 한글 등은 약 2단위로 간주
                    length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    width = max(width, min(50, length + 2))
            ws.column_dimensions[get_column_letter(col)].width = width

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"check_join_results.json_{timestamp}.xlsx"
        artifacts_dir = target_project / ".applycrypto" / "artifacts"
        artifacts_dir.mkdir(parents=True, exist_ok=True)
        out_path = artifacts_dir / out_name
        wb.save(out_path)
        return out_path

