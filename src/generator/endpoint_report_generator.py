"""
End Point 목록 보고서 생성기 (Endpoint Report Generator)

변경된 파일의 메소드들이 어떤 End Point에 매핑되는지를 Excel로 생성합니다.
"""

import os
import json
import glob
import difflib
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Set, Tuple
from dataclasses import dataclass, field
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config.config_manager import Configuration
from parser.java_ast_parser import JavaASTParser
from analyzer.callgraph_endpoint_finder import (
    find_all_endpoints_for_method,
    find_endpoints_that_call_method
)

# =====================================
# 상수 정의
# =====================================

# 기본 설정
DEFAULT_FONT_NAME = '맑은 고딕'
DEFAULT_FONT_SIZE = 10
DATE_FORMAT = '%Y%m%d'
DEFAULT_ARTIFACTS_DIR = '.applycrypto/artifacts'
CALL_GRAPH_PATH = '.applycrypto/results/call_graph.json'

# 파일 처리
FILE_ENCODINGS = ['utf-8', 'euc-kr', 'cp949', 'latin-1']
FILE_ENCODING_WITH_BOM = 'utf-8-sig'
BOM_MARKER = '\ufeff'

# Excel 색상 코드
COLOR_HEADER = 'DFDFDF'

# 열 정보
COLUMNS = {
    'A': {'width': 3, 'header': ''},
    'B': {'width': 45, 'header': '파일 경로'},
    'C': {'width': 30, 'header': '파일명'},
    'D': {'width': 35, 'header': '메소드명'},
    'E': {'width': 35, 'header': '엔드포인트'},
}

# =====================================
# 데이터 모델
# =====================================

@dataclass
class EndPointData:
    """엔드포인트 데이터 모델"""
    file_path: str          # 전체 파일 경로
    file_name: str          # 파일명 (확장자 포함)
    method_name: str        # 메소드명
    endpoint: str           # HTTP_METHOD /path (e.g., "GET /getBList.do")
    
    def __lt__(self, other):
        """정렬용 비교 연산자"""
        if self.file_path != other.file_path:
            return self.file_path < other.file_path
        if self.file_name != other.file_name:
            return self.file_name < other.file_name
        if self.method_name != other.method_name:
            return self.method_name < other.method_name
        return self.endpoint < other.endpoint


# =====================================
# 진입점(Entry point)
# =====================================

def generate_endpoint_report(config: Configuration):
    """
    End Point 목록 보고서를 생성합니다.
    
    Args:
        config (Configuration): ApplyCrypto 설정 객체
        
    Returns:
        None: 워크북을 파일로 저장합니다.
    """
    logger = logging.getLogger(__name__)
    
    target_project = config.target_project
    old_code_path = config.artifact_generation.old_code_path if config.artifact_generation else None

    if not target_project:
        raise ValueError('config must include target_project')
    
    if not old_code_path:
        raise ValueError('config must include artifact_generation.old_code_path')

    # 경로 설정
    tp = Path(target_project)
    op = Path(old_code_path) if old_code_path else None

    if not tp.exists():
        raise FileNotFoundError(f"타겟 프로젝트를 찾을 수 없습니다: {target_project}")
    if op is None or not op.exists():
        raise FileNotFoundError(f"원본 백업을 찾을 수 없습니다: {old_code_path}")

    # Call Graph JSON 로드
    call_graph_file = tp / CALL_GRAPH_PATH
    if not call_graph_file.exists():
        raise FileNotFoundError(f"Call Graph JSON을 찾을 수 없습니다: {call_graph_file}")

    logger.info(f"Call Graph 로드: {call_graph_file}")
    call_graph = load_call_graph(str(call_graph_file))

    # 변경된 파일 추출
    logger.info("변경된 파일 비교 시작...")
    changed_files = get_changed_java_files(str(tp), str(op))
    
    if not changed_files:
        logger.warning("변경된 Java 파일이 없습니다. 전체 프로젝트 파일을 사용합니다.")
        changed_files = get_all_java_files(str(tp))
    
    logger.info(f"비교할 파일 수: {len(changed_files)}")

    # 메소드 추출 및 엔드포인트 매칭
    logger.info("메소드 및 엔드포인트 추출 시작...")
    endpoint_data_list = []
    
    for file_path in changed_files:
        try:
            methods = extract_changed_methods(file_path, str(tp), str(op))
            
            if methods:
                logger.debug(f"파일 {file_path}: 추출된 메소드 = {methods}")
            
            # target_project 기준의 상대 경로 계산 (디렉토리만)
            rel_path = os.path.relpath(file_path, str(tp))
            dir_path = os.path.dirname(rel_path)
            # 경로 구분자를 \ 로 통일 및 마지막에 \ 추가
            dir_path = dir_path.replace('/', '\\')
            if not dir_path.endswith('\\'):
                dir_path += '\\'
            
            for method_name in methods:
                # find_endpoint_for_method가 리스트 반환 (여러 엔드포인트 가능)
                endpoints = find_endpoint_for_method(method_name, call_graph)
                
                if endpoints:
                    # 여러 엔드포인트가 있으면 각각에 대해 행 생성
                    logger.debug(f"메소드 {method_name}: {len(endpoints)}개 엔드포인트 발견")
                    for endpoint in endpoints:
                        endpoint_data = EndPointData(
                            file_path=dir_path,
                            file_name=os.path.basename(file_path),
                            method_name=method_name,
                            endpoint=endpoint
                        )
                        endpoint_data_list.append(endpoint_data)
                else:
                    # 엔드포인트가 없으면 ""으로 1행 생성
                    logger.debug(f"엔드포인트 미매칭: {method_name}")
                    endpoint_data = EndPointData(
                        file_path=dir_path,
                        file_name=os.path.basename(file_path),
                        method_name=method_name,
                        endpoint=""
                    )
                    endpoint_data_list.append(endpoint_data)
        except Exception as e:
            logger.warning(f"파일 처리 오류 ({file_path}): {e}")
            continue
    
    if not endpoint_data_list:
        logger.warning("추출된 엔드포인트가 없습니다.")
        endpoint_data_list = []

    # 정렬
    logger.info(f"총 {len(endpoint_data_list)}개 엔드포인트 추출")
    endpoint_data_list.sort()

    # 메소드명 검증 (Excel 생성 전)
    logger.info("\n" + "="*80)
    logger.info("[메소드명 검증]")
    logger.info("="*80)
    # 프로젝트의 실제 존재하는 메소드 맵 구축
    actual_methods_map = build_project_method_map(str(tp))
    validate_and_print_method_names(endpoint_data_list, actual_methods_map, str(tp))

    # Excel 생성
    logger.info("\n" + "="*80)
    logger.info("Excel 생성 시작...")
    out_dir = tp / DEFAULT_ARTIFACTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    
    today = datetime.now().strftime(DATE_FORMAT)
    output_file = out_dir / f"EndPoint_Report_{today}.xlsx"
    
    create_endpoint_workbook(endpoint_data_list, str(output_file))
    
    logger.info(f"End Point 보고서 생성 완료: {output_file}")
    try:
        print(f"✓ End Point 보고서 생성 완료: {output_file}")
    except UnicodeEncodeError:
        print(f"[OK] End Point 보고서 생성 완료: {output_file}")


# =====================================
# Call Graph 처리
# =====================================

def load_call_graph(call_graph_file: str) -> Dict:
    """
    Call Graph JSON을 로드합니다.
    
    Args:
        call_graph_file: Call Graph JSON 파일 경로
        
    Returns:
        Dict: Call Graph 데이터
    """
    with open(call_graph_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def find_endpoint_for_method(method_signature: str, call_graph: Dict) -> List[str]:
    """
    메소드 시그니처에 대응하는 모든 엔드포인트를 찾습니다.
    
    2단계 검색:
    1. endpoints에서 직접 검색 (Controller 메소드)
       → 1개 엔드포인트 반환
    2. call_trees에서 역추적 검색 (Service/Mapper/Interceptor 메소드)
       → 여러 엔드포인트 반환 (이 메소드를 호출하는 모든 엔드포인트)
    
    예시:
    - EmpController.addEmpByGet → ['/api/emps/addEmpByGet']
    - Employee.getDayOfBirth → [
        '/api/emps/addEmpByGet',
        '/api/emps/addEmpByPost',
        '/api/emps/editEmpByGet',
        '/api/emps/editEmpByPost',
        '/emp/addEmpByGet',
        '/emp/editEmpByGet',
        '/emp/editEmpByPost',
        ... (총 8개)
      ]
    
    Args:
        method_signature: 메소드 시그니처 (e.g., "Employee.getDayOfBirth")
        call_graph: Call Graph 데이터 (endpoints + call_trees)
        
    Returns:
        List[str]: 엔드포인트 경로들 (e.g., ['/api/emps/addEmpByGet', '/api/emps/addEmpByPost', ...])
                   매칭되는 엔드포인트가 없으면 빈 리스트
    """
    logger = logging.getLogger(__name__)
    
    # Step 1: endpoints에서 직접 검색 (Controller 메소드)
    results = find_all_endpoints_for_method(
        method_signature=method_signature,
        call_graph_data=call_graph,
        return_type="dict"
    )
    
    formatted_endpoints = []
    if results:
        for result in results:
            # HTTP_METHOD /path 형식으로 변환 (다중 경로 지원)
            http_method = result.get("http_method", "GET")
            paths = result.get("path", [])
            if isinstance(paths, str):
                paths = [paths]
                
            for p in paths:
                formatted_endpoints.append(f"{http_method} {p}")
            
        logger.debug(f"[직접 매칭] {method_signature} -> {formatted_endpoints}")
        return formatted_endpoints

    # Step 2: call_trees에서 역추적 (Service/Mapper 메소드)
    # 이 메소드를 호출하는 모든 엔드포인트 찾기
    matching_endpoints = find_endpoints_that_call_method(
        method_signature=method_signature,
        call_graph_data=call_graph,
        return_type="dict"
    )

    if matching_endpoints:
        formatted_endpoints = []
        for ep in matching_endpoints:
            http_method = ep.get("http_method", "GET")
            paths = ep.get("path", [])
            if isinstance(paths, str):
                paths = [paths]
            
            for p in paths:
                formatted_endpoints.append(f"{http_method} {p}")
                
        logger.debug(
            f"[역추적 매칭] {method_signature} -> {len(matching_endpoints)}개 엔드포인트: {formatted_endpoints}"
        )
        # 중복 제거 (리스트 순서 유지보다 중복 제거가 중요)
        return sorted(list(set(formatted_endpoints)))
    
    logger.debug(f"[엔드포인트 없음] {method_signature}")
    return []


# =====================================
# 파일 비교 및 메소드 추출
# =====================================

def extract_class_name(file_path: str) -> str:
    """
    Java 파일에서 클래스명을 추출합니다.
    JavaASTParser를 사용하여 정확하게 추출합니다.
    
    Args:
        file_path: Java 파일 경로
        
    Returns:
        str: 클래스명 (추출 실패 시 '')
    """
    try:
        parser = JavaASTParser()
        tree, error = parser.parse_file(Path(file_path), remove_comments=False)
        
        if error or tree is None:
            # AST 파싱 실패 시 정규식으로 간단히 추출
            return extract_class_name_regex(file_path)
        
        class_infos = parser.extract_class_info(tree, Path(file_path))
        
        if class_infos:
            # 첫 번째 public 클래스 또는 첫 번째 클래스 반환
            return class_infos[0].name
        
        return ''
    except Exception as e:
        # 오류 발생 시 정규식으로 재시도
        return extract_class_name_regex(file_path)


def extract_class_name_regex(file_path: str) -> str:
    """
    정규식을 사용하여 Java 파일에서 클래스명을 추출합니다.
    (JavaASTParser 실패 시 fallback)
    
    Args:
        file_path: Java 파일 경로
        
    Returns:
        str: 클래스명 (추출 실패 시 '')
    """
    import re
    
    try:
        content = read_file_safe(file_path)
        # 정규식: public/private/protected class ClassName { ... }
        match = re.search(r'(?:public|private|protected)?\s*(?:abstract)?\s*class\s+(\w+)', content)
        if match:
            return match.group(1)
        return ''
    except Exception:
        return ''


def get_changed_java_files(target_project: str, old_code_path: str) -> List[str]:
    """
    변경된 Java 파일을 찾습니다.
    
    Args:
        target_project: 타겟 프로젝트 경로
        old_code_path: 원본 프로젝트 경로
        
    Returns:
        List[str]: 변경된 Java 파일 경로 목록
    """
    logger = logging.getLogger(__name__)
    changed_files = []
    target_java_files = {}
    old_java_files = {}

    # src 디렉토리 자체에서 Java 파일 찾기 (src/main/java 형식도 지원)
    target_src = target_project
    old_src = old_code_path
    
    # 만약 target_project 내에 src/main/java가 존재하면 그것을 사용
    if os.path.exists(os.path.join(target_project, "src", "main", "java")):
        target_src = os.path.join(target_project, "src", "main", "java")
        old_src = os.path.join(old_code_path, "src", "main", "java")
        logger.debug(f"Maven 구조 감지: {target_src}")
    else:
        # 그 외의 경우 src 디렉토리 직접 사용
        target_src = target_project
        old_src = old_code_path
        logger.debug(f"표준 구조 감지: {target_src}")
    
    # 타겟 프로젝트의 모든 Java 파일 수집
    for root, dirs, files in os.walk(target_src):
        # .applycrypto 폴더는 제외
        dirs[:] = [d for d in dirs if d != '.applycrypto']
        
        for file in files:
            if file.endswith('.java'):
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, target_src)
                target_java_files[rel_path] = full_path
    
    # 원본 프로젝트의 모든 Java 파일 수집
    if os.path.exists(old_src):
        for root, dirs, files in os.walk(old_src):
            dirs[:] = [d for d in dirs if d != '.applycrypto']
            
            for file in files:
                if file.endswith('.java'):
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, old_src)
                    old_java_files[rel_path] = full_path
    
    logger.debug(f"타겟 Java 파일: {len(target_java_files)}개, 원본 Java 파일: {len(old_java_files)}개")
    
    # 변경된 파일 비교
    for rel_path, target_path in target_java_files.items():
        if rel_path in old_java_files:
            old_path = old_java_files[rel_path]
            if file_has_changed(target_path, old_path):
                logger.debug(f"변경됨: {rel_path}")
                changed_files.append(target_path)
        else:
            # 새로 생성된 파일
            logger.debug(f"신규: {rel_path}")
            changed_files.append(target_path)
    
    return changed_files


def get_all_java_files(target_project: str) -> List[str]:
    """
    프로젝트의 모든 Java 파일을 찾습니다.
    
    Args:
        target_project: 타겟 프로젝트 경로
        
    Returns:
        List[str]: 모든 Java 파일 경로 목록
    """
    java_files = []
    
    for root, dirs, files in os.walk(target_project):
        dirs[:] = [d for d in dirs if d != '.applycrypto']
        
        for file in files:
            if file.endswith('.java'):
                full_path = os.path.join(root, file)
                java_files.append(full_path)
    
    return java_files


def file_has_changed(target_file: str, old_file: str) -> bool:
    """
    두 파일의 내용이 변경되었는지 확인합니다.
    
    Args:
        target_file: 타겟 파일 경로
        old_file: 원본 파일 경로
        
    Returns:
        bool: 변경 여부
    """
    try:
        target_content = read_file_safe(target_file)
        old_content = read_file_safe(old_file)
        
        return target_content != old_content
    except Exception:
        return False


def read_file_safe(file_path: str) -> str:
    """
    파일을 안전하게 읽습니다. 다양한 인코딩을 시도합니다.
    
    Args:
        file_path: 파일 경로
        
    Returns:
        str: 파일 내용
    """
    for encoding in FILE_ENCODINGS:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
                if content.startswith(BOM_MARKER):
                    content = content[1:]
                return content
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    # 모든 인코딩 실패 시 binary 읽기
    with open(file_path, 'rb') as f:
        return f.read().decode('utf-8', errors='ignore')


def extract_changed_methods(file_path: str, target_project: str, old_code_path: str) -> Set[str]:
    """
    변경된 파일에서 변경된 메소드들을 AST parser를 이용해 추출합니다.
    
    방법: difflib(라인 변경 감지) + AST parser(메소드 범위 파악) + 메소드 블럭 비교(검증)
    
    1. difflib로 변경된 라인 식별
    2. AST parser로 메소드 범위 파악
    3. 변경된 라인이 속한 메소드 식별
    4. AST 메소드 블럭 비교로 변경 확인
    
    Args:
        file_path: 변경된 파일 경로 (타겟 프로젝트 기준)
        target_project: 타겟 프로젝트 경로
        old_code_path: 원본 프로젝트 경로
        
    Returns:
        Set[str]: 변경된 메소드 ('클래스명.메소드명' 형태)
    """
    logger = logging.getLogger(__name__)
    
    # 상대 경로 계산
    rel_path = os.path.relpath(file_path, target_project)
    old_file_path = os.path.join(old_code_path, rel_path)
    
    # 파일 읽기
    try:
        target_content = read_file_safe(file_path)
    except Exception as e:
        logger.warning(f"파일 읽기 실패: {file_path} - {e}")
        return set()
    
    # 클래스명 추출 (AST parser 우선)
    class_name = extract_class_name(file_path)
    if not class_name:
        logger.warning(f"클래스명 추출 실패: {file_path}")
        return set()
    
    # 새로 생성된 파일 처리
    if not os.path.exists(old_file_path):
        logger.debug(f"새 파일 {file_path}: 모든 메소드 추출")
        methods = extract_methods_with_ast(file_path)
        return {f"{class_name}.{m}" for m in methods}
    
    # 원본 파일 읽기
    try:
        old_content = read_file_safe(old_file_path)
    except Exception as e:
        logger.warning(f"원본 파일 읽기 실패: {old_file_path} - {e}")
        # 원본을 읽을 수 없으면 전체 파일의 메소드 추출
        methods = extract_methods_with_ast(file_path)
        return {f"{class_name}.{m}" for m in methods}
    
    # 동일한 경우 빈 집합 반환
    if target_content == old_content:
        logger.debug(f"파일 동일: {file_path}")
        return set()
    
    # ===== 핵심 로직: difflib + AST 기반 메소드 추출 =====
    target_lines = target_content.split('\n')
    old_lines = old_content.split('\n')
    
    # Step 1: difflib로 변경된 라인 식별
    changed_line_indices = identify_changed_lines(target_lines, old_lines)
    
    if not changed_line_indices:
        logger.debug(f"변경된 라인 없음: {file_path}")
        return set()
    
    # Step 2: AST parser로 현재 파일의 메소드 범위 파악
    method_ranges = extract_method_ranges_with_ast(file_path)
    
    if not method_ranges:
        logger.warning(f"메소드 범위 파악 실패: {file_path}")
        return set()
    
    # Step 3: 변경된 라인이 속한 메소드 식별
    changed_methods = map_changed_lines_to_methods(changed_line_indices, method_ranges)
    
    if changed_methods:
        logger.debug(f"변경된 메소드 식별: {file_path} -> {changed_methods}")
        return {f"{class_name}.{m}" for m in changed_methods}
    
    # Step 4: 메소드 블럭 비교 (검증)
    # as-is와 to-be 메소드 블럭 비교로 변경 확인
    changed_methods = compare_method_blocks_with_ast(file_path, old_file_path)
    
    if changed_methods:
        logger.debug(f"메소드 블럭 비교로 추출: {file_path} -> {changed_methods}")
        return {f"{class_name}.{m}" for m in changed_methods}
    
    logger.debug(f"변경된 메소드 미발견: {file_path}")
    return set()


def identify_changed_lines(target_lines: List[str], old_lines: List[str]) -> Set[int]:
    """
    difflib를 이용해 변경된 라인을 식별합니다.
    
    Args:
        target_lines: 타겟 파일 라인 목록
        old_lines: 원본 파일 라인 목록
        
    Returns:
        Set[int]: 변경된 라인 인덱스 (0-based)
    """
    changed_indices = set()
    
    seq = difflib.SequenceMatcher(None, old_lines, target_lines)
    
    for tag, i1, i2, j1, j2 in seq.get_opcodes():
        if tag != 'equal':
            # 추가/삭제/변경된 라인
            changed_indices.update(range(j1, j2))
    
    return changed_indices


def extract_methods_with_ast(file_path: str) -> Set[str]:
    """
    AST parser를 이용해 Java 파일의 모든 메소드를 추출합니다.
    
    Args:
        file_path: Java 파일 경로
        
    Returns:
        Set[str]: 메소드명 집합
    """
    logger = logging.getLogger(__name__)
    
    try:
        parser = JavaASTParser()
        tree, error = parser.parse_file(Path(file_path), remove_comments=False)
        
        if error or tree is None:
            logger.warning(f"AST 파싱 실패: {file_path}")
            return set()
        
        class_infos = parser.extract_class_info(tree, Path(file_path))
        methods = set()
        
        for class_info in class_infos:
            for method in class_info.methods:
                methods.add(method.name)
        
        return methods
    except Exception as e:
        logger.warning(f"AST 메소드 추출 오류: {file_path} - {e}")
        return set()


def extract_method_ranges_with_ast(file_path: str) -> List[Tuple[Tuple[int, int], str]]:
    """
    AST parser를 이용해 메소드의 라인 범위를 추출합니다.
    
    Args:
        file_path: Java 파일 경로
        
    Returns:
        List[Tuple[Tuple[int, int], str]]: [((시작라인, 종료라인), 메소드명)] 리스트 (라인은 0-based)
    """
    logger = logging.getLogger(__name__)
    
    try:
        parser = JavaASTParser()
        tree, error = parser.parse_file(Path(file_path), remove_comments=False)
        
        if error or tree is None:
            logger.warning(f"AST 파싱 실패: {file_path}")
            return []
        
        class_infos = parser.extract_class_info(tree, Path(file_path))
        method_ranges = []
        
        for class_info in class_infos:
            for method in class_info.methods:
                # Method 객체의 line_start, line_end 사용 (0-based)
                if method.line_start >= 0 and method.line_end >= 0:
                    method_ranges.append(((method.line_start, method.line_end), method.name))
        
        return method_ranges
    except Exception as e:
        logger.warning(f"AST 메소드 범위 추출 오류: {file_path} - {e}")
        return []


def map_changed_lines_to_methods(changed_lines: Set[int], method_ranges: List[Tuple[Tuple[int, int], str]]) -> Set[str]:
    """
    변경된 라인이 속한 메소드를 식별합니다.
    
    Args:
        changed_lines: 변경된 라인 인덱스 (0-based)
        method_ranges: [((시작라인, 종료라인), 메소드명)] 리스트
        
    Returns:
        Set[str]: 변경된 메소드명 집합
    """
    changed_methods = set()
    
    for line_idx in changed_lines:
        for (start_line, end_line), method_name in method_ranges:
            if start_line <= line_idx+1 <= end_line:
                changed_methods.add(method_name)
                break
    
    return changed_methods


def compare_method_blocks_with_ast(target_file: str, old_file: str) -> Set[str]:
    """
    AST parser를 이용해 메소드 블럭을 비교합니다.
    as-is와 to-be의 메소드 블럭이 다른 경우 변경된 메소드로 식별합니다.
    
    Args:
        target_file: 타겟 파일 경로
        old_file: 원본 파일 경로
        
    Returns:
        Set[str]: 변경된 메소드명 집합
    """
    logger = logging.getLogger(__name__)
    
    try:
        parser = JavaASTParser()
        
        # 타겟 파일 파싱
        tree_target, error_target = parser.parse_file(Path(target_file), remove_comments=False)
        if error_target or tree_target is None:
            logger.warning(f"타겟 파일 AST 파싱 실패: {target_file}")
            return set()
        
        # 원본 파일 파싱
        tree_old, error_old = parser.parse_file(Path(old_file), remove_comments=False)
        if error_old or tree_old is None:
            logger.warning(f"원본 파일 AST 파싱 실패: {old_file}")
            return set()
        
        # 클래스 정보 추출
        target_classes = parser.extract_class_info(tree_target, Path(target_file))
        old_classes = parser.extract_class_info(tree_old, Path(old_file))
        
        # 메소드 맵 생성 {메소드명: 메소드 객체}
        target_methods = {}
        old_methods = {}
        
        for class_info in target_classes:
            for method in class_info.methods:
                target_methods[method.name] = method
        
        for class_info in old_classes:
            for method in class_info.methods:
                old_methods[method.name] = method
        
        changed_methods = set()
        
        # 1. 새로 추가된 메소드
        for method_name in target_methods:
            if method_name not in old_methods:
                changed_methods.add(method_name)
        
        # 2. 제거된 메소드
        for method_name in old_methods:
            if method_name not in target_methods:
                changed_methods.add(method_name)
        
        # 3. 기존 메소드 중 내용이 변경된 경우
        for method_name in target_methods:
            if method_name in old_methods:
                target_method = target_methods[method_name]
                old_method = old_methods[method_name]
                
                # 메소드 소스 코드 비교 (source 속성 사용)
                target_source = target_method.source if target_method.source else ""
                old_source = old_method.source if old_method.source else ""
                
                # 소스 정규화 (공백/줄바꿈 제거)
                target_normalized = re.sub(r'\s+', ' ', target_source).strip()
                old_normalized = re.sub(r'\s+', ' ', old_source).strip()
                
                if target_normalized != old_normalized:
                    changed_methods.add(method_name)
        
        return changed_methods
    except Exception as e:
        logger.warning(f"메소드 블럭 비교 오류: {e}")
        return set()


def extract_all_methods(content: str) -> Set[str]:
    """
    Java 파일에서 모든 메소드를 추출합니다.
    
    Args:
        content: Java 파일 내용
        
    Returns:
        Set[str]: 메소드명 집합
    """
    import re
    
    methods = set()
    
    # 메소드 선언 정규식 (개선된 버전)
    # 패턴: [접근제어자] [static] [return_type] method_name(...)
    patterns = [
        r'(?:public|protected|private)\s+(?:static\s+)?(\w+)\s+(\w+)\s*\(',
        r'(?:public|protected|private)\s+(\w+)\s+(\w+)\s*\(',
        r'\b(\w+)\s+(\w+)\s*\(',  # 접근제어자 없는 경우
    ]
    
    for pattern in patterns:
        for match in re.finditer(pattern, content):
            # 마지막 캡처 그룹이 메소드명
            if len(match.groups()) >= 2:
                method_name = match.group(2)
            else:
                method_name = match.group(1)
                
            # 필터링: 생성자, 클래스명 제외
            if method_name and not method_name[0].isupper():
                methods.add(method_name)
    
    return methods


def remove_comments_and_strings(line: str) -> str:
    """
    Java 코드에서 주석과 문자열 리터럴을 제거합니다.
    
    Args:
        line: Java 소스 코드 라인
        
    Returns:
        str: 주석과 문자열이 제거된 라인 (괄호는 공백으로 대체)
    """
    result = []
    i = 0
    while i < len(line):
        # 블록 주석 처리 /* ... */
        if i < len(line) - 1 and line[i:i+2] == '/*':
            end = line.find('*/', i + 2)
            if end != -1:
                i = end + 2
            else:
                break
            continue
        
        # 라인 주석 처리 // ...
        if i < len(line) - 1 and line[i:i+2] == '//':
            break
        
        # 문자열 처리 (큰따옴표)
        if line[i] == '"':
            result.append(' ')
            i += 1
            while i < len(line):
                if line[i] == '\\' and i + 1 < len(line):
                    i += 2
                elif line[i] == '"':
                    i += 1
                    break
                else:
                    i += 1
            continue
        
        # 문자 처리 (작은따옴표)
        if line[i] == "'":
            result.append(' ')
            i += 1
            while i < len(line):
                if line[i] == '\\' and i + 1 < len(line):
                    i += 2
                elif line[i] == "'":
                    i += 1
                    break
                else:
                    i += 1
            continue
        
        result.append(line[i])
        i += 1
    
    return ''.join(result)


def build_method_ranges(lines: List[str]) -> List[Tuple[Tuple[int, int], str]]:
    """
    Java 파일의 모든 메소드 범위를 구성합니다.
    각 메소드의 시작 라인과 종료 라인을 식별합니다.
    
    Args:
        lines: 파일 라인 목록
        
    Returns:
        List[Tuple[Tuple[int, int], str]]: [((시작행, 종료행), 메소드명)] 리스트
    """
    import re
    
    method_ranges = []
    method_pattern = r'(?:public|protected|private|abstract|static|synchronized|native|default).*?\b(\w+)\s*\('
    
    brace_level = 0
    method_start = None
    method_name = None
    in_method = False
    
    for line_idx, line in enumerate(lines):
        stripped = line.strip()
        
        # 주석 줄 무시
        if stripped.startswith('//'):
            continue
        
        # 주석과 문자열 제거
        cleaned_line = remove_comments_and_strings(line)
        
        # 메소드 선언 찾기
        if not in_method and method_start is None:
            match = re.search(method_pattern, line)
            if match:
                potential_method = match.group(1)
                if potential_method and not potential_method[0].isupper():
                    method_start = line_idx
                    method_name = potential_method
                    in_method = True
                    brace_level = 0
                    # 메소드 시작 라인의 { 를 카운트
                    brace_level += cleaned_line.count('{') - cleaned_line.count('}')
        
        # 메소드 범위 추적
        elif in_method:
            brace_level += cleaned_line.count('{') - cleaned_line.count('}')
            
            if brace_level == 0 and in_method:
                if method_name and method_start is not None:
                    method_ranges.append(((method_start, line_idx), method_name))
                in_method = False
                method_start = None
                method_name = None
    
    return method_ranges


def extract_methods_from_diff(target_lines: List[str], old_lines: List[str]) -> Set[str]:
    """
    Diff 결과에서 변경된 메소드를 추출합니다.
    변경된 각 라인이 속한 메소드를 정확하게 식별합니다.
    
    Args:
        target_lines: 타겟 파일 라인 목록
        old_lines: 원본 파일 라인 목록
        
    Returns:
        Set[str]: 변경된 메소드명 집합
    """
    changed_methods = set()
    
    # 라인별 비교를 통해 변경된 라인 찾기
    seq = difflib.SequenceMatcher(None, old_lines, target_lines)
    
    changed_line_indices = set()
    for tag, i1, i2, j1, j2 in seq.get_opcodes():
        if tag != 'equal':
            changed_line_indices.update(range(j1, j2))
    
    if not changed_line_indices:
        return set()
    
    # 메소드 범위 맵 생성
    method_ranges = build_method_ranges(target_lines)
    
    # 변경된 각 라인이 속한 메소드 찾기
    for line_idx in changed_line_indices:
        for (start, end), method_name in method_ranges:
            if start <= line_idx <= end:
                changed_methods.add(method_name)
                break
    
    return changed_methods


def extract_method_body(method_name: str, lines: List[str]) -> str:
    """
    특정 메소드의 본문을 추출합니다.
    한 줄에 여러 메소드가 있는 경우도 처리합니다.
    
    Args:
        method_name: 메소드명
        lines: 파일 라인 목록
        
    Returns:
        str: 메소드 본문 (정규화됨)
    """
    import re
    
    # 전체 파일 내용을 한 줄로 통합 (라인 단위 처리의 한계 극복)
    content = '\n'.join(lines)
    
    # 메소드 시그니처 찾기: method_name(...) { ... }
    # 패턴: method_name( ... ) { ... } (중괄호 레벨 0)
    pattern = rf'\b{re.escape(method_name)}\s*\([^)]*\)\s*\{{'
    
    match = re.search(pattern, content)
    if not match:
        return ""
    
    start_pos = match.end() - 1  # opening { 위치
    
    # 중괄호 매칭으로 메소드 본문 끝 찾기
    brace_level = 1
    pos = start_pos + 1
    
    while pos < len(content) and brace_level > 0:
        if content[pos] == '{':
            brace_level += 1
        elif content[pos] == '}':
            brace_level -= 1
        pos += 1
    
    if brace_level != 0:
        return ""  # 매칭 실패
    
    # 메소드 본문 추출 및 정규화
    method_body = content[start_pos:pos]
    body_normalized = re.sub(r'\s+', ' ', method_body).strip()
    
    return body_normalized


# =====================================
# 메소드명 검증 (Excel 생성 전)
# =====================================

def build_project_method_map(target_project: str) -> Set[str]:
    """
    프로젝트의 모든 Java 파일을 스캔하여 실제 존재하는 메소드를 맵핑합니다.
    
    Args:
        target_project: 타겟 프로젝트 경로
        
    Returns:
        Set[str]: 실제 존재하는 메소드 ('ClassName.methodName' 형태)
    """
    logger = logging.getLogger(__name__)
    actual_methods = set()
    
    # 프로젝트의 모든 Java 파일 찾기
    java_files = []
    for root, dirs, files in os.walk(target_project):
        dirs[:] = [d for d in dirs if d != '.applycrypto']
        
        for file in files:
            if file.endswith('.java'):
                java_files.append(os.path.join(root, file))
    
    logger.info(f"스캔할 Java 파일: {len(java_files)}개")
    
    # 각 Java 파일에서 클래스명과 메소드명 추출
    parser = JavaASTParser()
    
    for java_file in java_files:
        try:
            tree, error = parser.parse_file(Path(java_file), remove_comments=False)
            
            if error or tree is None:
                logger.debug(f"AST 파싱 실패 (skip): {java_file}")
                continue
            
            # 클래스 정보 추출
            class_infos = parser.extract_class_info(tree, Path(java_file))
            
            for class_info in class_infos:
                class_name = class_info.name
                
                for method in class_info.methods:
                    method_name = method.name
                    # ClassName.methodName 형태로 추가
                    full_method_name = f"{class_name}.{method_name}"
                    actual_methods.add(full_method_name)
        
        except Exception as e:
            logger.debug(f"메소드 맵 구축 중 오류 (skip): {java_file} - {e}")
            continue
    
    logger.info(f"실제 존재하는 메소드: {len(actual_methods)}개")
    return actual_methods


def validate_and_print_method_names(endpoint_data_list: List[EndPointData], actual_methods_map: Set[str], target_project: str):
    """
    엔드포인트의 메소드명을 검증하고 콘솔에 출력합니다.
    
    2단계 검증:
    1단계: 형식 검증 (ClassName.methodName)
    2단계: 교차검증 - AST parser 결과를 정규식으로 재검증
    
    Args:
        endpoint_data_list: EndPointData 리스트
        actual_methods_map: 프로젝트에 실제 존재하는 메소드명 집합 (AST parser로 추출)
        target_project: 타겟 프로젝트 경로 (정규식 재검증용)
    """
    logger = logging.getLogger(__name__)
    
    if not endpoint_data_list:
        logger.warning("엔드포인트 데이터가 없습니다.")
        return
    
    # 메소드명 수집
    method_names: Dict[str, int] = {}  # method_name -> count
    format_ok_count = 0  # 형식 OK
    format_ng_count = 0  # 형식 NG
    
    format_ng_methods: List[str] = []
    
    for endpoint_data in endpoint_data_list:
        method_name = endpoint_data.method_name
        
        if method_name in method_names:
            method_names[method_name] += 1
        else:
            method_names[method_name] = 1
        
        # 1단계: 형식 검증 (ClassName.methodName)
        if '.' in method_name:
            format_ok_count += 1
        else:
            format_ng_count += 1
            format_ng_methods.append(method_name)
    
    # 2단계: 교차검증 (정규식으로 AST 결과 재검증)
    cross_validate_results = cross_validate_methods_by_regex(actual_methods_map, target_project)
    
    # 결과 출력
    print(f"\n{'='*80}")
    print(f"메소드명 검증 결과")
    print(f"{'='*80}")
    print(f"총 메소드명: {len(method_names)}개 (중복 포함: {len(endpoint_data_list)}개)")
    print(f"\n[단계 1] 형식 검증 (ClassName.methodName):")
    print(f"  [OK] 올바른 형식: {format_ok_count}개")
    print(f"  [NG] 잘못된 형식: {format_ng_count}개")
    
    print(f"\n[단계 2] 교차검증 (정규식으로 AST 결과 재확인):")
    print(f"  AST parser 발견: {len(actual_methods_map)}개 메소드")
    print(f"  정규식 재검증:")
    print(f"    [OK] AST/정규식 모두 일치: {cross_validate_results['both_ok']}개")
    print(f"    [NG] AST만 발견 (정규식 미확인): {cross_validate_results['ast_only']}개")
    print(f"  신뢰도: {cross_validate_results['confidence']:.1f}%")
    
    # 형식 오류 메소드 출력
    if format_ng_count > 0:
        print(f"\n{'-'*80}")
        print(f"[NG] 형식 오류 메소드명 (점(.) 구분자 없음): {format_ng_count}개")
        print(f"{'-'*80}")
        
        format_ng_methods_unique = sorted(set(format_ng_methods))
        for idx, invalid_method in enumerate(format_ng_methods_unique, 1):
            count = method_names.get(invalid_method, 1)
            count_str = f"({count}건)" if count > 1 else ""
            print(f"  {idx:3d}. {invalid_method:<60s} {count_str}")
    
    # AST만 발견된 메소드 경고
    if cross_validate_results['ast_only'] > 0:
        print(f"\n{'-'*80}")
        print(f"[경고] 정규식 확인 실패 메소드: {cross_validate_results['ast_only']}개")
        print(f"       (AST는 발견했으나 정규식으로 재검증 안 됨)")
        print(f"{'-'*80}")
        
        for idx, method_name in enumerate(cross_validate_results['ast_only_methods'], 1):
            count = method_names.get(method_name, 1)
            count_str = f"({count}건)" if count > 1 else ""
            print(f"  {idx:3d}. {method_name:<60s} {count_str}")
    
    # 최종 요약
    print(f"\n{'='*80}")
    has_error = format_ng_count > 0 or cross_validate_results['ast_only'] > 0
    if not has_error:
        print(f"[OK] 모든 메소드가 프로젝트에 존재하고 교차검증도 통과했습니다!")
    else:
        if format_ng_count > 0:
            print(f"[NG] 형식 오류: {format_ng_count}개")
        if cross_validate_results['ast_only'] > 0:
            print(f"[경고] 교차검증 실패: {cross_validate_results['ast_only']}개")
    print(f"{'='*80}\n")
    
    # 로그 기록
    logger.info(f"메소드명 검증 완료: 총 {len(method_names)}개, 형식OK {format_ok_count}개")
    logger.info(f"교차검증: AST발견 {len(actual_methods_map)}개, 정규식도 확인 {cross_validate_results['both_ok']}개")
    if format_ng_count > 0:
        logger.warning(f"형식 오류 메소드명 {format_ng_count}개")
    if cross_validate_results['ast_only'] > 0:
        logger.warning(f"정규식 미확인 메소드 {cross_validate_results['ast_only']}개")


def cross_validate_methods_by_regex(actual_methods_map: Set[str], target_project: str) -> Dict:
    """
    정규식을 사용하여 AST parser로 추출한 메소드들을 재검증합니다.
    (교차검증: AST 결과의 신뢰도 확인)
    
    Args:
        actual_methods_map: AST parser로 추출한 메소드 집합
        target_project: 타겟 프로젝트 경로
        
    Returns:
        Dict: {
            'both_ok': AST/정규식 모두 확인된 메소드 개수,
            'ast_only': AST만 발견, 정규식 미확인된 메소드 개수,
            'ast_only_methods': AST만 발견된 메소드 목록,
            'confidence': 신뢰도 (%)
        }
    """
    logger = logging.getLogger(__name__)
    
    # 정규식으로 재검증할 메소드 맵 구축
    regex_validated_methods = set()
    
    # AST에서 발견한 메소드들을 클래스별로 그룹화
    methods_by_class: Dict[str, Set[str]] = {}  # ClassName -> {methodName, ...}
    for full_method in actual_methods_map:
        parts = full_method.split('.')
        if len(parts) == 2:
            class_name, method_name = parts
            if class_name not in methods_by_class:
                methods_by_class[class_name] = set()
            methods_by_class[class_name].add(method_name)
    
    # 프로젝트의 모든 Java 파일에서 정규식으로 재검증
    java_files = []
    for root, dirs, files in os.walk(target_project):
        dirs[:] = [d for d in dirs if d != '.applycrypto' and d != '__pycache__' and d != 'venv']
        
        for file in files:
            if file.endswith('.java'):
                java_files.append(os.path.join(root, file))
    
    logger.info(f"정규식 재검증: {len(java_files)}개 파일 스캔")
    
    # 각 Java 파일 스캔
    for java_file in java_files:
        try:
            content = read_file_safe(java_file)
            
            # 클래스명 추출 (정규식) - class와 interface 모두 지원
            class_match = re.search(r'public\s+(?:abstract\s+)?(?:class|interface)\s+(\w+)', content)
            if not class_match:
                class_match = re.search(r'(?:class|interface)\s+(\w+)', content)
            
            if not class_match:
                continue
            
            class_name = class_match.group(1)
            
            # 이 클래스에 대해 검증할 메소드 목록
            if class_name not in methods_by_class:
                continue
            
            methods_to_find = methods_by_class[class_name]
            
            # 정규식으로 메소드 찾기 (인터페이스 메소드 포함)
            # 매우 간단한 패턴: methodName 뒤에 (... 있으면 메소드로 간주
            for method_name in methods_to_find:
                # 패턴: methodName + ( : 매우 단순한 메소드 시그니처
                simple_pattern = rf'\b{re.escape(method_name)}\s*\('
                
                if re.search(simple_pattern, content):
                    full_method = f"{class_name}.{method_name}"
                    regex_validated_methods.add(full_method)
        
        except Exception as e:
            logger.debug(f"정규식 재검증 중 오류 (skip): {java_file} - {e}")
            continue
    
    # 결과 비교
    both_ok = len(regex_validated_methods)
    ast_only = len(actual_methods_map) - both_ok
    ast_only_methods = sorted(actual_methods_map - regex_validated_methods)
    
    # 신뢰도 계산
    confidence = (both_ok / len(actual_methods_map) * 100) if actual_methods_map else 0
    
    logger.info(f"교차검증 완료: AST발견 {len(actual_methods_map)}개, 정규식도 확인 {both_ok}개, 신뢰도 {confidence:.1f}%")
    
    return {
        'both_ok': both_ok,
        'ast_only': ast_only,
        'ast_only_methods': ast_only_methods,
        'confidence': confidence
    }


# =====================================
# Excel 생성
# =====================================

def create_endpoint_workbook(endpoint_data_list: List[EndPointData], output_file: str):
    """
    엔드포인트 목록 Excel 워크북을 생성합니다.
    
    Args:
        endpoint_data_list: EndPointData 리스트
        output_file: 출력 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'End Point 목록'
    
    # 열 너비 설정
    ws.column_dimensions['A'].width = COLUMNS['A']['width']
    ws.column_dimensions['B'].width = COLUMNS['B']['width']
    ws.column_dimensions['C'].width = COLUMNS['C']['width']
    ws.column_dimensions['D'].width = COLUMNS['D']['width']
    ws.column_dimensions['E'].width = COLUMNS['E']['width']
    
    # 헤더 설정
    header_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 2행: 헤더
    headers = ['', '파일 경로', '파일명', '메소드명', '엔드포인트']
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[2].height = 16
    
    # 데이터 추가
    data_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    data_alignment_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    current_file_path = None
    current_file_name = None
    current_method_name = None
    
    file_path_merge_start = 3
    file_name_merge_start = 3
    method_merge_start = 3
    
    for row_idx, data in enumerate(endpoint_data_list, 3):
        # 데이터 입력
        ws.cell(row=row_idx, column=2).value = data.file_path
        ws.cell(row=row_idx, column=3).value = data.file_name
        ws.cell(row=row_idx, column=4).value = data.method_name
        ws.cell(row=row_idx, column=5).value = data.endpoint
        
        # 스타일 적용
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            if col == 2:  # 파일 경로 - wrap_text
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = data_alignment_left
        
        ws.row_dimensions[row_idx].height = 16
        
        # 계층적 병합 처리
        # 1. 파일경로가 변경된 경우
        if data.file_path != current_file_path:
            # 이전 그룹들 병합
            if file_path_merge_start < row_idx:
                merge_rows(ws, file_path_merge_start, row_idx - 1, [2])
            if file_name_merge_start < row_idx:
                merge_rows(ws, file_name_merge_start, row_idx - 1, [3])
            if method_merge_start < row_idx:
                merge_rows(ws, method_merge_start, row_idx - 1, [4])
            
            current_file_path = data.file_path
            current_file_name = data.file_name
            current_method_name = data.method_name
            file_path_merge_start = row_idx
            file_name_merge_start = row_idx
            method_merge_start = row_idx
        
        # 2. 파일경로는 같지만 파일명이 변경된 경우
        elif data.file_name != current_file_name:
            # 이전 파일명과 메소드 그룹 병합
            if file_name_merge_start < row_idx:
                merge_rows(ws, file_name_merge_start, row_idx - 1, [3])
            if method_merge_start < row_idx:
                merge_rows(ws, method_merge_start, row_idx - 1, [4])
            
            current_file_name = data.file_name
            current_method_name = data.method_name
            file_name_merge_start = row_idx
            method_merge_start = row_idx
        
        # 3. 파일경로와 파일명은 같지만 메소드가 변경된 경우
        elif data.method_name != current_method_name:
            # 이전 메소드 그룹 병합
            if method_merge_start < row_idx:
                merge_rows(ws, method_merge_start, row_idx - 1, [4])
            
            current_method_name = data.method_name
            method_merge_start = row_idx
    
    # 마지막 그룹 병합
    if endpoint_data_list:
        last_row = 2 + len(endpoint_data_list)
        if file_path_merge_start < last_row:
            merge_rows(ws, file_path_merge_start, last_row, [2])
        if file_name_merge_start < last_row:
            merge_rows(ws, file_name_merge_start, last_row, [3])
        if method_merge_start < last_row:
            merge_rows(ws, method_merge_start, last_row, [4])
    
    # 파일 저장 (기존 파일이 있으면 삭제 후 저장)
    logger = logging.getLogger(__name__)
    actual_output_file = save_workbook_with_fallback(wb, output_file, logger)




def save_workbook_with_fallback(wb: Workbook, output_file: str, logger: logging.Logger) -> str:
    """
    Excel 워크북을 파일로 저장합니다.
    기존 파일이 잠겨있으면 타임스탬프를 추가하여 다른 파일명으로 저장합니다.
    
    Args:
        wb: Workbook 객체
        output_file: 기본 저장 경로
        logger: Logger 객체
        
    Returns:
        str: 실제 저장된 파일 경로
    """
    # Step 1: 기존 파일 삭제 시도
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
            logger.debug(f"기존 파일 삭제 완료: {output_file}")
        except PermissionError:
            logger.warning(f"기존 파일이 이미 다른 프로그램에서 열려있습니다: {output_file}")
            # Step 2: 파일이 잠겨있으면 타임스탐프 추가하여 다른 이름으로 저장
            base_path, ext = os.path.splitext(output_file)
            timestamp = datetime.now().strftime('%H%M%S')
            fallback_file = f"{base_path}_{timestamp}{ext}"
            
            try:
                wb.save(fallback_file)
                logger.info(f"파일이 이미 열려있어 다른 이름으로 저장됨: {fallback_file}")
                return fallback_file
            except Exception as e:
                logger.error(f"대체 파일 저장도 실패했습니다: {fallback_file} - {e}")
                raise
    
    # Step 3: 정상 저장
    try:
        wb.save(output_file)
        logger.debug(f"파일 저장 완료: {output_file}")
        return output_file
    except PermissionError as e:
        logger.error(f"파일 저장 권한 오류: {output_file}")
        raise


def merge_rows(ws: Worksheet, start_row: int, end_row: int, columns: List[int]):
    """
    행을 병합합니다.
    
    Args:
        ws: 워크시트
        start_row: 시작 행
        end_row: 종료 행
        columns: 병합할 열 번호 리스트
    """
    for col in columns:
        if start_row < end_row:
            col_letter = chr(64 + col)  # A=65, B=66, ...
            ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
