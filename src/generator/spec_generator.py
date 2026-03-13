"""
Java 클래스별 Excel 사양서를 AST 파싱으로 생성하는 도구입니다.
(tree-sitter 기반 정확한 메서드 추출)
"""

import re
import os
import json
from typing import Optional, Set, Dict, Any, List
import zipfile
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl import Workbook, load_workbook
from config.config_manager import Configuration
from parser.java_ast_parser import JavaASTParser
from persistence.cache_manager import CacheManager

# 공통 상수
DEFAULT_FONT = Font(name='맑은 고딕', size=10)

# 정규식 패턴 상수
TITLE_PATTERNS = [
    r"\bTITLE\b\s*[:=\-]\s*(.*)",
    r"제목\s*[:=\-]\s*(.*)",
    r"프로그램\s*개요\s*[:=\-]\s*(.*)",
    r"프로그램명\s*[:=\-]\s*(.*)",
    r"\bSUMMARY\b\s*[:=\-]\s*(.*)"
]
AUTHOR_PATTERNS = [
    r"@?author\b[:\s]*(.*)",
    r"작성자\s*[:=\-]\s*(.*)"
]
HISTORY_KEYWORDS = [r'History', r'이력', r'변경이력']
DATE_PATTERNS = [
    r"(\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})",
    r"(\d{4}년\s*\d{1,2}월\s*\d{1,2}일)",
    r"(\d{4}[.\-]\d{2})"
]

# -----------------------------
# 진입점(Entry point)
# -----------------------------

def _filter_methods_by_names(methods: List[Dict[str, Any]], method_names: Set[str]) -> List[Dict[str, Any]]:
    """메소드 이름 집합으로 메소드 리스트를 필터링합니다.
    
    Args:
        methods: 메소드 딕셔너리 리스트
        method_names: 포함할 메소드 이름 집합
        
    Returns:
        List[Dict]: 필터링된 메소드 리스트
    """
    if not method_names:
        return methods
    return [m for m in methods if m.get('name') in method_names]


def _get_changed_java_files_flexible(target_project: str, old_code_path: str) -> List[str]:
    """
    경로 구조에 관계없이 변경된 Java 파일을 찾습니다.
    
    Args:
        target_project: 타겟 프로젝트 경로
        old_code_path: 원본 프로젝트 경로
        
    Returns:
        List[str]: 변경된 Java 파일 경로 목록
    """
    import hashlib
    
    changed_files = []
    
    # 타겟 프로젝트의 모든 Java 파일 수집 (상대경로 기준)
    target_java_files = {}  # {rel_path: full_path}
    for root, dirs, files in os.walk(target_project):
        # .applycrypto 폴더는 제외
        dirs[:] = [d for d in dirs if d != '.applycrypto']
        
        for file in files:
            if file.endswith('.java'):
                full_path = os.path.join(root, file)
                # src 이후의 상대경로로 기준화
                parts = full_path.split(os.sep)
                try:
                    src_idx = parts.index('src')
                    rel_path = os.sep.join(parts[src_idx:])
                except ValueError:
                    # src가 없으면 target_project 기준의 상대경로 사용
                    rel_path = os.path.relpath(full_path, target_project)
                
                target_java_files[rel_path] = full_path
    
    # 원본 프로젝트의 모든 Java 파일 수집
    old_java_files = {}
    for root, dirs, files in os.walk(old_code_path):
        dirs[:] = [d for d in dirs if d != '.applycrypto']
        
        for file in files:
            if file.endswith('.java'):
                full_path = os.path.join(root, file)
                # src 이후의 상대경로로 기준화
                parts = full_path.split(os.sep)
                try:
                    src_idx = parts.index('src')
                    rel_path = os.sep.join(parts[src_idx:])
                except ValueError:
                    rel_path = os.path.relpath(full_path, old_code_path)
                
                old_java_files[rel_path] = full_path
    
    # 변경된 파일 비교
    for rel_path, target_path in target_java_files.items():
        if rel_path in old_java_files:
            old_path = old_java_files[rel_path]
            # 파일 내용 비교 (해시 기반)
            try:
                with open(target_path, 'rb') as f:
                    target_hash = hashlib.md5(f.read()).hexdigest()
                with open(old_path, 'rb') as f:
                    old_hash = hashlib.md5(f.read()).hexdigest()
                
                if target_hash != old_hash:
                    changed_files.append(target_path)
            except Exception:
                # 읽기 실패 시 변경된 것으로 간주
                changed_files.append(target_path)
        else:
            # 새로 생성된 파일
            changed_files.append(target_path)
    
    return changed_files


def generate_spec(config: Configuration, zip_output=False, diff_mode=False, llm_enabled=False):
    """Java 클래스들로부터 Excel 사양서를 생성합니다.

    Args:
        config (Configuration): ApplyCrypto 설정 객체
        zip_output (bool): 실행 결과를 zip으로 압축하여 생성할지 여부. 기본값은 False.
        diff_mode (bool): 변경된 메소드만 포함할지 여부. 기본값은 False.
        llm_enabled (bool): LLM 기반 메서드 요약 활성화. 기본값은 False.

    Returns:
        None
    """
    # LLM provider 초기화 (활성화된 경우)
    llm_provider = None
    if llm_enabled:
        try:
            from src.modifier.llm.llm_factory import create_llm_provider
            llm_provider = create_llm_provider(config.llm_provider)
            if llm_provider:
                print(f"INFO - LLM 프로바이더 초기화 완료: {type(llm_provider).__name__}")
        except Exception as e:
            print(f"경고: LLM 프로바이더 초기화 실패 - {e}. 메서드 요약 기능 비활성화")
    # diff_mode 검증
    if diff_mode:
        if not config.artifact_generation or not config.artifact_generation.old_code_path:
            raise ValueError(
                "diff_mode를 사용하려면 config에서 artifact_generation.old_code_path가 필요합니다"
            )
    target_project = config.target_project

    java_files = []
    first_map = {}

    # `target_project`가 제공되면 ChangedFileList_*.txt에서 읽음
    if target_project:
        java_files, first_map = read_changedFileList_excel(target_project)

    if not target_project:
        raise ValueError("target_project is required for output directory")

    # diff_mode일 때 변경된 파일만 필터링 (ChangedFileList 기준 유지)
    if diff_mode:
        old_code_path = config.artifact_generation.old_code_path
        changed_files_all = _get_changed_java_files_flexible(target_project, old_code_path)
        # ChangedFileList의 파일 중에서 실제로 변경된 파일만 필터링
        java_files = [f for f in java_files if f in changed_files_all]
        print(f'Filtered to {len(java_files)} changed file(s) in diff mode...')

    # 출력 디렉터리 설정
    out_dir = os.path.join(target_project, '.applycrypto', 'artifacts')
    os.makedirs(out_dir, exist_ok=True)

    print(f'Processing {len(java_files)} Java file(s)...')

    # JavaASTParser 초기화 (캐시 매니저 사용)
    cache_dir = Path(target_project) / '.applycrypto' / 'cache'
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_manager = CacheManager(cache_dir)
    java_parser = JavaASTParser(cache_manager=cache_manager)

    # ZIP 출력이 요청된 경우, ZIP writer를 열어 생성된 워크북을 메모리에서 바로 스트리밍합니다
    if zip_output:
        out_path = Path(out_dir)
        out_path.mkdir(parents=True, exist_ok=True)
        date_str = datetime.now().strftime('%Y%m%d')
        # 서비스명으로 명시된 target_project 이름을 우선 사용합니다; 없으면 out_dir의 부모 폴더명 또는 out_path.name을 사용합니다
        try:
            if target_project:
                service_name = Path(target_project).name
            else:
                # out_dir의 부모를 시도해봅니다 (일반 패턴: <project>/.applycrypto/spec)
                if out_path.parent and out_path.parent.name:
                    service_name = out_path.parent.name
                else:
                    service_name = out_path.name
        except Exception:
            service_name = out_path.name
        zip_name = f"{service_name}_specs_{date_str}.zip"
        zip_path = out_path / zip_name
        try:
            with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                for jf in java_files:
                    try:
                        source = read_java_file(jf)
                    except Exception:
                        continue

                    # JavaASTParser를 사용한 정확한 메서드 추출
                    tree, error = java_parser.parse_file(jf)
                    if error or not tree:
                        print(f'Warning: AST parsing failed for {jf}: {error}')
                        continue
                    
                    classes = java_parser.extract_class_info(tree, jf)
                    if not classes:
                        continue
                    
                    class_info = classes[0]  # 첫 번째 클래스 사용
                    
                    # 필요한 정보 추출
                    package = class_info.package
                    imports = extract_imports(source)
                    class_name = class_info.name
                    extends = class_info.superclass or ''
                    implements = class_info.interfaces if class_info.interfaces else []
                    annotations = class_info.annotations
                    
                    # Method 객체를 extract_methods 호환 형식으로 변환
                    methods = _convert_method_objects_to_dict(class_info.methods, source)

                    # diff_mode일 때 변경된 메소드만 필터링
                    if diff_mode:
                        try:
                            from generator.endpoint_report_generator import extract_changed_methods
                            old_code_path = config.artifact_generation.old_code_path
                            changed_methods_full = extract_changed_methods(jf, target_project, old_code_path)
                            # "ClassName.methodName" 형식에서 메소드 이름만 추출
                            changed_method_names = {m.split('.')[-1] for m in changed_methods_full}
                            original_count = len(methods)
                            methods = _filter_methods_by_names(methods, changed_method_names)
                            filtered_count = len(methods)
                            # 디버그: 필터링 결과 상세 출력
                            # print(f'    [DIFF] {class_name}: extract_changed_methods={changed_methods_full}')
                            # print(f'    [DIFF] {class_name}: changed_names={changed_method_names}')
                            # print(f'    [DIFF] {class_name}: 메소드 필터링 {original_count}→{filtered_count}')
                            if not methods:
                                continue  # 변경된 메소드가 없으면 스킵
                        except Exception as e:
                            print(f'    [ERROR] {class_name}: extract_changed_methods 실패 - {e}')
                            import traceback
                            traceback.print_exc()
                            # 에러 시 원래 메소드 리스트 사용

                    try:
                        if class_name:
                            class_body = _get_class_body(source, class_name)
                            inner_anns = re.findall(r'@\w+(?:\([^)]*\))?', class_body)
                            seen = set(annotations or [])
                            merged = list(annotations or [])
                            for ann in inner_anns:
                                if ann not in seen:
                                    seen.add(ann)
                                    merged.append(ann)
                            annotations = merged
                    except Exception:
                        pass

                    if class_name:
                        write_excel_for_class(
                            class_name,
                            jf,
                            package,
                            imports,
                            extends,
                            implements,
                            annotations,
                            methods,
                            out_dir,
                            first_changed_map=first_map,
                            zip_writer=zf,
                            llm_provider=llm_provider,
                        )
        except Exception:
            pass

        return

    else:
        # 기본 동작: 개별 .xlsx 파일을 디스크에 저장
        for jf in java_files:
            try:
                source = read_java_file(jf)
            except Exception as e:
                print(f"Could not read {jf}: {e}")
                continue

            # JavaASTParser를 사용한 정확한 메서드 추출
            try:
                tree, error = java_parser.parse_file(jf)
                if error or not tree:
                    print(f'Warning: AST parsing failed for {jf}: {error}')
                    continue
                
                classes = java_parser.extract_class_info(tree, jf)
            except RecursionError as e:
                print(f'Error: Maximum recursion depth in parsing {jf}. Skipping.')
                continue
            except Exception as e:
                print(f'Error: Exception during AST extraction for {jf}: {str(e)[:100]}. Skipping.')
                continue
            
            if not classes:
                continue

            
            class_info = classes[0]  # 첫 번째 클래스 사용
            
            # 필요한 정보 추출
            package = class_info.package
            imports = extract_imports(source)
            class_name = class_info.name
            extends = class_info.superclass or ''
            implements = class_info.interfaces if class_info.interfaces else []
            annotations = class_info.annotations
            
            # Method 객체를 extract_methods 호환 형식으로 변환
            methods = _convert_method_objects_to_dict(class_info.methods, source)

            # diff_mode일 때 변경된 메소드만 필터링
            if diff_mode:
                try:
                    from generator.endpoint_report_generator import extract_changed_methods
                    old_code_path = config.artifact_generation.old_code_path
                    changed_methods_full = extract_changed_methods(jf, target_project, old_code_path)
                    # "ClassName.methodName" 형식에서 메소드 이름만 추출
                    changed_method_names = {m.split('.')[-1] for m in changed_methods_full}
                    original_count = len(methods)
                    methods = _filter_methods_by_names(methods, changed_method_names)
                    filtered_count = len(methods)
                    # 디버그: 필터링 결과 상세 출력
                    # print(f'    [DIFF] {class_name}: extract_changed_methods={changed_methods_full}')
                    # print(f'    [DIFF] {class_name}: changed_names={changed_method_names}')
                    # print(f'    [DIFF] {class_name}: 메소드 필터링 {original_count}→{filtered_count}')
                    if not methods:
                        continue  # 변경된 메소드가 없으면 스킵
                except Exception as e:
                    print(f'    [ERROR] {class_name}: extract_changed_methods 실패 - {e}')
                    import traceback
                    traceback.print_exc()
                    # 에러 시 원래 메소드 리스트 사용

            # 클래스 내부에서 발견되는 어노테이션을 기존 어노테이션에 병합하여 보강
            try:
                if class_name:
                    class_body = _get_class_body(source, class_name)
                    inner_anns = re.findall(r'@\w+(?:\([^)]*\))?', class_body)
                # 순서와 중복 제거를 보장 (기존 어노테이션을 우선으로 유지)
                    seen = set(annotations or [])
                    merged = list(annotations or [])
                    for ann in inner_anns:
                        if ann not in seen:
                            seen.add(ann)
                            merged.append(ann)
                    annotations = merged
            except Exception:
                pass

            if class_name:
                write_excel_for_class(
                    class_name, 
                    jf, 
                    package, 
                    imports, 
                    extends, 
                    implements, 
                    annotations, 
                    methods, 
                    out_dir, 
                    first_changed_map=first_map,
                    llm_provider=llm_provider
                )


# -----------------------------
# 워크북 / 시트 생성기
# -----------------------------
def create_specification_workbook_from_scratch(first_changed_path: str = ''):
    """서식과 포맷을 갖춘 Excel 사양서 워크북을 새로 생성합니다.

    Args:
        first_changed_path: 변경 항목의 기준이 되는 첫 경로(표제부 프로젝트명 추출용)

    Returns:
        wb: 생성된 openpyxl.Workbook 인스턴스
    """
   
    wb = Workbook()
   
    # 기본 시트 제거
    wb.remove(wb.active)
   
    # 템플릿 의존 없이 시트를 생성합니다 (값 채우기는 이후에 수행)
    create_cover_sheet(wb)
    create_object_definition_sheet(wb)
    create_object_declaration_sheet(wb)
    create_method_template_sheet(wb)
   
    return wb


def _extract_project_name(path_str: str) -> str:
    """주어진 경로 문자열에서 프로젝트명(표제부 표기용)을 추출합니다.

    Args:
        path_str: 경로 문자열

    Returns:
        프로젝트명 문자열 또는 빈 문자열
    """
    if not path_str:
        return ''
    s = str(path_str).replace('\\', '/').lstrip('/')
    parts = [p for p in s.split('/') if p]
    if not parts:
        return ''
    # Windows 절대경로(C:/...) 형태 처리
    if parts[0].endswith(':'):
        if len(parts) >= 4:
            return parts[3]
        if len(parts) >= 3:
            return parts[2]
        if len(parts) >= 2:
            return parts[1]
        return parts[0]
    # 경로가 프로젝트 루트 폴더로 시작하는 경우(예: /book-ssm-new/...), 첫 번째 요소를 반환
    return parts[0]


def _sanitize_sheet_title(t, existing_titles):
    """엑셀 시트 제목으로 안전하게 사용 가능한 문자열을 생성합니다.

    Args:
        t: 원본 제목 문자열
        existing_titles: 이미 존재하는 시트명 리스트(중복 방지)

    Returns:
        엑셀 규칙에 맞게 정리된 시트 제목
    """
    # 시트 제목 정리(엑셀 제한 및 금지 문자 처리)
    if not t:
        t = 'sheet'
    # 금지 문자(\ / ? * [ ]) 및 제어 문자 제거
    t = re.sub(r'[:\\/?\*\[\]]+', '_', t)
    t = re.sub(r'[\x00-\x1f]+', '', t)
    t = t.strip()
    # 31자 초과 시 잘라냄
    if len(t) > 31:
        t = t[:31]
    base = t or 'sheet'
    candidate = base
    i = 1
    while candidate in existing_titles:
        suffix = f"_{i}"
        # 총 길이가 31자를 넘지 않도록 보정
        candidate = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    return candidate


def create_cover_sheet(wb):
    """표제부(Cover) 시트의 레이아웃과 기본 포맷을 생성합니다.

    Args:
        wb: 대상 워크북

    Returns:
        None
    """
   
    from openpyxl.styles import PatternFill, Border, Side, Font
   
    ws = wb.create_sheet("표제부")
   
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
    # 참고: openpyxl은 시트 수준의 기본 폰트를 제공하지 않으므로 셀 단위로 설정합니다
   
    # 열 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 27
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 27
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 27
   
    # 행 높이 설정: 1행=20, 6행=150, 나머지=16
    for row in range(1, 8):
        if row == 1:
            ws.row_dimensions[row].height = 20
        elif row == 6:
            ws.row_dimensions[row].height = 150
        else:
            ws.row_dimensions[row].height = 16
   
    # 병합 셀 적용
    ws.merge_cells('A1:F1')  # 1행: A1~F1 병합
    ws.merge_cells('B4:F4')  # 4행: B4~F4 병합
    ws.merge_cells('D5:F5')  # 5행: D5~F5 병합
    ws.merge_cells('B6:F6')  # 6행: B6~F6 병합
   
    # 폰트 스타일 생성
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    label_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
   
    # 정렬 스타일 생성
    left_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
   
    # 1행: 제목(배경색 + 굵은 글씨)
    ws['A1'] = "프로그램 사양서"
    ws['A1'].font = bold_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
   
    # 2행: 시스템 정보 레이블 및 플레이스홀더
    ws['A2'] = "ⓐ 시스템명"
    ws['A2'].font = bold_font
    ws['A2'].alignment = left_align
    ws['A2'].fill = label_fill
   
    ws['C2'] = "ⓑ 시스템"
    ws['C2'].font = bold_font
    ws['C2'].alignment = left_align
    ws['C2'].fill = label_fill

    # 프로젝트명 셀은 비워둡니다; `fill_cover_sheet`에서 채웁니다.
    c = ws.cell(row=2, column=6, value='')
    c.font = DEFAULT_FONT
    c.alignment = left_align
   
    ws['E2'] = "ⓒ 서브시스템"
    ws['E2'].font = bold_font
    ws['E2'].alignment = left_align
    ws['E2'].fill = label_fill
   
    # 3행: 프로그램 정보 레이블 및 플레이스홀더
    ws['A3'] = "ⓓ 프로그램 ID"
    ws['A3'].font = bold_font
    ws['A3'].alignment = left_align
    ws['A3'].fill = label_fill
   
    ws['C3'] = "ⓔ 작성자"
    ws['C3'].font = bold_font
    ws['C3'].alignment = left_align
    ws['C3'].fill = label_fill
   
    ws['E3'] = "ⓕ 작성일"
    ws['E3'].font = bold_font
    ws['E3'].alignment = left_align
    ws['E3'].fill = label_fill
   
    # 4행: 프로그램명 레이블
    ws['A4'] = "ⓖ 프로그램명"
    ws['A4'].font = bold_font
    ws['A4'].alignment = left_align
    ws['A4'].fill = label_fill
   
    # 5행: 개발 정보 레이블 및 플레이스홀더
    ws['A5'] = "ⓗ 개발 유형"
    ws['A5'].font = bold_font
    ws['A5'].alignment = left_align
    ws['A5'].fill = label_fill
   
    ws['C5'] = "ⓘ 프로그램 유형"
    ws['C5'].font = bold_font
    ws['C5'].alignment = left_align
    ws['C5'].fill = label_fill
   
    # 6행: 프로그램 개요 레이블
    ws['A6'] = "ⓙ 프로그램 개요"
    ws['A6'].font = bold_font
    ws['A6'].alignment = left_align
    ws['A6'].fill = label_fill
   
    # A1:F6 범위의 모든 셀에 테두리 적용
    for row in range(1, 7):  # 1~6행
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.border = thin_border


def create_object_definition_sheet(wb):
    """Object정의(Object Definition) 시트를 생성합니다."""
   
    from openpyxl.styles import PatternFill, Border, Side, Font
   
    ws = wb.create_sheet("Object정의")
   
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
   
    # 열 너비 설정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15   # 작업구분
   
    # 행 높이 설정: 처음 50행을 기본 16으로 설정
    for row in range(1, 50):  # 처음 50행 설정
        ws.row_dimensions[row].height = 16
   
    # 테두리 스타일 생성
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
   
    # 폰트 스타일 생성
    header_font = Font(name='맑은 고딕', size=10, bold=True)
   
    # 채우기 스타일 생성
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    subheader_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
   
    # 정렬 스타일 생성
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
   
    # 1행: 제목
    ws['A1'] = "ⓚ Object 리스트"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')
    ws['A1'].border = thin_border
   
    # 2행: 컬럼 헤더
    headers = ["순번", "Object ID", "주요 기능", "작업구분"]
    for i, header in enumerate(headers, 1):
        col_letter = chr(64 + i)  # A, B, C, D 열
        ws[f'{col_letter}2'] = header
        ws[f'{col_letter}2'].font = header_font
        ws[f'{col_letter}2'].alignment = center_align
        ws[f'{col_letter}2'].fill = subheader_fill
        ws[f'{col_letter}2'].border = thin_border


def create_object_declaration_sheet(wb):
    """Object선언(Object Declaration) 시트를 생성합니다."""
   
    from openpyxl.styles import Font, PatternFill
   
    ws = wb.create_sheet("Object선언")
   
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
   
    # 열 너비 설정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
   
    # 행 높이 설정: 처음 50행을 기본 16으로 설정
    for row in range(1, 50):  # 처음 50행 설정
        ws.row_dimensions[row].height = 16
   
    # 폰트 스타일 생성
    header_font = Font(name='맑은 고딕', size=10, bold=True)
   
    # 채우기 스타일 생성
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
   
    # 정렬 스타일 생성
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
   
    # 1행: 제목 "ⓛ Object 선언" (배경색 및 굵은 글씨)
    ws['A1'] = "ⓛ Object 선언"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')


def create_method_template_sheet(wb):
    """메서드 템플릿(예: 로그인) 시트를 새로운 구조로 생성합니다.
    
    Row 1: ■ 프로그램 구성별 상세 사양서 작성(설계자)
    Row 2-5: 설계자 항목 (Object ID, Input Parameter, Return type, 상세 Logic)
    Row 6: 공백
    Row 7: ■ 프로그램 구성별 상세 사양서 작성(개발자)
    Row 8-12: 개발자 항목 (Object ID, Description, Input Parameter, Return type, 상세 Logic)
    """
   
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
   
    ws = wb.create_sheet("login")
   
    # 열 너비 설정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
   
    # 행 높이 설정
    # 섹션 헤더(1, 7행): 20
    # 데이터 행(2,3,4,5,8,9,10,11,12): 자동 맞춤
    # 공백(6행): 16
    for row in range(1, 13):
        if row in [1, 7]:
            ws.row_dimensions[row].height = 20
        elif row == 6:
            ws.row_dimensions[row].height = 16
        else:
            ws.row_dimensions[row].height = None  # 자동 높이
   
    # 테두리 스타일 생성
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
   
    # 폰트 스타일 생성
    header_font = Font(name='맑은 고딕', size=10, bold=True)
    default_font = Font(name='맑은 고딕', size=10)
   
    # 배경색 스타일
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
   
    # 정렬 스타일 생성
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    left_align_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False)
   
    # ===== 설계자 섹션 =====
    # Row 1: ■ 프로그램 구성별 상세 사양서 작성(설계자) - A1:B1 병합
    ws['A1'] = "■ 프로그램 구성별 상세 사양서 작성(설계자)"
    ws['B1'] = ""
    ws['A1'].font = header_font
    ws['A1'].alignment = left_align_no_wrap
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['A1'].border = thin_border
    ws['B1'].border = thin_border
    ws.merge_cells('A1:B1')
   
    # Row 2: ○ Object ID
    ws['A2'] = "○ Object ID"
    ws['B2'] = ""
    ws['A2'].font = header_font
    ws['A2'].alignment = left_align
    ws['A2'].border = thin_border
    ws['B2'].font = default_font
    ws['B2'].alignment = left_align
    ws['B2'].border = thin_border
   
    # Row 3: ○ Input Parameter
    ws['A3'] = "○ Input Parameter"
    ws['B3'] = ""
    ws['A3'].font = header_font
    ws['A3'].alignment = left_align
    ws['A3'].border = thin_border
    ws['B3'].font = default_font
    ws['B3'].alignment = left_align
    ws['B3'].border = thin_border
   
    # Row 4: ○ Return type
    ws['A4'] = "○ Return type"
    ws['B4'] = ""
    ws['A4'].font = header_font
    ws['A4'].alignment = left_align
    ws['A4'].border = thin_border
    ws['B4'].font = default_font
    ws['B4'].alignment = left_align
    ws['B4'].border = thin_border
   
    # Row 5: ○ 상세 Logic
    ws['A5'] = "○ 상세 Logic"
    ws['B5'] = ""
    ws['A5'].font = header_font
    ws['A5'].alignment = left_align
    ws['A5'].border = thin_border
    ws['B5'].font = default_font
    ws['B5'].alignment = left_align
    ws['B5'].border = thin_border
   
    # Row 6: 공백
    ws['A6'] = ""
    ws['B6'] = ""
    ws['A6'].border = thin_border
    ws['B6'].border = thin_border
   
    # ===== 개발자 섹션 =====
    # Row 7: ■ 프로그램 구성별 상세 사양서 작성(개발자) - A7:B7 병합
    ws['A7'] = "■ 프로그램 구성별 상세 사양서 작성(개발자)"
    ws['B7'] = ""
    ws['A7'].font = header_font
    ws['A7'].alignment = left_align_no_wrap
    ws['A7'].fill = header_fill
    ws['B7'].fill = header_fill
    ws['A7'].border = thin_border
    ws['B7'].border = thin_border
    ws.merge_cells('A7:B7')
   
    # Row 8: ○ Object ID
    ws['A8'] = "○ Object ID"
    ws['B8'] = ""
    ws['A8'].font = header_font
    ws['A8'].alignment = left_align
    ws['A8'].border = thin_border
    ws['B8'].font = default_font
    ws['B8'].alignment = left_align
    ws['B8'].border = thin_border
   
    # Row 9: ○ Description
    ws['A9'] = "○ Description"
    ws['B9'] = ""
    ws['A9'].font = header_font
    ws['A9'].alignment = left_align
    ws['A9'].border = thin_border
    ws['B9'].font = default_font
    ws['B9'].alignment = left_align
    ws['B9'].border = thin_border
   
    # Row 10: ○ Input Parameter
    ws['A10'] = "○ Input Parameter"
    ws['B10'] = ""
    ws['A10'].font = header_font
    ws['A10'].alignment = left_align
    ws['A10'].border = thin_border
    ws['B10'].font = default_font
    ws['B10'].alignment = left_align
    ws['B10'].border = thin_border
   
    # Row 11: ○ Return type
    ws['A11'] = "○ Return type"
    ws['B11'] = ""
    ws['A11'].font = header_font
    ws['A11'].alignment = left_align
    ws['A11'].border = thin_border
    ws['B11'].font = default_font
    ws['B11'].alignment = left_align
    ws['B11'].border = thin_border
   
    # Row 12: ○ 상세 Logic
    ws['A12'] = "○ 상세 Logic"
    ws['B12'] = ""
    ws['A12'].font = header_font
    ws['A12'].alignment = left_align
    ws['A12'].border = thin_border
    ws['B12'].font = default_font
    ws['B12'].alignment = left_align
    ws['B12'].border = thin_border


def create_other_definitions_sheet(wb):
    """사용자 사양에 맞춰 '기타정의사항' 시트를 생성합니다.

    레이아웃:
    - 시트명: '기타정의사항'
    - 기본 폰트: 맑은 고딕 10
    - A열 너비: 80
    - 행: 제목/라벨 행과 플레이스홀더(N/A)
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb.create_sheet("기타정의사항")

    # 기본 폰트

    # 열 너비
    ws.column_dimensions['A'].width = 80

    # 행 높이(기본값)
    for r in range(1, 18):
        ws.row_dimensions[r].height = 20 if r in (1,) else 16

    # 스타일
    header_font = Font(name='맑은 고딕', size=10, bold=True)
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 사용자 명세에 따른 행 내용
    entries = [
        (1, "ⓝ 기타 정의 사항", True),
        (2, "① 메시지 처리", False),
        (3, "N/A", False),
        (4, "② Code 마스터 테이블 정의(CBO Configuration 테이블)", False),
        (5, "N/A", False),
        (6, "③ 권한 체크", False),
        (7, "N/A", False),
        (8, "④ 관련 사업장 (법인/사업장 특화)", False),
        (9, "N/A", False),
        (10, "⑤ 타 모듈 영향도", False),
        (11, "N/A", False),
        (12, "⑥ 테스트시, 주의사항", False),
        (13, "N/A", False),
        (14, "⑦ 운영 이관 시, 주의사항", False),
        (15, "N/A", False),
        (16, "⑧ 기타", False),
        (17, "N/A", False),
    ]

    for row, text, is_header in entries:
        cell = ws.cell(row=row, column=1, value=text)
        if is_header:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_align
        else:
            cell.font = DEFAULT_FONT
            cell.alignment = left_align
        # 셀에 테두리 적용
        try:
            cell.border = thin_border
        except Exception:
            pass

    # 시트에서 수직 중앙 정렬 보장
    for row in ws.iter_rows(min_row=1, max_row=17, min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                cell.alignment = left_align


def create_change_history_sheet(wb):
    """워크북 끝에 '변경이력' 시트를 생성하고 기본 레이아웃을 적용합니다.

    Args:
        wb: 대상 워크북

    Returns:
        None
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # 기존 시트가 있으면 제거하여 워크북 끝에 추가되도록 합니다
    try:
        if '변경이력' in wb.sheetnames:
            del wb['변경이력']
    except Exception:
        pass

    ws = wb.create_sheet('변경이력')

    # 기본 폰트
    bold_font = Font(name='맑은 고딕', size=10, bold=True)

    # 4개 컬럼에 대한 열 너비 설정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 70

    # 행 높이
    ws.row_dimensions[1].height = 16
    for r in range(2, 11):
        ws.row_dimensions[r].height = 16

    # 스타일
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 1행: A1:D1 병합
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = 'ⓞ 변경이력'
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # 2행: 헤더
    headers = ['변경일자','CR NO','프로그램 개발ID','설계자','변경내용']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = DEFAULT_FONT
        c.alignment = center_align
        c.border = thin_border

    # 최초 변경 항목: 3행에 오늘 날짜와 초기 변경 설명 입력
    try:
        date_str = datetime.now().strftime('%Y-%m-%d')
    except Exception:
        date_str = ''

    for col in range(1, 6):
        c = ws.cell(row=3, column=col, value='')
        c.font = DEFAULT_FONT
        c.alignment = left_align if col == 2 else center_align
        c.border = thin_border

    c = ws.cell(row=3, column=1, value=date_str)
    c = ws.cell(row=3, column=5, value=' AI활용 개인정보 암호화 적용')

    # 4~10행: 테두리 포함 빈 행들
    for r in range(4, 11):
        for col in range(1, 6):
            c = ws.cell(row=r, column=col, value='')
            c.font = DEFAULT_FONT
            c.alignment = left_align if col == 2 else center_align
            c.border = thin_border


# -----------------------------
# 파서 / 추출기
# -----------------------------

def read_java_file(file_path):
    """Java 파일을 다양한 인코딩으로 안전하게 읽어 텍스트를 반환합니다.

    Args:
        file_path: 읽을 Java 파일 경로

    Returns:
        파일의 전체 텍스트 문자열 또는 빈 문자열
    """
    encodings = ['utf-8', 'euc-kr', 'cp1252']
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, LookupError):
            continue
    # 최후의 수단
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()


def extract_package(source):
    """소스에서 `package` 선언을 추출합니다.

    Args:
        source: Java 소스 문자열

    Returns:
        패키지 이름 문자열(없으면 빈 문자열)
    """
    match = re.search(r'^\s*package\s+([\w.]+)\s*;', source, re.MULTILINE)
    return match.group(1) if match else ''


def extract_imports(source):
    """소스에서 모든 `import` 구문을 추출해 리스트로 반환합니다.

    Args:
        source: Java 소스 문자열

    Returns:
        import 대상 문자열 리스트
    """
    imports = []
    pattern = r'^\s*import\s+([\w.*]+)\s*;'
    for line in source.split('\n'):
        match = re.match(pattern, line)
        if match:
            imports.append(match.group(1))
    return imports


def extract_class_declaration(source):
    """클래스 선언에서 이름, extends, implements, 직전 어노테이션을 추출합니다.

    Args:
        source: Java 소스 문자열

    Returns:
        (class_name, extends, implements_list, annotations_list)
    """
    # 클래스 선언 찾기
    pattern = r'(?:public\s+)?(?:abstract\s+)?(?:final\s+)?class\s+(\w+)(?:\s+extends\s+([\w.]+))?(?:\s+implements\s+([\w\s,]+))?'
    match = re.search(pattern, source)
   
    if not match:
        return None, None, None, []
   
    class_name = match.group(1)
    extends = match.group(2) if match.group(2) else None
    implements_str = match.group(3) if match.group(3) else ''
    implements = [i.strip() for i in implements_str.split(',')] if implements_str else []
   
    # 클래스 선언 앞의 어노테이션 추출
    ann_pattern = r'@\w+(?:\([^)]*\))?'
    annotations = re.findall(ann_pattern, source[:match.start()])
    annotations = annotations[-5:] if annotations else []
   
    return class_name, extends, implements, annotations


def _get_class_body(source, class_name):
    """지정한 클래스 이름의 본문(중괄호 포함)을 반환합니다.

    Args:
        source: 전체 Java 소스 문자열
        class_name: 대상 클래스 이름

    Returns:
        클래스 본문 문자열(없으면 빈 문자열)
    """
    try:
        pat = r'(?:public\s+)?(?:abstract\s+)?(?:final\s+)?class\s+' + re.escape(class_name)
        m = re.search(pat, source)
        if not m:
            return ''
        # 매치 이후 첫 번째 '{' 위치 찾기
        pos = source.find('{', m.end())
        if pos == -1:
            return ''
        # 대응하는 닫는 중괄호 찾기
        depth = 1
        i = pos + 1
        in_string = False
        string_char = None
        while i < len(source) and depth > 0:
            ch = source[i]
            if in_string:
                if ch == '\\' and i + 1 < len(source):
                    i += 2
                    continue
                if ch == string_char:
                    in_string = False
            else:
                if ch in ('"', "'"):
                    in_string = True
                    string_char = ch
                elif ch == '{':
                    depth += 1
                elif ch == '}':
                    depth -= 1
            i += 1
        return source[pos:i]
    except Exception:
        return ''


def _convert_method_objects_to_dict(method_objects, source: str):
    """JavaASTParser의 Method 객체들을 메서드 딕셔너리 형식으로 변환합니다.
    
    AST에서 추출한 메타정보(파라미터, 어노테이션, 예외 등)를 모두 포함합니다.
    
    Args:
        method_objects: JavaASTParser에서 추출한 Method 객체 리스트
        source: 원본 Java 소스 코드 문자열
        
    Returns:
        메서드 딕셔너리 리스트 (키: 'name','return_type','comment','source',
                              'parameters','annotations','exceptions','modifiers')
    """
    if not method_objects:
        return []
        
    methods = []
    
    for method in method_objects:
        # AST에서 직접 추출한 정보 사용
        method_name = getattr(method, 'name', '')
        return_type = getattr(method, 'return_type', '')
        
        if not method_name:
            continue
        
        # 접근 제어자 정보
        access_modifier = getattr(method, 'access_modifier', 'package')
        
        # 파라미터 정보 추출
        parameters = getattr(method, 'parameters', [])
        param_list = []
        if parameters:
            for param in parameters:
                param_name = getattr(param, 'name', '')
                param_type = getattr(param, 'type', '')
                is_varargs = getattr(param, 'is_varargs', False)
                param_list.append({
                    'name': param_name,
                    'type': param_type,
                    'is_varargs': is_varargs
                })
        
        # 메서드 특성 정보
        is_static = getattr(method, 'is_static', False)
        is_final = getattr(method, 'is_final', False)
        is_abstract = getattr(method, 'is_abstract', False)
        
        modifiers = []
        if is_static:
            modifiers.append('static')
        if is_final:
            modifiers.append('final')
        if is_abstract:
            modifiers.append('abstract')
        
        # 어노테이션 정보
        annotations = getattr(method, 'annotations', [])
        
        # 예외 정보
        exceptions = getattr(method, 'exceptions', [])
        
        # 소스 코드에서 메서드의 정확한 위치를 찾아 comment와 body 추출
        # 파라미터 정보를 사용하여 정확한 메서드를 구분 (오버로딩 메서드 지원)
        
        # 1) 파라미터를 포함한 정규식으로 정확한 메서드 찾기
        if param_list:
            # 파라미터별로 타입과 이름 매칭 (어노테이션도 허용)
            params_pattern = r'\s*,\s*'.join(
                r'(?:@\w+(?:\([^)]*\))?\s+)*' + re.escape(p['type']) + r'\s+' + re.escape(p['name'])
                for p in param_list
            )
            method_sig_pattern = (
                r'(?:public|protected|private)\s+'
                r'(?:static\s+)?'
                r'(?:final\s+)?'
                r'(?:synchronized\s+)?'
                r'(?:[\w<>?,\s]+?)\s+'
                + re.escape(method_name) + r'\s*\(\s*'
                + params_pattern
                + r'\s*\)(?:\s*(?:throws\s+[\w,.\s]+)?)?\s*\{'
            )
        else:
            # 파라미터가 없는 경우: 빈 괄호로 매칭
            method_sig_pattern = (
                r'(?:public|protected|private)\s+'
                r'(?:static\s+)?'
                r'(?:final\s+)?'
                r'(?:synchronized\s+)?'
                r'(?:[\w<>?,\s]+?)\s+'
                + re.escape(method_name) + r'\s*\(\s*\)'
                r'\s*(?:throws\s+[\w,.\s]+)?'
                r'\s*\{'
            )
        
        match = re.search(method_sig_pattern, source)
        if not match:
            # fallback: 간단한 메서드명 + 파라미터명 검색
            if param_list:
                # 최소한 첫 번째 파라미터는 포함
                fallback_pattern = (
                    r'\b' + re.escape(method_name) + r'\s*\(\s*'
                    + r'(?:@[^(]*?)?' + re.escape(param_list[0]['type']) + r'\s+' + re.escape(param_list[0]['name'])
                    + r'[^)]*\s*\)'
                )
                match = re.search(fallback_pattern, source)
            else:
                # 파라미터 없음: 빈 괄호로만 매칭
                fallback_pattern = r'\b' + re.escape(method_name) + r'\s*\(\s*\)\s*\{'
                match = re.search(fallback_pattern, source)
        
        if match:
            start_pos = match.start()
            brace_start = match.end() - 1
            
            # 어노테이션 + 메서드 선언부 + 본문 함께 추출
            source_code = extract_method_with_annotations_exact_match(source, method_name, param_list)
            
            # 실패 시 본문만 추출
            if not source_code:
                source_code = extract_method_body(source, brace_start)
            
            # 메서드 위 주석 추출
            comment = extract_comment_before_method(source, start_pos)
        else:
            # 위치 기반 정보가 없으면 기본값 사용
            source_code = ''
            comment = ''
        
        # 메서드 시그니처 생성 (파라미터 정보 포함)
        param_signature = ', '.join(
            f"{p['type']} {p['name']}" 
            for p in param_list
        )
        
        methods.append({
            'name': method_name,
            'return_type': return_type,
            'comment': comment,
            'source': source_code,
            'parameters': param_list,  # 파라미터 상세 정보
            'param_signature': param_signature,  # 파라미터 시그니처 (메서드 시트용)
            'annotations': annotations,  # 어노테이션 목록
            'exceptions': exceptions,  # throws 예외 목록
            'modifiers': modifiers,  # static, final, abstract 등
            'access_modifier': access_modifier  # public, protected 등
        })
    
    return methods


def extract_methods(source):
    """[DEPRECATED] 소스에서 메서드들을 추출합니다.
    
    NOTE: 현재는 JavaASTParser 기반 추출을 권장합니다.
    이 함수는 호환성을 위해 간단한 regex 기반 추출만 수행합니다.
    정확한 메서드 추출을 위해서는 _convert_method_objects_to_dict()를 사용하세요.

    Args:
        source: Java 소스 문자열

    Returns:
        메서드 딕셔너리 리스트 (키: 'name','return_type','comment','source')
    """
    methods = []
   
    # 제외할 Java 제어 키워드 목록 (메서드명이 제어문으로 오인되지 않도록)
    keywords = {'if', 'else', 'for', 'while', 'do', 'switch', 'try', 'catch', 'finally', 'synchronized', 'return', 'throw', 'new', 'instanceof'}
   
    # 메서드 선언을 찾기 위한 정규식 패턴
    # 형식: [접근제한자] [static] [final] [synchronized] 반환타입 메서드명(파라미터) [throws ...] {
    # 그룹1: 반환타입, 그룹2: 메서드명, 그룹3: 파라미터
    method_pattern = (
        r'(?:public|protected|private)\s+'           # 접근 제한자 (필수)
        r'(?:static\s+)?'                            # static (선택)
        r'(?:final\s+)?'                             # final (선택)
        r'(?:synchronized\s+)?'                      # synchronized (선택)
        r'([\w<>?,\s]+?)\s+'                         # 반환타입 (그룹1)
        r'(\w+)\s*'                                  # 메서드명 (그룹2)
        r'\(\s*([^)]*?)\s*\)'                        # 파라미터 (그룹3)
        r'\s*(?:throws\s+[\w,.\s]+)?'                # throws (선택)
        r'\s*\{'                                     # 여는 중괄호
    )
   
    for match in re.finditer(method_pattern, source):
        return_type = match.group(1).strip()
        method_name = match.group(2)
       
        # 메서드명이 Java 키워드나 제어문이면 건너뜀
        if method_name in keywords:
            continue
       
        start_pos = match.start()
       
        # 여는 중괄호부터 닫는 중괄호까지의 메서드 소스 추출
        brace_start = match.end() - 1
        source_code = extract_method_body(source, brace_start)
       
        # 메서드 위의 주석(설명)을 추출
        comment = extract_comment_before_method(source, start_pos)
       
        methods.append({
            'name': method_name,
            'return_type': return_type,
            'comment': comment,
            'source': source_code
        })
   
    return methods


def extract_method_body(source, brace_start):
    """여는 중괄호 위치부터 대응하는 닫는 중괄호까지 메서드 본문을 추출합니다.

    Args:
        source: 전체 Java 소스 문자열
        brace_start: 본문 시작 위치(여는 중괄호 인덱스)

    Returns:
        메서드 본문 문자열(들여쓰기 정규화 포함)
    """
    if brace_start >= len(source) or source[brace_start] != '{':
        return ''
   
    depth = 1
    pos = brace_start + 1
    in_string = False
    string_char = None
   
    while pos < len(source) and depth > 0:
        ch = source[pos]
       
        # 문자열 관련 예외 처리
        if in_string:
            if ch == '\\' and pos + 1 < len(source):
                pos += 2
                continue
            elif ch == string_char:
                in_string = False
        else:
            if ch in ('"', "'"):
                in_string = True
                string_char = ch
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
       
        pos += 1
   
    method_code = source[brace_start:pos]

    # 후처리: 들여쓰기 정규화
    # 1) 탭을 4스페이스로 확장
    # 2) 비어있지 않은 라인들 중 최소 선행 공백(min_indent)을 계산
    # 3) 모든 라인에서 min_indent를 제거하여 블록을 좌측 정렬
    # 4) 남은 들여쓰기를 4스페이스 단위로 정규화(상대 중첩은 보존)
    lines = [ln.expandtabs(4) for ln in method_code.split('\n')]

    # 비어있지 않은 라인들 중 최소 들여쓰기 찾기
    non_empty = [ln for ln in lines if ln.strip()]
    if non_empty:
        min_indent = min((len(ln) - len(ln.lstrip())) for ln in non_empty)
    else:
        min_indent = 0

    if min_indent > 0:
        lines = [ln[min_indent:] if len(ln) >= min_indent else ln.lstrip() for ln in lines]

    # 첫 번째 본문 라인이 1 레벨(4스페이스)이 되도록 조정할 shift 결정
    # '{'만 있는 것이 아닌 첫 번째 비어있지 않은 라인 인덱스 찾기
    first_body_idx = 0
    for i, ln in enumerate(lines):
        if ln.strip() == '{':
            continue
        if ln.strip():
            first_body_idx = i
            break

    # 해당 첫 본문 라인의 현재 들여쓰기 레벨 계산
    if first_body_idx < len(lines) and lines[first_body_idx].strip():
        first_indent = len(lines[first_body_idx]) - len(lines[first_body_idx].lstrip())
        first_level = first_indent // 4
    else:
        first_level = 1

    # first_level을 1로 만들기 위해 필요한 shift 계산
    shift = max(0, first_level - 1)

    # 들여쓰기를 4스페이스 단위로 정규화하되 상대적 중첩은 보존
    # 그리고 전체 깊이를 줄이기 위해 shift를 적용
    normalized = []
    for ln in lines:
        stripped = ln.lstrip()
        indent = len(ln) - len(stripped)
        level = indent // 4
        new_level = max(0, level - shift)
        normalized.append(('    ' * new_level) + stripped)

    method_code = '\n'.join(normalized)
    return method_code


def extract_comment_before_method(source, method_start):
    """메서드 선언 앞 JavaDoc에서 설명 추출.

    Args:
        source: Java 소스 문자열
        method_start: 메서드 시작 인덱스

    Returns:
        설명 문자열 (없으면 빈 문자열)
    """
    lines = source[:method_start].rstrip().split('\n')
   
    if not lines:
        return ''
   
    comment_lines = []
    idx = len(lines) - 1
   
    # 빈 라인과 어노테이션(@RequestMapping, @ResponseBody 등)을 건너뜀
    while idx >= 0:
        line = lines[idx].strip()
        if not line or line.startswith('@'):
            idx -= 1
            continue
        break
   
    # JavaDoc 블록 주석(/** */)만 대상
    # 단일 라인 주석(//)은 코드일 가능성이 높으므로 처리하지 않음
    if idx >= 0 and lines[idx].strip().endswith('*/'):
        # JavaDoc 주석인지 확인(/**로 시작하는지)
        temp_idx = idx
        found_javadoc = False
       
        # 주석 라인 수집
        while temp_idx >= 0:
            line = lines[temp_idx].strip()
            comment_lines.insert(0, lines[temp_idx])
            if line.startswith('/**'):
                found_javadoc = True
                break
            elif line.startswith('/*'):
                # 일반 블록 주석(즉, JavaDoc이 아님)
                break
            temp_idx -= 1
       
        # JavaDoc(/**)이 아니면 무시
        if not found_javadoc:
            comment_lines = []
   
    if not comment_lines:
        return ''
   
    # JavaDoc 주석에서 설명 파트 추출
    full_comment = '\n'.join(comment_lines)
   
    # 주석 구분자(/** */) 제거
    cleaned = re.sub(r'/\*\*|\*/', '', full_comment)
   
    # 각 라인의 선행 '*' 문자 제거
    cleaned = re.sub(r'^\s*\*\s*', '', cleaned, flags=re.MULTILINE)
   
    # 라인 단위로 분리
    lines_cleaned = cleaned.split('\n')
   
    # 설명 추출(빈 라인이 아니고 @태그도 아니며 <pre> 내부가 아닌 첫 줄)
    description = ''
    in_pre = False
    for line in lines_cleaned:
        stripped = line.strip()
       
        # <pre> 태그 유무 확인
        if '<pre>' in stripped.lower():
            in_pre = True
        if '</pre>' in stripped.lower():
            in_pre = False
            continue
       
        # 빈 라인, @태그, <pre> 내부 내용은 건너뜀
        if not stripped or stripped.startswith('@') or in_pre:
            continue
       
        # 설명에서 HTML 태그 제거
        stripped = re.sub(r'<[^>]+>', '', stripped)
       
        if stripped:
            description = stripped
            break
   
    return description


def scan_dir_for_java(root_dir):
    """디렉터리에서 .java 파일 경로 리스트 반환.

    Args:
        root_dir: 시작 디렉터리

    Returns:
        파일 경로 리스트
    """
    p = Path(root_dir)
    return [str(pf) for pf in p.rglob('*.java')]


def read_changedFileList_excel(target_project_path):
    """타겟 프로젝트의 ChangedFileList에서 변경된 Java 파일 목록 반환.

    Args:
        target_project_path: 프로젝트 루트 경로

    Returns:
        (java_files, first_map): 파일 리스트와 매핑
    """
    target_path = Path(target_project_path)
    project_name = target_path.name  # 프로젝트명 동적 추출
   
    applycrypto_dir = target_path / '.applycrypto'
   
    if not applycrypto_dir.exists():
        print(f'Warning: .applycrypto directory not found in {target_project_path}')
        return [], {}
   
    # 우선 ChangedFileList_*.txt 파일을 찾습니다
    changedFileList_txt_files = list(applycrypto_dir.glob('ChangedFileList_*.txt'))
   
    if changedFileList_txt_files:
        # TXT 파일 사용
        txt_file = changedFileList_txt_files[0]
        if len(changedFileList_txt_files) > 1:
            print(f'Warning: Multiple ChangedFileList Text files found. Using: {txt_file.name}')

        try:
            with open(txt_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            java_files = []
            first_map = {}
            for line in lines:
                line = line.strip()
                if line and not line.startswith('#'):  # 주석은 건너뜀
                    # 잘못된 경로 보정: 선행 슬래시 제거
                    if line.startswith('/'):
                        line = line[1:]

                    # Make absolute path
                    line_path = Path(line)
                    if not line_path.is_absolute():
                        abs_path = str(target_path / line)
                    else:
                        abs_path = line

                    java_files.append(abs_path)

            print(f'Found {len(java_files)} files in Text')
            return java_files, first_map

        except Exception as e:
            print(f'Error reading Text file {txt_file}: {e}')
            return [], {}
   
    print(f'Warning: No ChangedFileList_*.txt file found in {applycrypto_dir}')
    return [], {}


def clean_comment_text(raw):
    """JavaDoc/블록 주석에서 불필요한 부분 제거.

    Args:
        raw: 원본 주석 텍스트

    Returns:
        정리된 텍스트
    """
    # Remove comment delimiters and leading '*' and @9tagulcs
    cleaned = re.sub(r'/\*\*?|\*/', '', raw)
    cleaned = re.sub(r'^\s*\*\s?', '', cleaned, flags=re.MULTILINE)
    lines = []
    for line in cleaned.split('\n'):
        s = line.strip()
        if not s or s.startswith('@') or re.match(r'^\*+$', s):  # Skip lines that are only asterisks
            continue
        # remove simple HTML tags
        s = re.sub(r'<[^>]+>', '', s)
        lines.append(s)
    return '\n'.join(lines).strip()


def extract_class_javadoc(source, class_start_pos):
    """클래스 선언 앞 JavaDoc 추출 및 요약.

    Args:
        source: Java 소스 문자열
        class_start_pos: 클래스 시작 인덱스

    Returns:
        클래스 설명 문자열
    """
    # Look backwards from class declaration for JavaDoc or other nearby comments
    before_class = source[:class_start_pos]

    # 1) 표준 JavaDoc(/** ... */) 우선 검색
    javadoc_pattern = r'/\*\*(.*?)\*/'
    matches = list(re.finditer(javadoc_pattern, before_class, re.DOTALL))
    if matches:
        last_match = matches[-1]
        comment_text = last_match.group(1)
        lines = []
        for line in comment_text.split('\n'):
            line = line.strip()
            if line.startswith('*'):
                line = line[1:].strip()
            if line and not line.startswith('@'):
                lines.append(line)

        def pick_title_from_lines(lines_list):
            """라인 목록에서 제목으로 사용할 최적의 한 줄을 선택합니다.

            Args:
                lines_list: 주석 라인 문자열 리스트

            Returns:
                제목으로 적합한 문자열(없으면 빈 문자열)
            """
            # Common explicit labels
            label_patterns = [r"\bTITLE\b\s*[:=\-]\s*(.*)", r"제목\s*[:=\-]\s*(.*)",
                              r"프로그램\s*개요\s*[:=\-]\s*(.*)", r"프로그램명\s*[:=\-]\s*(.*)",
                              r"\bSUMMARY\b\s*[:=\-]\s*(.*)"]
            for l in lines_list:
                for pat in label_patterns:
                    m = re.search(pat, l, re.IGNORECASE)
                    if m and m.group(1).strip():
                        return m.group(1).strip()

            # Look for short, descriptive lines (prefer upper-case or Korean keywords)
            for l in lines_list:
                if len(l) <= 80 and not re.match(r'^\*+$', l) and not l.startswith('@'):
                    # If line contains Korean or mixed alphabets and spaces, prefer it
                    if re.search(r'[가-힣]', l) or re.search(r'[A-Z]{2,}', l):
                        return l.strip()

            # As a final fallback, return the first non-empty line
            for l in lines_list:
                if l.strip():
                    return l.strip()
            return ''

        title = pick_title_from_lines(lines)
        # build body: remove the matched title line if present
        body_lines = []
        for l in lines:
            if title and title == l:
                continue
            # also skip lines that contain the title as suffix/prefix
            if title and title in l and len(l) - len(title) < 6:
                continue
            body_lines.append(l)

        if title:
            if body_lines:
                return title + '\n' + '\n'.join(body_lines)
            return title
        return '\n'.join(lines).strip() or ''

    # 2) 일반 블록 주석(/* ... */)으로 대체 검색
    block_pattern = r'/\*(?!\*)(.*?)\*/'
    blocks = list(re.finditer(block_pattern, before_class, re.DOTALL))
    if blocks:
        last_block = blocks[-1]
        comment_text = last_block.group(1)
        # clean similarly to javadoc
        lines = []
        for line in comment_text.split('\n'):
            s = line.strip()
            # remove leading '*' characters sometimes used
            if s.startswith('*'):
                s = s[1:].strip()
            if s and not s.startswith('@'):
                lines.append(s)

        def pick_title_from_lines_block(lines_list):
            """블록 주석 라인에서 제목을 선택하는 보조 헬퍼입니다.

            Args:
                lines_list: 블록 주석 라인 리스트

            Returns:
                제목 문자열 또는 None
            """
            # 레이블 패턴으로 제목 추출
            for line in lines_list:
                for pattern in TITLE_PATTERNS:
                    m = re.search(pattern, line, re.IGNORECASE)
                    if m and m.group(1).strip():
                        return m.group(1).strip()
            
            # 짧고 의미있는 라인 찾기 (한글 또는 대문자 포함)
            for line in lines_list:
                if len(line) <= 80 and (re.search(r'[가-힣]', line) or re.search(r'[A-Z]{2,}', line)):
                    return line.strip()
            
            # 최후의 수단: 첫 번째 라인
            return lines_list[0] if lines_list else None

        title = pick_title_from_lines_block(lines)
        body_lines = [l for l in lines if not (title and (l == title or title in l and len(l) - len(title) < 6))]
        if title:
            return title + ('\n' + '\n'.join(body_lines) if body_lines else '')
        return '\n'.join(lines).strip() or ''

    # 3) Lastly, check for consecutive line comments (// ...) in the nearby lines
    #    Allow annotations (@Service) between the comment and the class declaration.
    lines = before_class.splitlines()
    max_lookback = 20
    tail = lines[-max_lookback:] if len(lines) > max_lookback else lines
    collected = []
    seen_comment = False
    for line in reversed(tail):
        s = line.strip()
        if s.startswith('//'):
            collected.append(s[2:].strip())
            seen_comment = True
            continue
        # allow annotations and blank lines before comments
        if not seen_comment and (s.startswith('@') or s == ''):
            continue
        # if we've already collected comment lines and encounter a non-comment, stop
        if seen_comment:
            break
        # otherwise, continue searching upwards
        continue

    if collected:
        # prefer TITLE-like in collected line comments
        def pick_from_collected(col):
            """수집된 라인 주석들에서 제목 후보를 추출합니다.

            Args:
                col: 역순으로 수집된 라인 주석 리스트

            Returns:
                추출된 제목 문자열
            """
            for l in col:
                m = re.search(r"\bTITLE\b\s*[:=\-]\s*(.*)", l, re.IGNORECASE)
                if m:
                    return m.group(1).strip()
            for l in col:
                if re.search(r'[가-힣]', l) or re.search(r'[A-Z]{2,}', l):
                    return l.strip()
            return ' '.join(reversed(col)).strip()

        collected_title_or_line = pick_from_collected(list(reversed(collected))).strip()
        # if the collected title is one of the collected lines, separate body
        if collected_title_or_line:
            # make body by removing the title-like line from collected
            rev = list(reversed(collected))
            body = [l for l in rev if l != collected_title_or_line]
            if body:
                return collected_title_or_line + '\n' + '\n'.join(body)
            return collected_title_or_line
        return ''

    return ''


def parse_comment_fields(source, class_start_pos):
    """클래스 선언 직전 주석을 구조화된 필드로 파싱.

    Args:
        source: Java 소스 문자열
        class_start_pos: 클래스 선언 시작 위치

    Returns:
        dict: {'title','authors','history','dates','tags','body','raw'}
    """
    # reuse logic to find the nearest comment block (javadoc, block, or line comments)
    before_class = source[:class_start_pos]

    # helper to normalize lines
    def norm_lines(raw_text):
        """주석 블록 텍스트를 라인 단위로 정규화하여 리스트로 반환합니다.

        Args:
            raw_text: 원본 주석 블록 텍스트

        Returns:
            정규화된 라인 문자열 리스트
        """
        txt = re.sub(r'/\*\*?|\*/', '', raw_text)
        txt = re.sub(r'^\s*\*\s?', '', txt, flags=re.MULTILINE)
        lines = [l.strip() for l in txt.splitlines()]
        # keep even short lines; filter empties
        return [l for l in lines if l and not re.match(r'^\*+$', l)]

    comment_text = None
    raw_block = ''

    # 1) javadoc
    javadoc_pattern = r'/\*\*(.*?)\*/'
    m = list(re.finditer(javadoc_pattern, before_class, re.DOTALL))
    if m:
        raw_block = m[-1].group(1)
        lines = norm_lines(raw_block)
    else:
        # 2) general block
        block_pattern = r'/\*(?!\*)(.*?)\*/'
        m2 = list(re.finditer(block_pattern, before_class, re.DOTALL))
        if m2:
            raw_block = m2[-1].group(1)
            lines = norm_lines(raw_block)
        else:
            # 3) line comments (//)
            lines_all = before_class.splitlines()
            tail = lines_all[-40:] if len(lines_all) > 40 else lines_all
            collected = []
            seen = False
            for line in reversed(tail):
                s = line.strip()
                if s.startswith('//'):
                    collected.append(s[2:].strip())
                    seen = True
                    continue
                if not seen and (s.startswith('@') or s == ''):
                    continue
                if seen:
                    break
            lines = list(reversed([l for l in collected if l and not re.match(r'^\*+$', l)]))

    title = ''
    authors = []
    history = []
    dates = []
    tags = []
    body_lines = []

    # 모듈 레벨 상수 사용

    i = 0
    while i < len(lines):
        l = lines[i]
        # title
        if not title:
            for pat in TITLE_PATTERNS:
                m = re.search(pat, l, re.IGNORECASE)
                if m and m.group(1).strip():
                    title = m.group(1).strip()
                    break
            if title:
                i += 1
                continue

        # author
        for pat in AUTHOR_PATTERNS:
            m = re.search(pat, l, re.IGNORECASE)
            if m and m.group(1).strip():
                authors.append(m.group(1).strip())
        # history start
        if any(kw.lower() in l.lower() for kw in HISTORY_KEYWORDS):
            # collect following lines until blank or next label-like line
            j = i + 1
            h = []
            while j < len(lines) and lines[j].strip() and not re.search(r'^[A-Z0-9]{2,}\b', lines[j]):
                h.append(lines[j].strip())
                j += 1
            if h:
                history.extend(h)
                i = j
                continue

        # dates
        for dp in DATE_PATTERNS:
            m = re.search(dp, l)
            if m:
                dates.append(m.group(1))

        # tags (simple: lines with 'TODO' or 'FIXME')
        if re.search(r'\b(TODO|FIXME|NOTE)\b', l, re.IGNORECASE):
            tags.append(l)

        body_lines.append(l)
        i += 1

    # if no explicit title, try to infer short descriptive line
    if not title:
        for l in lines:
            if len(l) <= 100 and (re.search(r'[가-힣]', l) or re.search(r'[A-Z]{2,}', l)):
                title = l
                break

    # remove duplicate title-like lines from body_lines
    def normalize_for_compare(s):
        """비교 목적으로 문자열을 정규화합니다(알파벳/숫자/한글만 남기는 등).

        Args:
            s: 입력 문자열

        Returns:
            비교용 정규화 문자열
        """
        if not s:
            return ''
        return re.sub(r'[^0-9a-zA-Z가-힣]+', '', s).lower()

    if title:
        tn = normalize_for_compare(title)
        new_body = []
        for idx, bl in enumerate(body_lines):
            if normalize_for_compare(bl) == tn:
                continue
            # if first line starts with the title text, strip the title from start
            if idx == 0 and tn and normalize_for_compare(bl).startswith(tn):
                # 원문에서 제목(substring)을 제거 시도합니다 (최선 시도)
                stripped = re.sub(re.escape(title), '', bl, flags=re.IGNORECASE).strip(' -:—–')
                if stripped:
                    new_body.append(stripped)
                continue
            new_body.append(bl)
        body_text = '\n'.join(new_body).strip()
    else:
        body_text = '\n'.join(body_lines).strip()

    result = {
        'title': title,
        'authors': authors,
        'history': history,
        'dates': dates,
        'tags': tags,
        'body': body_text,
        'raw': raw_block.strip() if raw_block else ''
    }

    return result


# -----------------------------
# 텍스트 정리 함수(셀 안전화)
# -----------------------------

def sanitize_for_excel(s, max_len=400):
    """Excel 셀에 안전하게 넣기 위해 텍스트 정리.

    Args:
        s: 원본 문자열
        max_len: 최대 길이

    Returns:
        정리된 문자열
    """
    if not s:
        return ''
    # remove control chars
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', s)
    # remove long runs of non-word decorative chars
    s = re.sub(r'^[^\w가-힣]+', '', s)
    s = re.sub(r'[^\w가-힣]+$', '', s)
    s = re.sub(r'([=\-\*_#]{2,})', ' ', s)
    # 주로 문장 부호로 이루어진 박스형 라인들은 제거
    lines = [ln.strip() for ln in s.splitlines() if not re.match(r'^[=\-\*_\s#]{3,}$', ln.strip())]
    s = '\n'.join(lines)
    # 연속된 공백들을 하나로 축약
    s = re.sub(r'\s+', ' ', s).strip()
    # 최대 길이로 자르되 가능하면 단어 단위로 잘라냄
    if len(s) > max_len:
        s = s[:max_len]
        # 마지막 공백 위치에서 잘라내도록 시도
        idx = s.rfind(' ')
        if idx > int(max_len * 0.6):
            s = s[:idx]
    # Avoid Excel treating text as formula: prefix with apostrophe if starts with = + - @
    if s and s[0] in ('=', '+', '-', '@'):
        s = "'" + s
    # Hard cap at Excel cell limit
    if len(s) > 32767:
        s = s[:32767]
    return s


def sanitize_preserve_newlines(s, max_len=32767):
    """코드 블록용 텍스트 정리 (줄바꿈 보존).

    Args:
        s: 원본 문자열
        max_len: 최대 길이

    Returns:
        정리된 문자열 (Excel 안전화)
    """
    if not s:
        return ''
    # remove control chars except newline and tab
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', s)
    # remove long runs of decorative chars on their own lines
    lines = [ln for ln in s.splitlines() if not re.match(r'^[=\-\*_\s#]{3,}$', ln.strip())]
    s = '\n'.join(lines)
    # For Excel safety, prefix the entire cell with a single apostrophe regardless
    # This prevents Excel from treating the content as a formula or special token.
    if not s.startswith("'"):
        s = "'" + s
    # trim to max_len preserving tail
    if len(s) > max_len:
        s = s[:max_len]
    return s


def determine_program_type(class_name, file_path):
    """클래스명/경로 기반 프로그램 유형 판단.

    Args:
        class_name: 클래스 이름
        file_path: 파일 경로

    Returns:
        유형 문자열 (e.g., Controller)
    """
    lower_name = class_name.lower()
    lower_path = file_path.lower()
   
    if 'controller' in lower_name or 'controller' in lower_path:
        return 'Controller'
    elif 'serviceimpl' in lower_name or 'serviceimpl' in lower_path:
        return 'serviceImpl'
    elif 'service' in lower_name or 'service' in lower_path:
        return 'Service'
    elif 'dao' in lower_name or 'dao' in lower_path or 'repository' in lower_name:
        return 'DAO'
    elif 'mapper' in lower_name or 'mapper' in lower_path:
        return 'Mapper'
    else:
        return ''


# -----------------------------
# 시트 채움 함수들
# -----------------------------

def fill_cover_sheet(ws, class_name, file_path, source, package, first_changed_path: str = ''):
    """표제부 시트를 템플릿에 따라 채움.

    Args:
        ws: 표제부 워크시트
        class_name: 클래스 이름
        file_path: 파일 경로
        source: 클래스 소스 텍스트
        package: 패키지 이름
        first_changed_path: ChangedFileList 원본 경로
    """
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
    font_style = DEFAULT_FONT
    # 클래스/인터페이스/enum의 JavaDoc이 있으면 추출
    # allow annotations between comment and declaration
    class_pattern = r'(?:public\s+)?(?:abstract\s+)?(?:final\s+)?(?:class|interface|enum|record)\s+' + re.escape(class_name)
    match = re.search(class_pattern, source)
    class_overview = ''
    if match:
        # use structured parser to fill additional fields
        fields = parse_comment_fields(source, match.start())
        # title + body for overview
        if fields.get('title'):
            t = sanitize_for_excel(fields.get('title'))
            b = sanitize_for_excel(fields.get('body'))
            parts = []
            # Always show explicit TITLE label
            parts.append(f"TITLE : {t}")
            if b:
                parts.append(b)
            # Append HISTORY if present in parsed fields
            if fields.get('history'):
                hist_text = '; '.join([h for h in fields.get('history') if h])
                if hist_text:
                    parts.append(f"HISTORY: {hist_text}")
            class_overview = '\n'.join(parts)
        else:
            # No explicit title: use body or javadoc, and include HISTORY if present
            body_or_javadoc = fields.get('body', '') or extract_class_javadoc(source, match.start())
            class_overview = sanitize_for_excel(body_or_javadoc)
            if fields.get('history'):
                hist_text = '; '.join([h for h in fields.get('history') if h])
                if hist_text:
                    if class_overview:
                        class_overview = class_overview + '\n' + f"HISTORY: {hist_text}"
                    else:
                        class_overview = f"HISTORY: {hist_text}"

        # fill author if extracted
        if fields.get('authors'):
            authors_text = sanitize_for_excel('; '.join(fields.get('authors')))
            ws.cell(row=3, column=4, value=authors_text).font = font_style

    # Determine program type
    program_type = determine_program_type(class_name, file_path)

    # Current date
    current_date = datetime.now().strftime('%Y.%m.%d')

    # Define font style (size 10, 맑은 고딕)

    # Fill according to template structure
    # Row 1: Title (already set in template)
    # Row 2: 시스템명, 시스템, 서브시스템 (leave as template defaults or blank)
    ws.cell(row=2, column=2, value='').font = font_style  # 시스템명 - blank
    ws.cell(row=2, column=4, value='온라인').font = font_style  # 시스템 - online
    # 서브시스템: populate from first_changed_path if provided and cell empty
    try:
        existing = ws.cell(row=2, column=6).value
        if (not existing or str(existing).strip() == '') and first_changed_path:
            proj = _extract_project_name(first_changed_path)
            if proj:
                ws.cell(row=2, column=6, value=proj).font = font_style
        else:
            ws.cell(row=2, column=6).font = font_style
    except Exception:
        pass

    # Row 3: 프로그램 ID, 작성자, 작성일
    ws.cell(row=3, column=2, value=class_name).font = font_style  # 프로그램 ID
    ws.cell(row=3, column=4, value='').font = font_style  # 작성자 - blank
    ws.cell(row=3, column=6, value=current_date).font = font_style  # 작성일

    # Row 4: 프로그램명
    ws.cell(row=4, column=2, value=class_name).font = font_style  # 프로그램명

    # Row 5: 개발 유형, 프로그램 유형
    ws.cell(row=5, column=2, value='java').font = font_style  # 개발 유형
    ws.cell(row=5, column=4, value=program_type if program_type else '').font = font_style  # 프로그램 유형

    # Row 6: 프로그램 개요
    # Write overview with wrapping and adjust row height so it shows as multi-line
    cell = ws.cell(row=6, column=2, value=class_overview if class_overview else '')
    cell.font = font_style
    from openpyxl.styles import Alignment
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    # estimate required row height (approx 1 line = 16)
    def _estimate_height(text, approx_chars_per_line=60):
        """텍스트 길이에 기반한 행 높이 계산.

        Args:
            text: 입력 텍스트
            approx_chars_per_line: 행당 문자 수 근사값

        Returns:
            행 높이 (픽셀)
        """
        if not text:
            return 18
        lines = 0
        for ln in str(text).split('\n'):
            ln = ln.strip()
            if not ln:
                lines += 1
            else:
                lines += max(1, (len(ln) + approx_chars_per_line - 1) // approx_chars_per_line)
        return min(max(18, lines * 16), 400)

    try:
        ws.row_dimensions[6].height = _estimate_height(class_overview)
    except Exception:
        pass

    print(f'- 표제부 완료: {class_name} (유형: {program_type or "미정의"})')


def extract_definition_from_summary(summary_text):
    """
    LLM 생성 summary에서 "정의:" 부분만 추출합니다.
    
    Args:
        summary_text: "* 정의: ... * 처리절차: ..." 형식의 텍스트
    
    Returns:
        "정의:" 부분의 텍스트 (예: "게시판 목록 조회.") 또는 빈 문자열
    """
    if not summary_text:
        return ''
    
    # "* 정의:" 패턴 찾기
    match = re.search(r'\*\s*정의\s*:\s*([^*]+?)(?:\*|$)', summary_text, re.DOTALL)
    if match:
        definition = match.group(1).strip()
        # 줄바꿈을 공백으로 치환
        definition = re.sub(r'\s+', ' ', definition).strip()
        return definition
    
    return ''


def fill_object_definition_sheet(ws, methods, file_path=None, config=None, source=None, llm_provider=None):
    """Object정의 시트를 메서드 목록으로 채움.

    Args:
        ws: Object정의 워크시트
        methods: 메서드 리스트 (딕셔너리)
        file_path: 클래스 파일 경로 (변경/신규 판단용)
        config: 설정 객체
        source: 클래스 전체 소스 코드 (메서드 추출용)
        llm_provider: LLM 프로바이더 (요약 생성용)
    """
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
    font_style = DEFAULT_FONT
   
    # Create border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
   
    # Define font style (size 10, 맑은 고딕)
    font_style = DEFAULT_FONT
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
   
    # Fill method data starting from row 3
    for idx, method in enumerate(methods, start=1):
        row_num = idx + 2  # Start from row 3
       
        # Column A: 순번 (center aligned)
        ws.cell(row=row_num, column=1, value=idx).alignment = center_alignment
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=1).font = font_style
       
        # Column B: Object ID (method name)
        ws.cell(row=row_num, column=2, value=method['name']).alignment = left_alignment
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=2).font = font_style
       
        # Column C: 주요 기능 (LLM summary의 "정의:" 부분 추출)
        description = ''
        
        # 1) LLM으로 메서드 요약 생성 시도
        if llm_provider and source:
            try:
                method_name = method['name']
                parameters = method.get('parameters', [])
                # 메서드 소스 추출
                method_source = extract_method_with_annotations_exact_match(source, method_name, parameters)
                if not method_source:
                    method_source = method.get('source', '')
                
                if method_source and method_source.strip():
                    # LLM으로 메서드 요약 생성
                    summary = generate_method_summary(method_source, method_name, llm_provider)
                    if summary and summary.strip() and summary != "~":
                        # "정의:" 부분만 추출
                        description = extract_definition_from_summary(summary)
            except Exception as e:
                print(f"    [Object정의 LLM 요약 실패] {method.get('name', '?')}: {str(e)[:60]}", flush=True)
        
        # 2) LLM 실패시 기존 comment에서 추출
        if not description and method.get('comment'):
            comment_lines = method['comment'].strip().split('\n')
            for line in comment_lines:
                cleaned = line.strip().lstrip('/*').lstrip('*').strip()
                if cleaned and not cleaned.startswith('@'):
                    description = cleaned
                    break
        
        ws.cell(row=row_num, column=3, value=description).alignment = left_alignment
        ws.cell(row=row_num, column=3).border = thin_border
        ws.cell(row=row_num, column=3).font = font_style
       
        # Column D: 작업구분 (determine by comparing target vs original sources)
        status = ''
        try:
            # attempt to determine class name from provided file_path
            class_name = None
            if file_path:
                try:
                    class_name = Path(file_path).stem
                except Exception:
                    class_name = None

            # load config to find project roots
            if config is not None:
                target_root = config.target_project if config.target_project else None
                old_code_path = config.artifact_generation.old_code_path if config.artifact_generation else None
            else:
                target_root = None
                old_code_path = None
       
            def find_java_file(root, clsname):
                """루트 디렉터리에서 주어진 클래스 이름에 해당하는 Java 파일 경로를 찾습니다.

                Args:
                    root: 탐색 시작 디렉터리
                    clsname: 클래스 이름(확장자 없이)

                Returns:
                    찾은 파일의 절대 경로 문자열 또는 None
                """
                if not root or not clsname:
                    return None
                target_name = clsname + '.java'
                for dirpath, _, filenames in os.walk(root):
                    if target_name in filenames:
                        return os.path.join(dirpath, target_name)
                return None

            # Match typical Java method declarations only: optional modifiers, return type, method name, params, then '{'
            # This avoids matching control statements like if(...), for(...), etc.
            METHOD_SIG_RE = re.compile(
                r'^\s*(?:public|protected|private|static|final|synchronized|abstract|native|transient|strictfp)?\s*'
                r'([\w<>,\s\[\]@?&]+?)\s+'  # return type (group 1)
                r'([A-Za-z_]\w*)\s*'          # method name (group 2)
                r'\(([^)]*)\)\s*\{',
                re.MULTILINE)

            def extract_methods_from_file(path):
                """주어진 Java 파일에서 메서드 시그니처를 추출하여 간단한 맵을 생성합니다.

                Args:
                    path: Java 파일 경로

                Returns:
                    methods_map: {method_name: {'sig': signature_str}} 딕셔너리
                """
                methods_map = {}
                if not path or not os.path.exists(path):
                    return methods_map
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        src = f.read()
                        for m in METHOD_SIG_RE.finditer(src):
                            name = m.group(2)
                            params = m.group(3).strip()
                            param_types = []
                            if params:
                                for p in params.split(','):
                                    p = p.strip()
                                    if p:
                                        parts = p.split()
                                        if len(parts) > 1:
                                            param_type = ' '.join(parts[:-1])
                                        else:
                                            param_type = parts[0] if parts else ''
                                        param_types.append(param_type)
                            sig = f"{name}({','.join(param_types)})"
                            methods_map[name] = {'sig': sig}
                except Exception:
                    pass
                return methods_map

            target_java = find_java_file(target_root, class_name) if class_name else None
            orig_java = find_java_file(old_code_path, class_name) if class_name else None
            target_methods = extract_methods_from_file(target_java) if target_java else {}
            original_methods = extract_methods_from_file(orig_java) if orig_java else {}

            name = method['name']
            in_target = name in target_methods
            in_original = name in original_methods
            if target_java == orig_java:
                status = ''
            elif in_target and not in_original:
                status = '신규'
            elif in_target and in_original:
                if target_methods[name]['sig'] == original_methods[name]['sig']:
                    status = ''
                else:
                    status = '변경'
            elif in_original and not in_target:
                status = '삭제'
            else:
                status = ''

        except Exception:
            status = ''

        ws.cell(row=row_num, column=4, value=status).alignment = center_alignment
        ws.cell(row=row_num, column=4).border = thin_border
        ws.cell(row=row_num, column=4).font = font_style
   
    print(f'  Object정의 완료: {len(methods)}개 메소드')


def fill_object_declaration_sheet(ws, package, imports, annotations, extends, implements, source=None):
    """Object선언 시트를 패키지/임포트/선언 정보로 채움.

    Args:
        ws: Object선언 워크시트
        package: 패키지 이름
        imports: import 리스트
        annotations: 어노테이션 리스트
        extends: extends 클래스명
        implements: implements 리스트
        source: 전체 소스 (필드 추출용)
    """
    from copy import copy
   
    # 시트 기본 폰트 설정(크기 10, 맑은 고딕)
    font_style = DEFAULT_FONT
   
    # Clear all data rows after row 1 (keep header row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
   
    current_row = 2
   
    # 템플릿 형식을 적용하여 행 추가를 돕는 내부 헬퍼 함수
    def add_row_with_formatting(row_num, col_a_value='', col_b_value=''):
        """행을 추가하고 기본 서식을 설정합니다."""
        # Set values with font
        if col_a_value:
            ws.cell(row=row_num, column=1, value=col_a_value).font = font_style
        if col_b_value:
            ws.cell(row=row_num, column=2, value=col_b_value).font = font_style
   
    # 섹션 1: Package(패키지)
    add_row_with_formatting(current_row, col_a_value='○ Package')
    current_row += 1
   
    if package:
        add_row_with_formatting(current_row, col_b_value=f'package {package};')
    current_row += 1
   
    # 섹션 2: Import(임포트)
    add_row_with_formatting(current_row, col_a_value='○ Import')
    current_row += 1
   
    if imports:
        for imp in imports:
            # Skip commented imports
            imp_stripped = imp.strip()
            if imp_stripped.startswith('//') or imp_stripped.startswith('/*'):
                continue
           
            # Format: import xxx;
            import_line = imp if imp.startswith('import ') else f'import {imp}'
            if not import_line.endswith(';'):
                import_line += ';'
           
            add_row_with_formatting(current_row, col_b_value=import_line)
            current_row += 1
   
    # 빈 행 추가
    current_row += 1
   
    # 섹션 3: Annotations(어노테이션)
    add_row_with_formatting(current_row, col_a_value='○ Service annotation')
    current_row += 1

    # Next, scan class body for annotated field declarations (e.g., @Autowired + private ...;)
    if source:
        try:
            # Pattern: one or more annotations followed by optional whitespace/newlines and a field declaration ending with ;
            field_pattern = re.compile(r'(?:@\w+(?:\([^)]*\))?\s*)+\s*(?:private|protected|public)\s+[\w<>\[\],\s]+\s+\w+\s*;', re.MULTILINE)
            found = []
            seen = set()
            for m in field_pattern.finditer(source):
                snippet = m.group(0)
                # Normalize lines and remove empty ones
                lines = [ln for ln in snippet.splitlines() if ln.strip()]
                if not lines:
                    continue
                # Last line should be the field declaration
                decl = lines[-1].strip()
                if not decl.endswith(';'):
                    continue
                anns = [a.strip() for a in lines[:-1]] if len(lines) > 1 else []

                key = (tuple(anns), decl)
                if key in seen:
                    continue
                seen.add(key)
                found.append((anns, decl))

            for anns, decl in found:
                # write annotation lines first (each as its own row)
                for ann in anns:
                    # skip commented annotation lines
                    if ann.startswith('//') or ann.startswith('/*'):
                        continue
                    add_row_with_formatting(current_row, col_b_value=ann)
                    current_row += 1
                # then the field declaration line
                add_row_with_formatting(current_row, col_b_value=decl)
                current_row += 1
        except Exception:
            pass

    # 빈 행 추가
    current_row += 1
   
    # 섹션 4: extends 클래스
    add_row_with_formatting(current_row, col_a_value='○ extends class')
    current_row += 1
   
    if extends:
        add_row_with_formatting(current_row, col_b_value=f'extends {extends}')
        current_row += 1
   
    # 빈 행 추가
    current_row += 1
   
    # 섹션 5: implements 인터페이스
    add_row_with_formatting(current_row, col_a_value='○ implements class')
    current_row += 1
   
    if implements:
        # implements가 문자열일 수도 있고 리스트일 수도 있으므로 처리
        if isinstance(implements, str):
            # 쉼표로 구분된 문자열인 경우
            impl_list = [i.strip() for i in implements.split(',')] if implements else []
        else:
            # 리스트인 경우
            impl_list = implements if isinstance(implements, list) else []
        
        for impl in impl_list:
            if impl:  # 빈 문자열 제외
                add_row_with_formatting(current_row, col_b_value=f'implements {impl}')
                current_row += 1
   
    print(f'  Object선언 완료: Package={package}, Imports={len(imports) if imports else 0}, Annotations={len(annotations) if annotations else 0}')


def copy_worksheet(source_ws, target_wb, new_name):
    """템플릿 시트를 복사하여 포맷 유지.

    Args:
        source_ws: 원본 워크시트
        target_wb: 대상 워크북
        new_name: 새 시트 이름

    Returns:
        복사된 워크시트
    """
    from copy import copy
   
    sheet_title = _sanitize_sheet_title(new_name, set(target_wb.sheetnames))
    # 새 시트 생성
    target_ws = target_wb.create_sheet(title=sheet_title)
   
    # 열 너비(칼럼 치수) 복사
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
   
    # 행 높이(로우 치수) 복사
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height
   
    # 병합된 셀 범위 복사
    for merged_cell_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_cell_range))
   
    # 모든 서식을 포함해 셀을 복사 (필수 속성만)
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
           
            # 값 복사
            if cell.value:
                target_cell.value = cell.value
           
            # 폰트, 테두리, 채우기, 정렬 복사 (가장 중요한 속성들)
            if cell.font:
                target_cell.font = copy(cell.font)
            if cell.border:
                target_cell.border = copy(cell.border)
            if cell.fill:
                target_cell.fill = copy(cell.fill)
            if cell.alignment:
                target_cell.alignment = copy(cell.alignment)
   
    return target_ws


# ------- JavaDoc/주석 추출 함수 -------
def extract_javadoc_and_comments(method_source):
    """JavaDoc과 주석에서 핵심 기능 설명을 추출합니다.
    
    Returns:
        (javadoc_text, inline_comments) - JavaDoc 내용과 인라인 주석 리스트
    """
    javadoc_text = ""
    inline_comments = []
    
    lines = method_source.split('\n')
    in_javadoc = False
    javadoc_lines = []
    
    for line in lines:
        stripped = line.strip()
        
        # JavaDoc 시작 (/** 그리고 */ 없음)
        if '/**' in line and '*/' not in line:
            in_javadoc = True
            start_idx = line.find('/**')
            if start_idx != -1:
                after_javadoc = line[start_idx+3:].strip()
                if after_javadoc and not after_javadoc.startswith('*'):
                    javadoc_lines.append(after_javadoc)
            continue
        
        # JavaDoc 종료 (*/)
        if in_javadoc and '*/' in line:
            in_javadoc = False
            end_idx = line.find('*/')
            before_end = line[:end_idx].strip()
            if before_end and not before_end.startswith('*'):
                javadoc_lines.append(before_end)
            if javadoc_lines:
                javadoc_text = ' '.join([l.lstrip('*').strip() for l in javadoc_lines if l.strip()])
            javadoc_lines = []
            continue
        
        # JavaDoc 내부 라인
        if in_javadoc:
            cleaned = stripped.lstrip('*').strip()
            if cleaned:
                javadoc_lines.append(cleaned)
            continue
        
        # 단일 라인 JavaDoc (/** ... */)
        if '/**' in stripped and '*/' in stripped:
            content = stripped.replace('/**', '').replace('*/', '').strip()
            if content:
                javadoc_text = content
            continue
        
        # 인라인 주석 (// 뒤의 텍스트)
        if '//' in line and not stripped.startswith('//'):
            idx = line.find('//')
            comment = line[idx+2:].strip()
            if comment and len(comment) > 5:
                inline_comments.append(comment)
        elif stripped.startswith('//') and len(stripped) > 5:
            comment = stripped[2:].strip()
            if comment:
                inline_comments.append(comment)
    
    return javadoc_text, inline_comments


def generate_method_summary_from_comments(javadoc_text, inline_comments):
    """JavaDoc과 주석에서 메서드 요약을 생성합니다.
    
    Returns:
        추출된 요약 또는 None
    """
    if javadoc_text:
        sentences = javadoc_text.split('.')
        if sentences:
            first_sentence = sentences[0].strip()
            if len(first_sentence) > 150:
                first_sentence = first_sentence[:147] + "..."
            if first_sentence:
                return first_sentence
    
    if inline_comments:
        relevant_comments = [c for c in inline_comments if any(
            keyword in c.lower() for keyword in 
            ['조회', '저장', '수정', '삭제', '암호화', '복호화', '처리', '반환', '데이터', '개인정보', '페이지']
        )]
        if relevant_comments:
            return relevant_comments[0]
    
    return None


# ------- 메서드 요약 생성 함수 -------
def _sanitize_json_string(text: str) -> str:
    """JSON 파싱을 위해 제어 문자를 제거합니다.
    
    Args:
        text: 정제할 문자열
        
    Returns:
        제어 문자가 제거된 문자열
    """
    # 0x00-0x1F (제어 문자) 제거, 단 \n, \r, \t는 유지
    result = []
    for char in text:
        code = ord(char)
        # 제어 문자 제외 (탭, 줄바꿈, 캐리지 리턴은 유지)
        if code < 0x20 and code not in (0x09, 0x0A, 0x0D):  # \t, \n, \r
            # 제어 문자를 공백으로 치환
            result.append(' ')
        else:
            result.append(char)
    return ''.join(result)


def _extract_json_from_response(text: str) -> str:
    """응답에서 JSON 객체를 추출합니다.
    
    Args:
        text: 응답 텍스트
        
    Returns:
        추출된 JSON 문자열
    """
    # 마크다운 코드블록 제거
    text = re.sub(r'```(?:json)?\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    
    # JSON 객체의 시작과 끝 찾기
    start_idx = text.find('{')
    if start_idx == -1:
        return text.strip()
    
    # 끝에서부터 }를 찾기 (마지막 닫는 괄호)
    end_idx = text.rfind('}')
    if end_idx == -1 or end_idx <= start_idx:
        return text.strip()
    
    # JSON 부분만 추출
    json_str = text[start_idx:end_idx + 1]
    return _sanitize_json_string(json_str)


def generate_method_summary(method_source: str, method_name: str, llm_provider) -> str:
    """메서드 요약을 생성합니다. LLM이 있으면 우선 LLM으로, 실패시 휴리스틱으로 분석합니다."""
    if not method_source or not method_source.strip():
        return "기능 없음"
    
    # LLM이 있으면 LLM으로 시도
    if llm_provider:
        try:
            prompt = {
                'task': 'summarize_java_method',
                'instruction': """Java 메서드의 핵심 기능을 한국어로 간결하게 요약하세요.
                                
[우선순위]
1) 보안/암호화 (decrypt, encrypt, CryptoService)
2) 트랜잭션 관리 (@Transactional, DB 접근)
3) 의존성 (주입받은 Service/DAO)
4) 예외 처리 (throws 선언)
5) 데이터 접근 (조회/저장/수정)
6) 배치/반복 처리 (루프 포함)
7) 변환/가공

[응답 형식]: 오직 JSON 객체만을 응답하세요 (다른 텍스트, 마크다운, 코드블록 없음)
{"summary": "* 정의: ... * 처리절차: ... * 핵심기능: ... * 특징: ..."}

[주의사항]
- 응답은 유효한 JSON이어야 함
- 줄바꿈, 탭, 제어 문자 사용 금지
- 특수문자 @,#,$,% 제외
- summary 값은 한 줄의 문자열
""",
                'method': {'name': method_name, 'source': method_source}
            }
            
            response = llm_provider.call(
                json.dumps(prompt, ensure_ascii=False),
                max_tokens=5000,
                temperature=0.0
            )
            
            content = response.get('content', '') if isinstance(response, dict) else str(response)
            
            # 응답에서 JSON 추출 및 정제
            content = _extract_json_from_response(content)
            
            # JSON 파싱
            result = json.loads(content)
            summary = result.get('summary', '').strip() if result.get('summary') else ''
            if summary:
                # print(f"    [메서드 요약 LLM] {method_name}: {summary[:80]}", flush=True)
                return summary
        except json.JSONDecodeError as e:
            print(f"    [메서드 요약 LLM 실패] {method_name}: JSON 파싱 오류 ({str(e)[:60]}) - 휴리스틱으로 폴백", flush=True)
            return ""
        except Exception as e:
            print(f"    [메서드 요약 LLM 실패] {method_name}: {str(e)[:80]} - 휴리스틱으로 폴백", flush=True)
            return ""
    
    # LLM 미활성화 또는 실패한 경우 기본값 반환
    return ""


def format_summary_multiline(summary: str) -> str:
    """요약 텍스트를 여러 줄로 포맷팅합니다.
    
    "* 정의: ... * 처리절차: ..." 형식을 여러 줄로 변환합니다.
    예: "* 정의: ...\n* 처리절차: ...\n* 핵심기능: ...\n* 특징: ..."
    
    Args:
        summary: 한 줄의 요약 텍스트
        
    Returns:
        여러 줄로 포맷팅된 텍스트
    """
    if not summary:
        return summary
    
    # "* " 패턴 앞에 줄바꿈 추가 (첫 번째 "* "는 제외)
    # 정의, 처리절차, 핵심기능, 특징 항목들을 줄바꿈으로 분리
    result = re.sub(
        r'(?<!\n)(\s*\*\s+(?:정의|처리절차|핵심기능|특징)\s*:)',
        r'\n\1',
        summary
    )
    
    # 맨 앞의 줄바꿈 제거
    result = result.lstrip('\n')
    
    return result


# 메서드 시트 관련 유틸리티
# -----------------------------

def set_cell_value_safe(ws, cell_ref, value):
    """병합된 셀을 포함하여 안전하게 셀 값을 설정합니다.
    
    Args:
        ws: 워크시트
        cell_ref: 셀 참조 (예: 'B5')
        value: 설정할 값
    """
    try:
        # 병합된 셀인지 확인
        if cell_ref in ws.merged_cells:
            # 병합된 셀의 범위 찾기
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    ws.unmerge_cells(str(merged_range))
                    break
        # 값 설정
        ws[cell_ref].value = value
    except Exception as e:
        print(f"    [셀 쓰기 실패] {cell_ref}: {str(e)[:60]}", flush=True)


def create_method_sheet(wb, method, source, template_ws, llm_provider=None):
    """메서드별 템플릿 시트 복사 및 내용 채움.
    
    JavaASTParser 메타정보를 활용하여 메서드 시그니처, 파라미터, 
    어노테이션, 예외 정보를 자동으로 채웁니다.

    Args:
        wb: 대상 워크북
        method: 메서드 딕셔너리
        source: Java 소스
        template_ws: 템플릿 워크시트
        llm_provider: LLM 프로바이더 (옵션)
    """
    method_name = method['name']
    return_type = method['return_type']
    description = method.get('comment', '')
    
    # AST에서 추출한 메타정보 (있으면)
    parameters = method.get('parameters', [])
    annotations = method.get('annotations', [])
    exceptions = method.get('exceptions', [])
    modifiers = method.get('modifiers', [])
    param_signature = method.get('param_signature', '')
   
    # Copy template sheet with all formatting (sheet title sanitized inside function)
    ws = copy_worksheet(template_ws, wb, method_name)
   
    # Now replace the content while preserving formatting
    # 새로운 템플릿 구조:
    # Row 1: ■ 프로그램 구성별 상세 사양서 작성(설계자)
    # Row 2: ○ Object ID
    # Row 3: ○ Input Parameter
    # Row 4: ○ Return type
    # Row 5: ○ 상세 Logic
    # Row 6: 공백
    # Row 7: ■ 프로그램 구성별 상세 사양서 작성(개발자)
    # Row 8: ○ Object ID
    # Row 9: ○ Description
    # Row 10: ○ Input Parameter
    # Row 11: ○ Return type
    # Row 12: ○ 상세 Logic
   
    # ===== 설계자 섹션 (Designer) =====
    # Row 2: ○ Object ID (메서드 이름)
    set_cell_value_safe(ws, 'B2', method_name)
   
    # Row 3: ○ Input Parameter (파라미터 정보)
    param_display = param_signature if param_signature else "(없음)"
    set_cell_value_safe(ws, 'B3', sanitize_for_excel(param_display))
   
    # Row 4: ○ Return type
    set_cell_value_safe(ws, 'B4', return_type if return_type else '')
   
    # Row 5: ○ 상세 Logic (설계자) - JavaDoc 주석 내용
    # 메서드 특성과 파라미터 정보를 포함하여 구성
    detail_logic = []
    
    # 1) 메서드 시그니처
    method_sig = f"{method_name}({param_signature})"
    # detail_logic.append(f"Method: {method_sig}")
    
    # 2) 메서드 특성 (modifier 정보)
    if modifiers:
        detail_logic.append(f"Modifiers: {''.join(modifiers)}\n")
    
    # 3) 어노테이션
    # if annotations:
        # detail_logic.append(f"Annotations: {''.join(annotations)}\n")
    
    # 4) 예외
    # if exceptions:
        # detail_logic.append(f"Exceptions: throws {''.join(exceptions)}\n")
    
    # 5) JavaDoc 주석
    if description:
        # detail_logic.append(f"Description:\n{description}\n")
        detail_logic.append(f"{description}\n")
    
    # 빈 메서드인 경우 최소한 JavaDoc 주석이라도 표시
    if not detail_logic and not description:
        detail_logic.append("")
    
    set_cell_value_safe(ws, 'B5', sanitize_preserve_newlines(''.join(detail_logic) if detail_logic else ''))
    
    # ===== 개발자 섹션 (Developer) =====
    # 메서드 소스 추출 (우선순위: AST source > 정규식 추출 > 기본 추출)
    # 1. AST에서 추출한 source 먼저 시도 (가장 정확함)
    method_source = method.get('source', '')
    
    # 2. AST source가 없으면 정규식 기반 추출 시도
    if not method_source:
        method_source = extract_method_with_annotations_exact_match(source, method_name, parameters)
    
    # 3. 정규식 추출도 실패하면 기본 추출 함수 사용
    if not method_source:
        method_source = extract_method_with_annotations(source, method_name)

    # LLM 으로 메서드 요약 생성
    method_summary = ""
    if method_source and method_source.strip():
        method_summary = generate_method_summary(method_source, method_name, llm_provider)
    
    # Row 8: ○ Object ID (개발자)
    set_cell_value_safe(ws, 'B8', method_name)
    
    # Row 9: ○ Description (LLM 메서드 요약) - 여러 줄로 포맷팅
    formatted_summary = format_summary_multiline(method_summary) if method_summary else ""
    set_cell_value_safe(ws, 'B9', sanitize_preserve_newlines(formatted_summary))

    # Row 10: ○ Input Parameter (개발자)
    set_cell_value_safe(ws, 'B10', sanitize_for_excel(param_display))
   
    # Row 11: ○ Return type (개발자)
    set_cell_value_safe(ws, 'B11', return_type if return_type else '')
   
    # Row 12: ○ 상세 Logic (개발자) - Description + "\n" + 메서드 소스 코드
    combined_content = ""
    if method_summary and method_source:
        combined_content = format_summary_multiline(method_summary) + "\n\n" + method_source
    elif method_summary:
        combined_content = format_summary_multiline(method_summary)
    elif method_source:
        combined_content = method_source
    else:
        combined_content = ""
    set_cell_value_safe(ws, 'B12', sanitize_preserve_newlines(combined_content))
   
    return ws


def extract_method_with_annotations(source, method_name):
    """메서드 이름으로 선언부/본문 추출 (어노테이션 포함).
    
    패키지 프라이빗 메소드(접근제어자 없음)도 캡처합니다.

    Args:
        source: Java 소스 코드
        method_name: 메서드 이름

    Returns:
        메서드 소스 문자열
    """
    # Pattern to find method with its annotations
    # 접근제어자는 선택사항 (public, protected, private, 또는 없음)
    # 하지만 메소드는 반드시 반환타입 + 메소드명 + 괄호 + 중괄호 형식
    # 예: void sortUserList(...) { }
    method_pattern = (
        r'(?:public|protected|private)?\s*'  # 접근제어자 (선택)
        r'(?:static\s+)?(?:final\s+)?(?:synchronized\s+)?'  # 제어자들 (선택)
        r'(?:abstract\s+)?'  # abstract 추가 (선택)
        r'[\w<>?,\s]+?\s+'  # 반환타입
        + re.escape(method_name)  # 메소드명
        + r'\s*\([^)]*\)'  # 파라미터
        r'\s*(?:throws\s+[\w,.\s]+)?'  # throws (선택)
        r'\s*\{'  # 시작 브레이스
    )
   
    match = re.search(method_pattern, source)
    if not match:
        return ''
   
    method_start = match.start()
   
    # Look backwards for annotations (skip blank lines and comments)
    lines = source[:method_start].split('\n')
    annotation_start_idx = len(lines) - 1
    annotations = []
   
    # Collect annotations going backwards
    idx = len(lines) - 1
    while idx >= 0:
        line = lines[idx].strip()
       
        # Skip empty lines
        if not line:
            idx -= 1
            continue
       
        # Skip single-line comments (but don't skip annotations)
        if line.startswith('//') and not line.startswith('@'):
            idx -= 1
            continue
       
        # If we hit an annotation, collect it
        if line.startswith('@'):
            annotations.insert(0, lines[idx].rstrip())
            idx -= 1
            continue
       
        # If we hit something else (not annotation, not empty, not comment), stop
        break
   
    # 메서드 본문 추출
    brace_start = match.end() - 1
    method_body = extract_method_body(source, brace_start)
   
    # Combine annotations + method declaration + body
    result_parts = []
   
    # Add annotations (excluding commented ones) - strip leading whitespace for alignment
    for ann in annotations:
        ann_stripped = ann.strip()
        if not ann_stripped.startswith('//') and not ann_stripped.startswith('/*'):
            result_parts.append(ann_stripped)  # Use stripped version to remove leading spaces
   
    # Add method declaration and body - strip leading whitespace from method declaration
    method_declaration_and_body = source[match.start():match.end() - 1] + method_body
    method_declaration_and_body = method_declaration_and_body.lstrip()  # Remove leading whitespace
    result_parts.append(method_declaration_and_body)
   
    result = '\n'.join(result_parts)
    
    # 어노테이션 정리: @ 기호와 그 다음 라인의 단어를 한 줄로 합치기
    # 예: @ \n Override → @Override (여러 줄 어노테이션 문제 해결)
    result = re.sub(r'(@)\s*\n\s*(\w+)', r'\1\2', result)
    
    return result



def extract_method_with_annotations_exact_match(source, method_name, param_list):
    """파라미터를 포함하여 정확한 메서드 추출 (어노테이션 포함).
    
    같은 메서드명이 여러 개 있을 때 (오버로딩), 파라미터로 정확하게 구분합니다.
    패키지 프라이빗 메소드(접근제어자 없음)도 캡처합니다.
    
    Args:
        source: Java 소스 코드
        method_name: 메서드 이름
        param_list: 파라미터 리스트 [{'type': '...', 'name': '...'}, ...]
    
    Returns:
        메서드 소스 문자열 (어노테이션 + 메서드 선언 + 본문)
    """
    # 파라미터를 포함한 정규식 패턴 생성
    if param_list:
        # 여러 파라미터인 경우: 파라미터 어노테이션도 허용
        # 배열 타입(e.g. long[], String[])을 포함하여 처리
        params_patterns = []
        for p in param_list:
            # 타입이 배열인 경우도 안전하게 처리
            ptype = p['type']
            pname = p['name']
            # 배열 브래킷도 포함하여 이스케이프
            param_pattern = (
                r'(?:@\w+(?:\([^)]*\))?\s+)*'  # 파라미터 어노테이션 (선택)
                + re.escape(ptype) +  # 타입 (배열 [] 포함)
                r'\s+' +
                re.escape(pname)  # 파라미터 이름
            )
            params_patterns.append(param_pattern)
        
        params_pattern_str = r'\s*,\s*'.join(params_patterns)
        method_pattern = (
            r'(?:public|protected|private)?\s*'  # 접근제어자 (선택)
            r'(?:static\s+)?'
            r'(?:final\s+)?'
            r'(?:synchronized\s+)?'
            r'(?:abstract\s+)?'  # abstract 추가
            r'(?:[\w<>\[\],.\s]+?)\s+'  # return type (배열 [] 포함)
            + re.escape(method_name) + r'\s*\(\s*'
            + params_pattern_str
            + r'\s*\)'
            r'(?:\s+throws\s+[\w<>,.\s]+)?'  # throws 절 (선택)
            r'\s*\{'
        )
    else:
        # 파라미터 없는 경우: 빈 괄호
        method_pattern = (
            r'(?:public|protected|private)?\s*'  # 접근제어자 (선택)
            r'(?:static\s+)?'
            r'(?:final\s+)?'
            r'(?:synchronized\s+)?'
            r'(?:abstract\s+)?'  # abstract 추가
            r'(?:[\w<>\[\],.\s]+?)\s+'  # return type (배열 [] 포함)
            + re.escape(method_name) + r'\s*\(\s*\)'
            r'(?:\s+throws\s+[\w<>,.\s]+)?'  # throws 절 (선택)
            r'\s*\{'
        )
    
    match = re.search(method_pattern, source)
    if not match:
        return ''
    
    method_start = match.start()
    
    # 메서드 선언 이전에서 어노테이션 찾기
    lines = source[:method_start].split('\n')
    annotations = []
    
    # 역순으로 어노테이션 수집
    idx = len(lines) - 1
    while idx >= 0:
        line = lines[idx].strip()
        
        # 빈 줄 무시
        if not line:
            idx -= 1
            continue
        
        # 단일 라인 주석 무시
        if line.startswith('//'):
            idx -= 1
            continue
        
        # 어노테이션 수집
        if line.startswith('@'):
            annotations.insert(0, lines[idx].rstrip())
            idx -= 1
            continue
        
        # 다른 것을 만나면 중지
        break
    
    # 메서드 본문 추출
    brace_start = match.end() - 1
    method_body = extract_method_body(source, brace_start)
    
    # 어노테이션 + 메서드 선언부 + 본문 조합
    result_parts = []
    
    # 어노테이션 추가
    for ann in annotations:
        ann_stripped = ann.strip()
        if not ann_stripped.startswith('//') and not ann_stripped.startswith('/*'):
            result_parts.append(ann_stripped)
    
    # 메서드 선언부 + 본문 추가
    method_declaration_and_body = source[match.start():match.end() - 1] + method_body
    method_declaration_and_body = method_declaration_and_body.lstrip()
    result_parts.append(method_declaration_and_body)
    
    result = '\n'.join(result_parts)
    
    # 어노테이션 정리: @ 기호와 그 다음 라인의 단어를 한 줄로 합치기
    # 예: @ \n Override → @Override (여러 줄 어노테이션 문제 해결)
    result = re.sub(r'(@)\s*\n\s*(\w+)', r'\1\2', result)
    
    return result



# 쓰기(Writer) / 파일 출력
# -------------------------

def write_excel_for_class(class_name, file_path, package, imports, extends, implements, annotations, methods, out_dir, first_changed_map: dict = None, zip_writer: zipfile.ZipFile = None, llm_provider=None):
    """템플릿 사용 또는 새로 생성하여 Excel 사양서 기록."""
    # DEBUG: 메소드 개수 추적
    initial_method_count = len(methods) if methods else 0
    
    # Create artifact directory
    out_dir_path = Path(out_dir)
    out_dir_path.mkdir(parents=True, exist_ok=True)
   
    # Prepare artifact file path
    safe_name = class_name or Path(file_path).stem
    date_str = datetime.now().strftime('%Y%m%d')

    # Determine project name from file_path if possible (look for 'src' segment)
    fp = Path(file_path)
    project_name = ''
    try:
        if 'src' in fp.parts:
            idx = fp.parts.index('src')
            if idx > 0:
                project_name = fp.parts[idx - 1]
    except Exception:
        project_name = ''

    # Fallbacks if project_name still not found
    if not project_name:
        if len(fp.parents) >= 3:
            project_name = fp.parents[2].name
        else:
            project_name = fp.parent.name

    # Prepare artifact file path (per-class original behavior)
    out_file = out_dir_path / f"{safe_name}_{date_str}.xlsx"
   
    try:
        # Read original Java source for JavaDoc extraction
        source = read_java_file(file_path)
       
        # Create workbook from scratch (no template dependency)
        first_changed_value = ''
        try:
            if first_changed_map:
                first_changed_value = first_changed_map.get(str(Path(file_path)), '')
        except Exception:
            first_changed_value = ''
        if not first_changed_value:
            first_changed_value = file_path
        wb = create_specification_workbook_from_scratch(first_changed_value)
       
        # Fill 표제부 sheet
        if '표제부' in wb.sheetnames:
            ws_cover = wb['표제부']
            fill_cover_sheet(ws_cover, class_name, file_path, source, package, first_changed_value)
       
        # Fill Object정의 sheet
        if 'Object정의' in wb.sheetnames:
            ws_obj_def = wb['Object정의']
            # DEBUG: fill_object_definition_sheet 직전 메소드 개수
            before_fill = len(methods) if methods else 0
            fill_object_definition_sheet(ws_obj_def, methods, file_path=file_path, config=Configuration, source=source, llm_provider=llm_provider)
            after_fill = len(methods) if methods else 0
            if before_fill != after_fill:
                print(f'    [DEBUG] {class_name}: fill_object_definition_sheet 후 메소드 개수 변화 {before_fill}→{after_fill}')
       
        # Fill Object선언 sheet
        if 'Object선언' in wb.sheetnames:
            ws_obj_decl = wb['Object선언']
            fill_object_declaration_sheet(ws_obj_decl, package, imports, annotations, extends, implements, source)
       
        # Create method sheets for each method using template
        if 'login' in wb.sheetnames:
            template_method_ws = wb['login']
            # Remove template example method sheets
            template_method_sheets = ['login', 'getUserInfo']
            for sheet_name in template_method_sheets:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
            
            # DEBUG: 메소드 시트 생성 직전 메소드 개수 확인
            final_method_count = len(methods) if methods else 0
            if initial_method_count != final_method_count:
                print(f'    [DEBUG] {class_name}: 메소드 개수 변화 {initial_method_count}→{final_method_count}')
           
            for method in methods:
                create_method_sheet(wb, method, source, template_method_ws, llm_provider)
            print(f'  메소드 시트 생성 완료: {len(methods)}개')
            # Ensure '기타정의사항' is created after all method sheets so it appears last
            try:
                # Remove existing sheet if present to avoid duplicates
                if '기타정의사항' in wb.sheetnames:
                    del wb['기타정의사항']
            except Exception:
                pass
            try:
                create_other_definitions_sheet(wb)
            except Exception:
                pass
            try:
                populate_other_definitions_sheet(wb, source, class_name, file_path)
            except Exception:
                pass
            try:
                create_change_history_sheet(wb)
            except Exception:
                pass
        else:
            print(f'Warning: Cannot create method sheets - template sheet (login) not found')
       
            # Populate '기타정의사항' sheet based on source analysis
            if '기타정의사항' in wb.sheetnames:
                try:
                    populate_other_definitions_sheet(wb, source, class_name, file_path)
                except Exception:
                    pass
       
        # Remove any stray default sheet 'Sheet1' if present (prevent accidental extra sheet)
        try:
            if 'Sheet1' in wb.sheetnames:
                del wb['Sheet1']
        except Exception:
            pass

        # Save the workbook: either write into provided ZipFile (in-memory) or to disk
        if zip_writer is not None:
            try:
                from io import BytesIO
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                arcname = f"{safe_name}_{date_str}.xlsx"
                try:
                    zip_writer.writestr(arcname, buf.getvalue())
                except Exception as e:
                    # ZIP 쓰기 실패는 로그로 남기되 계속 진행
                    print(f'- ZIP 쓰기 실패: {e}', flush=True)
                # 출력 파일 정보는 항상 출력 (zip_writer.filename은 str로 변환)
                try:
                    zfn = getattr(zip_writer, 'filename', None)
                    print(f'- 출력 파일: {str(zfn)}', flush=True)
                except Exception:
                    print(f'- 출력 파일: (unknown)', flush=True)
            except Exception as e:
                print(f'- ZIP 처리 실패: {e}', flush=True)
        else:
            try:
                wb.save(out_file)
                print(f'- 출력 파일: {out_file}', flush=True)
            except Exception as e:
                print(f'저장 실패 {out_file}: {e}', flush=True)
       
    except Exception as e:
        print(f'저장 실패 {out_file}: {e}', flush=True)
        import traceback
        traceback.print_exc()


def populate_other_definitions_sheet(wb, source: str, class_name: str, file_path: str):
    """Java 소스 코드를 간단한 휴리스틱으로 분석하여 '기타정의사항' 시트를 채웁니다.

    각 섹션마다 관련 키워드 또는 패턴을 찾아 요약을 작성합니다. 찾지 못하면 'N/A'로 남깁니다.
    보수적인 처리를 사용하여 짧고 사람이 읽기 좋은 요약을 반환합니다.
    """
    import re

    ws = wb['기타정의사항']
    if not source:
        return

    def first_join(matches):
        """리스트 항목을 중복 제거 후 정렬하여 쉼표로 결합한 문자열을 반환합니다.

        Args:
            matches: 문자열 리스트

        Returns:
            결합된 문자열 또는 빈 문자열
        """
        s = sorted(set(matches))
        return ', '.join(s) if s else ''

    # ① 메시지 처리
    msg_patterns = [r'JmsTemplate', r'KafkaTemplate', r'@KafkaListener', r'@JmsListener', r'sendMessage\(', r'publish\(', r'Producer', r'Consumer']
    msgs = []
    for p in msg_patterns:
        if re.search(p, source, re.IGNORECASE):
            msgs.append(p.replace('\\', ''))
    msg_text = first_join(msgs) or 'N/A'
    ws.cell(row=3, column=1, value=msg_text)

    # ② Code 마스터 테이블 정의 (CBO)
    cbo_patterns = [r'\bCBO\b', r'CodeMaster', r'codeMaster', r'Configuration', r'config', r'CODE_']
    cbos = [m.group(0) for pat in cbo_patterns for m in re.finditer(pat, source, re.IGNORECASE)]
    cbo_text = first_join(cbos) or 'N/A'
    ws.cell(row=5, column=1, value=cbo_text)

    # ③ 권한 체크
    auth_patterns = [r'@PreAuthorize', r'@Secured', r'hasRole\(', r'ROLE_\w+', r'SecurityContextHolder', r'isAuthenticated\(']
    auths = [m.group(0) for pat in auth_patterns for m in re.finditer(pat, source, re.IGNORECASE)]
    auth_text = first_join(auths) or 'N/A'
    ws.cell(row=7, column=1, value=auth_text)

    # ④ 관련 사업장
    biz_patterns = [r'corp', r'company', r'branch', r'site', r'office', r'business']
    bizs = []
    for p in biz_patterns:
        if re.search(p, source, re.IGNORECASE):
            bizs.append(p)
    biz_text = first_join(bizs) or 'N/A'
    ws.cell(row=9, column=1, value=biz_text)

    # ⑤ 타 모듈 영향도 - 외부 import 기반 간단 추정
    try:
        imports = extract_imports(source)
    except Exception:
        imports = []
    base_pkg = extract_package(source) or ''
    external = []
    for imp in imports:
        root = imp.split('.')[0]
        if base_pkg and not imp.startswith(base_pkg):
            external.append(root)
    impact_text = first_join(external) or 'N/A'
    ws.cell(row=11, column=1, value=impact_text)

    # ⑥ 테스트시 주의사항
    test_patterns = [r'TODO', r'FIXME', r'테스트', r'주의', r'@Transactional', r'transaction', r'concurrency', r'timeout', r'rollback']
    tests = [m.group(0) for pat in test_patterns for m in re.finditer(pat, source, re.IGNORECASE)]
    test_text = first_join(tests) or 'N/A'
    ws.cell(row=13, column=1, value=test_text)

    # ⑦ 운영 이관 시, 주의사항
    migr_patterns = [r'migration', r'migrate', r'db migrate', r'schema', r'ALTER\s+TABLE', r'마이그레이션']
    migrs = [m.group(0) for pat in migr_patterns for m in re.finditer(pat, source, re.IGNORECASE)]
    migr_text = first_join(migrs) or 'N/A'
    ws.cell(row=15, column=1, value=migr_text)

    # ⑧ 기타 - collect TODO/FIXME/comments
    other_matches = re.findall(r'//.*(TODO|FIXME|주의)|/\*.*?(TODO|FIXME|주의).*?\*/', source, re.IGNORECASE | re.DOTALL)
    others = [x for tup in other_matches for x in tup if x]
    other_text = first_join(others) or 'N/A'
    ws.cell(row=17, column=1, value=other_text)

    # Apply font and alignment to filled cells
    font = DEFAULT_FONT
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for r in (3,5,7,9,11,13,15,17):
        c = ws.cell(row=r, column=1)
        if c.value is None:
            c.value = 'N/A'
        c.font = font
        c.alignment = align
